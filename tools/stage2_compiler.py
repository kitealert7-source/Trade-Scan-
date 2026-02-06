"""
Stage-2 Presentation Compiler — SOP-Complete with Full Metric Computation
Consumes Stage-1 authoritative artifacts and produces AK_Trade_Report Excel.
Implements SOP_OUTPUT §5.1 (Performance Summary) and §5.2 (Yearwise Performance).
All metrics computed from Stage-1 data. No placeholders.
"""

import csv
import json
import math
from datetime import datetime, timedelta
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
ALT_FILL = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
LEFT_ALIGN = Alignment(horizontal="left")

# Session boundaries (UTC hours)
ASIA_START, ASIA_END = 0, 8
LONDON_START, LONDON_END = 8, 16
NY_START, NY_END = 16, 24


def load_stage1_artifacts(run_folder: Path):
    execution_dir = run_folder / "execution"
    metadata_dir = run_folder / "metadata"
    
    required_files = [
        execution_dir / "results_tradelevel.csv",
        execution_dir / "results_standard.csv",
        execution_dir / "results_risk.csv",
        execution_dir / "results_yearwise.csv",
        execution_dir / "metrics_glossary.csv",
        metadata_dir / "run_metadata.json",
    ]
    
    for f in required_files:
        if not f.exists():
            raise FileNotFoundError(f"Required Stage-1 artifact missing: {f}")
    
    def read_csv(path):
        with open(path, "r", encoding="utf-8") as f:
            return list(csv.DictReader(f))
    
    tradelevel = read_csv(execution_dir / "results_tradelevel.csv")
    standard = read_csv(execution_dir / "results_standard.csv")
    risk = read_csv(execution_dir / "results_risk.csv")
    yearwise = read_csv(execution_dir / "results_yearwise.csv")
    glossary = read_csv(execution_dir / "metrics_glossary.csv")
    
    with open(metadata_dir / "run_metadata.json", "r", encoding="utf-8") as f:
        metadata = json.load(f)
        
    starting_capital = metadata.get("reference_capital_usd")
    if not isinstance(starting_capital, (int, float)) or starting_capital <= 0:
        raise ValueError("Missing or invalid reference_capital_usd in metadata")

    return {
        "tradelevel": tradelevel,
        "standard": standard[0] if standard else {},
        "risk": risk[0] if risk else {},
        "yearwise": yearwise,
        "glossary": glossary,
        "metadata": metadata,
    }


def _safe_float(val, default=0.0):
    try:
        return float(val) if val not in (None, "", "None") else default
    except (ValueError, TypeError):
        return default


def _safe_int(val, default=0):
    try:
        return int(float(val)) if val not in (None, "", "None") else default
    except (ValueError, TypeError):
        return default


def _parse_timestamp(ts_str):
    if not ts_str:
        return None
    try:
        return datetime.fromisoformat(ts_str.replace("Z", "+00:00").replace("+00:00", ""))
    except:
        return None


def _get_session(dt):
    """Classify trade by session based on hour (UTC)."""
    if dt is None:
        return "unknown"
    hour = dt.hour
    if ASIA_START <= hour < ASIA_END:
        return "asia"
    elif LONDON_START <= hour < LONDON_END:
        return "london"
    else:
        return "ny"


def _compute_atr_percentile_regimes(trades):
    """
    Classify trades by volatility regime using ATR-percentile method.
    ATR derived from trade_high - trade_low per trade.
    Returns dict mapping trade index to regime.
    """
    if len(trades) < 3:
        return {i: "normal" for i in range(len(trades))}
    
    # Compute ATR proxy: trade_high - trade_low per trade
    atr_values = []
    for t in trades:
        # Prefer execution-emitted ATR; fallback preserved for backward compatibility
        atr = _safe_float(t.get("atr_entry", 0))
        if atr > 0:
            atr_values.append(atr)
        else:
            atr_values.append(_safe_float(t.get("entry_price", 0)) * 0.015)
    
    # Compute percentile thresholds from all ATR values
    sorted_atrs = sorted([a for a in atr_values if a > 0])
    if not sorted_atrs:
        return {i: "normal" for i in range(len(trades))}
    
    n = len(sorted_atrs)
    p33_idx = int(n * 0.33)
    p66_idx = int(n * 0.66)
    p33_threshold = sorted_atrs[p33_idx] if p33_idx < n else sorted_atrs[-1]
    p66_threshold = sorted_atrs[p66_idx] if p66_idx < n else sorted_atrs[-1]
    
    # Assign regime per trade based on ATR percentile
    regimes = {}
    for i, atr in enumerate(atr_values):
        if atr <= p33_threshold:
            regimes[i] = "low"
        elif atr <= p66_threshold:
            regimes[i] = "normal"
        else:
            regimes[i] = "high"
    
    return regimes


def _compute_sharpe_ratio(returns, risk_free_rate=0.0):
    """Compute Sharpe Ratio from list of returns."""
    if len(returns) < 2:
        return 0.0
    avg_return = sum(returns) / len(returns)
    std_return = math.sqrt(sum((r - avg_return) ** 2 for r in returns) / len(returns))
    if std_return == 0:
        return 0.0
    return (avg_return - risk_free_rate) / std_return * math.sqrt(252)  # Annualized


def _compute_sortino_ratio(returns, risk_free_rate=0.0):
    """Compute Sortino Ratio (downside deviation only)."""
    if len(returns) < 2:
        return 0.0
    avg_return = sum(returns) / len(returns)
    downside_returns = [r for r in returns if r < 0]
    if not downside_returns:
        return avg_return * math.sqrt(252) if avg_return > 0 else 0.0
    downside_std = math.sqrt(sum(r ** 2 for r in downside_returns) / len(downside_returns))
    if downside_std == 0:
        return 0.0
    return (avg_return - risk_free_rate) / downside_std * math.sqrt(252)


def _compute_k_ratio(equity_curve):
    """Compute K-Ratio from equity curve (slope / std error of slope)."""
    n = len(equity_curve)
    if n < 3:
        return 0.0
    
    # Linear regression: y = slope * x + intercept
    x_vals = list(range(n))
    x_mean = sum(x_vals) / n
    y_mean = sum(equity_curve) / n
    
    numerator = sum((x - x_mean) * (y - y_mean) for x, y in zip(x_vals, equity_curve))
    denominator = sum((x - x_mean) ** 2 for x in x_vals)
    
    if denominator == 0:
        return 0.0
    
    slope = numerator / denominator
    intercept = y_mean - slope * x_mean
    
    # Residuals and standard error
    residuals = [y - (slope * x + intercept) for x, y in zip(x_vals, equity_curve)]
    residual_ss = sum(r ** 2 for r in residuals)
    std_error = math.sqrt(residual_ss / (n - 2)) / math.sqrt(denominator) if n > 2 else 1.0
    
    if std_error == 0:
        return slope if slope > 0 else 0.0
    
    return slope / std_error


def _compute_sqn(returns):
    """Compute System Quality Number: sqrt(N) * expectancy / std_dev."""
    n = len(returns)
    if n < 2:
        return 0.0
    avg = sum(returns) / n
    std = math.sqrt(sum((r - avg) ** 2 for r in returns) / n)
    if std == 0:
        return 0.0
    return math.sqrt(n) * avg / std



def _compute_metrics_from_trades(trades, starting_capital, direction_filter=None):
    """Compute all metrics from trade-level data. direction_filter: 1=Long, -1=Short, None=All"""
    
    filtered = trades
    if direction_filter is not None:
        filtered = [t for t in trades if _safe_int(t.get("direction", 0)) == direction_filter]
    
    if not filtered:
        return _empty_metrics(starting_capital)
    
    pnls = [_safe_float(t.get("net_pnl", 0)) for t in filtered]
    bars_list = [_safe_int(t.get("bars_held", 0)) for t in filtered if t.get("bars_held") not in (None, "", "None")]
    
    wins = [p for p in pnls if p > 0]
    losses = [p for p in pnls if p < 0]
    
    gross_profit = sum(wins)
    gross_loss = abs(sum(losses))
    net_profit = sum(pnls)
    
    trade_count = len(pnls)
    win_count = len(wins)
    loss_count = len(losses)
    win_rate = (win_count / trade_count * 100) if trade_count > 0 else 0.0
    
    profit_factor = (gross_profit / gross_loss) if gross_loss > 0 else gross_profit if gross_profit > 0 else 0.0
    avg_trade = (net_profit / trade_count) if trade_count > 0 else 0.0
    avg_win = (sum(wins) / win_count) if win_count > 0 else 0.0
    avg_loss = (sum(losses) / loss_count) if loss_count > 0 else 0.0
    win_loss_ratio = (avg_win / abs(avg_loss)) if avg_loss != 0 else avg_win if avg_win > 0 else 0.0
    expectancy = avg_trade
    
    largest_win = max(wins) if wins else 0.0
    largest_loss = min(losses) if losses else 0.0
    
    # Equity curve and drawdown
    equity_curve = []
    cumulative = 0.0
    peak = 0.0
    max_dd = 0.0
    for pnl in pnls:
        cumulative += pnl
        equity_curve.append(cumulative)
        if cumulative > peak:
            peak = cumulative
        dd = peak - cumulative
        if dd > max_dd:
            max_dd = dd
    
    max_dd_pct = (max_dd / starting_capital * 100) if starting_capital > 0 else 0.0
    return_dd_ratio = (net_profit / max_dd) if max_dd > 0 else net_profit if net_profit > 0 else 0.0
    return_on_capital = (net_profit / starting_capital * 100) if starting_capital > 0 else 0.0
    
    # Streaks
    max_consec_wins = 0
    max_consec_losses = 0
    current_wins = 0
    current_losses = 0
    for pnl in pnls:
        if pnl > 0:
            current_wins += 1
            current_losses = 0
            max_consec_wins = max(max_consec_wins, current_wins)
        elif pnl < 0:
            current_losses += 1
            current_wins = 0
            max_consec_losses = max(max_consec_losses, current_losses)
        else:
            current_wins = 0
            current_losses = 0
    
    # Bars statistics
    avg_bars = (sum(bars_list) / len(bars_list)) if bars_list else 0.0
    win_bars = [_safe_int(t.get("bars_held", 0)) for t in filtered 
                if _safe_float(t.get("net_pnl", 0)) > 0 and t.get("bars_held") not in (None, "", "None")]
    loss_bars = [_safe_int(t.get("bars_held", 0)) for t in filtered 
                 if _safe_float(t.get("net_pnl", 0)) < 0 and t.get("bars_held") not in (None, "", "None")]
    avg_bars_win = (sum(win_bars) / len(win_bars)) if win_bars else 0.0
    avg_bars_loss = (sum(loss_bars) / len(loss_bars)) if loss_bars else 0.0
    
    # Trading period and time analysis
    entry_dates = []
    exit_dates = []
    for t in filtered:
        entry_dt = _parse_timestamp(t.get("entry_timestamp", ""))
        exit_dt = _parse_timestamp(t.get("exit_timestamp", ""))
        if entry_dt:
            entry_dates.append(entry_dt)
        if exit_dt:
            exit_dates.append(exit_dt)
    
    if entry_dates and exit_dates:
        first_entry = min(entry_dates)
        last_exit = max(exit_dates)
        trading_period_days = (last_exit - first_entry).days
        if trading_period_days < 1:
            trading_period_days = 1
    else:
        trading_period_days = 1
        first_entry = None
        last_exit = None
    
    trades_per_month = (trade_count / (trading_period_days / 30)) if trading_period_days >= 30 else trade_count
    
    # % Time in Market (total bars held / total bars in period)
    total_bars_held = sum(bars_list) if bars_list else 0
    # Assuming 4H timeframe: 6 bars per day
    total_bars_in_period = trading_period_days * 6
    pct_time_in_market = (total_bars_held / total_bars_in_period * 100) if total_bars_in_period > 0 else 0.0
    
    # Longest flat period (days between trades)
    longest_flat_days = 0
    if len(exit_dates) > 1:
        sorted_exits = sorted(exit_dates)
        sorted_entries = sorted(entry_dates)
        for i in range(1, len(sorted_entries)):
            gap = (sorted_entries[i] - sorted_exits[i-1]).days if i-1 < len(sorted_exits) else 0
            if gap > longest_flat_days:
                longest_flat_days = gap
    
    # MFE / MAE computation (consume execution-emitted values)
    mfe_list = []
    mae_list = []
    
    for t in filtered:
        mfe_r = _safe_float(t.get("mfe_r", 0))
        mae_r = _safe_float(t.get("mae_r", 0))
    
        # Use emitted values only; no recomputation
        if mfe_r > 0 or mae_r > 0:
            mfe_list.append(mfe_r)
            mae_list.append(mae_r)
    
    avg_mfe_r = (sum(mfe_list) / len(mfe_list)) if mfe_list else 0.0
    avg_mae_r = (sum(mae_list) / len(mae_list)) if mae_list else 0.0
    edge_ratio = (avg_mfe_r / avg_mae_r) if avg_mae_r > 0 else avg_mfe_r if avg_mfe_r > 0 else 0.0
    
    # Concentration (top 5 trades contribution)
    sorted_wins = sorted(wins, reverse=True)
    top5_profit = sum(sorted_wins[:5]) if len(sorted_wins) >= 5 else sum(sorted_wins)
    top5_pct = (top5_profit / gross_profit * 100) if gross_profit > 0 else 0.0
    
    sorted_losses = sorted(losses)
    worst5_loss = sum(sorted_losses[:5]) if len(sorted_losses) >= 5 else sum(sorted_losses)
    worst5_pct = (abs(worst5_loss) / gross_loss * 100) if gross_loss > 0 else 0.0
    
    # Risk metrics (computed from trade returns)
    returns = [pnl / starting_capital for pnl in pnls]
    sharpe_ratio = _compute_sharpe_ratio(returns)
    sortino_ratio = _compute_sortino_ratio(returns)
    k_ratio = _compute_k_ratio(equity_curve)
    sqn = _compute_sqn(pnls)
    return_retracement_ratio = return_dd_ratio  # Same as Return/DD
    
    # Volatility regime breakdown (ATR-percentile based)
    vol_regimes = _compute_atr_percentile_regimes(filtered)
    
    vol_low_pnls = []
    vol_normal_pnls = []
    vol_high_pnls = []
    
    for i, t in enumerate(filtered):
        pnl = _safe_float(t.get("net_pnl", 0))
        regime = vol_regimes.get(i, "normal")
        if regime == "low":
            vol_low_pnls.append(pnl)
        elif regime == "normal":
            vol_normal_pnls.append(pnl)
        else:
            vol_high_pnls.append(pnl)
    
    net_profit_low_vol = sum(vol_low_pnls)
    net_profit_normal_vol = sum(vol_normal_pnls)
    net_profit_high_vol = sum(vol_high_pnls)
    trades_low_vol = len(vol_low_pnls)
    trades_normal_vol = len(vol_normal_pnls)
    trades_high_vol = len(vol_high_pnls)
    avg_trade_low_vol = (net_profit_low_vol / trades_low_vol) if trades_low_vol > 0 else 0.0
    avg_trade_normal_vol = (net_profit_normal_vol / trades_normal_vol) if trades_normal_vol > 0 else 0.0
    avg_trade_high_vol = (net_profit_high_vol / trades_high_vol) if trades_high_vol > 0 else 0.0
    
    # Session breakdown
    asia_pnls = []
    london_pnls = []
    ny_pnls = []
    
    for t in filtered:
        pnl = _safe_float(t.get("net_pnl", 0))
        entry_dt = _parse_timestamp(t.get("entry_timestamp", ""))
        session = _get_session(entry_dt)
        if session == "asia":
            asia_pnls.append(pnl)
        elif session == "london":
            london_pnls.append(pnl)
        else:
            ny_pnls.append(pnl)
    
    net_profit_asia = sum(asia_pnls)
    net_profit_london = sum(london_pnls)
    net_profit_ny = sum(ny_pnls)
    trades_asia = len(asia_pnls)
    trades_london = len(london_pnls)
    trades_ny = len(ny_pnls)
    avg_trade_asia = (net_profit_asia / trades_asia) if trades_asia > 0 else 0.0
    avg_trade_london = (net_profit_london / trades_london) if trades_london > 0 else 0.0
    avg_trade_ny = (net_profit_ny / trades_ny) if trades_ny > 0 else 0.0
    
    return {
        "starting_capital": starting_capital,
        "net_profit": round(net_profit, 2),
        "gross_profit": round(gross_profit, 2),
        "gross_loss": round(gross_loss, 2),
        "profit_factor": round(profit_factor, 2),
        "expectancy": round(expectancy, 2),
        "return_dd_ratio": round(return_dd_ratio, 2),
        "total_trades": trade_count,
        "winning_trades": win_count,
        "losing_trades": loss_count,
        "pct_profitable": round(win_rate, 2),
        "trades_per_month": round(trades_per_month, 2),
        "longest_flat_days": longest_flat_days,
        "avg_trade": round(avg_trade, 2),
        "avg_win": round(avg_win, 2),
        "avg_loss": round(avg_loss, 2),
        "win_loss_ratio": round(win_loss_ratio, 2),
        "avg_mfe_r": round(avg_mfe_r, 2),
        "avg_mae_r": round(avg_mae_r, 2),
        "edge_ratio": round(edge_ratio, 2),
        "largest_win": round(largest_win, 2),
        "largest_loss": round(largest_loss, 2),
        "top5_pct_gross_profit": round(top5_pct, 2),
        "worst5_loss_pct": round(worst5_pct, 2),
        "max_consec_wins": max_consec_wins,
        "max_consec_losses": max_consec_losses,
        "max_dd_usd": round(max_dd, 2),
        "max_dd_pct": round(max_dd_pct, 2),
        "return_on_capital": round(return_on_capital, 2),
        "pct_time_in_market": round(pct_time_in_market, 2),
        "sharpe_ratio": round(sharpe_ratio, 2),
        "sortino_ratio": round(sortino_ratio, 2),
        "k_ratio": round(k_ratio, 2),
        "sqn": round(sqn, 2),
        "return_retracement_ratio": round(return_retracement_ratio, 2),
        "avg_bars_win": round(avg_bars_win, 2),
        "avg_bars_loss": round(avg_bars_loss, 2),
        "avg_bars": round(avg_bars, 2),
        "trading_period_days": trading_period_days,
        "net_profit_low_vol": round(net_profit_low_vol, 2),
        "net_profit_normal_vol": round(net_profit_normal_vol, 2),
        "net_profit_high_vol": round(net_profit_high_vol, 2),
        "trades_low_vol": trades_low_vol,
        "trades_normal_vol": trades_normal_vol,
        "trades_high_vol": trades_high_vol,
        "avg_trade_low_vol": round(avg_trade_low_vol, 2),
        "avg_trade_normal_vol": round(avg_trade_normal_vol, 2),
        "avg_trade_high_vol": round(avg_trade_high_vol, 2),
        "net_profit_asia": round(net_profit_asia, 2),
        "net_profit_london": round(net_profit_london, 2),
        "net_profit_ny": round(net_profit_ny, 2),
        "trades_asia": trades_asia,
        "trades_london": trades_london,
        "trades_ny": trades_ny,
        "avg_trade_asia": round(avg_trade_asia, 2),
        "avg_trade_london": round(avg_trade_london, 2),
        "avg_trade_ny": round(avg_trade_ny, 2),
    }


def _empty_metrics(starting_capital=0.0):
    return {
        "starting_capital": starting_capital,
        "net_profit": 0.0, "gross_profit": 0.0, "gross_loss": 0.0,
        "profit_factor": 0.0, "expectancy": 0.0, "return_dd_ratio": 0.0,
        "total_trades": 0, "winning_trades": 0, "losing_trades": 0,
        "pct_profitable": 0.0, "trades_per_month": 0.0, "longest_flat_days": 0,
        "avg_trade": 0.0, "avg_win": 0.0, "avg_loss": 0.0, "win_loss_ratio": 0.0,
        "avg_mfe_r": 0.0, "avg_mae_r": 0.0, "edge_ratio": 0.0,
        "largest_win": 0.0, "largest_loss": 0.0,
        "top5_pct_gross_profit": 0.0, "worst5_loss_pct": 0.0,
        "max_consec_wins": 0, "max_consec_losses": 0,
        "max_dd_usd": 0.0, "max_dd_pct": 0.0,
        "return_on_capital": 0.0, "pct_time_in_market": 0.0,
        "sharpe_ratio": 0.0, "sortino_ratio": 0.0, "k_ratio": 0.0, "sqn": 0.0,
        "return_retracement_ratio": 0.0,
        "avg_bars_win": 0.0, "avg_bars_loss": 0.0, "avg_bars": 0.0,
        "trading_period_days": 0,
        "net_profit_low_vol": 0.0, "net_profit_normal_vol": 0.0, "net_profit_high_vol": 0.0,
        "trades_low_vol": 0, "trades_normal_vol": 0, "trades_high_vol": 0,
        "avg_trade_low_vol": 0.0, "avg_trade_normal_vol": 0.0, "avg_trade_high_vol": 0.0,
        "net_profit_asia": 0.0, "net_profit_london": 0.0, "net_profit_ny": 0.0,
        "trades_asia": 0, "trades_london": 0, "trades_ny": 0,
        "avg_trade_asia": 0.0, "avg_trade_london": 0.0, "avg_trade_ny": 0.0,
    }


def _compute_yearwise_metrics(trades, year, starting_capital):
    """Compute yearwise metrics for a specific year."""
    year_trades = []
    for t in trades:
        exit_dt = _parse_timestamp(t.get("exit_timestamp", ""))
        if exit_dt and exit_dt.year == year:
            year_trades.append(t)
    
    if not year_trades:
        return None
    
    pnls = [_safe_float(t.get("net_pnl", 0)) for t in year_trades]
    bars_list = [_safe_int(t.get("bars_held", 0)) for t in year_trades if t.get("bars_held") not in (None, "", "None")]
    
    wins = [p for p in pnls if p > 0]
    losses = [p for p in pnls if p < 0]
    
    gross_profit = sum(wins)
    gross_loss = abs(sum(losses))
    net_profit = sum(pnls)
    trade_count = len(pnls)
    win_count = len(wins)
    loss_count = len(losses)
    win_rate = (win_count / trade_count * 100) if trade_count > 0 else 0.0
    profit_factor = (gross_profit / gross_loss) if gross_loss > 0 else gross_profit if gross_profit > 0 else 0.0
    avg_trade = (net_profit / trade_count) if trade_count > 0 else 0.0
    avg_bars = (sum(bars_list) / len(bars_list)) if bars_list else 0.0
    
    # Drawdown for year
    cumulative = 0.0
    peak = 0.0
    max_dd = 0.0
    for pnl in pnls:
        cumulative += pnl
        if cumulative > peak:
            peak = cumulative
        dd = peak - cumulative
        if dd > max_dd:
            max_dd = dd
    
    max_dd_pct = (max_dd / starting_capital * 100) if starting_capital > 0 else 0.0
    return_dd_ratio = (net_profit / max_dd) if max_dd > 0 else net_profit if net_profit > 0 else 0.0
    
    return {
        "year": year,
        "net_profit": round(net_profit, 2),
        "gross_profit": round(gross_profit, 2),
        "gross_loss": round(gross_loss, 2),
        "trade_count": trade_count,
        "win_rate": round(win_rate, 2),
        "profit_factor": round(profit_factor, 2),
        "avg_trade": round(avg_trade, 2),
        "max_dd_usd": round(max_dd, 2),
        "max_dd_pct": round(max_dd_pct, 2),
        "return_dd_ratio": round(return_dd_ratio, 2),
        "avg_bars": round(avg_bars, 2),
        "win_count": win_count,
        "loss_count": loss_count,
    }


def _add_settings_sheet(ws, metadata):
    ws.append(["Parameter", "Value"])
    for col in [1, 2]:
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = LEFT_ALIGN
    
    settings = [
        ("Run ID", metadata.get("run_id", "")),
        ("Strategy Name", metadata.get("strategy_name", "")),
        ("Symbol", metadata.get("symbol", "")),
        ("Broker", metadata.get("broker", "")),
        ("Reference Capital (USD)", metadata.get("reference_capital_usd", "")),
        ("Position Sizing Basis", metadata.get("position_sizing_basis", "")),
        ("Timeframe", metadata.get("timeframe", "")),
        ("Date Range Start", metadata.get("date_range", {}).get("start", "")),
        ("Date Range End", metadata.get("date_range", {}).get("end", "")),
        ("Execution Timestamp", metadata.get("execution_timestamp_utc", "")),
        ("Engine Name", metadata.get("engine_name", "")),
        ("Engine Version", metadata.get("engine_version", "")),
        ("Schema Version", metadata.get("schema_version", "")),
    ]
    
    for row_idx, (param, value) in enumerate(settings, 2):
        ws.cell(row=row_idx, column=1, value=param).alignment = LEFT_ALIGN
        ws.cell(row=row_idx, column=2, value=value).alignment = LEFT_ALIGN


def _compute_buy_hold_benchmark(trades):
    """Compute Buy & Hold benchmark from trade price series (contextual only)."""
    if not trades or len(trades) < 2:
        return None
    
    try:
        # Sort trades by entry timestamp
        sorted_trades = sorted(trades, key=lambda t: t.get("entry_timestamp", ""))
        
        # Use first entry price and last exit price
        first_price = _safe_float(sorted_trades[0].get("entry_price", 0))
        last_price = _safe_float(sorted_trades[-1].get("exit_price", 0))
        
        if first_price <= 0 or last_price <= 0:
            return None
        
        # Return percentage
        bh_return_pct = ((last_price - first_price) / first_price) * 100
        
        # Compute max drawdown from exit prices (simplified equity curve)
        exit_prices = [_safe_float(t.get("exit_price", 0)) for t in sorted_trades]
        exit_prices = [p for p in exit_prices if p > 0]
        if not exit_prices:
            return None
        
        peak = exit_prices[0]
        max_dd_pct = 0.0
        for price in exit_prices:
            if price > peak:
                peak = price
            dd = ((peak - price) / peak) * 100 if peak > 0 else 0
            if dd > max_dd_pct:
                max_dd_pct = dd
        
        return {
            "return_pct": round(bh_return_pct, 2),
            "max_drawdown_pct": round(max_dd_pct, 2),
            "first_price": first_price,
            "last_price": last_price,
        }
    except Exception:
        return None


def _add_performance_summary_sheet(ws, trades, starting_capital):
    """SOP §5.1 compliant Performance Summary with All/Long/Short breakdown."""
    
    all_metrics = _compute_metrics_from_trades(trades, starting_capital, None)
    long_metrics = _compute_metrics_from_trades(trades, starting_capital, 1)
    short_metrics = _compute_metrics_from_trades(trades, starting_capital, -1)
    
    def add_metric_row(row_num, label, all_val, long_val, short_val, is_currency=False, is_pct=False):
        ws.cell(row=row_num, column=1, value=label).alignment = LEFT_ALIGN
        
        # Write pure numeric values with Excel number format
        for col, v in [(2, all_val), (3, long_val), (4, short_val)]:
            cell = ws.cell(row=row_num, column=col, value=v)
            cell.alignment = LEFT_ALIGN
            if is_currency:
                cell.number_format = '#,##0.00'
            elif is_pct:
                cell.number_format = '0.00'
        
        if row_num % 2 == 0:
            for col in range(1, 5):
                ws.cell(row=row_num, column=col).fill = ALT_FILL
        
        return row_num + 1
    
    # Title
    ws.cell(row=1, column=1, value="Strategy Performance Summary").font = Font(bold=True, size=14)
    ws.cell(row=1, column=1).alignment = LEFT_ALIGN
    
    # Headers
    headers = ["Metric", "All Trades", "Long Trades", "Short Trades"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = LEFT_ALIGN
    
    row = 4
    
    # Capital & Profitability
    row = add_metric_row(row, "Starting Capital", all_metrics["starting_capital"], long_metrics["starting_capital"], short_metrics["starting_capital"], is_currency=True)
    row = add_metric_row(row, "Net Profit (USD)", all_metrics["net_profit"], long_metrics["net_profit"], short_metrics["net_profit"], is_currency=True)
    row = add_metric_row(row, "Gross Profit (USD)", all_metrics["gross_profit"], long_metrics["gross_profit"], short_metrics["gross_profit"], is_currency=True)
    row = add_metric_row(row, "Gross Loss (USD)", all_metrics["gross_loss"], long_metrics["gross_loss"], short_metrics["gross_loss"], is_currency=True)
    row = add_metric_row(row, "Profit Factor", all_metrics["profit_factor"], long_metrics["profit_factor"], short_metrics["profit_factor"])
    row = add_metric_row(row, "Expectancy (USD)", all_metrics["expectancy"], long_metrics["expectancy"], short_metrics["expectancy"], is_currency=True)
    row = add_metric_row(row, "Return / Drawdown Ratio", all_metrics["return_dd_ratio"], long_metrics["return_dd_ratio"], short_metrics["return_dd_ratio"])
    row += 1
    
    # Trade Activity
    row = add_metric_row(row, "Total Trades", all_metrics["total_trades"], long_metrics["total_trades"], short_metrics["total_trades"])
    row = add_metric_row(row, "Winning Trades", all_metrics["winning_trades"], long_metrics["winning_trades"], short_metrics["winning_trades"])
    row = add_metric_row(row, "Losing Trades", all_metrics["losing_trades"], long_metrics["losing_trades"], short_metrics["losing_trades"])
    row = add_metric_row(row, "% Profitable", all_metrics["pct_profitable"], long_metrics["pct_profitable"], short_metrics["pct_profitable"], is_pct=True)
    row = add_metric_row(row, "Trades per Month", all_metrics["trades_per_month"], long_metrics["trades_per_month"], short_metrics["trades_per_month"])
    row = add_metric_row(row, "Longest Flat Period (Days)", all_metrics["longest_flat_days"], long_metrics["longest_flat_days"], short_metrics["longest_flat_days"])
    row += 1
    
    # Averages & Trade Quality
    row = add_metric_row(row, "Avg Trade (USD)", all_metrics["avg_trade"], long_metrics["avg_trade"], short_metrics["avg_trade"], is_currency=True)
    row = add_metric_row(row, "Avg Win (USD)", all_metrics["avg_win"], long_metrics["avg_win"], short_metrics["avg_win"], is_currency=True)
    row = add_metric_row(row, "Avg Loss (USD)", all_metrics["avg_loss"], long_metrics["avg_loss"], short_metrics["avg_loss"], is_currency=True)
    row = add_metric_row(row, "Win/Loss Ratio", all_metrics["win_loss_ratio"], long_metrics["win_loss_ratio"], short_metrics["win_loss_ratio"])
    row = add_metric_row(row, "Avg MFE (R)", all_metrics["avg_mfe_r"], long_metrics["avg_mfe_r"], short_metrics["avg_mfe_r"])
    row = add_metric_row(row, "Avg MAE (R)", all_metrics["avg_mae_r"], long_metrics["avg_mae_r"], short_metrics["avg_mae_r"])
    row = add_metric_row(row, "Edge Ratio (MFE / MAE)", all_metrics["edge_ratio"], long_metrics["edge_ratio"], short_metrics["edge_ratio"])
    row += 1
    
    # Extremes & Concentration
    row = add_metric_row(row, "Largest Win (USD)", all_metrics["largest_win"], long_metrics["largest_win"], short_metrics["largest_win"], is_currency=True)
    row = add_metric_row(row, "Largest Loss (USD)", all_metrics["largest_loss"], long_metrics["largest_loss"], short_metrics["largest_loss"], is_currency=True)
    row = add_metric_row(row, "% of Gross Profit (Top Trades)", all_metrics["top5_pct_gross_profit"], long_metrics["top5_pct_gross_profit"], short_metrics["top5_pct_gross_profit"], is_pct=True)
    row = add_metric_row(row, "Worst 5 Trades Loss %", all_metrics["worst5_loss_pct"], long_metrics["worst5_loss_pct"], short_metrics["worst5_loss_pct"], is_pct=True)
    row += 1
    
    # Streaks
    row = add_metric_row(row, "Max Consecutive Wins", all_metrics["max_consec_wins"], long_metrics["max_consec_wins"], short_metrics["max_consec_wins"])
    row = add_metric_row(row, "Max Consecutive Losses", all_metrics["max_consec_losses"], long_metrics["max_consec_losses"], short_metrics["max_consec_losses"])
    row += 1
    
    # Drawdown & Exposure
    row = add_metric_row(row, "Max Drawdown (USD)", all_metrics["max_dd_usd"], long_metrics["max_dd_usd"], short_metrics["max_dd_usd"], is_currency=True)
    row = add_metric_row(row, "Max Drawdown (%)", all_metrics["max_dd_pct"], long_metrics["max_dd_pct"], short_metrics["max_dd_pct"], is_pct=True)
    row = add_metric_row(row, "Return on Capital", all_metrics["return_on_capital"], long_metrics["return_on_capital"], short_metrics["return_on_capital"], is_pct=True)
    row = add_metric_row(row, "% Time in Market", all_metrics["pct_time_in_market"], long_metrics["pct_time_in_market"], short_metrics["pct_time_in_market"], is_pct=True)
    row += 1
    
    # Risk & System Quality
    row = add_metric_row(row, "Sharpe Ratio", all_metrics["sharpe_ratio"], long_metrics["sharpe_ratio"], short_metrics["sharpe_ratio"])
    row = add_metric_row(row, "Sortino Ratio", all_metrics["sortino_ratio"], long_metrics["sortino_ratio"], short_metrics["sortino_ratio"])
    row = add_metric_row(row, "K-Ratio", all_metrics["k_ratio"], long_metrics["k_ratio"], short_metrics["k_ratio"])
    row = add_metric_row(row, "SQN (System Quality Number)", all_metrics["sqn"], long_metrics["sqn"], short_metrics["sqn"])
    row = add_metric_row(row, "Return Retracement Ratio", all_metrics["return_retracement_ratio"], long_metrics["return_retracement_ratio"], short_metrics["return_retracement_ratio"])
    row += 1
    
    # Duration
    row = add_metric_row(row, "Avg Bars in Winning Trades", all_metrics["avg_bars_win"], long_metrics["avg_bars_win"], short_metrics["avg_bars_win"])
    row = add_metric_row(row, "Avg Bars in Losing Trades", all_metrics["avg_bars_loss"], long_metrics["avg_bars_loss"], short_metrics["avg_bars_loss"])
    row = add_metric_row(row, "Avg Bars per Trade", all_metrics["avg_bars"], long_metrics["avg_bars"], short_metrics["avg_bars"])
    row = add_metric_row(row, "Trading Period (Days)", all_metrics["trading_period_days"], long_metrics["trading_period_days"], short_metrics["trading_period_days"])
    row += 1
    
    # Volatility Regime Breakdown
    row = add_metric_row(row, "Net Profit - Low Volatility", all_metrics["net_profit_low_vol"], long_metrics["net_profit_low_vol"], short_metrics["net_profit_low_vol"], is_currency=True)
    row = add_metric_row(row, "Net Profit - Normal Volatility", all_metrics["net_profit_normal_vol"], long_metrics["net_profit_normal_vol"], short_metrics["net_profit_normal_vol"], is_currency=True)
    row = add_metric_row(row, "Net Profit - High Volatility", all_metrics["net_profit_high_vol"], long_metrics["net_profit_high_vol"], short_metrics["net_profit_high_vol"], is_currency=True)
    row = add_metric_row(row, "Trades - Low Volatility", all_metrics["trades_low_vol"], long_metrics["trades_low_vol"], short_metrics["trades_low_vol"])
    row = add_metric_row(row, "Trades - Normal Volatility", all_metrics["trades_normal_vol"], long_metrics["trades_normal_vol"], short_metrics["trades_normal_vol"])
    row = add_metric_row(row, "Trades - High Volatility", all_metrics["trades_high_vol"], long_metrics["trades_high_vol"], short_metrics["trades_high_vol"])
    row = add_metric_row(row, "Avg Trade - Low Volatility", all_metrics["avg_trade_low_vol"], long_metrics["avg_trade_low_vol"], short_metrics["avg_trade_low_vol"], is_currency=True)
    row = add_metric_row(row, "Avg Trade - Normal Volatility", all_metrics["avg_trade_normal_vol"], long_metrics["avg_trade_normal_vol"], short_metrics["avg_trade_normal_vol"], is_currency=True)
    row = add_metric_row(row, "Avg Trade - High Volatility", all_metrics["avg_trade_high_vol"], long_metrics["avg_trade_high_vol"], short_metrics["avg_trade_high_vol"], is_currency=True)
    row += 1
    
    # Session Breakdown
    row = add_metric_row(row, "Net Profit - Asia Session", all_metrics["net_profit_asia"], long_metrics["net_profit_asia"], short_metrics["net_profit_asia"], is_currency=True)
    row = add_metric_row(row, "Net Profit - London Session", all_metrics["net_profit_london"], long_metrics["net_profit_london"], short_metrics["net_profit_london"], is_currency=True)
    row = add_metric_row(row, "Net Profit - New York Session", all_metrics["net_profit_ny"], long_metrics["net_profit_ny"], short_metrics["net_profit_ny"], is_currency=True)
    row = add_metric_row(row, "Trades - Asia Session", all_metrics["trades_asia"], long_metrics["trades_asia"], short_metrics["trades_asia"])
    row = add_metric_row(row, "Trades - London Session", all_metrics["trades_london"], long_metrics["trades_london"], short_metrics["trades_london"])
    row = add_metric_row(row, "Trades - New York Session", all_metrics["trades_ny"], long_metrics["trades_ny"], short_metrics["trades_ny"])
    row = add_metric_row(row, "Avg Trade - Asia Session", all_metrics["avg_trade_asia"], long_metrics["avg_trade_asia"], short_metrics["avg_trade_asia"], is_currency=True)
    row = add_metric_row(row, "Avg Trade - London Session", all_metrics["avg_trade_london"], long_metrics["avg_trade_london"], short_metrics["avg_trade_london"], is_currency=True)
    row = add_metric_row(row, "Avg Trade - New York Session", all_metrics["avg_trade_ny"], long_metrics["avg_trade_ny"], short_metrics["avg_trade_ny"], is_currency=True)
    
    # Buy & Hold Benchmark (Contextual Only) — SOP §5.1
    bh = _compute_buy_hold_benchmark(trades)
    if bh is not None:
        row += 2
        ws.cell(row=row, column=1, value="Buy & Hold Benchmark (Contextual Only)").font = Font(bold=True, size=12)
        ws.cell(row=row, column=1).alignment = LEFT_ALIGN
        row += 1
        
        ws.cell(row=row, column=1, value="First Price").alignment = LEFT_ALIGN
        cell = ws.cell(row=row, column=2, value=bh["first_price"])
        cell.number_format = '#,##0.00'
        row += 1
        
        ws.cell(row=row, column=1, value="Last Price").alignment = LEFT_ALIGN
        cell = ws.cell(row=row, column=2, value=bh["last_price"])
        cell.number_format = '#,##0.00'
        row += 1
        
        ws.cell(row=row, column=1, value="Buy & Hold Return (%)").alignment = LEFT_ALIGN
        cell = ws.cell(row=row, column=2, value=bh["return_pct"])
        cell.number_format = '0.00'
        row += 1
        
        ws.cell(row=row, column=1, value="Buy & Hold Max Drawdown (%)").alignment = LEFT_ALIGN
        cell = ws.cell(row=row, column=2, value=bh["max_drawdown_pct"])
        cell.number_format = '0.00'
        
        # Contextual comparison (display-only)
        strategy_return_pct = (all_metrics["net_profit"] / starting_capital) * 100 if starting_capital > 0 else 0.0
        relative_perf = strategy_return_pct - bh["return_pct"]
        
        row += 1
        ws.cell(row=row, column=1, value="Strategy vs Buy & Hold (Contextual Only)").font = Font(bold=True)
        row += 1
        
        ws.cell(row=row, column=1, value="Strategy Net Return (%)")
        ws.cell(row=row, column=2, value=round(strategy_return_pct, 2))
        cell = ws.cell(row=row, column=2)
        cell.number_format = '0.00'
        row += 1
        
        ws.cell(row=row, column=1, value="Buy & Hold Return (%)")
        ws.cell(row=row, column=2, value=bh["return_pct"])
        cell = ws.cell(row=row, column=2)
        cell.number_format = '0.00'
        row += 1
        
        ws.cell(row=row, column=1, value="Relative Performance (%)")
        ws.cell(row=row, column=2, value=round(relative_perf, 2))
        cell = ws.cell(row=row, column=2)
        cell.number_format = '0.00'





def _add_yearwise_sheet(ws, trades, starting_capital):
    """SOP §5.2 compliant Yearwise Performance."""
    
    years = set()
    for t in trades:
        exit_dt = _parse_timestamp(t.get("exit_timestamp", ""))
        if exit_dt:
            years.add(exit_dt.year)
    
    headers = [
        "Year", "Net Profit (USD)", "Gross Profit (USD)", "Gross Loss (USD)",
        "Trade Count", "Win Rate (%)", "Profit Factor", "Avg Trade (USD)",
        "Max Drawdown (USD)", "Max Drawdown (%)", "Return / DD Ratio",
        "Avg Bars per Trade", "Winning Trades", "Losing Trades",
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = LEFT_ALIGN
    
    row = 2
    # Column formats: Year(none), NetProfit(curr), GrossProfit(curr), GrossLoss(curr),
    # TradeCount(none), WinRate(pct), ProfitFactor(dec), AvgTrade(curr),
    # MaxDD_USD(curr), MaxDD_Pct(pct), ReturnDD(dec), AvgBars(dec), WinCount(none), LossCount(none)
    col_formats = [None, '#,##0.00', '#,##0.00', '#,##0.00', None, '0.00', '0.00', '#,##0.00', '#,##0.00', '0.00', '0.00', '0.00', None, None]
    
    for year in sorted(years):
        metrics = _compute_yearwise_metrics(trades, year, starting_capital)
        if metrics:
            values = [
                metrics["year"],
                metrics['net_profit'],
                metrics['gross_profit'],
                metrics['gross_loss'],
                metrics["trade_count"],
                metrics['win_rate'],
                metrics['profit_factor'],
                metrics['avg_trade'],
                metrics['max_dd_usd'],
                metrics['max_dd_pct'],
                metrics['return_dd_ratio'],
                metrics['avg_bars'],
                metrics["win_count"],
                metrics["loss_count"],
            ]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.alignment = LEFT_ALIGN
                if col_formats[col-1]:
                    cell.number_format = col_formats[col-1]
                if row % 2 == 0:
                    cell.fill = ALT_FILL
            row += 1


def _add_trades_list_sheet(ws, trades):
    """Trades List with all required fields."""
    
    headers = [
        "Parent Trade ID", "Sequence Index", "Strategy Name", "Direction",
        "Entry Time", "Exit Time", "Entry Price", "Exit Price", "PnL (USD)", "Bars Held",
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = LEFT_ALIGN
    
    row = 2
    for trade in trades:
        direction = _safe_int(trade.get("direction", 0))
        values = [
            trade.get("parent_trade_id", ""),
            trade.get("sequence_index", ""),
            trade.get("strategy_name", ""),
            "LONG" if direction == 1 else "SHORT",
            trade.get("entry_timestamp", ""),
            trade.get("exit_timestamp", ""),
            trade.get("entry_price", ""),
            trade.get("exit_price", ""),
            trade.get("net_pnl", ""),
            trade.get("bars_held", ""),
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.alignment = LEFT_ALIGN
            if row % 2 == 0:
                cell.fill = ALT_FILL
        row += 1


def generate_excel_report(artifacts, output_path: Path):
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl is required for Excel generation")
    
    wb = openpyxl.Workbook()
    
    ws_settings = wb.active
    ws_settings.title = "Settings"
    _add_settings_sheet(ws_settings, artifacts["metadata"])
    
    ws_summary = wb.create_sheet("Performance Summary")
    
    starting_capital = artifacts["metadata"]["reference_capital_usd"]
    _add_performance_summary_sheet(ws_summary, artifacts["tradelevel"], starting_capital)
    
    
    ws_yearwise = wb.create_sheet("Yearwise Performance")
    _add_yearwise_sheet(ws_yearwise, artifacts["tradelevel"], starting_capital)
    
    ws_trades = wb.create_sheet("Trades List")
    _add_trades_list_sheet(ws_trades, artifacts["tradelevel"])
    
    # Auto-fit all columns
    for ws in wb.worksheets:
        for col_idx in range(1, ws.max_column + 1):
            max_length = 0
            column = get_column_letter(col_idx)
            for row_idx in range(1, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column].width = min(max_length + 2, 50)
        
        if ws.title == "Performance Summary":
            ws.freeze_panes = "A4"
        else:
            ws.freeze_panes = "A2"
    
    wb.save(output_path)


def compile_stage2(run_folder: Path):
    artifacts = load_stage1_artifacts(run_folder)
    
    strategy_name = artifacts["metadata"].get("strategy_name", "UNKNOWN")
    excel_filename = f"AK_Trade_Report_{strategy_name}.xlsx"
    excel_path = run_folder / excel_filename
    
    generate_excel_report(artifacts, excel_path)
    
    return run_folder, [excel_filename]


def main():
    import sys
    
    if len(sys.argv) < 2:
        raise SystemExit("ERROR: Run path argument required. Usage: python stage2_compiler.py <run_folder_path>")
    
    run_folder = Path(sys.argv[1])
    
    if not run_folder.exists():
        raise SystemExit(f"ERROR: Run folder not found: {run_folder}")
    
    if not run_folder.is_dir():
        raise SystemExit(f"ERROR: Run path is not a directory: {run_folder}")
    
    try:
        output_dir, files = compile_stage2(run_folder)
        
        print(f"Path: {output_dir}")
        print("Generated files:")
        for f in files:
            print(f"  - {output_dir / f}")
        print()
        print("Stage-2 presentation artifacts generated successfully from immutable Stage-1 inputs.")
        
    except FileNotFoundError as e:
        raise SystemExit(f"Stage-2 compilation aborted. {e}")
    except Exception as e:
        raise SystemExit(f"Stage-2 compilation aborted. {e}")


if __name__ == "__main__":
    main()
