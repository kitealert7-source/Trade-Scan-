
"""
run_pipeline.py -- Master Execution Pipeline Orchestrator (v3.2 - State Gated)

Usage:
  python tools/run_pipeline.py <DIRECTIVE_ID>
  python tools/run_pipeline.py --all

Purpose:
  Orchestrates the deterministic multi-asset execution flow:

  0. Preflight (Safety & Governance Checks)
  1. Directive Parsing (Concurrency + Symbol Detection)
  2. Stage-1: Generation (RESEARCH Data Only)
  3. Stage-1.5: Portfolio Risk Constraints (Conditional)
  4. Stage-2: Compilation
  5. Stage-3: Aggregation
  6. Stage-4: Portfolio Evaluation

Execution Model -- Mandatory Compliance:

  - All execution gated by run_state.json (Audit Phase 7)
  - All Stage-1 executions MUST use RESEARCH market data.
  - CLEAN or derived datasets are non-authoritative and prohibited.
  - Directive must contain executable conditions.
  - STRATEGY_PLUGIN_CONTRACT.md must be satisfied.
  - SOP_INDICATOR.md must be enforced (repository-only indicators).
  - No inline indicator logic permitted.

Authority:
  governance/SOP/SOP_TESTING.md
  governance/SOP/STRATEGY_PLUGIN_CONTRACT.md
  governance/SOP/SOP_INDICATOR.md

"""


import sys
import subprocess
import time
import shutil
import os
import json
import hashlib
from datetime import datetime, timezone
from pathlib import Path

# Config
PROJECT_ROOT = Path(__file__).parent.parent
PYTHON_EXE = sys.executable
DIRECTIVES_DIR = PROJECT_ROOT / "backtest_directives"
ACTIVE_DIR = DIRECTIVES_DIR / "active"
COMPLETED_DIR = DIRECTIVES_DIR / "completed"

# Governance Imports
sys.path.insert(0, str(PROJECT_ROOT))
from tools.pipeline_utils import (
    generate_run_id, 
    PipelineStateManager, 
    DirectiveStateManager,
    get_engine_version,
    parse_directive
)

def run_command(cmd_list, step_name):
    print(f"\n{'='*40}")
    print(f"[{step_name}] Executing: {' '.join(cmd_list)}")
    print(f"{'='*40}")
    
    start_time = time.time()
    try:
        subprocess.run(cmd_list, cwd=PROJECT_ROOT, check=True)
        duration = time.time() - start_time
        print(f"[{step_name}] COMPLETED (Time: {duration:.2f}s)")
        return True
    except subprocess.CalledProcessError as e:
        if e.returncode != 2:
            print(f"\n[FATAL] {step_name} FAILED with exit code {e.returncode}")
        # Propagate error up
        raise e
    except Exception as e:
        print(f"\n[FATAL] {step_name} FAILED with error: {e}")
        raise e

def get_directive_path(directive_id):
    """Locate the directive file."""
    candidates = [
        ACTIVE_DIR / directive_id,
        ACTIVE_DIR / f"{directive_id}.txt"
    ]
    for c in candidates:
        if c.exists():
            return c
    
    print(f"[ERROR] Directive file not found for ID: {directive_id}")
    print(f"Searched in: {ACTIVE_DIR}")
    sys.exit(1)

def parse_concurrency_config(file_path):
    from tools.pipeline_utils import parse_directive
    config = parse_directive(file_path)
    # Extract symbols list -- support both cased key variants
    symbols = config.get("symbols", config.get("Symbols", []))
    if isinstance(symbols, str):
        symbols = [s.strip() for s in symbols.split(",") if s.strip()]
    elif not isinstance(symbols, list):
        symbols = []
    max_concurrent = config.get("max_concurrent_positions", len(symbols))
    if isinstance(max_concurrent, str) and max_concurrent.isdigit():
        max_concurrent = int(max_concurrent)
    return max_concurrent, len(symbols)

def run_single_directive(directive_id):
    """Execution logic for a single directive."""
    # 1. Parsing
    d_path = get_directive_path(directive_id)
    clean_id = d_path.stem 
    
    print(f"[CONFIG] Directive: {d_path.name}")
    
    # 1.1 Generate Deterministic Run ID (Early Binding for State)
    # We need a run ID. For single directive, usually we have multiple symbols.
    # The current batch harness generates ONE run_id PER SYMBOL.
    # This creates a complexity: 
    #   - Directive = "SPX04"
    #   - Symbols = ["SPX500", "NAS100"]
    #   - Runs = 2 separate IDs?
    #
    # Current run_stage1.py architecture iterates symbols and emits artifacts.
    # It generates run_id inside the loop.
    # BUT Phase 7 requires ORCHESTRATOR to manage state.
    # If One Directive -> Multiple Runs, orchestrator must manage Multiple State Files?
    #
    # Audit Rule: "A run may start ONLY with a single explicit human directive." (Invariant 1)
    # But "One Directive -> One Atomic Run" (Invariant 1).
    # If a directive has multiple symbols, is it ONE run or MULTIPLE?
    # SOP_TESTING says: "Each iteration... is isolated... must independently satisfy RUN_COMPLETE".
    #
    # For Batch Mode (Multi-Symbol inside one directive), run_stage1 does the loop.
    # This implies run_stage1 is the "Run".
    # 
    # HOWEVER, Phase 7 requires "runs/<RUN_ID>/run_state.json".
    # If run_stage1 generates multiple RUN_IDs, then we have multiple state files?
    # Checks: run_stage1.py loops symbols.
    # If we want strict state gating *before* run_stage1, we need a parent Run ID for the *Directive Execution*?
    # OR we treat the Directive execution as the "Run" and the symbol-runs as sub-artifacts?
    #
    # Let's look at `pipeline_utils.generate_run_id`. It takes a `symbol`.
    # This implies Run ID is per-symbol.
    #
    # If run_stage1 loops 10 symbols, it generates 10 artifacts with 10 run_ids.
    #
    # PROBLEM: The Orchestrator calls run_stage1 ONCE for the directive.
    # If orchestrator must create state *before* stage1, it needs to know the IDs.
    #
    # SOLUTION:
    # We will enforce "Single Symbol per Run" architecture or "Master Run ID".
    # Given existing "Batch Harness", let's parse symbols here and managing state for each?
    #
    # BUT run_stage1.py does the lopp internally.
    # If we want strict State Gating, the Orchestrator should probably loop?
    #
    # Let's inspect `run_stage1.py` again. It loops `parsed_config.get("Symbols", [])`.
    #
    # Refactoring `run_stage1.py` to single-symbol execution is a large change.
    # ALTERNATIVE: The Orchestrator manages a "Directive Run State"?
    #
    # Re-reading Audit: 
    # "Create run_state.json inside runs/<RUN_ID>/."
    # "Only run_pipeline.py may mutate this file."
    #
    # IF run_stage1 produces multiple run_ids, the Orchestrator must know them ahead of time.
    #
    # Let's parsing symbols here.
    
    # We need the symbol list pre-Stage-0.25 only to pre-calculate Run IDs.
    # Use a raw yaml.safe_load() to extract symbols without invoking parse_directive()
    # strict validation (which requires test: wrapper, collisions, etc.).
    # This ensures non-canonical directives still reach Stage -0.25 rather than
    # failing here with a confusing INVALID DIRECTIVE STRUCTURE error.
    import yaml as _yaml_pre
    try:
        _raw_pre = _yaml_pre.safe_load(d_path.read_text(encoding="utf-8")) or {}
    except Exception as _pre_err:
        print(f"[FATAL] YAML_PARSE_ERROR (pre-Stage-0.25): {_pre_err}")
        sys.exit(1)
    # Support both canonical (test: wrapper) and flat directives for symbol extraction only
    _test_block_pre = _raw_pre.get("test", {})
    symbols = (
        _raw_pre.get("symbols")
        or _raw_pre.get("Symbols")
        or _test_block_pre.get("symbols")
        or _test_block_pre.get("Symbols")
        or []
    )
    if isinstance(symbols, str):
        symbols = [symbols]

    if not symbols:
        print("[FATAL] No symbols found in directive.")
        sys.exit(1)

    print(f"[ORCHESTRATOR] Found {len(symbols)} symbols: {symbols}")

    # ----------------------------------------------------------
    # STAGE -0.25: DIRECTIVE CANONICALIZATION GATE
    # Must run before any state initialization or pipeline stage.
    # ----------------------------------------------------------
    from tools.canonicalizer import canonicalize, CanonicalizationError
    import yaml as _yaml

    try:
        raw_yaml = d_path.read_text(encoding="utf-8")
        parsed_raw = _yaml.safe_load(raw_yaml)
        canonical, canonical_yaml, diff_lines, violations, has_drift = \
            canonicalize(parsed_raw)
    except CanonicalizationError as e:
        print(f"\n[FATAL] STAGE -0.25 CANONICALIZATION FAILED:")
        print(f"  {e}")
        sys.exit(1)
    except _yaml.YAMLError as e:
        print(f"\n[FATAL] YAML_PARSE_ERROR: {e}")
        sys.exit(1)

    if violations:
        print("[STAGE -0.25] Structural changes detected:")
        for level, msg in violations:
            print(f"  [{level}] {msg}")

    if has_drift:
        print("\n[STAGE -0.25] STRUCTURAL DRIFT -- directive is not canonical.")
        print("  --- Unified Diff ---")
        for line in diff_lines:
            print(f"  {line}", end="")
        tmp_path = Path("/tmp") / f"{clean_id}_canonical.yaml"
        tmp_path.write_text(canonical_yaml, encoding="utf-8")
        print(f"\n  Corrected YAML written to: {tmp_path}")
        print("  Human must review and approve overwrite.")
        print("[HALT] Pipeline stopped. Fix directive and re-run.")
        sys.exit(1)
    else:
        print("[STAGE -0.25] Directive is in canonical form. [OK]")

    # Stage -0.25 passed. Now safe to call parse_directive() with strict validation.
    # This is the earliest correct point: canonical structure is confirmed.
    # Fix 3: Previously parse_directive() ran before Stage -0.25, so non-canonical
    # directives failed with INVALID DIRECTIVE STRUCTURE before reaching the gate.
    from tools.pipeline_utils import parse_directive
    p_conf = parse_directive(d_path)
    # Authoritative symbol resolution from fully parsed config
    symbols = p_conf.get("Symbols", p_conf.get("symbols", symbols))
    if isinstance(symbols, str):
        symbols = [symbols]

    dir_state_mgr = DirectiveStateManager(clean_id)
    dir_state_mgr.initialize()
    
    current_dir_state = dir_state_mgr.get_state()
    print(f"[ORCHESTRATOR] Directive State: {current_dir_state}")
    
    # Resume Safety Logic
    if current_dir_state == "PORTFOLIO_COMPLETE":
         print(f"[ORCHESTRATOR] Directive {clean_id} is already COMPLETE. Aborting.")
         return
    elif current_dir_state == "FAILED":
         print(f"[ORCHESTRATOR] Directive {clean_id} is FAILED.")
         print(f"[ORCHESTRATOR] To reset, run: python tools/reset_directive.py {clean_id} --reason \"<justification>\"")
         sys.exit(1)
    
    # --provision-only flag
    provision_only = "--provision-only" in sys.argv

    # 1. Initialize State for All Symbols
    print("[ORCHESTRATOR] Initializing symbol states...") 

    run_ids = []
    for symbol in symbols:
        run_id, _ = generate_run_id(d_path, symbol)
        run_ids.append(run_id)
        # Init individual run state (unless we are resuming later stages, but init is idempotent mostly)
        if current_dir_state not in ["SYMBOL_RUNS_COMPLETE", "PORTFOLIO_COMPLETE"]:
             print(f"[ORCHESTRATOR] Managing Run ID: {run_id} ({symbol})")
             state_mgr = PipelineStateManager(run_id, directive_id=clean_id)
             state_mgr.initialize()

    try:
        # Resume Check
        if dir_state_mgr.get_state() == "SYMBOL_RUNS_COMPLETE":  # live fetch -- not stale var
             print("[ORCHESTRATOR] Resuming at Stage-4 (Portfolio)...")
        else:
            # 2. Preflight (Directive Level - Runs Once)
            live_state = dir_state_mgr.get_state()
            _PREFLIGHT_SKIP = {"PREFLIGHT_COMPLETE", "PREFLIGHT_COMPLETE_SEMANTICALLY_VALID",
                               "SYMBOL_RUNS_COMPLETE", "PORTFOLIO_COMPLETE"}
            if live_state not in _PREFLIGHT_SKIP:
                print("[ORCHESTRATOR] Starting Preflight Checks...")
                
                # Execute Preflight
                try:
                    run_command([PYTHON_EXE, "tools/exec_preflight.py", clean_id], "Preflight")
                except subprocess.CalledProcessError as e:
                    if e.returncode == 2:
                        print("\n============================================================")
                        print("[ADMISSION GATE] STRATEGY REQUIRES HUMAN IMPLEMENTATION")
                        print("============================================================")
                        print("ACTION: Please open the generated strategy file and implement `check_entry` and `check_exit`.")
                        print("STATE: Pipeline paused cleanly. Rerun after implementation.")
                        sys.exit(0)
                    else:
                        raise e
                # Transition ALL to PREFLIGHT_COMPLETE
                for rid in run_ids:
                    PipelineStateManager(rid).transition_to("PREFLIGHT_COMPLETE")
                
                dir_state_mgr.transition_to("PREFLIGHT_COMPLETE")
            else:
                 print(f"[ORCHESTRATOR] Preflight already complete (state={live_state}). Checking Semantic Status...")

            # --- STAGE 0.5: SEMANTIC VALIDATION ---
            # Gate: Must be PREFLIGHT_COMPLETE to enter.
            # Exit: PREFLIGHT_COMPLETE_SEMANTICALLY_VALID
            
            # Semantic validation gate uses live state (fetched below)
            # No stale variable dependency.
            
            check_state = dir_state_mgr.get_state()
            if check_state == "PREFLIGHT_COMPLETE":
                print("[ORCHESTRATOR] Starting Stage-0.5: Semantic Validation...")
                from tools.semantic_validator import validate_semantic_signature
                
                try:
                    validate_semantic_signature(str(d_path))
                    
                    # Transition ALL to PREFLIGHT_COMPLETE_SEMANTICALLY_VALID
                    for rid in run_ids:
                        PipelineStateManager(rid).transition_to("PREFLIGHT_COMPLETE_SEMANTICALLY_VALID")
                        
                    dir_state_mgr.transition_to("PREFLIGHT_COMPLETE_SEMANTICALLY_VALID")
                    print("[ORCHESTRATOR] Semantic Validation PASSED.")
                    
                except Exception as e:
                    if "PROVISION_REQUIRED" in str(e):
                        print("\n============================================================")
                        print("[ADMISSION GATE] STRATEGY REQUIRES HUMAN IMPLEMENTATION")
                        print("============================================================")
                        print(f"Details: {e}")
                        print("ACTION: Please open the generated strategy.py file and implement `check_entry` and `check_exit`.")
                        print("STATE: Pipeline paused cleanly. Rerun after implementation.")
                        sys.exit(0)
                    else:
                        print(f"[FATAL] Semantic Validation FAILED: {e}")
                        # Transition to FAILED
                        for rid in run_ids:
                            try: PipelineStateManager(rid).transition_to("FAILED")
                            except Exception as e_cleanup: print(f"[WARN] Failed to mark {rid} as FAILED: {e_cleanup}")
                        try: dir_state_mgr.transition_to("FAILED")
                        except Exception as e_cleanup: print(f"[WARN] Failed to mark directive as FAILED: {e_cleanup}")
                        sys.exit(1)
            elif check_state == "PREFLIGHT_COMPLETE_SEMANTICALLY_VALID":
                print("[ORCHESTRATOR] Semantic Validation already COMPLETE. Resuming...")
            elif check_state not in ["SYMBOL_RUNS_COMPLETE", "PORTFOLIO_COMPLETE"]:
                pass

            # --- STAGE 0.55: SEMANTIC COVERAGE CHECK ---
            # Gate: Runs after semantic validation passes.
            # Verifies all behavioral directive parameters are referenced in strategy.py.
            # Fix 2: Use parsed strategy_id from p_conf, not clean_id, so this gate
            # is not silently skipped when directive ID differs from strategy name.
            _cov_state = dir_state_mgr.get_state()
            if _cov_state in ("PREFLIGHT_COMPLETE_SEMANTICALLY_VALID",):
                _cov_strategy_id = p_conf.get("Strategy", p_conf.get("strategy")) or clean_id
                _cov_strategy_path = PROJECT_ROOT / "strategies" / _cov_strategy_id / "strategy.py"
                if _cov_strategy_path.exists():
                    try:
                        from governance.semantic_coverage_checker import check_semantic_coverage
                        check_semantic_coverage(str(d_path), str(_cov_strategy_path))
                        print("[ORCHESTRATOR] Stage-0.55 Semantic Coverage Check PASSED.")
                    except RuntimeError as e:
                        if "SEMANTIC_COVERAGE_FAILURE" in str(e):
                            print(f"\n[FATAL] {e}")
                            for rid in run_ids:
                                try: PipelineStateManager(rid).transition_to("FAILED")
                                except Exception: pass
                            try: dir_state_mgr.transition_to("FAILED")
                            except Exception: pass
                            sys.exit(1)
                        raise

            # --- PROVISION-ONLY EXIT POINT ---
            # Stops after Stage-0.5 (semantic validation + Admission Gate enforced).
            if provision_only:
                strategy_path = PROJECT_ROOT / "strategies" / clean_id / "strategy.py"
                print(f"[PROVISION-ONLY] Strategy provisioned at: {strategy_path}")
                print(f"[PROVISION-ONLY] Human review required before execution.")
                print(f"[PROVISION-ONLY] Re-run without --provision-only after review.")
                return

            # --- STAGE 0.75: STRATEGY DRY-RUN VALIDATION ---
            live_state_075 = dir_state_mgr.get_state()
            if live_state_075 == "PREFLIGHT_COMPLETE_SEMANTICALLY_VALID":
                from tools.strategy_dryrun_validator import validate_strategy_dryrun
                dryrun_ok = validate_strategy_dryrun(clean_id, symbols[0], d_path)
                if not dryrun_ok:
                    print("[FATAL] Stage-0.75 Dry-Run Validation FAILED.")
                    for rid in run_ids:
                        try: PipelineStateManager(rid).transition_to("FAILED")
                        except Exception: pass
                    dir_state_mgr.transition_to("FAILED")
                    sys.exit(1)

            # Clean legacy summary CSV only if Stage-1 will rerun for at least one symbol
            _any_stage1_rerun = any(
                PipelineStateManager(rid).get_state_data()["current_state"]
                in ("IDLE", "PREFLIGHT_COMPLETE", "PREFLIGHT_COMPLETE_SEMANTICALLY_VALID")
                for rid in run_ids
            )
            summary_csv = PROJECT_ROOT / "backtests" / f"batch_summary_{clean_id}.csv"
            if summary_csv.exists() and _any_stage1_rerun:
                summary_csv.unlink()

            # --- STAGE-0.9: PRE-EXECUTION ATOMIC SNAPSHOTTING ---
            # Snapshot strategy for ALL symbols simultaneously before execution
            # This mathematically eliminates mid-pipeline mutation.
            strategy_id = p_conf.get("Strategy", p_conf.get("strategy"))
            if strategy_id:
                source_strategy_path = PROJECT_ROOT / "strategies" / strategy_id / "strategy.py"
                if source_strategy_path.exists():
                    print("[ORCHESTRATOR] Performing atomic preemptive strategy snapshots...")
                    for rid in run_ids:
                        mgr = PipelineStateManager(rid)
                        target_path = mgr.run_dir / "strategy.py"
                        if not target_path.exists():
                            target_path.parent.mkdir(parents=True, exist_ok=True)
                            shutil.copy(str(source_strategy_path), str(target_path))

            # Orchestrate Atomic Stage-1 Execution
            print("[ORCHESTRATOR] Launching Stage-1 Generator (Atomic)...")
            
            for symbol, rid in zip(symbols, run_ids):
                mgr = PipelineStateManager(rid)
                
                # Skip if symbol has already progressed past Stage-1.
                # Guard must cover all forward states, not just COMPLETE,
                # to prevent illegal re-execution on resume.
                _SKIP_STATES = {
                    "STAGE_1_COMPLETE", "STAGE_2_COMPLETE",
                    "STAGE_3_COMPLETE", "STAGE_3A_COMPLETE", "COMPLETE"
                }
                st = mgr.get_state_data()["current_state"]
                if st in _SKIP_STATES:
                    continue

                try:
                    # 1. Execute Atomic Stage-1
                    cmd = [
                        PYTHON_EXE, 
                        "tools/run_stage1.py", 
                        clean_id, 
                        "--symbol", symbol,
                        "--run_id", rid
                    ]
                    run_command(cmd, f"Stage-1: {symbol}")
                    
                    # --- STAGE-1 ARTIFACT GATE ---
                    out_folder = PROJECT_ROOT / "backtests" / f"{clean_id}_{symbol}"
                    if not (out_folder / "raw" / "results_tradelevel.csv").exists():
                        raise RuntimeError(f"[FATAL] Stage-1 artifact missing for {symbol}. (Probable NO_TRADES).")
                    
                    # 3. Transition to Complete
                    mgr.transition_to("STAGE_1_COMPLETE")
                    
                except Exception as e:
                    print(f"[ERROR] Stage-1 Failed for {symbol}: {e}")
                    try:
                        mgr.transition_to("FAILED")
                    except Exception as e_cleanup:
                        print(f"[WARN] Failed to mark {symbol} run as FAILED: {e_cleanup}")
                    raise e
            
            # End of Stage 1 Loop - Move to Stage 2
            
            # Stage 2
            run_command([PYTHON_EXE, "-m", "engine_dev.universal_research_engine.v1_4_0.stage2_compiler", "--scan", clean_id], "Stage-2 Compilation")
            
            # Per-run_id artifact existence gate (idempotent -- re-verifies on resume)
            for rid, symbol in zip(run_ids, symbols):
                mgr = PipelineStateManager(rid)
                current = mgr.get_state_data()["current_state"]
                if current in ("STAGE_1_COMPLETE", "STAGE_2_COMPLETE"):
                    # Verify AK_Trade_Report exists before marking/confirming complete
                    run_folder = PROJECT_ROOT / "backtests" / f"{clean_id}_{symbol}"
                    ak_reports = list(run_folder.glob("AK_Trade_Report_*.xlsx"))
                    if ak_reports:
                        if current != "STAGE_2_COMPLETE":
                            mgr.transition_to("STAGE_2_COMPLETE")
                    else:
                        print(f"[WARN] Stage-2 artifact missing for {symbol} ({rid[:8]}). Marking FAILED.")
                        mgr.transition_to("FAILED")

            # Stage 3
            run_command([PYTHON_EXE, "tools/stage3_compiler.py", clean_id], "Stage-3 Aggregation")
            
            # --- STAGE-3 ARTIFACT GATE (idempotent) ---
            master_filter_path = PROJECT_ROOT / "backtests" / "Strategy_Master_Filter.xlsx"
            if not master_filter_path.exists():
                print(f"[FATAL] Stage-3 artifact missing: {master_filter_path}")
                for rid in run_ids:
                    try: PipelineStateManager(rid).transition_to("FAILED")
                    except Exception: pass
                dir_state_mgr.transition_to("FAILED")
                sys.exit(1)
            
            import openpyxl
            wb = openpyxl.load_workbook(master_filter_path, read_only=True)
            ws = wb.active
            
            # Resolve strategy column index dynamically
            try:
                headers = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
                strategy_idx = headers.index("strategy")
            except Exception as e:
                print(f"[FATAL] Failed to resolve 'strategy' column in Master Filter: {e}")
                for rid in run_ids:
                    try: PipelineStateManager(rid).transition_to("FAILED")
                    except Exception: pass
                dir_state_mgr.transition_to("FAILED")
                sys.exit(1)
                
            # Count rows whose strategy column starts with this directive's clean_id
            actual_count = sum(
                1 for row in ws.iter_rows(min_row=2, values_only=True)
                if row and len(row) > strategy_idx and row[strategy_idx] and str(row[strategy_idx]).startswith(clean_id)
            )
            wb.close()
            expected_count = len(symbols)
            
            if actual_count != expected_count:
                print(f"[FATAL] Stage-3 cardinality mismatch: expected {expected_count}, found {actual_count} for {clean_id}")
                for rid in run_ids:
                    try: PipelineStateManager(rid).transition_to("FAILED")
                    except Exception: pass
                dir_state_mgr.transition_to("FAILED")
                sys.exit(1)
            
            print(f"[GATE] Stage-3 artifact verified: {actual_count}/{expected_count} rows for {clean_id}")
            # --- END GATE ---
            
            for rid, symbol in zip(run_ids, symbols):
                mgr = PipelineStateManager(rid)
                current = mgr.get_state_data()["current_state"]
                
                if current == "STAGE_2_COMPLETE":
                    mgr.transition_to("STAGE_3_COMPLETE")
                    
                    # Stage 3A (Snapshot)
                    # mgr.transition_to("STAGE_3A_START") # Removed
                    
                    # 1. Verify Snapshot Existence
                    snapshot_path = mgr.run_dir / "strategy.py"
                    if not snapshot_path.exists():
                        mgr.transition_to("FAILED")
                        raise RuntimeError(f"Snapshot missing for {rid}")

                    # 2. Verify Snapshot Integrity (Hash Check)
                    strategy_id = p_conf.get("Strategy", p_conf.get("strategy"))
                    source_path = PROJECT_ROOT / "strategies" / strategy_id / "strategy.py"
                    
                    if not source_path.exists():
                        mgr.transition_to("FAILED")
                        raise RuntimeError(f"Source strategy missing: {source_path}")

                    def get_file_hash(p):
                        return hashlib.sha256(p.read_bytes()).hexdigest()

                    if get_file_hash(snapshot_path) != get_file_hash(source_path):
                        mgr.transition_to("FAILED")
                        raise RuntimeError(f"Snapshot Integrity Mismatch! {rid}/strategy.py != strategies/{strategy_id}/strategy.py")
                    
                    print(f"[ORCHESTRATOR] Snapshot Verified: {rid} matches source.")
                    
                    # Phase 9: Log Snapshot
                    mgr._append_audit_log("SNAPSHOT_VERIFIED", {
                        "strategy_hash": get_file_hash(snapshot_path),
                        "source_hash": get_file_hash(source_path)
                    })
                    
                    # 3. Artifact Hash Binding (Phase 8)
                    bt_dir = PROJECT_ROOT / "backtests" / f"{clean_id}_{symbol}"
                    required_artifacts = {
                        "results_tradelevel.csv": bt_dir / "raw" / "results_tradelevel.csv",
                        "results_standard.csv": bt_dir / "raw" / "results_standard.csv",
                        "batch_summary.csv": PROJECT_ROOT / "backtests" / f"batch_summary_{clean_id}.csv"
                    }
                    
                    artifacts_manifest = {}
                    for name, path in required_artifacts.items():
                        if not path.exists():
                            mgr.transition_to("FAILED")
                            raise RuntimeError(f"Missing required artifact for binding: {path}")
                        artifacts_manifest[name] = get_file_hash(path)

                    manifest = {
                        "run_id": rid,
                        "strategy_hash": get_file_hash(snapshot_path),
                        "artifacts": artifacts_manifest,
                        "timestamp": datetime.now(timezone.utc).isoformat()
                    }

                    manifest_path = mgr.run_dir / "STRATEGY_SNAPSHOT.manifest.json"
                    with open(manifest_path, "w") as f:
                        json.dump(manifest, f, indent=4)

                    print(f"[ORCHESTRATOR] Manifest Bound: {manifest_path}")

                    # Phase 9: Log Artifact Binding
                    mgr._append_audit_log("ARTIFACT_BOUND", {
                        "manifest_path": str(manifest_path),
                        "artifact_hashes": artifacts_manifest
                    })

                    mgr.transition_to("STAGE_3A_COMPLETE")
                    mgr.transition_to("COMPLETE")
                    
                    # Phase 9: Log Completion
                    mgr._append_audit_log("RUN_COMPLETE", {"status": "SUCCESS"})
            
            # End of Symbol Runs
            dir_state_mgr.transition_to("SYMBOL_RUNS_COMPLETE")


        # Verify Manifests before Stage-4
        print("[ORCHESTRATOR] Verifying Artifact Integrity before Portfolio Evaluation...")
        
        for rid, symbol in zip(run_ids, symbols):
             mgr = PipelineStateManager(rid)
             manifest_path = mgr.run_dir / "STRATEGY_SNAPSHOT.manifest.json"
             
             if not manifest_path.exists():
                mgr.transition_to("FAILED")
                raise RuntimeError(f"Manifest missing for run {rid}")

             with open(manifest_path, "r") as f:
                manifest = json.load(f)

             # Reconstruct paths
             bt_dir = PROJECT_ROOT / "backtests" / f"{clean_id}_{symbol}"
             required_artifacts = {
                "results_tradelevel.csv": bt_dir / "raw" / "results_tradelevel.csv",
                "results_standard.csv": bt_dir / "raw" / "results_standard.csv",
                "batch_summary.csv": PROJECT_ROOT / "backtests" / f"batch_summary_{clean_id}.csv"
             }
             
             # Check Artifacts
             manifest_keys = set(manifest["artifacts"].keys())
             required_keys = set(required_artifacts.keys())
             
             if manifest_keys != required_keys:
                 mgr.transition_to("FAILED")
                 raise RuntimeError(f"Manifest Tampering Detected! Key mismatch for run {rid}. Expected: {required_keys}, Found: {manifest_keys}")

             for name, expected_hash in manifest["artifacts"].items():
                 target_path = required_artifacts[name]
                 if not target_path.exists():
                     mgr.transition_to("FAILED")
                     raise RuntimeError(f"Artifact missing during verification: {target_path}")
                     
                 # Re-compute hash
                 current_hash = hashlib.sha256(target_path.read_bytes()).hexdigest()
                 if current_hash != expected_hash:
                     mgr.transition_to("FAILED")
                     raise RuntimeError(f"Artifact Tampering Detected! {name} hash mismatch for run {rid}.")
             
             print(f"[ORCHESTRATOR] Verified Integrity: {rid}")

        # Stage 4 (Portfolio)
        run_command([PYTHON_EXE, "tools/portfolio_evaluator.py", clean_id], "Stage-4 Evaluation")
        
        # --- STAGE-4 ARTIFACT GATE (idempotent) ---
        portfolio_ledger_path = PROJECT_ROOT / "strategies" / "Master_Portfolio_Sheet.xlsx"
        if not portfolio_ledger_path.exists():
            print(f"[FATAL] Stage-4 ledger artifact missing: {portfolio_ledger_path}")
            for rid in run_ids:
                try: PipelineStateManager(rid).transition_to("FAILED")
                except Exception: pass
            dir_state_mgr.transition_to("FAILED")
            sys.exit(1)
            
        import openpyxl
        wb = openpyxl.load_workbook(portfolio_ledger_path, read_only=True)
        ws = wb.active
        
        try:
            headers = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            pid_idx = headers.index("portfolio_id")
            runs_idx = headers.index("constituent_run_ids")
        except Exception as e:
            print(f"[FATAL] Failed to resolve columns in Master Ledger: {e}")
            for rid in run_ids:
                try: PipelineStateManager(rid).transition_to("FAILED")
                except Exception: pass
            dir_state_mgr.transition_to("FAILED")
            sys.exit(1)
            
        # Find the appended row for this portfolio
        matching_rows = []
        row_count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(cell is not None for cell in row):
                row_count += 1
            if row and len(row) > max(pid_idx, runs_idx) and str(row[pid_idx]) == clean_id:
                matching_rows.append(row)
        wb.close()
        
        if row_count == 0:
            print(f"[FATAL] Stage-4 validation failed: {portfolio_ledger_path.name} is empty (0 data rows).")
            for rid in run_ids:
                try: PipelineStateManager(rid).transition_to("FAILED")
                except Exception: pass
            dir_state_mgr.transition_to("FAILED")
            sys.exit(1)
            
        if len(matching_rows) != 1:
            print(f"[FATAL] Stage-4 validation failed: Expected exactly 1 row for {clean_id} in Master Ledger, found {len(matching_rows)}")
            for rid in run_ids:
                try: PipelineStateManager(rid).transition_to("FAILED")
                except Exception: pass
            dir_state_mgr.transition_to("FAILED")
            sys.exit(1)
            
        portfolio_row = matching_rows[0]
            
        # Optional check: component cardinality (whitespace & null safe)
        raw_runs_str = str(portfolio_row[runs_idx]) if portfolio_row[runs_idx] is not None else ""
        saved_runs = [r.strip() for r in raw_runs_str.split(",") if r.strip()]
        
        if len(saved_runs) != len(symbols):
            print(f"[FATAL] Stage-4 validation failed: Expected {len(symbols)} constituent runs but found {len(saved_runs)}")
            for rid in run_ids:
                try: PipelineStateManager(rid).transition_to("FAILED")
                except Exception: pass
            dir_state_mgr.transition_to("FAILED")
            sys.exit(1)
            
        print(f"[GATE] Stage-4 artifact verified: {clean_id} present in Master Ledger with {len(saved_runs)} runs.")
        # --- END GATE ---
        
        # --- STAGE 5A & 5B: DETERMINISTIC REPORT GENERATION ---
        try:
            from tools.report_generator import generate_backtest_report, generate_strategy_portfolio_report
            backtest_root = PROJECT_ROOT / "backtests"
            strategy_id = p_conf.get("Strategy", p_conf.get("strategy"))
            print("[ORCHESTRATOR] Generating Deterministic Markdown Reports...")
            generate_backtest_report(clean_id, backtest_root)
            # Portfolio report: directory is under clean_id (directive name), not strategy_id
            generate_strategy_portfolio_report(clean_id, PROJECT_ROOT)
            if strategy_id and strategy_id != clean_id:
                generate_strategy_portfolio_report(strategy_id, PROJECT_ROOT)
        except Exception as rep_err:
            import traceback
            traceback.print_exc()
            print(f"[ERROR] REPORT_GENERATION_FAILURE: {rep_err}")
            print(f"[WARN] Non-authoritative step failed. Directive state unaffected.")
        # --- END REPORT GENERATION ---
        
        dir_state_mgr.transition_to("PORTFOLIO_COMPLETE")

        # --- STEP 8: CAPITAL WRAPPER (Deployable Artifact Emission) ---
        # Non-authoritative: failure does NOT invalidate PORTFOLIO_COMPLETE.
        # Classify as CAPITAL_WRAPPER_FAILURE and report — do not sys.exit.
        try:
            print("[ORCHESTRATOR] Running Step 8: Capital Wrapper...")
            run_command([PYTHON_EXE, "tools/capital_wrapper.py", clean_id], "Capital Wrapper")
            print("[ORCHESTRATOR] Step 8: Capital Wrapper COMPLETE.")
        except Exception as cw_err:
            print(f"[ERROR] CAPITAL_WRAPPER_FAILURE: {cw_err}")
            print("[WARN] Capital wrapper failed. Directive state is unaffected (PORTFOLIO_COMPLETE).")
            print("[WARN] Re-run manually: python tools/capital_wrapper.py " + clean_id)

        # --- STEP 9: DEPLOYABLE ARTIFACT VERIFICATION ---
        # Non-authoritative: verifies the capital wrapper outputs are structurally sound.
        # Reports DEPLOYABLE_INTEGRITY_FAILURE without invalidating directive state.
        try:
            print("[ORCHESTRATOR] Running Step 9: Deployable Artifact Verification...")
            deploy_root = PROJECT_ROOT / "strategies" / clean_id / "deployable"
            profiles = ["CONSERVATIVE_V1", "AGGRESSIVE_V1"]
            step9_failures = []

            for prof in profiles:
                d = deploy_root / prof
                if not d.exists():
                    step9_failures.append(f"  [{prof}] Profile directory missing: {d}")
                    continue

                # 1. All 5 artifacts present
                required_files = [
                    "equity_curve.csv",
                    "equity_curve.png",
                    "deployable_trade_log.csv",
                    "summary_metrics.json",
                ]
                for fname in required_files:
                    if not (d / fname).exists():
                        step9_failures.append(f"  [{prof}] Missing artifact: {fname}")

                # 2. summary_metrics.json equity math check
                metrics_path = d / "summary_metrics.json"
                if metrics_path.exists():
                    import json as _json
                    m = _json.loads(metrics_path.read_text(encoding="utf-8"))
                    diff = abs(m.get("final_equity", 0) - (m.get("starting_capital", 0) + m.get("realized_pnl", 0)))
                    if diff >= 0.01:
                        step9_failures.append(f"  [{prof}] Equity math mismatch: diff={diff:.4f}")
                    if m.get("final_equity", 0) <= 0:
                        step9_failures.append(f"  [{prof}] Final equity is zero or negative")

                # 3. equity_curve.csv has equity column with no negative values
                eq_path = d / "equity_curve.csv"
                if eq_path.exists():
                    import csv as _csv
                    with open(eq_path, newline="", encoding="utf-8") as cf:
                        reader = _csv.DictReader(cf)
                        for row_num, row in enumerate(reader, 1):
                            eq_val = float(row.get("equity", 1))
                            if eq_val <= 0:
                                step9_failures.append(f"  [{prof}] Negative equity at row {row_num}")
                                break

                # 4. Trade log row count matches summary_metrics total_accepted
                tl_path = d / "deployable_trade_log.csv"
                if tl_path.exists() and metrics_path.exists():
                    with open(tl_path, newline="", encoding="utf-8") as cf:
                        tl_rows = sum(1 for _ in cf) - 1  # subtract header
                    expected = m.get("total_accepted", -1)
                    if tl_rows != expected:
                        step9_failures.append(f"  [{prof}] Trade log count {tl_rows} != total_accepted {expected}")

                if not step9_failures:
                    print(f"[ORCHESTRATOR] Step 9 [{prof}]: All artifacts verified.")

            if step9_failures:
                print("[ERROR] DEPLOYABLE_INTEGRITY_FAILURE:")
                for line in step9_failures:
                    print(line)
                print("[WARN] Deployable verification failed. Directive state unaffected (PORTFOLIO_COMPLETE).")
            else:
                print("[ORCHESTRATOR] Step 9: Deployable Artifact Verification COMPLETE.")

        except Exception as dv_err:
            print(f"[ERROR] DEPLOYABLE_INTEGRITY_FAILURE: {dv_err}")
            print("[WARN] Deployable verification failed. Directive state unaffected (PORTFOLIO_COMPLETE).")
        # --- END STEPS 8 & 9 ---

    except Exception as e:
        print(f"[ORCHESTRATOR] Execution Failed: {e}")
        
        # Directive Commit FAILED
        try:
            dir_state_mgr.transition_to("FAILED")
        except Exception as e_cleanup:
            print(f"[WARN] Failed to mark directive as FAILED: {e_cleanup}")

        # Crash Handling: Mark all non-terminal runs as FAILED
        print("[ORCHESTRATOR] Performing fail-safe state cleanup...")
        for rid in run_ids:
            try:
                mgr = PipelineStateManager(rid)
                if mgr.state_file.exists():
                     with open(mgr.state_file, 'r') as f:
                         data = json.load(f)
                     if data['current_state'] != 'COMPLETE' and data['current_state'] != 'FAILED':
                         print(f"[CLEANUP] Marking run {rid} as FAILED")
                         mgr.transition_to("FAILED")
            except Exception as cleanup_err:
                print(f"[WARN] Failed to cleanup run {rid}: {cleanup_err}")
                
        sys.exit(1)


def run_batch_mode():
    """Sequential Batch Execution."""
    active_dir = PROJECT_ROOT / "backtest_directives" / "active"
    completed_dir = PROJECT_ROOT / "backtest_directives" / "completed"
    
    if not active_dir.exists():
        print(f"[BATCH] Active directory not found: {active_dir}")
        return

    directives = sorted(active_dir.glob("*.txt"))
    if not directives:
        print("[BATCH] No directives found in active/")
        return

    print(f"[BATCH] Found {len(directives)} directives: {[d.name for d in directives]}")
    completed_dir.mkdir(parents=True, exist_ok=True)

    try:
        for idx, d_path in enumerate(directives):
            d_name = d_path.name
            d_id = d_path.stem
            print(f"\n[BATCH] Processing Directive {idx+1}/{len(directives)}: {d_name}")
            try:
                run_single_directive(d_id)
                if "--provision-only" not in sys.argv:
                    final_dst = completed_dir / d_name
                    if final_dst.exists(): 
                        os.remove(final_dst)
                    shutil.move(str(d_path), str(final_dst))
                    print(f"[BATCH] Completed: {d_name} -> {completed_dir}")
                else:
                    print(f"[BATCH] Provision-only: {d_name} remains in active/")
            except Exception as e:
                print(f"[BATCH] FAILED: {d_name} - {e}")
                print(f"[FAIL-FAST] Stopping batch execution.")
                sys.exit(1)
        print("\n[BATCH] All directives processed successfully.")
    except Exception as e:
        print(f"\n[BATCH-ABORT] Batch aborted due to error: {e}")
        sys.exit(1)

def main():
    if len(sys.argv) < 2:
        print("Usage: python tools/run_pipeline.py <DIRECTIVE_ID> | --all")
        sys.exit(1)

    arg = sys.argv[1]

    if arg == "--all":
        run_batch_mode()
    else:
        directive_id = arg.replace(".txt", "")
        print(f"MASTER PIPELINE EXECUTION -- {directive_id}")
        run_single_directive(directive_id)
        print("\n[SUCCESS] Pipeline Completed Successfully.")

if __name__ == "__main__":
    main()
