"""
Governance-Authorized Basket Directive Reset Tool

Usage:
    python tools/basket_reset.py <DIRECTIVE_ID> --reason "<justification>"

Authority: AGENT.md -- Protected Infrastructure Policy
Purpose: Purge ALL artifacts associated with a basket directive so it can
be re-run cleanly through the canonical pipeline. Parallel to
`tools/reset_directive.py` for normal strategies.

Why a basket-specific tool exists:
    Basket directives use a different state machine than normal strategies.
    Their runs terminate at `BASKET_COMPLETE` status in the run_registry
    rather than transitioning the directive's `directive_state.json` to
    `PORTFOLIO_COMPLETE`. So `reset_directive.py` errors with "Only FAILED
    or PORTFOLIO_COMPLETE can be reset" on basket directives that are in
    `IDLE` state (the directive-state-machine sees them as never executed,
    while the pipeline's uniqueness guard sees them as already executed via
    the registry).

    This tool resolves that mismatch by treating the run_registry as the
    source of truth for basket completion and purging all associated
    artifacts atomically. Re-runs then go through the normal pipeline
    (`python tools/run_pipeline.py --all`).

What gets purged:
    1. run_registry entries (TradeScan_State/registry/run_registry.json)
       matching this directive_id
    2. Backtest artifacts (TradeScan_State/backtests/<id>_<basket_id>/)
    3. Vault artifacts (DRY_RUN_VAULT/baskets/<id>/)
    4. Run state folders (TradeScan_State/runs/<run_id>/ for each old run)
    5. MPS Baskets sheet row (Master_Portfolio_Sheet.xlsx) where
       directive_id matches
    6. basket_runs.csv row (TradeScan_State/research/basket_runs.csv)
       where directive_id matches
    7. Directive .txt: moved from completed/ (or active_backup/, active/)
       back to INBOX/ so the pipeline picks it up; any stale .admitted
       marker is removed

All actions are audited to governance/reset_audit_log.csv with a complete
list of purged artifact paths.

This tool MUST NOT be called by the agent autonomously. Operator
authorization is required per invocation. Each call requires --reason.
"""

from __future__ import annotations

import argparse
import csv
import json
import shutil
import sys
import yaml
from datetime import datetime, timezone
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

from config.path_authority import TRADE_SCAN_STATE, DRY_RUN_VAULT  # noqa: E402

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------

REGISTRY_PATH = TRADE_SCAN_STATE / "registry" / "run_registry.json"
BACKTESTS_DIR = TRADE_SCAN_STATE / "backtests"
VAULT_BASKETS_DIR = DRY_RUN_VAULT / "baskets"
RUNS_DIR = TRADE_SCAN_STATE / "runs"
MPS_PATH = TRADE_SCAN_STATE / "strategies" / "Master_Portfolio_Sheet.xlsx"
BASKET_RUNS_CSV = TRADE_SCAN_STATE / "research" / "basket_runs.csv"
AUDIT_LOG = PROJECT_ROOT / "governance" / "reset_audit_log.csv"

DIRECTIVE_SEARCH_DIRS = [
    PROJECT_ROOT / "backtest_directives" / "INBOX",
    PROJECT_ROOT / "backtest_directives" / "active",
    PROJECT_ROOT / "backtest_directives" / "active_backup",
    PROJECT_ROOT / "backtest_directives" / "completed",
]
INBOX_DIR = PROJECT_ROOT / "backtest_directives" / "INBOX"


# ---------------------------------------------------------------------------
# Validation
# ---------------------------------------------------------------------------


def _find_directive_file(directive_id: str) -> Path | None:
    """Find the .txt for this directive in any of the lifecycle folders."""
    for d in DIRECTIVE_SEARCH_DIRS:
        candidate = d / f"{directive_id}.txt"
        if candidate.exists():
            return candidate
    return None


def _is_basket_directive(directive_path: Path) -> tuple[bool, str | None]:
    """Returns (is_basket, basket_id). basket_id is None if not a basket."""
    try:
        data = yaml.safe_load(directive_path.read_text(encoding="utf-8"))
    except Exception as e:
        raise ValueError(f"Could not parse directive YAML: {e}")
    if not isinstance(data, dict):
        return (False, None)
    basket_cfg = data.get("basket")
    if not isinstance(basket_cfg, dict):
        return (False, None)
    basket_id = basket_cfg.get("basket_id")
    if not basket_id:
        return (False, None)
    return (True, str(basket_id))


# ---------------------------------------------------------------------------
# Purge operations — each returns a list of strings describing what it did
# ---------------------------------------------------------------------------


def _purge_registry(directive_id: str) -> tuple[list[str], list[str]]:
    """Remove run_registry entries matching directive_id.
    Returns (purged_run_ids, log_lines).
    """
    if not REGISTRY_PATH.exists():
        return ([], [f"registry not found at {REGISTRY_PATH} (skipped)"])
    reg = json.loads(REGISTRY_PATH.read_text(encoding="utf-8"))
    purged_run_ids = [
        rid for rid, entry in reg.items()
        if isinstance(entry, dict) and entry.get("directive_id") == directive_id
    ]
    if not purged_run_ids:
        return ([], [f"no registry entries found for {directive_id}"])
    # Backup before mutating
    backup = REGISTRY_PATH.with_suffix(".json.bak.basket_reset")
    shutil.copy2(REGISTRY_PATH, backup)
    for rid in purged_run_ids:
        del reg[rid]
    REGISTRY_PATH.write_text(json.dumps(reg, indent=2), encoding="utf-8")
    log = [
        f"purged {len(purged_run_ids)} registry entries: {', '.join(rid[:8] for rid in purged_run_ids)}",
        f"registry backup written to {backup.name}",
    ]
    return (purged_run_ids, log)


def _purge_backtest_dir(directive_id: str, basket_id: str) -> list[str]:
    """Delete TradeScan_State/backtests/<directive_id>_<basket_id>/."""
    target = BACKTESTS_DIR / f"{directive_id}_{basket_id}"
    if not target.exists():
        return [f"backtest dir not found: {target.name} (skipped)"]
    shutil.rmtree(target)
    return [f"purged backtest dir: {target}"]


def _purge_vault_dir(directive_id: str) -> list[str]:
    """Delete DRY_RUN_VAULT/baskets/<directive_id>/."""
    target = VAULT_BASKETS_DIR / directive_id
    if not target.exists():
        return [f"vault dir not found: {target} (skipped)"]
    shutil.rmtree(target)
    return [f"purged vault dir: {target}"]


def _purge_run_dirs(run_ids: list[str]) -> list[str]:
    """Delete TradeScan_State/runs/<run_id>/ for each run_id."""
    if not run_ids:
        return ["no run dirs to purge (no registry entries existed)"]
    log = []
    for rid in run_ids:
        target = RUNS_DIR / rid
        if not target.exists():
            log.append(f"run dir not found: {rid[:8]} (skipped)")
            continue
        shutil.rmtree(target)
        log.append(f"purged run dir: runs/{rid[:8]}...")
    return log


def _purge_mps_baskets_row(directive_id: str) -> list[str]:
    """Remove row from MPS Baskets sheet where directive_id matches."""
    if not MPS_PATH.exists():
        return [f"MPS not found at {MPS_PATH} (skipped)"]
    try:
        import pandas as pd
        from openpyxl import load_workbook
    except ImportError as e:
        return [f"required Excel libs not available ({e}); MPS Baskets row NOT purged"]
    try:
        # Read all sheets to preserve them
        wb = load_workbook(MPS_PATH)
        if "Baskets" not in wb.sheetnames:
            return ["MPS has no 'Baskets' sheet (skipped)"]
        # Read into DataFrame for filtering
        df = pd.read_excel(MPS_PATH, sheet_name="Baskets")
        if "directive_id" not in df.columns:
            return ["MPS Baskets has no 'directive_id' column (skipped)"]
        before = len(df)
        df = df[df["directive_id"] != directive_id]
        after = len(df)
        if before == after:
            return [f"no MPS Baskets row found for {directive_id} (skipped)"]
        # Write back via pandas + openpyxl preserving other sheets
        with pd.ExcelWriter(MPS_PATH, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            df.to_excel(writer, sheet_name="Baskets", index=False)
        return [f"purged {before - after} MPS Baskets row(s) for {directive_id}"]
    except PermissionError:
        raise PermissionError(
            f"Cannot write to {MPS_PATH} — close Excel if it's open and retry"
        )


def _purge_basket_runs_csv_row(directive_id: str) -> list[str]:
    """Remove row from basket_runs.csv where directive_id matches."""
    if not BASKET_RUNS_CSV.exists():
        return [f"basket_runs.csv not found at {BASKET_RUNS_CSV} (skipped)"]
    try:
        import pandas as pd
    except ImportError:
        return ["pandas not available; basket_runs.csv row NOT purged"]
    df = pd.read_csv(BASKET_RUNS_CSV)
    if "directive_id" not in df.columns:
        return [f"basket_runs.csv has no 'directive_id' column (skipped)"]
    before = len(df)
    df = df[df["directive_id"] != directive_id]
    after = len(df)
    if before == after:
        return [f"no basket_runs.csv row found for {directive_id} (skipped)"]
    df.to_csv(BASKET_RUNS_CSV, index=False)
    return [f"purged {before - after} basket_runs.csv row(s) for {directive_id}"]


def _restore_directive_to_inbox(directive_id: str) -> list[str]:
    """Move directive .txt back to INBOX/, removing any stale .admitted marker."""
    src = _find_directive_file(directive_id)
    if src is None:
        raise FileNotFoundError(
            f"Directive file {directive_id}.txt not found in any of: "
            f"{[str(d) for d in DIRECTIVE_SEARCH_DIRS]}"
        )
    INBOX_DIR.mkdir(parents=True, exist_ok=True)
    dest = INBOX_DIR / f"{directive_id}.txt"
    log = []
    if src.parent == INBOX_DIR:
        log.append(f"directive already in INBOX (no move needed)")
    else:
        if dest.exists():
            dest.unlink()
        shutil.move(str(src), str(dest))
        log.append(f"moved directive: {src.parent.name}/ -> INBOX/")
    # Remove stale .admitted marker if present in any lifecycle dir
    for d in DIRECTIVE_SEARCH_DIRS:
        marker = d / f"{directive_id}.txt.admitted"
        if marker.exists():
            marker.unlink()
            log.append(f"removed stale marker: {d.name}/{marker.name}")
    return log


def _write_audit(
    directive_id: str,
    previous_status: str,
    reason: str,
    log_lines: list[str],
) -> None:
    AUDIT_LOG.parent.mkdir(parents=True, exist_ok=True)
    write_header = not AUDIT_LOG.exists()
    now = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
    summary = " | ".join(log_lines)
    with open(AUDIT_LOG, "a", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        if write_header:
            writer.writerow(
                ["timestamp", "directive_id", "previous_state", "new_state", "reason"]
            )
        writer.writerow(
            [now, directive_id, previous_status, f"BASKET_RESET ({summary})", reason]
        )


# ---------------------------------------------------------------------------
# Main reset entry point
# ---------------------------------------------------------------------------


def basket_reset(directive_id: str, reason: str) -> dict:
    """Atomic purge of all basket artifacts for a directive.

    Returns a dict summarizing what was done. Idempotent: running twice
    on the same directive only does the file-move step the second time
    (everything else reports 'skipped').
    """
    # Find directive file
    directive_path = _find_directive_file(directive_id)
    if directive_path is None:
        raise FileNotFoundError(
            f"Directive file {directive_id}.txt not found in any lifecycle folder"
        )

    # Validate it's a basket directive
    is_basket, basket_id = _is_basket_directive(directive_path)
    if not is_basket:
        raise ValueError(
            f"{directive_id} is not a basket directive (no basket: section in YAML). "
            f"Use tools/reset_directive.py for normal strategies."
        )

    print(f"[BASKET_RESET] {directive_id} (basket_id={basket_id})")
    print(f"[BASKET_RESET] reason: {reason}")

    all_log: list[str] = []
    previous_status = "UNKNOWN"

    # 1. Purge registry — also captures previous status before purge
    if REGISTRY_PATH.exists():
        try:
            reg_peek = json.loads(REGISTRY_PATH.read_text(encoding="utf-8"))
            statuses = [
                entry.get("status")
                for entry in reg_peek.values()
                if isinstance(entry, dict) and entry.get("directive_id") == directive_id
            ]
            if statuses:
                previous_status = statuses[-1] or "UNKNOWN"
        except Exception:
            pass

    purged_run_ids, log = _purge_registry(directive_id)
    all_log.extend(log)
    for line in log:
        print(f"  - {line}")

    # 2. Purge backtest dir
    log = _purge_backtest_dir(directive_id, basket_id)
    all_log.extend(log)
    for line in log:
        print(f"  - {line}")

    # 3. Purge vault dir
    log = _purge_vault_dir(directive_id)
    all_log.extend(log)
    for line in log:
        print(f"  - {line}")

    # 4. Purge run dirs
    log = _purge_run_dirs(purged_run_ids)
    all_log.extend(log)
    for line in log:
        print(f"  - {line}")

    # 5. Purge MPS Baskets row
    log = _purge_mps_baskets_row(directive_id)
    all_log.extend(log)
    for line in log:
        print(f"  - {line}")

    # 6. Purge basket_runs.csv row
    log = _purge_basket_runs_csv_row(directive_id)
    all_log.extend(log)
    for line in log:
        print(f"  - {line}")

    # 7. Restore directive to INBOX
    log = _restore_directive_to_inbox(directive_id)
    all_log.extend(log)
    for line in log:
        print(f"  - {line}")

    # 8. Audit
    _write_audit(directive_id, previous_status, reason, all_log)
    print(f"[AUDIT] entry written to {AUDIT_LOG}")
    print(f"[DONE] {directive_id} ready for re-run via run_pipeline.py")

    return {
        "directive_id": directive_id,
        "basket_id": basket_id,
        "previous_status": previous_status,
        "purged_run_ids": purged_run_ids,
        "log": all_log,
    }


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------


def main():
    parser = argparse.ArgumentParser(
        description=(
            "Purge basket directive artifacts so it can be re-run cleanly "
            "through the canonical pipeline."
        )
    )
    parser.add_argument("directive_id", help="Basket directive id (e.g. 90_PORT_H2_5M_RECYCLE_S03_V1_P00)")
    parser.add_argument(
        "--reason",
        required=True,
        help="Governance reason for the reset (mandatory; logged to reset_audit_log.csv)",
    )
    args = parser.parse_args()

    try:
        basket_reset(args.directive_id, args.reason)
    except FileNotFoundError as e:
        print(f"[ERROR] {e}", file=sys.stderr)
        sys.exit(1)
    except ValueError as e:
        print(f"[ERROR] {e}", file=sys.stderr)
        sys.exit(1)
    except PermissionError as e:
        print(f"[ERROR] {e}", file=sys.stderr)
        sys.exit(2)
    except Exception as e:
        print(f"[ERROR] Unexpected: {e}", file=sys.stderr)
        sys.exit(3)


if __name__ == "__main__":
    main()
