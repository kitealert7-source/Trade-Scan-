"""
control_panel.py — CLI for recording portfolio control decisions.

Records user intent in portfolio_control table. Does NOT execute
promote/disable — that is the interpreter's job.

Usage:
  python tools/control_panel.py --list
  python tools/control_panel.py --status
  python tools/control_panel.py --select <portfolio_id> [--profile X]
  python tools/control_panel.py --burn <portfolio_id>
  python tools/control_panel.py --drop <portfolio_id> --reason "..."
  python tools/control_panel.py --deselect <portfolio_id>
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

from tools.ledger_db import (
    _connect, create_tables,
    upsert_portfolio_control, read_portfolio_control,
    update_control_status, delete_portfolio_control,
    log_control_action, read_control_log,
    read_mps, query_mps, read_master_filter,
    LEDGER_DB_PATH,
)

TS_EXEC_ROOT = PROJECT_ROOT.parent / "TS_Execution"
PORTFOLIO_YAML = TS_EXEC_ROOT / "portfolio.yaml"
CANDIDATES_PATH = PROJECT_ROOT.parent / "TradeScan_State" / "candidates" / "Filtered_Strategies_Passed.xlsx"
REGISTRY_PATH = TS_EXEC_ROOT / "burn_in_registry.yaml"


# ---------------------------------------------------------------------------
# Validation helpers
# ---------------------------------------------------------------------------

def _portfolio_exists_in_mps(portfolio_id: str) -> bool:
    """Check if portfolio_id exists in portfolio_sheet (MPS)."""
    df = read_mps()
    if df.empty:
        return False
    return portfolio_id in df["portfolio_id"].values


def _strategy_exists_in_master_filter(portfolio_id: str) -> bool:
    """Check if portfolio_id matches a strategy in master_filter (exact or prefix)."""
    df = read_master_filter()
    if df.empty or "strategy" not in df.columns:
        return False
    for s in df["strategy"].unique():
        if s == portfolio_id or s.startswith(portfolio_id + "_"):
            return True
    return False


# ---------------------------------------------------------------------------
# Commands
# ---------------------------------------------------------------------------

def cmd_list() -> int:
    """Print current control table."""
    df = read_portfolio_control()
    if df.empty:
        print("  portfolio_control: (empty)")
        return 0
    # Format for display
    display_cols = ["portfolio_id", "selected", "burn", "status", "profile", "reason", "last_updated"]
    cols = [c for c in display_cols if c in df.columns]
    print(df[cols].to_string(index=False))
    print(f"\n  Total: {len(df)} entries")
    return 0


def cmd_status() -> int:
    """Health check — compare all 4 stores."""
    import yaml

    errors = []

    # 1. portfolio_control
    df_ctrl = read_portfolio_control()
    ctrl_burnin = set(df_ctrl[df_ctrl["status"] == "BURN_IN"]["portfolio_id"])
    print(f"  portfolio_control:    {len(ctrl_burnin)} BURN_IN")

    # 2. portfolio.yaml
    yaml_ids = set()
    if PORTFOLIO_YAML.exists():
        try:
            with open(PORTFOLIO_YAML, encoding="utf-8") as f:
                data = yaml.safe_load(f)
            for s in (data.get("portfolio") or {}).get("strategies") or []:
                if s.get("enabled", True):
                    yaml_ids.add(s["id"])
        except Exception as e:
            errors.append(f"portfolio.yaml read error: {e}")
    print(f"  portfolio.yaml:       {len(yaml_ids)} entries")

    # 3. ledger.db IN_PORTFOLIO
    conn = _connect()
    db_count = conn.execute(
        'SELECT COUNT(*) FROM master_filter WHERE "IN_PORTFOLIO" = 1'
    ).fetchone()[0]
    conn.close()
    print(f"  ledger.db IN_PORTFOLIO: {db_count} run_ids")

    # 4. FSP BURN_IN
    fsp_burnin = set()
    if CANDIDATES_PATH.exists():
        try:
            from openpyxl import load_workbook
            wb = load_workbook(str(CANDIDATES_PATH), read_only=True, data_only=True)
            ws = wb["Sheet1"]
            header = [c.value for c in ws[1]]
            si = header.index("strategy")
            sti = header.index("candidate_status")
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[sti] == "BURN_IN" and row[si]:
                    fsp_burnin.add(row[si])
            wb.close()
        except Exception as e:
            errors.append(f"FSP read error: {e}")
    print(f"  FSP BURN_IN:          {len(fsp_burnin)} strategies")

    # 5. burn_in_registry
    reg_count = 0
    if REGISTRY_PATH.exists():
        try:
            with open(REGISTRY_PATH, encoding="utf-8") as f:
                reg = yaml.safe_load(f) or {}
            reg_count = len(reg.get("primary") or []) + len(reg.get("coverage") or [])
        except Exception as e:
            errors.append(f"Registry read error: {e}")
    print(f"  burn_in_registry:     {reg_count} entries")

    # Verdict
    print()
    if errors:
        for e in errors:
            print(f"  [ERROR] {e}")
        print("  VERDICT: ERRORS DETECTED")
        return 1

    # Check alignment: control BURN_IN count should match others
    all_match = (
        len(ctrl_burnin) == db_count == len(fsp_burnin) == reg_count
    )
    if all_match:
        print(f"  VERDICT: ALL ALIGNED ({len(ctrl_burnin)} BURN_IN)")
        return 0
    else:
        print("  VERDICT: DRIFT DETECTED")
        if len(ctrl_burnin) != db_count:
            print(f"    control ({len(ctrl_burnin)}) != DB IN_PORTFOLIO ({db_count})")
        if len(fsp_burnin) != reg_count:
            print(f"    FSP ({len(fsp_burnin)}) != registry ({reg_count})")
        if len(ctrl_burnin) != len(fsp_burnin):
            print(f"    control ({len(ctrl_burnin)}) != FSP ({len(fsp_burnin)})")
        return 1


def cmd_select(portfolio_id: str, profile: str) -> int:
    """Mark a portfolio for burn-in consideration."""
    # Validate: must exist in MPS or master_filter
    if not _portfolio_exists_in_mps(portfolio_id) and not _strategy_exists_in_master_filter(portfolio_id):
        print(f"  [ERROR] {portfolio_id} not found in MPS or master_filter.")
        print(f"  Run the pipeline for this strategy first.")
        return 1

    conn = _connect()
    create_tables(conn)
    # Check if already exists
    existing = conn.execute(
        "SELECT status FROM portfolio_control WHERE portfolio_id = ?",
        (portfolio_id,),
    ).fetchone()
    if existing:
        status = existing[0]
        if status == "BURN_IN":
            print(f"  [SKIP] {portfolio_id} is already BURN_IN.")
            conn.close()
            return 0
        if status == "SELECTED":
            print(f"  [SKIP] {portfolio_id} is already SELECTED.")
            conn.close()
            return 0

    upsert_portfolio_control(
        conn, portfolio_id,
        selected=1, burn=0,
        status="SELECTED",
        profile=profile,
        updated_by="user",
    )
    log_control_action(conn, portfolio_id, "select",
                       status_before=None, status_after="SELECTED",
                       detail=f"profile={profile}")
    conn.close()
    print(f"  [OK] {portfolio_id} -> SELECTED (profile={profile})")
    return 0


def cmd_burn(portfolio_id: str) -> int:
    """Mark a SELECTED portfolio for burn-in promotion."""
    conn = _connect()
    existing = conn.execute(
        "SELECT status, burn FROM portfolio_control WHERE portfolio_id = ?",
        (portfolio_id,),
    ).fetchone()

    if not existing:
        print(f"  [ERROR] {portfolio_id} not in portfolio_control. Run --select first.")
        conn.close()
        return 1

    status = existing[0]
    if status == "BURN_IN":
        print(f"  [SKIP] {portfolio_id} is already BURN_IN.")
        conn.close()
        return 0
    if status not in ("SELECTED", "RBIN"):
        print(f"  [ERROR] {portfolio_id} has status={status}. Must be SELECTED or RBIN.")
        conn.close()
        return 1

    upsert_portfolio_control(
        conn, portfolio_id,
        burn=1, selected=1,
        status=status,  # interpreter will transition to BURN_IN
        updated_by="user",
    )
    log_control_action(conn, portfolio_id, "burn",
                       status_before=status, status_after=status,
                       detail="burn=1, awaiting interpreter")
    conn.close()
    print(f"  [OK] {portfolio_id} -> burn=1 (awaiting interpreter)")
    return 0


def cmd_drop(portfolio_id: str, reason: str) -> int:
    """Mark a BURN_IN portfolio for removal."""
    if not reason:
        print("  [ERROR] --reason is required for --drop.")
        return 1

    conn = _connect()
    existing = conn.execute(
        "SELECT status FROM portfolio_control WHERE portfolio_id = ?",
        (portfolio_id,),
    ).fetchone()

    if not existing:
        print(f"  [ERROR] {portfolio_id} not in portfolio_control.")
        conn.close()
        return 1

    status = existing[0]
    if status == "RBIN":
        print(f"  [SKIP] {portfolio_id} is already RBIN.")
        conn.close()
        return 0
    if status != "BURN_IN":
        print(f"  [ERROR] {portfolio_id} has status={status}. Can only drop BURN_IN entries.")
        conn.close()
        return 1

    upsert_portfolio_control(
        conn, portfolio_id,
        burn=0,
        reason=reason,
        updated_by="user",
    )
    log_control_action(conn, portfolio_id, "drop",
                       status_before="BURN_IN", status_after="BURN_IN",
                       detail=f"burn=0, reason={reason}")
    conn.close()
    print(f"  [OK] {portfolio_id} -> burn=0 (awaiting interpreter)")
    print(f"  Reason: {reason}")
    return 0


def cmd_deselect(portfolio_id: str) -> int:
    """Remove from control table. Only if SELECTED (not promoted)."""
    conn = _connect()
    existing = conn.execute(
        "SELECT status FROM portfolio_control WHERE portfolio_id = ?",
        (portfolio_id,),
    ).fetchone()

    if not existing:
        print(f"  [SKIP] {portfolio_id} not in portfolio_control.")
        conn.close()
        return 0

    status = existing[0]
    if status != "SELECTED":
        print(f"  [ERROR] {portfolio_id} has status={status}. Can only deselect SELECTED entries.")
        conn.close()
        return 1

    log_control_action(conn, portfolio_id, "deselect",
                       status_before="SELECTED", status_after=None,
                       detail="removed from control table")
    delete_portfolio_control(conn, portfolio_id)
    conn.close()
    print(f"  [OK] {portfolio_id} removed from portfolio_control.")
    return 0


# ---------------------------------------------------------------------------
# Interactive menu
# ---------------------------------------------------------------------------

def _pick_from_list(items: list[str], prompt: str) -> str | None:
    """Show numbered list, return selected item or None on cancel."""
    if not items:
        print("  (none available)")
        return None
    for i, item in enumerate(items, 1):
        print(f"  {i}. {item}")
    print(f"  0. Cancel")
    try:
        choice = input(f"\n{prompt}: ").strip()
        if not choice or choice == "0":
            return None
        idx = int(choice) - 1
        if 0 <= idx < len(items):
            return items[idx]
        print("  Invalid choice.")
        return None
    except (ValueError, EOFError):
        return None


def _pick_from_grouped_list(
    portfolios: list[str], singles: list[str], prompt: str,
) -> str | None:
    """Show numbered list grouped by Portfolios / Single-Asset, return selected."""
    combined = []
    print()
    if portfolios:
        print(f"  --- Portfolios ({len(portfolios)}) ---")
        for pid in portfolios:
            combined.append(pid)
            print(f"  {len(combined):>3}. {pid}")
    if singles:
        print(f"\n  --- Single-Asset Composites ({len(singles)}) ---")
        for pid in singles:
            combined.append(pid)
            print(f"  {len(combined):>3}. {pid}")
    if not combined:
        print("  (none available)")
        return None
    print(f"    0. Cancel")
    try:
        choice = input(f"\n  {prompt}: ").strip()
        if not choice or choice == "0":
            return None
        idx = int(choice) - 1
        if 0 <= idx < len(combined):
            return combined[idx]
        print("  Invalid choice.")
        return None
    except (ValueError, EOFError):
        return None


def _confirm(action_desc: str) -> bool:
    """Prompt user for y/n confirmation. Returns True if confirmed."""
    try:
        ans = input(f"\n  Are you sure you want to {action_desc}? (y/n): ").strip().lower()
        return ans in ("y", "yes")
    except (EOFError, KeyboardInterrupt):
        return False


def _run_interpreter(dry_run: bool = False) -> None:
    """Run portfolio_interpreter as subprocess."""
    import subprocess as _sp
    cmd = [sys.executable, str(Path(__file__).parent / "portfolio_interpreter.py")]
    if dry_run:
        cmd.append("--dry-run")
    label = "dry-run" if dry_run else "applying"
    print(f"\n  Running interpreter ({label})...")
    _sp.run(cmd, cwd=str(PROJECT_ROOT))


def interactive_menu() -> int:
    """Interactive control panel — menu-driven, no arguments needed."""
    while True:
        print(f"\n{'=' * 60}")
        print("  PORTFOLIO CONTROL PANEL")
        print(f"{'=' * 60}")
        print("  1. Show portfolio")
        print("  2. System health")
        print("  3. Select strategy")
        print("  4. Promote to burn-in")
        print("  5. Drop from burn-in")
        print("  6. Remove selection")
        print("  7. Apply pending changes")
        print("  8. Dry-run changes")
        print("  9. View audit log")
        print("  0. Exit")
        print()

        try:
            choice = input("  Choose [0-9]: ").strip()
        except (EOFError, KeyboardInterrupt):
            print()
            return 0

        if choice == "0":
            return 0

        elif choice == "1":
            cmd_list()

        elif choice == "2":
            cmd_status()

        elif choice == "3":
            # Select — show available strategies grouped by tab
            print("\n  Loading available strategies from MPS...")
            df_ctrl = read_portfolio_control()
            existing = set(df_ctrl["portfolio_id"]) if not df_ctrl.empty else set()

            df_port = query_mps(sheet="Portfolios")
            df_single = query_mps(sheet="Single-Asset Composites")

            port_avail = sorted(
                pid for pid in (df_port["portfolio_id"].unique() if not df_port.empty else [])
                if pid not in existing
            )
            single_avail = sorted(
                pid for pid in (df_single["portfolio_id"].unique() if not df_single.empty else [])
                if pid not in existing
            )

            if not port_avail and not single_avail:
                print("  No new strategies available (all already in control table or MPS empty).")
                continue

            pid = _pick_from_grouped_list(port_avail, single_avail, "Select number")
            if not pid:
                continue

            profile = input("  Profile [CONSERVATIVE_V1]: ").strip() or "CONSERVATIVE_V1"
            cmd_select(pid, profile)

        elif choice == "4":
            # Promote — show SELECTED entries, confirm, set burn=1, auto-run interpreter
            df_ctrl = read_portfolio_control()
            candidates = []
            if not df_ctrl.empty:
                mask = df_ctrl["status"].isin(["SELECTED", "RBIN"])
                candidates = sorted(df_ctrl.loc[mask, "portfolio_id"].tolist())
            if not candidates:
                print("  No SELECTED entries to promote. Use option 3 first.")
                continue

            print(f"\n  SELECTED strategies:")
            pid = _pick_from_list(candidates, "Promote number")
            if not pid:
                continue
            if not _confirm(f"promote {pid}"):
                print("  Cancelled.")
                continue
            cmd_burn(pid)
            _run_interpreter()
            # Auto-refresh: show updated portfolio
            print()
            cmd_list()

        elif choice == "5":
            # Drop — show BURN_IN entries, confirm, set burn=0, auto-run interpreter
            df_ctrl = read_portfolio_control()
            candidates = []
            if not df_ctrl.empty:
                mask = df_ctrl["status"] == "BURN_IN"
                candidates = sorted(df_ctrl.loc[mask, "portfolio_id"].tolist())
            if not candidates:
                print("  No BURN_IN entries to drop.")
                continue

            print(f"\n  BURN_IN strategies:")
            pid = _pick_from_list(candidates, "Drop number")
            if not pid:
                continue
            reason = input("  Reason: ").strip()
            if not reason:
                print("  [ERROR] Reason is required.")
                continue
            if not _confirm(f"drop {pid}"):
                print("  Cancelled.")
                continue
            cmd_drop(pid, reason)
            _run_interpreter()
            # Auto-refresh: show updated portfolio
            print()
            cmd_list()

        elif choice == "6":
            # Deselect — show SELECTED entries
            df_ctrl = read_portfolio_control()
            candidates = []
            if not df_ctrl.empty:
                mask = df_ctrl["status"] == "SELECTED"
                candidates = sorted(df_ctrl.loc[mask, "portfolio_id"].tolist())
            if not candidates:
                print("  No SELECTED entries to deselect.")
                continue

            print(f"\n  SELECTED strategies:")
            pid = _pick_from_list(candidates, "Deselect number")
            if pid:
                cmd_deselect(pid)

        elif choice == "7":
            # Apply pending changes
            _run_interpreter()
            # Auto-refresh
            print()
            cmd_list()

        elif choice == "8":
            # Dry-run
            _run_interpreter(dry_run=True)

        elif choice == "9":
            # Audit log
            df = read_control_log(limit=20)
            if df.empty:
                print("  (no log entries)")
            else:
                display = ["id", "portfolio_id", "action", "status_before", "status_after", "detail", "timestamp"]
                cols = [c for c in display if c in df.columns]
                print(df[cols].to_string(index=False))

        else:
            print("  Invalid choice.")


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main() -> int:
    # If no arguments → interactive menu
    if len(sys.argv) == 1:
        return interactive_menu()

    parser = argparse.ArgumentParser(
        description="Portfolio Control Panel — record intent for promote/disable"
    )
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument("--list", action="store_true", help="Show control table")
    group.add_argument("--status", action="store_true", help="Health check across all stores")
    group.add_argument("--log", nargs="?", const="__all__", metavar="ID",
                        help="Show audit log (optionally for a specific portfolio_id)")
    group.add_argument("--select", metavar="ID", help="Select portfolio for burn-in consideration")
    group.add_argument("--burn", metavar="ID", help="Mark for burn-in promotion")
    group.add_argument("--drop", metavar="ID", help="Mark for removal from burn-in")
    group.add_argument("--deselect", metavar="ID", help="Remove from control table (SELECTED only)")

    parser.add_argument("--profile", default="CONSERVATIVE_V1",
                        help="Deployment profile (default: CONSERVATIVE_V1)")
    parser.add_argument("--reason", default="", help="Reason for drop (required with --drop)")

    args = parser.parse_args()

    if args.list:
        return cmd_list()
    elif args.status:
        return cmd_status()
    elif args.log is not None:
        pid = None if args.log == "__all__" else args.log
        df = read_control_log(portfolio_id=pid)
        if df.empty:
            print("  (no log entries)")
        else:
            display = ["id", "portfolio_id", "action", "status_before", "status_after", "detail", "timestamp"]
            cols = [c for c in display if c in df.columns]
            print(df[cols].to_string(index=False))
        return 0
    elif args.select:
        return cmd_select(args.select, args.profile)
    elif args.burn:
        return cmd_burn(args.burn)
    elif args.drop:
        return cmd_drop(args.drop, args.reason)
    elif args.deselect:
        return cmd_deselect(args.deselect)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
