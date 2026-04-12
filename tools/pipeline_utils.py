"""
pipeline_utils.py — Shared Logic for Trade_Scan Governance
Purpose: Centralize Run ID generation, Canonical Hashing, and State Management.
Authority: SOP_TESTING, SOP_AGENT_ENGINE_GOVERNANCE
"""

import os
import sys
import json
import hashlib
import time
import shutil
from pathlib import Path
from datetime import datetime, timezone
import importlib.util
from config.state_paths import RUNS_DIR, STRATEGIES_DIR
from config.status_enums import RUN_TERMINAL_STATES, RUN_ABORTED

import yaml
from yaml.loader import SafeLoader
from yaml.resolver import BaseResolver

# Backward compatibility: legacy strategy stubs imported FilterStack from
# tools.pipeline_utils. Canonical location is engines.filter_stack.
from engines.filter_stack import FilterStack

# ==============================================================================
# CONFIGURATION
# ==============================================================================

from dataclasses import dataclass, field

PROJECT_ROOT = Path(__file__).parent.parent
# RUNS_DIR imported from config.state_paths

# ==============================================================================
# PIPELINE CONTEXT
# ==============================================================================

@dataclass
class PipelineContext:
    """
    Unified execution context passed to StageRunner and all Stage units.
    Decouples stages from orchestrator globals.
    """
    directive_id: str
    directive_path: Path
    project_root: Path
    python_exe: str
    provision_only: bool = False
    
    # Mutable fields populated during Bootstrap or execution
    directive_config: dict = field(default_factory=dict)
    run_ids: list[str] = field(default_factory=list)
    symbols: list[str] = field(default_factory=list)
    planned_runs: list[dict] = field(default_factory=list)
    registry_path: Path | None = None
    current_state: str = "INITIALIZED"
    # State managers — injected by bootstrap so stages never instantiate directly
    directive_state_manager: object | None = None
    # Stage idempotency tracking — stages that have already completed in this run
    # Populated by StageRunner; persists across crash-restart via context reconstruction.
    completed_stages: set = field(default_factory=set)

    @staticmethod
    def from_directive_id(directive_id: str, active_dir: Path, project_root: Path, python_exe: str) -> 'PipelineContext':
        """Standard factory for creating a context before Bootstrap."""
        from tools.orchestration.pre_execution import find_directive_path
        d_path = find_directive_path(active_dir, directive_id)
        if not d_path:
            raise FileNotFoundError(f"Directive {directive_id} not found in {active_dir}")
        
        return PipelineContext(
            directive_id=directive_id,
            directive_path=d_path,
            project_root=project_root,
            python_exe=python_exe
        )

# ==============================================================================
# CANONICAL HASHING & RUN ID
# ==============================================================================



# ---------------------------------------------------------------------------
# Strict Duplicate-Key YAML Loader
# ---------------------------------------------------------------------------

class NoDuplicateSafeLoader(SafeLoader):
    """SafeLoader that raises ValueError on duplicate mapping keys."""
    pass


def _construct_mapping_no_duplicates(loader, node, deep=False):
    mapping = {}
    for key_node, value_node in node.value:
        key = loader.construct_object(key_node, deep=deep)
        if key in mapping:
            raise ValueError(
                f"DUPLICATE DIRECTIVE KEY DETECTED: '{key}'"
            )
        value = loader.construct_object(value_node, deep=deep)
        mapping[key] = value
    return mapping


NoDuplicateSafeLoader.add_constructor(
    BaseResolver.DEFAULT_MAPPING_TAG,
    _construct_mapping_no_duplicates,
)


def parse_directive(file_path: Path) -> dict:
    """
    Load a directive using YAML-safe parsing with strict duplicate-key detection.

    Contract:
    - Fails fast on invalid YAML, duplicate keys, or non-dict root.
    - Preserves full nested YAML structure.
    - Mirrors test: sub-keys into root for backward-compatible access.
      Raises ValueError on collision (never silently overwrites a root key).

    Returns:
        dict: Parsed directive config with test: keys hoisted to root level.
    """
    with open(file_path, "r", encoding="utf-8") as f:
        try:
            data = yaml.load(f, Loader=NoDuplicateSafeLoader)
        except ValueError:
            raise  # Re-raise duplicate key errors verbatim
        except yaml.YAMLError as e:
            raise ValueError(f"INVALID DIRECTIVE STRUCTURE: {e}")

    if not isinstance(data, dict):
        raise ValueError(
            "INVALID DIRECTIVE STRUCTURE: root element must be a YAML mapping"
        )

    # YAML auto-parses bare date literals (e.g. 2024-01-01) as datetime.date
    # objects instead of strings. This breaks json.dumps() in get_canonical_hash()
    # and would cause run ID drift. Convert recursively to ISO string.
    import datetime as _dt

    def _stringify_dates(obj):
        if isinstance(obj, (_dt.date, _dt.datetime)):
            return obj.isoformat()
        if isinstance(obj, dict):
            return {k: _stringify_dates(v) for k, v in obj.items()}
        if isinstance(obj, list):
            return [_stringify_dates(i) for i in obj]
        return obj

    data = _stringify_dates(data)

    # STRICT: test: wrapper is mandatory (flat directives no longer supported)
    if "test" not in data:
        raise ValueError(
            "INVALID DIRECTIVE STRUCTURE: 'test:' wrapper block is required. "
            "Flat directives are no longer supported."
        )

    # Mirror test: sub-keys into root for backward-compatible downstream access.
    # Collision detection: raises if a root key would be silently overwritten.
    test_block = data.get("test", {})
    if isinstance(test_block, dict):
        # Case-insensitive collision check to match downstream get_key_ci() reads.
        # Prevents ambiguous duplicates like 'Strategy' vs 'strategy' coexisting.
        existing_keys_lower = {ek.lower() for ek in data.keys()}
        for k, v in test_block.items():
            if k.lower() in existing_keys_lower:
                raise ValueError(
                    f"KEY COLLISION during test: merge: "
                    f"'{k}' conflicts with an existing root-level key (case-insensitive check)"
                )
            data[k] = v
            existing_keys_lower.add(k.lower())

    return data

def get_canonical_hash(parsed_data: dict) -> str:
    """Generate SHA256 hash of canonical JSON representation."""
    canonical_str = json.dumps(parsed_data, sort_keys=True, ensure_ascii=True)
    return hashlib.sha256(canonical_str.encode()).hexdigest()[:8]

def get_engine_version(engine_path=None):
    """
    Dynamically import engine module and read __version__.
    Default path: engine_dev/universal_research_engine/v1_5_4/main.py
    """
    if not engine_path:
        engine_path = PROJECT_ROOT / "engine_dev/universal_research_engine/v1_5_4/main.py"
        
    if not engine_path.exists():
        raise RuntimeError(f"Engine main.py not found at {engine_path}")

    spec = importlib.util.spec_from_file_location("universal_research_engine", engine_path)
    if spec is None or spec.loader is None:
        raise RuntimeError("Failed to load engine spec")
        
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)

    if not hasattr(module, "__version__"):
        raise RuntimeError("Engine module missing __version__ attribute")

    return module.__version__

def generate_run_id(directive_path: Path, symbol: str, attempt_id: str = "attempt_01") -> tuple[str, str]:
    """
    Generate Deterministic Run ID based on Governance Rules.
    Returns: (run_id, content_hash)
    """
    # Public parse for downstream field access (correct YAML structure).
    parsed_config = parse_directive(directive_path)

    # Resolve broker/timeframe for the lineage string.
    broker = parsed_config.get("Broker", parsed_config.get("broker", "OctaFx"))
    timeframe = parsed_config.get("Timeframe", parsed_config.get("timeframe", "1d"))

    # -------------------------------------------------------------------------
    # RUN ID HASH — computed from the LEGACY flat-parser output, frozen.
    #
    # \u26a0\ufe0f  DO NOT CHANGE THIS LOGIC. Changing it invalidates ALL existing run IDs,
    #     requiring a full re-run of every directive. The legacy flat parser is
    #     kept here as an internal private function purely for hash stability.
    #     It is NOT used for directive config access anywhere else.
    # -------------------------------------------------------------------------
    def _legacy_flat_parse(file_path: Path) -> dict:
        """Verbatim copy of the original flat key:value parser — hash use only."""
        with open(file_path, 'r', encoding='utf-8') as _f:
            lines = _f.readlines()
        _parsed = {}
        _current_key = None
        for _line in lines:
            _line = _line.strip()
            if not _line or _line.startswith("#"):
                continue
            if _line.startswith("-") and _current_key:
                _val = _line[1:].strip()
                if not isinstance(_parsed[_current_key], list):
                    _parsed[_current_key] = []
                _parsed[_current_key].append(_val)
                continue
            if ":" in _line:
                _parts = _line.split(":", 1)
                _key = _parts[0].strip()
                _val = _parts[1].strip()
                if not _val:
                    _parsed[_key] = []
                    _current_key = _key
                else:
                    _parsed[_key] = _val
                    _current_key = _key
            else:
                if _current_key and isinstance(_parsed.get(_current_key), list):
                    _parsed[_current_key].append(_line)
        return _parsed

    _legacy_config = _legacy_flat_parse(directive_path)
    _legacy_config.update({
        "BROKER": broker,
        "TIMEFRAME": timeframe,
        "START_DATE": _legacy_config.get("Start Date", _legacy_config.get("start_date", "2015-01-01")),
        "END_DATE": _legacy_config.get("End Date", _legacy_config.get("end_date", "2026-01-31")),
    })
    content_hash = get_canonical_hash(_legacy_config)
    # -------------------------------------------------------------------------

    engine_ver = get_engine_version()

    # test.name is mirrored to root by parse_directive(); may carry a run-context
    # suffix (e.g. __E152) that distinguishes re-runs under different engines.
    # Include it in the lineage string so suffix-tagged runs produce distinct run IDs.
    test_name = str(parsed_config.get("name", "")).strip()

    # Lineage String
    lineage_str = f"{content_hash}_{symbol}_{timeframe}_{broker}_{engine_ver}_{test_name}_{attempt_id}"
    run_id = hashlib.sha256(lineage_str.encode()).hexdigest()[:24]

    return run_id, content_hash


# ==============================================================================
# RUN ID LOOKUP (SHARED UTILITY)
# ==============================================================================

def find_run_id_for_directive(directive_id: str) -> str:
    """Find the run_id for a directive using a 3-level fallback chain.

    Resolution order:
      1. Scan ``runs/*/run_state.json`` for matching directive_id (original)
      2. Read ``runs/<directive_id>/directive_state.json`` → latest attempt ``run_ids``
      3. Read ``run_id`` column from Strategy_Master_Filter.xlsx

    Only returns run_ids whose folder still exists on disk.
    Returns run_id string, or empty string if not found.
    """
    if not RUNS_DIR.exists():
        return ""

    # --- Fallback 1 (primary): scan run_state.json files ---
    for d in sorted(RUNS_DIR.iterdir()):
        rs = d / "run_state.json"
        if rs.exists():
            try:
                data = json.loads(rs.read_text(encoding="utf-8"))
                if data.get("directive_id") == directive_id:
                    return data.get("run_id", d.name)
            except Exception:
                continue

    # --- Fallback 2: directive_state.json → run_ids array ---
    ds_file = RUNS_DIR / directive_id / "directive_state.json"
    if ds_file.exists():
        try:
            ds_data = json.loads(ds_file.read_text(encoding="utf-8"))
            latest = ds_data.get("latest_attempt", "attempt_01")
            attempt = ds_data.get("attempts", {}).get(latest, {})
            run_ids = attempt.get("run_ids", [])
            for rid in run_ids:
                if (RUNS_DIR / rid).exists():
                    return rid
            # If none on disk, return the first one anyway (caller decides)
            if run_ids:
                return run_ids[0]
        except Exception:
            pass

    # --- Fallback 3: Strategy_Master_Filter.xlsx → run_id column ---
    try:
        from config.state_paths import MASTER_FILTER_PATH
        if MASTER_FILTER_PATH.exists():
            import openpyxl
            wb = openpyxl.load_workbook(MASTER_FILTER_PATH, read_only=True, data_only=True)
            ws = wb.active
            headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
            rid_col = headers.index("run_id") if "run_id" in headers else None
            name_col = headers.index("strategy_name") if "strategy_name" in headers else None
            if rid_col is not None and name_col is not None:
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[name_col] == directive_id and row[rid_col]:
                        wb.close()
                        return str(row[rid_col])
            wb.close()
    except Exception:
        pass

    return ""


# ==============================================================================
# STATE MANAGEMENT (SINGLE WRITER, ATOMIC)
# ==============================================================================

class PipelineStateManager:
    """
    Manages the run_state.json file.
    Only the Orchestrator should use the write methods (initialize, transition).
    Stage scripts should use verify_state().
    """
    
    ALLOWED_TRANSITIONS = {
        "IDLE": ["PREFLIGHT_COMPLETE", "FAILED"],
        "PREFLIGHT_COMPLETE": ["PREFLIGHT_COMPLETE_SEMANTICALLY_VALID", "STAGE_1_COMPLETE", "FAILED"],
        "PREFLIGHT_COMPLETE_SEMANTICALLY_VALID": ["STAGE_1_COMPLETE", "FAILED", "ABORTED"],
        "STAGE_1_COMPLETE": ["STAGE_2_COMPLETE", "FAILED", "ABORTED"],
        "STAGE_2_COMPLETE": ["STAGE_3_COMPLETE", "FAILED", "ABORTED"],
        "STAGE_3_COMPLETE": ["STAGE_3A_COMPLETE", "FAILED", "ABORTED"],
        "STAGE_3A_COMPLETE": ["COMPLETE", "FAILED", "ABORTED"],
        "COMPLETE": [],
        "FAILED": [],
        "ABORTED": [],  # Terminal — watchdog / recovery only
    }

    def __init__(self, run_id: str, directive_id: str = None):
        self.run_id = run_id
        self.directive_id = directive_id
        self.run_dir = RUNS_DIR / run_id
        self.state_file = self.run_dir / "run_state.json"
        self.audit_log = self.run_dir / "audit.log"
        
    def _write_atomic(self, data: dict):
        """Atomic Write: Create temp, flush, rename."""
        self.run_dir.mkdir(parents=True, exist_ok=True)
        temp_file = self.state_file.with_suffix(".tmp")
        
        with open(temp_file, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=4)
            f.flush()
            os.fsync(f.fileno())
            
        shutil.move(str(temp_file), str(self.state_file))

    def initialize(self, metadata: dict | None = None):
        """Creates the run directory and initial state file with Audit Log.

        Args:
            metadata: Optional dict of non-authoritative run-context fields
                      (e.g. engine_version, engine_status, engine_model).
                      Written into the state file under the 'metadata' key.
                      Append-only; does not participate in governance gates.
        """
        # Idempotency guard: only skip reset when state is exactly
        # PREFLIGHT_COMPLETE_SEMANTICALLY_VALID \u2014 the state provision-only
        # leaves runs at. All other non-IDLE states fall through to the
        # existing reset-to-IDLE logic, preserving the re-init contract
        # tested by TestInitializeResetHistory.
        current = self.get_state_data().get("current_state")
        if current == "PREFLIGHT_COMPLETE_SEMANTICALLY_VALID":
            return  # Provision-only resume \u2014 do not reset

        self.run_dir.mkdir(parents=True, exist_ok=True)

        # 1. State File
        initial_data = {
            "run_id": self.run_id,
            "directive_id": self.directive_id,
            "current_state": "IDLE",
            "history": [],
            "last_updated": datetime.now(timezone.utc).isoformat()
        }
        if metadata:
            initial_data["metadata"] = metadata
        
        # Atomic Write (New Runs Only)
        if not self.state_file.exists():
            self._write_atomic(initial_data) # Use existing _write_atomic
        else:
             # Reset to IDLE if re-running (Governance allowed for same run_id/content)
             # This implies updating an existing state file, which _write_atomic does if given the full data.
             # To reset to IDLE, we need to load existing data and update it.
             try:
                 with open(self.state_file, 'r', encoding='utf-8') as f:
                     existing_data = json.load(f)
                 # Capture old_state BEFORE overwriting current_state so history
                 # records the true prior state (e.g. STAGE_3_COMPLETE \u2192 IDLE)
                 # not the post-mutation value (IDLE \u2192 IDLE).
                 old_state = existing_data.get("current_state", "UNKNOWN")
                 existing_data["history"].append({
                     "from": old_state,
                     "to": "IDLE",
                     "timestamp": datetime.now(timezone.utc).isoformat() + "Z"
                 })
                 existing_data["current_state"] = "IDLE"
                 existing_data["last_updated"] = datetime.now(timezone.utc).isoformat()
                 self._write_atomic(existing_data)
             except Exception as e:
                 print(f"[WARNING] Could not reset existing state file for {self.run_id}: {e}. Initializing as new.")
                 self._write_atomic(initial_data)


        # 2. Audit Log (Phase 9)
        if not self.audit_log.exists():
            self._append_audit_log("RUN_INITIALIZED", {"initial_state": "IDLE"})

    def _append_audit_log(self, event_type: str, details: dict):
        """Phase 9: Append-Only Immutable Audit Log."""
        entry = {
            "timestamp": datetime.now(timezone.utc).isoformat(),
            "event": event_type,
            "run_id": self.run_id,
            **details
        }
        # Ensure run directory exists for audit log
        self.run_dir.mkdir(parents=True, exist_ok=True)
        # strict append mode
        with open(self.audit_log, "a", encoding="utf-8") as f:
            f.write(json.dumps(entry) + "\n")

    def transition_to(self, new_state: str):
        """
        Updates the state machine to new_state if the transition is valid.
        Records history and updates timestamp.
        """
        # Guard: if the run directory or state file doesn't exist, the run was
        # deleted or never fully initialized. Log a warning and skip instead of
        # crashing the orchestrator with an unhandled FileNotFoundError.
        if not self.run_dir.exists() or not self.state_file.exists():
            print(
                f"[WARN] Run state missing for {self.run_id} — "
                "run directory or state file not found. Skipping transition to avoid orchestrator crash."
            )
            self._append_audit_log("MISSING_STATE_SKIP", {
                "attempted_transition": new_state,
                "reason": "run_state.json not found"
            })
            return
             
        with open(self.state_file, 'r', encoding='utf-8') as f:
            data = json.load(f)
            
        old_state = data["current_state"]
        
        # STRICT ENFORCEMENT
        allowed = self.ALLOWED_TRANSITIONS.get(old_state, [])
        if new_state not in allowed:
            # Log failure attempt?
            self._append_audit_log("ILLEGAL_TRANSITION_ATTEMPT", {
                "from": old_state,
                "to": new_state, 
                "allowed": allowed
            })
            raise RuntimeError(f"[FATAL] Illegal State Transition: {old_state} -> {new_state}. Allowed: {allowed}")
            
        # Update
        data["history"].append({
            "from": old_state,
            "to": new_state,
            "timestamp": datetime.now(timezone.utc).isoformat()
        })
        data["current_state"] = new_state
        data["last_transition"] = datetime.now(timezone.utc).isoformat()
        
        self._write_atomic(data)
        
        # Phase 9: Audit Log
        self._append_audit_log("STATE_TRANSITION", {
            "from": old_state,
            "to": new_state
        })
        print(f"[STATE] Transition {self.run_id}: {old_state} -> {new_state}")

    def abort(self, reason: str = "WATCHDOG_TIMEOUT") -> bool:
        """Transition to ABORTED from any in-progress state.

        Unlike ``transition_to()``, this accepts a ``reason`` string that is
        persisted alongside the state for post-mortem diagnostics.

        Returns True if the transition succeeded, False if the run was already
        in a terminal state (COMPLETE, FAILED, ABORTED, IDLE) and was skipped.
        """
        if not self.state_file.exists():
            return False

        with open(self.state_file, 'r', encoding='utf-8') as f:
            data = json.load(f)

        old_state = data.get("current_state", "UNKNOWN")
        # Only abort from in-progress states
        terminal = RUN_TERMINAL_STATES | {"IDLE"}
        if old_state in terminal:
            return False

        allowed = self.ALLOWED_TRANSITIONS.get(old_state, [])
        if "ABORTED" not in allowed:
            self._append_audit_log("ABORT_SKIPPED", {
                "from": old_state,
                "abort_reason": reason,
                "detail": "ABORTED not in allowed transitions for this state",
            })
            return False

        data["history"].append({
            "from": old_state,
            "to": "ABORTED",
            "timestamp": datetime.now(timezone.utc).isoformat() + "Z",
        })
        data["current_state"] = "ABORTED"
        data["previous_state"] = old_state
        data["abort_reason"] = reason
        data["last_transition"] = datetime.now(timezone.utc).isoformat() + "Z"

        self._write_atomic(data)
        self._append_audit_log("STATE_TRANSITION", {
            "from": old_state,
            "to": "ABORTED",
            "abort_reason": reason,
        })
        print(f"[STATE] Abort {self.run_id}: {old_state} -> ABORTED (reason={reason})")
        return True

    def record_heartbeat(self):
        """Update heartbeat_ts to signify active RUNNING loop safely."""
        if not self.state_file.exists():
            return
            
        with open(self.state_file, 'r', encoding='utf-8') as f:
            data = json.load(f)
            
        # Only inject heartbeat for currently running tasks.
        if data.get("current_state") == "STAGE_1_COMPLETE":
             pass # Allowed to ping during transitions
             
        data["heartbeat_ts"] = datetime.now(timezone.utc).timestamp()
        
        # Immediate sync
        temp_file = self.state_file.with_suffix(".tmp")
        with open(temp_file, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=4)
        shutil.move(str(temp_file), str(self.state_file))
    def verify_state(self, expected_state: str):
        """
        Verify current state matches expected_state.
        Raises RuntimeError on mismatch or missing state file.
        Caller (run_stage1.py) handles failures via its existing try/except.
        """
        if not self.state_file.exists():
            raise RuntimeError(
                f"[FATAL] State file missing: {self.state_file}"
            )

        try:
            with open(self.state_file, "r", encoding="utf-8") as f:
                data = json.load(f)
        except Exception as e:
            raise RuntimeError(f"[FATAL] Corrupt state file: {e}")

        current = data.get("current_state")

        # Identity Check
        if data.get("run_id") != self.run_id:
            raise RuntimeError(
                f"[FATAL] Run Identity Mismatch. File: {data.get('run_id')} vs Arg: {self.run_id}"
            )

        if current != expected_state:
            # Allow forward states for re-running reports
            if (expected_state == "STAGE_2_COMPLETE" and current in ["STAGE_3_COMPLETE", "STAGE_3A_COMPLETE", "COMPLETE"]) or \
               (expected_state == "STAGE_1_COMPLETE" and current in ["STAGE_2_COMPLETE", "STAGE_3_COMPLETE", "STAGE_3A_COMPLETE", "COMPLETE"]):
                pass
            else:
                raise RuntimeError(
                    f"[FATAL] State Mismatch. Expected: {expected_state}, Found: {current}"
                )

        # print(f"[VERIFIED] State is {current}")

    def get_state_data(self) -> dict:
        """Reads and returns the full state data."""
        if not self.state_file.exists():
            return {"current_state": "IDLE"}
        
        try:
            with open(self.state_file, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception:
             return {"current_state": "IDLE"}

class DirectiveStateManager:
    """
    Manages the high-level lifecycle of a directive execution batch.
    Persists state to runs/<DIRECTIVE_ID>/directive_state.json.
    """
    
    ALLOWED_TRANSITIONS = {
        "INITIALIZED": ["PREFLIGHT_COMPLETE", "FAILED"],
        "PREFLIGHT_COMPLETE": ["PREFLIGHT_COMPLETE_SEMANTICALLY_VALID", "FAILED"],
        "PREFLIGHT_COMPLETE_SEMANTICALLY_VALID": ["SYMBOL_RUNS_COMPLETE", "FAILED"],
        "SYMBOL_RUNS_COMPLETE": ["PORTFOLIO_COMPLETE", "FAILED"],
        "PORTFOLIO_COMPLETE": ["FAILED"], # In case of post-completion failure/invalidation?
        "FAILED": ["INITIALIZED", "SYMBOL_RUNS_COMPLETE"] # Allow retry/reset with archival
    }

    def __init__(self, directive_id: str):
        self.directive_id = directive_id
        self.directive_dir = RUNS_DIR / directive_id
        self.state_file = self.directive_dir / "directive_state.json"
        self.audit_log = self.directive_dir / "directive_audit.log"

    def _write_atomic(self, data: dict):
        """Atomic Write with fsync for durability."""
        self.directive_dir.mkdir(parents=True, exist_ok=True)
        temp_file = self.state_file.with_suffix(".tmp")
        
        with open(temp_file, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=4)
            f.flush()
            os.fsync(f.fileno())
            
        shutil.move(str(temp_file), str(self.state_file))

    def _archive_current_state(self):
        """Archives current state and log before reset."""
        timestamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%S")
        
        # Archive State
        if self.state_file.exists():
            archive_state = self.directive_dir / f"directive_state.json.bak.{timestamp}"
            shutil.move(str(self.state_file), str(archive_state))
            print(f"[ARCHIVE] State moved to {archive_state.name}")

        # Archive Log
        if self.audit_log.exists():
            archive_log = self.directive_dir / f"directive_audit.log.bak.{timestamp}"
            shutil.move(str(self.audit_log), str(archive_log))
            print(f"[ARCHIVE] Log moved to {archive_log.name}")

    def _get_next_attempt_id(self, data: dict) -> str:
        attempts = data.get("attempts", {})
        if not attempts:
            return "attempt_01"
        attempt_number = len(attempts) + 1
        return f"attempt_{attempt_number:02d}"

    def initialize(self):
        """Creates directory and initializes state using attempt hierarchy."""
        self.directive_dir.mkdir(parents=True, exist_ok=True)
        
        initial_data = {
            "directive_id": self.directive_id,
            "latest_attempt": "attempt_01",
            "attempts": {
                "attempt_01": {
                    "status": "INITIALIZED",
                    "history": ["INITIALIZED"]
                }
            },
            "last_updated": datetime.now(timezone.utc).isoformat()
        }
        
        # Init State File
        if not self.state_file.exists():
            self._write_atomic(initial_data)
            self._append_audit_log("DIRECTIVE_INITIALIZED", {"attempt": "attempt_01"})
        else:
            # Check if we should reset? pipeline triggers initialize() at start.
            # If it exists, we might be resuming. 
            pass

    def create_new_attempt(self):
        """Rotates the FSM to a new attempt, preserving history."""
        if not self.state_file.exists():
            self.initialize()
            return
            
        with open(self.state_file, 'r', encoding='utf-8') as f:
            data = json.load(f)
            
        next_attempt = self._get_next_attempt_id(data)
        
        # Reset global state for the pipeline orchestrator loop
        data["latest_attempt"] = next_attempt
        if "attempts" not in data:
            data["attempts"] = {}
            
        data["attempts"][next_attempt] = {
            "status": "INITIALIZED",
            "history": ["INITIALIZED"]
        }
        
        # Purge legacy flat state fields if present
        if "current_state" in data:
            del data["current_state"]
        if "history" in data:
            del data["history"]
            
        data["last_updated"] = datetime.now(timezone.utc).isoformat()
        self._write_atomic(data)
        
        self._append_audit_log("NEW_ATTEMPT_CREATED", {"attempt": next_attempt})
        print(f"[ATTEMPT] Provisioning {next_attempt} for directive {self.directive_id}")

    def get_state(self) -> str:
        if not self.state_file.exists():
            return "IDLE"
        try:
            with open(self.state_file, 'r', encoding='utf-8') as f:
                data = json.load(f)
                latest_attempt = data.get("latest_attempt", "attempt_01")
                attempts = data.get("attempts", {})
                if latest_attempt in attempts:
                    return attempts[latest_attempt].get("status", "IDLE")
                # Fallback purely for safety 
                return data.get("current_state", "IDLE")
        except Exception:
            return "IDLE"
            
    def get_latest_attempt(self) -> str:
        if not self.state_file.exists():
            return "attempt_01"
        try:
            with open(self.state_file, 'r', encoding='utf-8') as f:
                return json.load(f).get("latest_attempt", "attempt_01")
        except Exception:
            return "attempt_01"

    def register_run_ids(self, run_ids: list[str]):
        """Registers generated run IDs onto the latest attempt payload."""
        if not self.state_file.exists():
            return
        with open(self.state_file, 'r', encoding='utf-8') as f:
            data = json.load(f)
        latest_attempt = data.get("latest_attempt", "attempt_01")
        attempts = data.setdefault("attempts", {})
        current_attempt = attempts.setdefault(latest_attempt, {"status": "INITIALIZED", "history": ["INITIALIZED"]})
        
        current_attempt["run_ids"] = run_ids
        if len(run_ids) == 1:
            current_attempt["run_id"] = run_ids[0]
            
        data["last_updated"] = datetime.now(timezone.utc).isoformat()
        self._write_atomic(data)

    def transition_to(self, new_state: str):
        """Transitions directive state with strict validation and logging."""
        
        # Special Handling: Reset from FAILED -> INITIALIZED
        if new_state == "INITIALIZED":
            current_state = self.get_state()
            if current_state == "FAILED":
                print(f"[RESET] Hard Reset triggered for {self.directive_id}")
                self._archive_current_state()
                self.initialize() # Re-create fresh
                return
            elif current_state == "IDLE":
                self.initialize()
                return

        if not self.state_file.exists():
            raise FileNotFoundError(f"Directive state not found: {self.directive_id}")

        with open(self.state_file, 'r', encoding='utf-8') as f:
            data = json.load(f)
            
        latest_attempt = data.get("latest_attempt", "attempt_01")
        attempts = data.setdefault("attempts", {})
        current_attempt = attempts.setdefault(latest_attempt, {"status": "INITIALIZED", "history": ["INITIALIZED"]})
        
        old_state = current_attempt.get("status", "INITIALIZED")
        
        # Validate
        allowed = self.ALLOWED_TRANSITIONS.get(old_state, [])
        if new_state not in allowed:
            self._append_audit_log("ILLEGAL_TRANSITION_ATTEMPT", {
                "from": old_state, "to": new_state
            })
            raise RuntimeError(f"Illegal Directive Transition: {old_state} -> {new_state}")
            
        # Clean up legacy fields
        if "current_state" in data:
            del data["current_state"]
        if "history" in data:
            del data["history"]
            
        # Update
        current_attempt["status"] = new_state
        if "history" not in current_attempt or not current_attempt["history"]:
            current_attempt["history"] = ["INITIALIZED"]
        elif current_attempt["history"][0] != "INITIALIZED":
            current_attempt["history"].insert(0, "INITIALIZED")

        if current_attempt["history"][-1] != new_state:
            current_attempt["history"].append(new_state)

        # Set protection flag and verify artifacts when reaching PORTFOLIO_COMPLETE
        if new_state == "PORTFOLIO_COMPLETE":
            data["protected"] = True
            run_ids = current_attempt.get("run_ids", [])
            self._verify_completion_artifacts(run_ids)

        data["last_updated"] = datetime.now(timezone.utc).isoformat()
        
        # Atomic Write
        self._write_atomic(data)
        
        # Log
        self._append_audit_log("STATE_TRANSITION", {
            "from": old_state, "to": new_state, "attempt": latest_attempt
        })
        print(f"[DIRECTIVE] Transition {self.directive_id} ({latest_attempt}): {old_state} -> {new_state}")

    def _verify_completion_artifacts(self, run_ids: list[str]) -> list[str]:
        """Non-blocking verification of artifacts at PORTFOLIO_COMPLETE.

        Returns a list of warning strings (empty if all OK).
        """
        warnings = []

        # 1. Check run folders and snapshots
        for rid in run_ids:
            run_dir = RUNS_DIR / rid
            if not run_dir.exists():
                warnings.append(f"run folder missing: {rid}")
                continue
            if not (run_dir / "strategy.py").exists():
                warnings.append(f"strategy.py snapshot missing: {rid}")
            if not (run_dir / "data").exists():
                warnings.append(f"data/ folder missing: {rid}")

        # 2. Check strategy.py authority copy
        strat_dir = PROJECT_ROOT / "strategies" / self.directive_id
        if not (strat_dir / "strategy.py").exists():
            warnings.append(f"authority strategy.py missing: strategies/{self.directive_id}/strategy.py")

        # 3. Check deployable folder
        deploy_dir = STRATEGIES_DIR / self.directive_id / "deployable"
        if not deploy_dir.exists() or not any(deploy_dir.iterdir()):
            warnings.append(f"deployable/ missing or empty: {self.directive_id}")

        # Log results
        if warnings:
            for w in warnings:
                print(f"[VERIFY-WARN] {w}")
            self._append_audit_log("COMPLETION_VERIFICATION", {
                "status": "WARNINGS",
                "warnings": warnings,
            })
        else:
            print(f"[VERIFY] All completion artifacts verified for {self.directive_id}")
            self._append_audit_log("COMPLETION_VERIFICATION", {
                "status": "OK",
            })

        return warnings

    def _append_audit_log(self, event: str, details: dict):
        entry = {
            "timestamp": datetime.now(timezone.utc).isoformat(),
            "event": event,
            "directive_id": self.directive_id,
            **details
        }
        with open(self.audit_log, 'a', encoding='utf-8') as f:
            f.write(json.dumps(entry) + "\n")


# ==============================================================================
# EXCEL FILE UTILITY
# ==============================================================================

def ensure_xlsx_writable(path: Path, timeout: int = 15) -> None:
    """
    On Windows, detect and kill Excel if it holds a lock on the xlsx file.

    Excel opens .xlsx files with an exclusive write lock. Python's to_excel()
    raises PermissionError in that case, even if a FileLock is held.

    This function:
      1. Probes whether the file is writable (tries to open in r+b mode).
      2. If PermissionError → kills all EXCEL.EXE processes via taskkill.
      3. Waits up to `timeout` seconds for the file to become writable.
      4. Raises RuntimeError if still locked after timeout.

    No-op on non-Windows platforms or when the file does not yet exist.
    """
    if os.name != "nt":
        return
    path = Path(path)
    if not path.exists():
        return

    def _is_locked() -> bool:
        try:
            with open(path, "r+b"):
                return False
        except PermissionError:
            return True

    if not _is_locked():
        return

    print(f"[XLSX] '{path.name}' is open in Excel — closing Excel...")
    import subprocess as _sp
    _sp.run(["taskkill", "/f", "/im", "EXCEL.EXE"], capture_output=True)

    deadline = time.time() + timeout
    while time.time() < deadline:
        time.sleep(0.5)
        if not _is_locked():
            print(f"[XLSX] '{path.name}' released.")
            return

    raise RuntimeError(
        f"XLSX_LOCK_TIMEOUT: '{path.name}' still locked after {timeout}s. "
        "Close Excel manually and retry."
    )
