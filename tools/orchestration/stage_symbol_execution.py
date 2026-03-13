"""Symbol execution stages (stage-0.9 through stage-3A)."""

from __future__ import annotations

import hashlib
import json
import shutil
import importlib
import pandas as pd
from datetime import datetime, timezone
from pathlib import Path

from tools.orchestration.pipeline_errors import PipelineExecutionError
from tools.orchestration.run_registry import (
    claim_next_planned_run,
    ensure_registry,
    list_runs,
    requeue_running_runs,
    update_run_state,
)
from tools.orchestration.transition_service import (
    transition_directive_state,
    transition_run_state,
    transition_run_state_sequence,
)
from tools.pipeline_utils import PipelineStateManager
from tools.system_registry import log_run_to_registry
from config.state_paths import RUNS_DIR, BACKTESTS_DIR, STRATEGIES_DIR, MASTER_FILTER_PATH
from config.engine_loader import get_active_engine


def run_symbol_execution_stages(
    *,
    clean_id: str,
    p_conf: dict,
    run_ids: list[str],
    symbols: list[str],
    project_root: Path,
    python_exe: str,
    run_command,
    registry_path: Path | None = None,
) -> None:
    """Execute stage-0.9 through stage-3A and close symbol runs."""
    strategy_id = p_conf.get("Strategy", p_conf.get("strategy")) or clean_id
    if registry_path is None:
        registry_path = RUNS_DIR / clean_id / "run_registry.json"

    if not registry_path.exists():
        ensure_registry(
            registry_path,
            clean_id,
            [
                {"run_id": rid, "strategy": strategy_id, "symbol": sym}
                for rid, sym in zip(run_ids, symbols)
            ],
        )

    reclaimed = requeue_running_runs(registry_path, clean_id)
    if reclaimed:
        print(f"[ORCHESTRATOR] Re-queued {reclaimed} interrupted RUNNING registry entries to PLANNED.")

    registry_runs = list_runs(registry_path, clean_id)
    if not registry_runs:
        raise PipelineExecutionError(
            f"Run registry has no planned runs: {registry_path}",
            directive_id=clean_id,
            run_ids=run_ids,
        )
    run_ids = [run["run_id"] for run in registry_runs]
    symbols = [run["symbol"] for run in registry_runs]

    any_stage1_rerun = any(
        PipelineStateManager(rid).get_state_data()["current_state"]
        in ("IDLE", "PREFLIGHT_COMPLETE", "PREFLIGHT_COMPLETE_SEMANTICALLY_VALID")
        for rid in run_ids
    )
    summary_csv = BACKTESTS_DIR / f"batch_summary_{clean_id}.csv"
    if summary_csv.exists() and any_stage1_rerun:
        summary_csv.unlink()

    if strategy_id:
        source_strategy_path = project_root / "strategies" / strategy_id / "strategy.py"
        if source_strategy_path.exists():
            print("[ORCHESTRATOR] Performing atomic preemptive strategy snapshots...")
            for rid in run_ids:
                mgr = PipelineStateManager(rid)
                target_path = mgr.run_dir / "strategy.py"
                if not target_path.exists():
                    target_path.parent.mkdir(parents=True, exist_ok=True)
                    shutil.copy(str(source_strategy_path), str(target_path))

    print("[ORCHESTRATOR] Launching Stage-1 Generator (Registry Worker)...")
    skip_states = {
        "STAGE_1_COMPLETE",
        "STAGE_2_COMPLETE",
        "STAGE_3_COMPLETE",
        "STAGE_3A_COMPLETE",
        "COMPLETE",
    }
    while True:
        claim = claim_next_planned_run(registry_path, clean_id)
        if claim is None:
            break

        rid = claim["run_id"]
        symbol = claim["symbol"]
        mgr = PipelineStateManager(rid)
        current = mgr.get_state_data()["current_state"]

        if current in skip_states:
            update_run_state(registry_path, clean_id, rid, "COMPLETE")
            continue

        if current == "FAILED":
            update_run_state(
                registry_path,
                clean_id,
                rid,
                "FAILED",
                last_error="Run state already FAILED before Stage-1 execution.",
            )
            raise RuntimeError(f"Run {rid} is already FAILED before Stage-1.")

        try:
            from tools.skill_loader import run_skill

            run_skill("backtest_execution", strategy=clean_id, symbol=symbol, run_id=rid)

            out_folder = BACKTESTS_DIR / f"{clean_id}_{symbol}"
            if not (out_folder / "raw" / "results_tradelevel.csv").exists():
                raise RuntimeError(f"[FATAL] Stage-1 artifact missing for {symbol}. (Probable NO_TRADES).")

            transition_run_state(rid, "STAGE_1_COMPLETE")
            update_run_state(registry_path, clean_id, rid, "COMPLETE")
        except Exception as err:
            print(f"[ERROR] Stage-1 Failed for {symbol}: {err}")
            try:
                transition_run_state(rid, "FAILED")
            except Exception as cleanup_err:
                print(f"[WARN] Failed to mark {symbol} run as FAILED: {cleanup_err}")
            try:
                update_run_state(registry_path, clean_id, rid, "FAILED", last_error=str(err))
            except Exception as reg_err:
                print(f"[WARN] Failed to update registry state for {rid}: {reg_err}")
            
            # Master Ledger Update
            log_run_to_registry(rid, "failed", clean_id)
            raise err

    # Engine Version Registry Resolution
    try:
        active_engine = get_active_engine()
    except Exception as e:
        raise PipelineExecutionError(f"Engine Resolution Failed: {e}", directive_id=clean_id)

    engine_module = f"engine_dev.universal_research_engine.{active_engine}.stage2_compiler"
    
    # Safety Check: Verify module existence via importlib
    try:
        importlib.import_module(engine_module)
    except ImportError:
        raise PipelineExecutionError(
            f"Configured engine version '{active_engine}' does not exist. (Module {engine_module} not found).",
            directive_id=clean_id
        )

    run_command(
        [python_exe, "-m", engine_module, "--scan", clean_id],
        "Stage-2 Compilation",
    )

    for rid, symbol in zip(run_ids, symbols):
        mgr = PipelineStateManager(rid)
        current = mgr.get_state_data()["current_state"]
        if current in ("STAGE_1_COMPLETE", "STAGE_2_COMPLETE"):
            run_folder = BACKTESTS_DIR / f"{clean_id}_{symbol}"
            ak_reports = list(run_folder.glob("AK_Trade_Report_*.xlsx"))
            if ak_reports:
                if current != "STAGE_2_COMPLETE":
                    transition_run_state(rid, "STAGE_2_COMPLETE")
            else:
                print(f"[WARN] Stage-2 artifact missing for {symbol} ({rid[:8]}). Marking FAILED.")
                transition_run_state(rid, "FAILED")
                update_run_state(
                    registry_path,
                    clean_id,
                    rid,
                    "FAILED",
                    last_error="Stage-2 artifact missing (AK_Trade_Report_*.xlsx).",
                )
                log_run_to_registry(rid, "failed", clean_id)

    run_command([python_exe, "tools/stage3_compiler.py", clean_id], "Stage-3 Aggregation")

    master_filter_path = MASTER_FILTER_PATH
    if not master_filter_path.exists():
        raise PipelineExecutionError(
            f"Stage-3 artifact missing: {master_filter_path}",
            directive_id=clean_id,
            run_ids=run_ids,
        )

    import openpyxl

    wb = openpyxl.load_workbook(master_filter_path, read_only=True)
    ws = wb.active
    try:
        headers = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        strategy_idx = headers.index("strategy")
    except Exception as err:
        wb.close()
        raise PipelineExecutionError(
            f"Failed to resolve 'strategy' column in Master Filter: {err}",
            directive_id=clean_id,
            run_ids=run_ids,
        ) from err

    actual_count = sum(
        1
        for row in ws.iter_rows(min_row=2, values_only=True)
        if row and len(row) > strategy_idx and row[strategy_idx] and str(row[strategy_idx]).startswith(clean_id)
    )
    wb.close()
    expected_count = len(symbols)
    if actual_count != expected_count:
        raise PipelineExecutionError(
            f"Stage-3 cardinality mismatch: expected {expected_count}, found {actual_count} for {clean_id}",
            directive_id=clean_id,
            run_ids=run_ids,
        )

    print(f"[GATE] Stage-3 artifact verified: {actual_count}/{expected_count} rows for {clean_id}")

    for rid, symbol in zip(run_ids, symbols):
        mgr = PipelineStateManager(rid)
        current = mgr.get_state_data()["current_state"]
        if current == "STAGE_2_COMPLETE":
            transition_run_state(rid, "STAGE_3_COMPLETE")

            snapshot_path = mgr.run_dir / "strategy.py"
            if not snapshot_path.exists():
                transition_run_state(rid, "FAILED")
                update_run_state(registry_path, clean_id, rid, "FAILED", last_error="Snapshot file missing.")
                log_run_to_registry(rid, "failed", clean_id)
                # 2. Binding (Verification of Hash)
            source_path = project_root / "strategies" / strategy_id / "strategy.py"
            if not source_path.exists():
                transition_run_state(rid, "FAILED")
                update_run_state(registry_path, clean_id, rid, "FAILED", last_error="Source strategy missing.")
                log_run_to_registry(rid, "failed", clean_id)
                raise RuntimeError(f"Source strategy missing: {source_path}")

            def get_file_hash(path: Path) -> str:
                return hashlib.sha256(path.read_bytes()).hexdigest()

            if get_file_hash(snapshot_path) != get_file_hash(source_path):
                transition_run_state(rid, "FAILED")
                update_run_state(
                    registry_path,
                    clean_id,
                    rid,
                    "FAILED",
                    last_error="Snapshot integrity mismatch against source strategy.",
                )
                log_run_to_registry(rid, "failed", clean_id)
                raise RuntimeError(
                    f"Snapshot Integrity Mismatch! {rid}/strategy.py != strategies/{strategy_id}/strategy.py"
                )

            print(f"[ORCHESTRATOR] Snapshot Verified: {rid} matches source.")
            mgr._append_audit_log(
                "SNAPSHOT_VERIFIED",
                {"strategy_hash": get_file_hash(snapshot_path), "source_hash": get_file_hash(source_path)},
            )

            bt_dir = RUNS_DIR / rid / "data"
            
            # --- MANDATORY ARTIFACT: EQUITY CURVE GENERATION ---
            trade_file = bt_dir / "results_tradelevel.csv"
            equity_file = bt_dir / "equity_curve.csv"
            if trade_file.exists() and not equity_file.exists():
                try:
                    df_t = pd.read_csv(trade_file)
                    if "pnl_usd" in df_t.columns:
                        # Deterministic cumulative PnL series starting from $10,000
                        pnl_series = df_t["pnl_usd"].fillna(0)
                        equity_series = 10000.0 + pnl_series.cumsum()
                        df_eq = pd.DataFrame({
                            "exit_timestamp": df_t.get("exit_timestamp", []),
                            "equity": equity_series
                        })
                        df_eq.to_csv(equity_file, index=False)
                        print(f"[ORCHESTRATOR] Generated local equity curve for {rid}")
                except Exception as e:
                    print(f"[WARN] Failed to auto-generate equity curve for {rid}: {e}")

            required_artifacts = {
                "results_tradelevel.csv": bt_dir / "results_tradelevel.csv",
                "results_standard.csv": bt_dir / "results_standard.csv",
                "equity_curve.csv": bt_dir / "equity_curve.csv",
                # "batch_summary.csv": bt_dir / "batch_summary.csv",
            }
            artifacts_manifest: dict[str, str] = {}
            for name, path in required_artifacts.items():
                if not path.exists():
                    transition_run_state(rid, "FAILED")
                    update_run_state(
                        registry_path,
                        clean_id,
                        rid,
                        "FAILED",
                        last_error=f"Missing required artifact for binding: {path}",
                    )
                    log_run_to_registry(rid, "failed", clean_id)
                    raise RuntimeError(f"Missing required artifact for binding: {path}")
                artifacts_manifest[name] = get_file_hash(path)

            manifest = {
                "run_id": rid,
                "strategy_hash": get_file_hash(snapshot_path),
                "artifacts": artifacts_manifest,
                "timestamp": datetime.now(timezone.utc).isoformat(),
            }
            manifest_path = mgr.run_dir / "manifest.json"
            
            # Manifest Freeze Guard
            if manifest_path.exists() and current == "COMPLETE":
                with open(manifest_path, "r", encoding="utf-8") as f:
                    existing_manifest = json.load(f)
                if existing_manifest.get("run_id") == rid and existing_manifest.get("artifacts") == artifacts_manifest:
                    print(f"[ORCHESTRATOR] Manifest verified (frozen): {manifest_path}")
                else:
                    raise RuntimeError(f"[FATAL] Manifest Immutability Violation: Attempted to modify manifest of COMPLETED run {rid}.")
            else:
                with open(manifest_path, "w", encoding="utf-8") as f:
                    json.dump(manifest, f, indent=4)
                print(f"[ORCHESTRATOR] Manifest Bound: {manifest_path}")
            mgr._append_audit_log(
                "ARTIFACT_BOUND",
                {"manifest_path": str(manifest_path), "artifact_hashes": artifacts_manifest},
            )

            transition_run_state_sequence(rid, ["STAGE_3A_COMPLETE", "COMPLETE"])
            mgr._append_audit_log("RUN_COMPLETE", {"status": "SUCCESS"})
            
            # SUCCESS Master Ledger update
            log_run_to_registry(rid, "complete", clean_id)

    transition_directive_state(clean_id, "SYMBOL_RUNS_COMPLETE")
