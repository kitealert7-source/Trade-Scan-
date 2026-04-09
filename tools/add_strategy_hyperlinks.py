"""
add_strategy_hyperlinks.py — Add clickable hyperlinks to strategy columns
in pipeline Excel artifacts.

Supports two targets (--target flag):

  candidates (default):
    File:   TradeScan_State/candidates/Filtered_Strategies_Passed.xlsx
    Column: C (strategy) — full strategy+symbol name
    Links:  ../backtests/{strategy}/

  portfolio:
    File:   TradeScan_State/strategies/Master_Portfolio_Sheet.xlsx
    Column: B (source_strategy) — base strategy name
    Links:  {source_strategy}/  (same directory as the xlsx)

Idempotent: re-running overwrites stale links, skips matching links.
Preserves all existing formatting (headers, widths, number formats).

Usage:
    python tools/add_strategy_hyperlinks.py                    # candidates
    python tools/add_strategy_hyperlinks.py --target portfolio # portfolio sheet
    python tools/add_strategy_hyperlinks.py --target all       # both
    python tools/add_strategy_hyperlinks.py --dry-run          # preview only
"""

import argparse
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font

PROJECT_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PROJECT_ROOT))

from config.state_paths import CANDIDATE_FILTER_PATH

PORTFOLIO_SHEET_PATH = PROJECT_ROOT.parent / "TradeScan_State" / "strategies" / "Master_Portfolio_Sheet.xlsx"

# Hyperlink font — blue underline, preserves size from formatter
_LINK_FONT = Font(color="0563C1", underline="single")


def _find_column(ws, name: str) -> int | None:
    """Find column index by header name. Returns None if not found."""
    for col_idx, cell in enumerate(ws[1], start=1):
        if str(cell.value or "").strip() == name:
            return col_idx
    return None


def _apply_hyperlinks(xlsx_path: Path, col_name: str, link_prefix: str,
                      dry_run: bool = False) -> dict:
    """Generic hyperlink applicator for any xlsx + column + prefix."""
    if not xlsx_path.exists():
        print(f"[ABORT] File not found: {xlsx_path}")
        return {"total": 0, "linked": 0, "skipped_match": 0,
                "skipped_empty": 0, "missing_targets": []}

    wb = load_workbook(xlsx_path)
    ws = wb.active

    col_idx = _find_column(ws, col_name)
    if col_idx is None:
        print(f"[ABORT] Column '{col_name}' not found in {xlsx_path.name}")
        return {"total": 0, "linked": 0, "skipped_match": 0,
                "skipped_empty": 0, "missing_targets": []}

    total = 0
    linked = 0
    skipped_empty = 0
    skipped_match = 0
    missing_targets = []

    for row_idx in range(2, ws.max_row + 1):
        total += 1
        cell = ws.cell(row=row_idx, column=col_idx)
        value = str(cell.value or "").strip()

        if not value:
            skipped_empty += 1
            continue

        target = f"{link_prefix}{value}/"

        if cell.hyperlink and cell.hyperlink.target == target:
            skipped_match += 1
            continue

        if not dry_run:
            cell.hyperlink = target
            cell.font = _LINK_FONT

        linked += 1

        abs_target = xlsx_path.parent / target
        if not abs_target.exists():
            missing_targets.append(value)

    if not dry_run and linked > 0:
        wb.save(xlsx_path)

    prefix = "[DRY RUN] " if dry_run else ""
    print(f"{prefix}{xlsx_path.name}:")
    print(f"{prefix}  Processed:      {total} rows")
    print(f"{prefix}  Linked:         {linked} rows")
    print(f"{prefix}  Skipped (match):{skipped_match} rows")
    print(f"{prefix}  Skipped (empty):{skipped_empty} rows")
    if missing_targets:
        print(f"{prefix}  Missing targets: {len(missing_targets)}")
        for mt in missing_targets[:5]:
            print(f"    {mt}")
        if len(missing_targets) > 5:
            print(f"    ... and {len(missing_targets) - 5} more")

    return {"total": total, "linked": linked, "skipped_match": skipped_match,
            "skipped_empty": skipped_empty, "missing_targets": missing_targets}


def link_candidates(dry_run: bool = False) -> dict:
    """Hyperlink Column C (strategy) → ../backtests/{strategy}/"""
    return _apply_hyperlinks(
        CANDIDATE_FILTER_PATH, "strategy", "../backtests/", dry_run
    )


def link_portfolio(dry_run: bool = False) -> dict:
    """Hyperlink Column A (portfolio_id) → {portfolio_id}/"""
    return _apply_hyperlinks(
        PORTFOLIO_SHEET_PATH, "portfolio_id", "", dry_run
    )


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Add clickable hyperlinks to strategy columns in pipeline Excel artifacts.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument("--target", choices=["candidates", "portfolio", "all"],
                        default="candidates",
                        help="Which file(s) to process (default: candidates)")
    parser.add_argument("--dry-run", action="store_true",
                        help="Preview without modifying the file")
    args = parser.parse_args()

    if args.target in ("candidates", "all"):
        link_candidates(dry_run=args.dry_run)
    if args.target in ("portfolio", "all"):
        link_portfolio(dry_run=args.dry_run)


if __name__ == "__main__":
    main()
