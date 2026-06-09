"""cointegration_excel.py — Phase 3: SQLite → Excel report.

Reads cointegration_daily + singles_daily from SQLite and writes a
multi-sheet workbook organized by ASSET CLASS (per 2026-05-21 cleanup):

  1. Summary                   — universe-level structural state
  2. Forex (incl. Metals)      — curated FX + XAU candidates only
  3. Crypto                    — curated BTC/ETH candidates only
  4. Indices & Stocks          — placeholder (deferred until equity universe)
  5. All Pairs (Diagnostic)    — full 210-pair-pair output, audit-only
  6. Singles (Diagnostic)      — full single-symbol ADF output, audit-only
  7. History                   — last 90 daily snapshots per pair-window
  8. Notes                     — doctrine, classifier rules, per-candidate
                                  rationale

The asset-class tabs (2-4) are operator-actionable; the diagnostic tabs
(5-6) preserve the full screener output as audit trail but are flagged
as not-to-be-acted-on-without-economic-rationale. The curated candidates
come from governance/cointegration_candidates.yaml — see that file for
the structural reasoning behind each candidate.

Per COINTEGRATION_SCREENER_V1_SPEC.md §8 (ranking) + §9 (sheet layout)
+ 2026-05-21 hypothesis-led curation amendment.

Conditional formatting is applied via explicit cell fills (computed in
Python) — simpler than openpyxl ConditionalFormatting rules for the
mixed categorical+numeric thresholds we use.

CLI:
    python tools/cointegration_excel.py --export
"""
from __future__ import annotations

import argparse
import math
import sqlite3
import sys
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parent.parent
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.filters import FilterColumn, Filters

from tools.capital.capital_broker_spec import load_broker_spec

import yaml

from config.path_authority import DATA_ROOT
from tools.cointegration_db import (
    HYSTERESIS_LOOKBACK,
    P_BREAKING,
    P_COINTEGRATED,
    SINGLES_TABLE_NAME,
    SQLITE_DB,
    TABLE_NAME,
    connect,
)
# Co-located with parquet + SQLite under SYSTEM_FACTORS — see
# cointegration_db.py for the 2026-05-20 location-move rationale.
EXCEL_PATH = DATA_ROOT / "SYSTEM_FACTORS" / "FX_COINTEGRATION" / "Cointegration_Screener.xlsx"
CANDIDATES_YAML = PROJECT_ROOT / "governance" / "cointegration_candidates.yaml"

# --- Colors (hex without #) --------------------------------------
COLOR_HEADER_BG    = "4472C4"  # matches existing format_excel convention
COLOR_HEADER_FG    = "FFFFFF"
COLOR_SECTION_BG   = "D9E1F2"  # section header in Summary
COLOR_GREEN_BG     = "C6EFCE"
COLOR_GREEN_FG     = "006100"
COLOR_YELLOW_BG    = "FFEB9C"
COLOR_YELLOW_FG    = "9C5700"
COLOR_RED_BG       = "FFC7CE"
COLOR_RED_FG       = "9C0006"
COLOR_GREY_BG      = "EAEAEA"
COLOR_BOOTSTRAP_BG = "FCE4D6"  # soft orange — bootstrap-classifier flag

THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


# ---------------------------------------------------------------------------
# Composite ranking score (spec §8)
# ---------------------------------------------------------------------------


def _half_life_quality(hl: float | None) -> float:
    """Peaks at 15 days; falls off below 3 or above 60 (spec §8)."""
    if hl is None or pd.isna(hl) or hl <= 0:
        return 0.0
    return float(math.exp(-abs(math.log(hl / 15.0))))


def _stability_persistence(conn: sqlite3.Connection,
                            pair_a: str, pair_b: str, lookback_days: int,
                            *, days: int = 90) -> float:
    """Fraction of last `days` snapshots in regime='cointegrated'.

    On day 1 returns 1.0 or 0.0 depending on today's regime alone.
    """
    rows = conn.execute(
        f"""SELECT regime FROM {TABLE_NAME}
            WHERE pair_a = ? AND pair_b = ? AND lookback_days = ?
            ORDER BY as_of DESC LIMIT ?""",
        (pair_a, pair_b, int(lookback_days), int(days)),
    ).fetchall()
    if not rows:
        return 0.0
    return sum(1 for r in rows if r["regime"] == "cointegrated") / len(rows)


def _excursion_containment(conn: sqlite3.Connection,
                            pair_a: str, pair_b: str, lookback_days: int,
                            *, days: int = 252) -> float:
    """Fraction of last `days` snapshots with |current_zscore| ≤ 3.0."""
    rows = conn.execute(
        f"""SELECT current_zscore FROM {TABLE_NAME}
            WHERE pair_a = ? AND pair_b = ? AND lookback_days = ?
            ORDER BY as_of DESC LIMIT ?""",
        (pair_a, pair_b, int(lookback_days), int(days)),
    ).fetchall()
    if not rows:
        return 0.0
    contained = sum(
        1 for r in rows
        if r["current_zscore"] is not None and abs(r["current_zscore"]) <= 3.0
    )
    return contained / len(rows)


def composite_score(conn: sqlite3.Connection, row: dict) -> float:
    """`stability_persistence * half_life_quality * excursion_containment`.

    Each component ∈ [0, 1]; NaN inputs → 0. (Spec §8.)
    """
    sp = _stability_persistence(conn, row["pair_a"], row["pair_b"],
                                  row["lookback_days"])
    hq = _half_life_quality(row["half_life_days"])
    ec = _excursion_containment(conn, row["pair_a"], row["pair_b"],
                                  row["lookback_days"])
    return float(sp * hq * ec)


# ---------------------------------------------------------------------------
# Friendly universe (rule-derived) + cointegration persistence tiers
# ---------------------------------------------------------------------------
# Two orthogonal axes surfaced on the diagnostic tabs:
#   tier             — backtest PERFORMANCE (elite/friendly), rule-derived from
#                      the MPS into governance/fx_fx_friendly.yaml by
#                      tools/derive_friendly.py. Refresh monthly/post-backtest.
#   persistence_tier — current cointegration DURABILITY (how long the 1d×252
#                      relationship has held), computed live from the DB over
#                      the COMPLETE universe. Lower bands double as a
#                      developing-relationship radar.

FX_FX_FRIENDLY_YAML = PROJECT_ROOT / "governance" / "fx_fx_friendly.yaml"

# Persistence bands — FIXED operational day-thresholds, NOT tied to any live
# metric. Originally informed by the ~7-day median daily half-life (2026-06)
# but deliberately decoupled: if the half-life regime shifts, revisit these as
# a conscious decision — they will not silently re-scale.
PERSISTENCE_BANDS: list[tuple[int, str]] = [   # (min_streak_days, label), desc
    (90, "entrenched"),
    (60, "mature"),
    (30, "established"),
    (15, "developing"),
    (10, "emerging"),
]


def _canonical_pair_key(pair_a: str, pair_b: str) -> str:
    """Order-independent pair key shared with tools/derive_friendly.py."""
    return "/".join(sorted([str(pair_a).strip(), str(pair_b).strip()]))


def load_friendly(path: Path = FX_FX_FRIENDLY_YAML) -> dict:
    """canonical_key -> record {tier, median_ret_dd, evaluable}. Empty if absent.

    Joined to screener rows by the canonical (order-independent) key so the
    flag attaches regardless of pair orientation across MPS / DB / yaml.
    """
    if not path.is_file():
        return {}
    data = yaml.safe_load(path.read_text(encoding="utf-8")) or {}
    return {r["pair"]: r for r in (data.get("pairs") or []) if r.get("pair")}


def friendly_tier(friendly: dict, pair_a: str, pair_b: str) -> str:
    """'elite' / 'friendly' / '' for a pair, via the canonical key."""
    rec = friendly.get(_canonical_pair_key(pair_a, pair_b))
    return rec["tier"] if rec else ""


def coint_streak_days(conn: sqlite3.Connection, pair_a: str, pair_b: str,
                      lookback_days: int = 252, *, lookback: int = 120) -> int:
    """Current consecutive cointegrated-day streak on the 1d × `lookback_days`
    window, counting back from the most recent snapshot.

    Smoothed against 1-day flickers: a day counts if regime=='cointegrated' OR
    (its 5d rolling-median p < 0.05 while regime != 'broken'), so a single raw
    p-value blip the screener already tagged 'breaking' does not reset a
    genuinely-stable relationship to zero.
    """
    rows = conn.execute(
        f"""SELECT regime, pvalue_rolling_median_5d FROM {TABLE_NAME}
            WHERE pair_a = ? AND pair_b = ? AND lookback_days = ? AND tf = '1d'
            ORDER BY as_of DESC LIMIT ?""",
        (pair_a, pair_b, int(lookback_days), int(lookback)),
    ).fetchall()
    streak = 0
    for r in rows:
        regime = r["regime"]
        pmed = r["pvalue_rolling_median_5d"]
        smoothed_ok = pmed is not None and pmed < 0.05 and regime != "broken"
        if regime == "cointegrated" or smoothed_ok:
            streak += 1
        else:
            break
    return streak


def persistence_tier(streak_days: int) -> str:
    """Map a streak (days) to its fixed band label; '' below 10d."""
    for lo, label in PERSISTENCE_BANDS:
        if streak_days >= lo:
            return label
    return ""


# ---------------------------------------------------------------------------
# Conditional-fill helpers
# ---------------------------------------------------------------------------


def _fill(bg: str, fg: str | None = None) -> tuple[PatternFill, Font]:
    return (PatternFill("solid", fgColor=bg),
            Font(color=fg) if fg else Font())


_REGIME_FILL = {
    "cointegrated": _fill(COLOR_GREEN_BG,  COLOR_GREEN_FG),
    "breaking":     _fill(COLOR_YELLOW_BG, COLOR_YELLOW_FG),
    "broken":       _fill(COLOR_RED_BG,    COLOR_RED_FG),
}
_AGREEMENT_FILL = {
    "BOTH":     _fill(COLOR_GREEN_BG,  COLOR_GREEN_FG),
    "252-only": _fill(COLOR_YELLOW_BG, COLOR_YELLOW_FG),
    "504-only": _fill(COLOR_YELLOW_BG, COLOR_YELLOW_FG),
    "NEITHER":  _fill(COLOR_RED_BG,    COLOR_RED_FG),
}
# Column-D ("coint_window") operator-facing relabel for the All Pairs
# (Diagnostic) sheet. The underlying _pivot_today() `agreement` field stays
# BOTH/252-only/504-only/NEITHER (Summary §4 + the pivot tests depend on it);
# these friendlier labels are display-only so the column-D AutoFilter dropdown
# reads as a regime selector (Regime 252 / Regime 504 / Both / Neither).
_AGREEMENT_TO_WINDOW = {
    "BOTH":     "Both",
    "252-only": "Regime 252",
    "504-only": "Regime 504",
    "NEITHER":  "Neither",
}
_COINT_WINDOW_FILL = {                       # same colour semantics as _AGREEMENT_FILL
    "Both":       _fill(COLOR_GREEN_BG,  COLOR_GREEN_FG),
    "Regime 252": _fill(COLOR_YELLOW_BG, COLOR_YELLOW_FG),
    "Regime 504": _fill(COLOR_YELLOW_BG, COLOR_YELLOW_FG),
    "Neither":    _fill(COLOR_RED_BG,    COLOR_RED_FG),
}
# Dual-TF agreement palette — 1d × 252 vs 4h × 1500. Same colors as
# _AGREEMENT_FILL (cross-window 252/504), different key set since the
# semantics are cross-timeframe rather than cross-window.
_DUAL_TF_FILL = {
    "BOTH":    _fill(COLOR_GREEN_BG,  COLOR_GREEN_FG),
    "1d-only": _fill(COLOR_YELLOW_BG, COLOR_YELLOW_FG),
    "4h-only": _fill(COLOR_YELLOW_BG, COLOR_YELLOW_FG),
    "NEITHER": _fill(COLOR_RED_BG,    COLOR_RED_FG),
}
def _half_life_color(hl) -> tuple[PatternFill, Font] | None:
    if hl is None or pd.isna(hl):
        return _fill(COLOR_RED_BG, COLOR_RED_FG)
    if 5 <= hl <= 30:
        return _fill(COLOR_GREEN_BG, COLOR_GREEN_FG)
    if (3 <= hl < 5) or (30 < hl <= 60):
        return _fill(COLOR_YELLOW_BG, COLOR_YELLOW_FG)
    return _fill(COLOR_RED_BG, COLOR_RED_FG)


def _zscore_color(z) -> tuple[PatternFill, Font] | None:
    if z is None or pd.isna(z):
        return None
    az = abs(z)
    if az >= 3.0:
        return _fill(COLOR_RED_BG, COLOR_RED_FG)
    if az >= 2.0:
        return _fill(COLOR_YELLOW_BG, COLOR_YELLOW_FG)
    return None


# ---------------------------------------------------------------------------
# Sheet writers
# ---------------------------------------------------------------------------


def _header_style(cell) -> None:
    cell.fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
    cell.font = Font(color=COLOR_HEADER_FG, bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER


def _section_header(ws, row: int, text: str, span: int = 4) -> int:
    ws.cell(row=row, column=1, value=text)
    ws.cell(row=row, column=1).font = Font(bold=True, size=12, color="000000")
    ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor=COLOR_SECTION_BG)
    ws.merge_cells(start_row=row, end_row=row, start_column=1, end_column=span)
    return row + 1


def _compute_neutral_basket(sym_a: str, sym_b: str,
                              beta: float | None) -> tuple[float | None, float | None]:
    """Return (lot_a, lot_b) for the minimum OctaFX β-neutral basket.

    Derivation: spread = B - β·A. To hold a position with no sensitivity to
    outright moves in A (β-neutral), the dollar P&L from the β·A "expected"
    component must cancel the dollar P&L from the B leg. With L_A lots of A
    and L_B lots of B, and `usd_per_pu_per_lot` from each broker spec
    (MT5-verified USD P&L per 1 price unit per 1.0 lot), this gives:

        L_A / L_B  =  |β|  ×  (usd_per_pu_per_lot_B  /  usd_per_pu_per_lot_A)

    The minimum tradable basket is constructed by setting the smaller-lot
    side to its broker min_lot, computing the other side, and rounding it
    to lot_step. Returns (None, None) if either broker spec is missing,
    β is invalid, or either usd_per_pu_per_lot is zero/missing.
    """
    if beta is None or not isinstance(beta, (int, float)) or beta == 0:
        return (None, None)
    if not math.isfinite(beta):
        return (None, None)
    try:
        spec_a = load_broker_spec(sym_a)
        spec_b = load_broker_spec(sym_b)
    except FileNotFoundError:
        return (None, None)
    cal_a = spec_a.get("calibration", {}) or {}
    cal_b = spec_b.get("calibration", {}) or {}
    usd_a = float(cal_a.get("usd_per_pu_per_lot", 0) or 0)
    usd_b = float(cal_b.get("usd_per_pu_per_lot", 0) or 0)
    if usd_a <= 0 or usd_b <= 0:
        return (None, None)
    min_a  = float(spec_a.get("min_lot",  0.01))
    min_b  = float(spec_b.get("min_lot",  0.01))
    step_a = float(spec_a.get("lot_step", 0.01))
    step_b = float(spec_b.get("lot_step", 0.01))
    ratio = abs(beta) * (usd_b / usd_a)   # = lots_A / lots_B
    if ratio <= 0 or not math.isfinite(ratio):
        return (None, None)
    if ratio >= 1.0:
        # A needs more lots than B per basket — set B at minimum, scale A up
        l_b = min_b
        l_a = max(min_a, round(ratio * l_b / step_a) * step_a)
    else:
        # B needs more lots than A per basket — set A at minimum, scale B up
        l_a = min_a
        l_b = max(min_b, round(l_a / ratio / step_b) * step_b)
    return (round(l_a, 4), round(l_b, 4))


def _apply_all_pairs_default_filter(ws) -> None:
    """Single-column regime-selector default filter for the All Pairs
    (Diagnostic) sheet. Header at row 1, data at rows 2+.

    Column D ("coint_window") tags each pair by the window(s) in which it is
    currently cointegrated, using operator-facing labels:

        Both        - cointegrated on BOTH the 252d and 504d windows
        Regime 252  - cointegrated on the 252d window only
        Regime 504  - cointegrated on the 504d window only
        Neither     - cointegrated on neither window

    The column-D AutoFilter dropdown IS the regime selector. The default view
    pins it to "Regime 252" (operator request 2026-06-05) so the file opens on
    the fresh-formation 252d-only pairs. The operator clicks the dropdown to
    switch to "Both" / "Regime 504" / "Neither", or multi-checks to combine
    (e.g. Regime 252 + Both to see every 252d-cointegrated pair). Every other
    column keeps its own independent dropdown for further narrowing; none is
    pre-filtered, so the regime choice alone drives the default visible set.

    (Supersedes the previous 4-filter "BOTH + both-windows-cointegrated +
    tradable half_life_504" default — those criteria conflict with a
    252d-only default, since a 252-only pair is by definition not
    504-cointegrated and has no meaningful 504d half-life.)
    """
    last_row = ws.max_row
    if last_row <= 1:
        return
    last_col = get_column_letter(ws.max_column)
    ws.auto_filter.ref = f"A1:{last_col}{last_row}"

    # Single value-list filter on column D. colId is 0-indexed from the ref's
    # first column (A), so D -> colId=3; the cell column index is 1-based: D=4.
    DEFAULT_WINDOW = "Regime 252"
    fc_d = FilterColumn(colId=3)
    fc_d.filters = Filters(filter=[DEFAULT_WINDOW])
    ws.auto_filter.filterColumn.append(fc_d)

    for r in range(2, last_row + 1):
        if ws.cell(row=r, column=4).value != DEFAULT_WINDOW:
            ws.row_dimensions[r].hidden = True


def _apply_default_filter_cointegrated(
    ws, header_row: int, regime_col_idx: int,
) -> None:
    """Add an Excel AutoFilter to the sheet's data range and pre-apply a
    'cointegrated'-only filter on the named regime column.

    Sets:
      * `auto_filter.ref` so Excel renders the filter dropdown UI
      * `add_filter_column` so the saved file remembers the cointegrated
        filter is active
      * `row_dimensions[r].hidden = True` for every non-cointegrated row so
        Excel renders the filtered view immediately on open (without this,
        the filter UI shows as "applied" but all rows still display until
        the user re-clicks the filter)

    Operator can click the filter dropdown and uncheck `(Show All)` /
    pick a different regime to expand the view manually.
    """
    last_row = ws.max_row
    if last_row <= header_row:
        return
    last_col = get_column_letter(ws.max_column)
    ws.auto_filter.ref = f"A{header_row}:{last_col}{last_row}"
    # add_filter_column wants a 0-indexed offset from auto_filter.ref's
    # first column; our ref starts at A so offset = (col_idx - 1)
    ws.auto_filter.add_filter_column(regime_col_idx - 1, ["cointegrated"])
    for r in range(header_row + 1, last_row + 1):
        if ws.cell(row=r, column=regime_col_idx).value != "cointegrated":
            ws.row_dimensions[r].hidden = True


# ---------------------------------------------------------------------------
# Curated candidates loader + per-candidate query
# ---------------------------------------------------------------------------


def load_candidates(path: Path = CANDIDATES_YAML) -> dict:
    """Read the hypothesis-led candidates YAML.

    Returns the raw `asset_classes` dict. Each top-level key is an
    asset-class slug (e.g. 'forex'); values carry `label`, `description`,
    `candidates` list (may be empty), and optional `deferred` bool.
    """
    if not path.is_file():
        return {}
    data = yaml.safe_load(path.read_text(encoding="utf-8")) or {}
    return data.get("asset_classes") or {}


def _query_candidate_rows(
    conn: sqlite3.Connection, candidate: dict, df_pairs: pd.DataFrame,
    df_singles: pd.DataFrame,
) -> dict:
    """Resolve one candidate to per-window result rows.

    Returns a dict with `regime`, `adf_pvalue`, `half_life_days`,
    `current_zscore` for each lookback window (252, 504), plus
    `route`, `rationale`, `type`, `subject` (display label).
    Empty values become None.
    """
    ctype = candidate["type"]

    def _cells(sub: pd.DataFrame) -> dict:
        out: dict[str, object] = {}
        # 1d × {252, 504} — original two windows. Filter by tf for safety
        # even though 252/504 lookbacks happen to be 1d-only in current data.
        for lb in (252, 504):
            r = sub[(sub.tf == "1d") & (sub.lookback_days == lb)]
            if r.empty:
                out[f"regime_{lb}"] = None
                out[f"p_{lb}"] = None
                out[f"hl_{lb}"] = None
                out[f"z_{lb}"] = None
            else:
                rr = r.iloc[0]
                out[f"regime_{lb}"] = rr["regime"]
                out[f"p_{lb}"] = float(rr["adf_pvalue"])
                out[f"hl_{lb}"] = float(rr["half_life_days"]) if pd.notna(
                    rr["half_life_days"]) else None
                out[f"z_{lb}"] = float(rr["current_zscore"]) if pd.notna(
                    rr["current_zscore"]) else None

        # 4h × 1500 — Tier B dual-TF surface (design doc Thread B baseline).
        r_4h = sub[(sub.tf == "4h") & (sub.lookback_days == 1500)]
        if r_4h.empty:
            out["regime_4h"] = None
            out["p_4h"] = None
        else:
            rr = r_4h.iloc[0]
            out["regime_4h"] = rr["regime"]
            out["p_4h"] = float(rr["adf_pvalue"])

        # Dual-TF agreement — 1d × 252 vs 4h × 1500. 252 is the canonical
        # 1d window for the Pine pair-research baseline (15m × N=30) per
        # COINTEGRATION_FILTER_DESIGN_2026-05-26.md §5.
        reg_1d = out.get("regime_252")
        reg_4h = out.get("regime_4h")
        if reg_1d == "cointegrated" and reg_4h == "cointegrated":
            out["dual_tf_agreement"] = "BOTH"
        elif reg_1d == "cointegrated":
            out["dual_tf_agreement"] = "1d-only"
        elif reg_4h == "cointegrated":
            out["dual_tf_agreement"] = "4h-only"
        else:
            out["dual_tf_agreement"] = "NEITHER"
        return out

    lot_a: float | None = None
    lot_b: float | None = None

    if ctype == "single":
        sym = candidate["symbol"]
        sub = df_singles[df_singles.symbol == sym]
        subject = sym
    elif ctype == "synthetic_ratio":
        sym = f"RATIO:{candidate['pair_a']}/{candidate['pair_b']}"
        sub = df_singles[df_singles.symbol == sym]
        subject = sym
    elif ctype == "pair":
        # Canonical orientation: alphabetical
        a, b = sorted([candidate["pair_a"], candidate["pair_b"]])
        sub = df_pairs[(df_pairs.pair_a == a) & (df_pairs.pair_b == b)]
        subject = f"{a}/{b}"
        # β-neutral lot basket via OctaFX broker specs. Use the 252d β
        # (shorter window, more responsive to current market structure)
        # to drive the sizing — the 504d β is an alternative the operator
        # can compute manually from the table if needed.
        beta_row = sub[sub.lookback_days == 252]
        if not beta_row.empty:
            beta = beta_row.iloc[0]["hedge_ratio"]
            beta = float(beta) if pd.notna(beta) else None
            lot_a, lot_b = _compute_neutral_basket(a, b, beta)
    else:
        sub = pd.DataFrame()
        subject = candidate.get("key", "?")

    cells = _cells(sub)
    cells.update({
        "key":       candidate["key"],
        "type":      ctype,
        "subject":   subject,
        "route":     candidate.get("route", "—"),
        "canonical": candidate.get("canonical"),
        "lot_a":     lot_a,
        "lot_b":     lot_b,
        "rationale": candidate.get("rationale", "").strip(),
    })
    return cells


# Asset-class membership for the pair_class column on the All Pairs
# (Diagnostic) + History sheets. Lets the operator filter the row firehose
# down to "FX pairs only" / "Index pairs only" / "Cross-asset only" via
# the auto-filter dropdown without typing pair_a/pair_b filters manually.
_FX_SYMBOLS = frozenset({
    "AUDUSD", "EURUSD", "GBPUSD", "NZDUSD", "USDCAD", "USDCHF", "USDJPY",
    "AUDJPY", "AUDNZD", "CADJPY", "CHFJPY", "EURAUD", "EURGBP", "EURJPY",
    "GBPAUD", "GBPJPY", "GBPNZD", "NZDJPY",
})
_IDX_SYMBOLS = frozenset({
    "SPX500", "NAS100", "US30", "UK100", "FRA40", "ESP35", "EUSTX50",
    "GER40", "JPN225", "AUS200",
})
_CC_SYMBOLS = frozenset({"XAUUSD", "BTCUSD", "ETHUSD"})


def _classify_pair(sym_a: str, sym_b: str) -> str:
    """Bucket a pair-pair into FX / IDX / CROSS for filter UX.

    FX    — both legs are FX pairs (18-symbol FX_UNIVERSE)
    IDX   — both legs are equity indices (10-symbol set)
    CROSS — anything else: FX × IDX, FX × commodity/crypto, IDX × CC,
            CC × CC. All cross-asset combinations bucket together since
            their economic mechanisms differ from pure within-class spreads.
    """
    a_fx = sym_a in _FX_SYMBOLS
    b_fx = sym_b in _FX_SYMBOLS
    a_idx = sym_a in _IDX_SYMBOLS
    b_idx = sym_b in _IDX_SYMBOLS
    if a_fx and b_fx:
        return "FX"
    if a_idx and b_idx:
        return "IDX"
    return "CROSS"


def _pivot_today(df: pd.DataFrame) -> pd.DataFrame:
    """Pivot from (pair, window) rows to (pair) rows with window columns."""
    if df.empty:
        return df
    pivoted = []
    for (a, b), grp in df.groupby(["pair_a", "pair_b"], sort=True):
        r252 = grp[grp.lookback_days == 252]
        r504 = grp[grp.lookback_days == 504]
        if r252.empty or r504.empty:
            continue
        r252 = r252.iloc[0]
        r504 = r504.iloc[0]
        regime_252 = r252["regime"]
        regime_504 = r504["regime"]
        if regime_252 == "cointegrated" and regime_504 == "cointegrated":
            agreement = "BOTH"
        elif regime_252 == "cointegrated":
            agreement = "252-only"
        elif regime_504 == "cointegrated":
            agreement = "504-only"
        else:
            agreement = "NEITHER"
        # methodology_version is shared across both windows (same screener run
        # produces both); take from 252 with 504 fallback. Surfaces the cohort
        # tag (v1_raw_adf legacy vs v2_log_eg post-correction) in the All Pairs
        # diagnostic so operators don't accidentally compare across cohorts.
        methodology = (
            r252["methodology_version"] if "methodology_version" in r252
            and pd.notna(r252["methodology_version"])
            else (r504["methodology_version"] if "methodology_version" in r504
                  and pd.notna(r504["methodology_version"]) else "v1_raw_adf")
        )
        pivoted.append({
            "pair_a": a, "pair_b": b,
            "pair_class": _classify_pair(a, b),
            "agreement": agreement,
            "regime_252": regime_252, "regime_504": regime_504,
            "adf_pvalue_252": r252["adf_pvalue"], "adf_pvalue_504": r504["adf_pvalue"],
            "half_life_days_252": r252["half_life_days"], "half_life_days_504": r504["half_life_days"],
            "current_zscore_252": r252["current_zscore"], "current_zscore_504": r504["current_zscore"],
            "hedge_ratio_252": r252["hedge_ratio"], "hedge_ratio_504": r504["hedge_ratio"],
            "history_depth": max(r252["history_depth"], r504["history_depth"]),
            "methodology": methodology,
        })
    return pd.DataFrame(pivoted)


def _write_summary(wb: Workbook, conn: sqlite3.Connection,
                    df_today: pd.DataFrame) -> None:
    ws = wb.create_sheet("Summary", 0)
    # Two distinct dates surface here:
    #   data_as_of — newest as_of in the per-pair-latest result set; this is
    #                the LATEST data cutoff used by the screener (e.g. on a
    #                Saturday for an FX-dominant universe this is the prior
    #                Friday because no new FX bars exist).
    #   run_date   — when THIS Excel was regenerated (the screener actually ran).
    # Showing both removes the 2026-06-06 confusion where the data_as_of date
    # (Fri 2026-06-05) made it look like the screener had not run on Saturday.
    if not df_today.empty:
        data_as_of = df_today["as_of"].max()
    else:
        data_as_of = "—"
    run_date = datetime.now(timezone.utc).date().isoformat()
    ws.cell(
        row=1, column=1,
        value=(
            f"Cointegration Screener — Summary  "
            f"(data as-of: {data_as_of} | run: {run_date} UTC)  "
            f"[methodology: v2_log_eg]"
        )
    ).font = Font(bold=True, size=14)
    ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=6)
    ws.cell(
        row=2, column=1,
        value="Math: log prices + Engle-Granger / MacKinnon criticals via statsmodels.coint"
    ).font = Font(italic=True, size=10)
    ws.merge_cells(start_row=2, end_row=2, start_column=1, end_column=6)

    row = 4

    # --- Section 1: Universe regime distribution
    row = _section_header(ws, row, "1. Universe regime distribution", span=5)
    headers = ["window", "cointegrated", "breaking", "broken", "total"]
    for c, h in enumerate(headers, start=1):
        _header_style(ws.cell(row=row, column=c, value=h))
    row += 1
    for lb in (252, 504):
        sub = df_today[df_today.lookback_days == lb]
        c1 = sub[sub.regime == "cointegrated"].shape[0]
        c2 = sub[sub.regime == "breaking"].shape[0]
        c3 = sub[sub.regime == "broken"].shape[0]
        ws.cell(row=row, column=1, value=f"{lb}d")
        for col_i, (val, fill_key) in enumerate(
            [(c1, "cointegrated"), (c2, "breaking"), (c3, "broken")], start=2):
            cell = ws.cell(row=row, column=col_i, value=val)
            cell.fill, cell.font = _REGIME_FILL[fill_key]
            cell.alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=5, value=c1 + c2 + c3).alignment = Alignment(horizontal="center")
        row += 1
    row += 1

    # --- Section 2: Half-life summary (cointegrated only)
    row = _section_header(ws, row, "2. Half-life days (cointegrated pairs only)", span=5)
    headers = ["window", "median", "mean", "min", "max"]
    for c, h in enumerate(headers, start=1):
        _header_style(ws.cell(row=row, column=c, value=h))
    row += 1
    for lb in (252, 504):
        sub = df_today[(df_today.lookback_days == lb) & (df_today.regime == "cointegrated")]
        if sub.empty or sub["half_life_days"].isna().all():
            ws.cell(row=row, column=1, value=f"{lb}d")
            ws.cell(row=row, column=2, value="—")
            row += 1
            continue
        hl = sub["half_life_days"].dropna()
        ws.cell(row=row, column=1, value=f"{lb}d")
        ws.cell(row=row, column=2, value=round(float(hl.median()), 2))
        ws.cell(row=row, column=3, value=round(float(hl.mean()), 2))
        ws.cell(row=row, column=4, value=round(float(hl.min()), 2))
        ws.cell(row=row, column=5, value=round(float(hl.max()), 2))
        for c in range(1, 6):
            ws.cell(row=row, column=c).alignment = Alignment(horizontal="center")
        row += 1
    row += 1

    # --- Section 3: Top recurring currencies in cointegrated pairs (252d)
    row = _section_header(ws, row, "3. Top currencies in cointegrated pairs (252d)", span=3)
    headers = ["currency", "appearances", "% of cointegrated pairs"]
    for c, h in enumerate(headers, start=1):
        _header_style(ws.cell(row=row, column=c, value=h))
    row += 1
    coint_252 = df_today[(df_today.lookback_days == 252) & (df_today.regime == "cointegrated")]
    counter: Counter[str] = Counter()
    for _, r in coint_252.iterrows():
        for sym in (r["pair_a"], r["pair_b"]):
            for currency in (sym[:3], sym[3:]):
                counter[currency] += 1
    total = coint_252.shape[0] * 2 or 1
    for currency, count in counter.most_common(10):
        ws.cell(row=row, column=1, value=currency).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=2, value=count).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=3, value=f"{count/total*100:.1f}%").alignment = Alignment(horizontal="center")
        row += 1
    row += 1

    # --- Section 4: Window agreement
    row = _section_header(ws, row, "4. Window agreement (252d × 504d cointegration overlap)", span=2)
    headers = ["agreement", "pair count"]
    for c, h in enumerate(headers, start=1):
        _header_style(ws.cell(row=row, column=c, value=h))
    row += 1
    pivoted = _pivot_today(df_today)
    agree_counts = (pivoted["agreement"].value_counts() if not pivoted.empty else pd.Series(dtype=int))
    for label in ("BOTH", "252-only", "504-only", "NEITHER"):
        ws.cell(row=row, column=1, value=label)
        if label in _AGREEMENT_FILL:
            ws.cell(row=row, column=1).fill, ws.cell(row=row, column=1).font = _AGREEMENT_FILL[label]
            ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=2, value=int(agree_counts.get(label, 0))).alignment = Alignment(horizontal="center")
        row += 1
    row += 1

    # --- Section 5: Regime changes vs prior snapshot
    row = _section_header(ws, row, "5. Regime changes vs prior snapshot", span=4)
    headers = ["transition", "count", "pair examples (up to 5)"]
    for c, h in enumerate(headers, start=1):
        _header_style(ws.cell(row=row, column=c, value=h))
    row += 1
    if df_today.empty:
        prior_as_of = None
    else:
        # Distinct as_of values, sorted descending
        prior_row = conn.execute(
            f"SELECT DISTINCT as_of FROM {TABLE_NAME} ORDER BY as_of DESC LIMIT 2"
        ).fetchall()
        prior_as_of = prior_row[1]["as_of"] if len(prior_row) >= 2 else None
    if prior_as_of is None:
        ws.cell(row=row, column=1, value="(no prior snapshot — first run)")
        row += 2
    else:
        prior_df = pd.read_sql_query(
            f"SELECT pair_a, pair_b, lookback_days, regime FROM {TABLE_NAME} WHERE as_of = ?",
            conn, params=(prior_as_of,))
        # Join today × prior on the key
        merged = df_today.merge(
            prior_df, on=["pair_a", "pair_b", "lookback_days"],
            suffixes=("_today", "_prior"))
        newly_broken = merged[(merged.regime_prior == "cointegrated")
                              & (merged.regime_today.isin(["broken", "breaking"]))]
        newly_recovered = merged[(merged.regime_prior.isin(["broken", "breaking"]))
                                 & (merged.regime_today == "cointegrated")]
        for label, sub in (("newly_broken", newly_broken),
                           ("newly_recovered", newly_recovered)):
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=len(sub)).alignment = Alignment(horizontal="center")
            examples = ", ".join(
                f"{r.pair_a}/{r.pair_b} ({r.lookback_days}d)"
                for r in sub.head(5).itertuples()
            ) or "—"
            ws.cell(row=row, column=3, value=examples)
            row += 1
    row += 1

    # --- Section 6: Bootstrap visibility
    row = _section_header(ws, row, "6. Bootstrap visibility (history_depth distribution)", span=3)
    headers = ["status", "row count", "% of universe"]
    for c, h in enumerate(headers, start=1):
        _header_style(ws.cell(row=row, column=c, value=h))
    row += 1
    n_total = len(df_today)
    n_bootstrap = (df_today["history_depth"] < HYSTERESIS_LOOKBACK).sum()
    n_hysteresis = (df_today["history_depth"] >= HYSTERESIS_LOOKBACK).sum()
    for label, count, fill in (
        (f"bootstrap (history_depth < {HYSTERESIS_LOOKBACK})", n_bootstrap, _fill(COLOR_BOOTSTRAP_BG, "000000")),
        (f"hysteresis-active (history_depth ≥ {HYSTERESIS_LOOKBACK})", n_hysteresis, _fill(COLOR_GREEN_BG, COLOR_GREEN_FG)),
    ):
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=1).fill, ws.cell(row=row, column=1).font = fill
        ws.cell(row=row, column=2, value=int(count)).alignment = Alignment(horizontal="center")
        pct = (count / n_total * 100) if n_total else 0
        ws.cell(row=row, column=3, value=f"{pct:.1f}%").alignment = Alignment(horizontal="center")
        row += 1
    row += 1

    # --- Section 7: 4h × 1500 universe + dual-TF intersection
    # Tier A of the 4h integration (per COINTEGRATION_FILTER_DESIGN_2026-05-26).
    # Surfaces (a) the 4h × 1500 regime distribution alongside the 1d Section 1
    # tally above and (b) the dual-TF intersection count at progressively
    # tighter raw-p thresholds, the cohort-size table the design doc §4 used
    # to size Thread A vs Thread B.
    row = _section_header(
        ws, row,
        "7. 4h × 1500 universe + dual-TF intersection (1d × 252 ∩ 4h × 1500)",
        span=5,
    )
    df_1d_252 = df_today[(df_today.tf == "1d") & (df_today.lookback_days == 252)]
    df_4h_1500 = df_today[(df_today.tf == "4h") & (df_today.lookback_days == 1500)]

    # Sub-table 7a — 4h × 1500 regime distribution (mirrors Section 1 layout)
    headers = ["surface", "cointegrated", "breaking", "broken", "total"]
    for c, h in enumerate(headers, start=1):
        _header_style(ws.cell(row=row, column=c, value=h))
    row += 1
    c_co = df_4h_1500[df_4h_1500.regime == "cointegrated"].shape[0]
    c_br = df_4h_1500[df_4h_1500.regime == "breaking"].shape[0]
    c_bk = df_4h_1500[df_4h_1500.regime == "broken"].shape[0]
    ws.cell(row=row, column=1, value="4h × 1500")
    for col_i, (val, fill_key) in enumerate(
        [(c_co, "cointegrated"), (c_br, "breaking"), (c_bk, "broken")], start=2):
        cell = ws.cell(row=row, column=col_i, value=val)
        cell.fill, cell.font = _REGIME_FILL[fill_key]
        cell.alignment = Alignment(horizontal="center")
    ws.cell(row=row, column=5, value=c_co + c_br + c_bk).alignment = Alignment(horizontal="center")
    row += 2

    # Sub-table 7b — cohort sizes by raw p-value threshold + dual-TF intersection
    headers = ["threshold", "1d × 252 count", "4h × 1500 count", "BOTH (dual-TF)"]
    for c, h in enumerate(headers, start=1):
        _header_style(ws.cell(row=row, column=c, value=h))
    row += 1
    for thr in (0.05, 0.01, 0.001):
        sub_1d = df_1d_252[df_1d_252.adf_pvalue < thr]
        sub_4h = df_4h_1500[df_4h_1500.adf_pvalue < thr]
        pairs_1d = set(zip(sub_1d.pair_a, sub_1d.pair_b))
        pairs_4h = set(zip(sub_4h.pair_a, sub_4h.pair_b))
        c_both = len(pairs_1d & pairs_4h)
        ws.cell(row=row, column=1, value=f"p < {thr}").alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=2, value=len(sub_1d)).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=3, value=len(sub_4h)).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=4, value=c_both).alignment = Alignment(horizontal="center")
        row += 1
    row += 1

    # Note: raw screener p-values carry the Engle-Granger bias documented in
    # COINTEGRATION_FILTER_DESIGN_2026-05-26.md §3. The effective conventional
    # cointegration threshold under correct math is p < 0.01 on the displayed
    # values (≈ true p < 0.05 after the ~4-5× bias correction).
    ws.cell(row=row, column=1, value=(
        "Note: thresholds are raw screener p-values (Engle-Granger bias not "
        "corrected). Treat displayed p < 0.01 as the effective conventional "
        "cointegration threshold (≈ true p < 0.05). See "
        "COINTEGRATION_FILTER_DESIGN_2026-05-26.md §3."))
    ws.cell(row=row, column=1).font = Font(italic=True, size=9, color="9C5700")
    ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=row, end_row=row, start_column=1, end_column=5)
    ws.row_dimensions[row].height = 30

    # Column widths
    for c, w in enumerate([26, 14, 14, 14, 14], start=1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["C"].width = 40   # examples + % cells


def _write_today(wb: Workbook, conn: sqlite3.Connection,
                  df_today: pd.DataFrame, position: int | None = None) -> None:
    if position is None:
        position = len(wb.sheetnames)
    ws = wb.create_sheet("Today", position)
    pivoted = _pivot_today(df_today)

    # Compute composite score per row from SQLite (uses 252d row only).
    scores = []
    for _, r in pivoted.iterrows():
        score_row = {"pair_a": r["pair_a"], "pair_b": r["pair_b"],
                     "lookback_days": 252,
                     "half_life_days": r["half_life_days_252"]}
        scores.append(composite_score(conn, score_row))
    pivoted["score"] = scores
    pivoted = pivoted.sort_values("score", ascending=False).reset_index(drop=True)

    # Column-D display relabel: BOTH/252-only/504-only/NEITHER -> friendly
    # regime-selector labels (display-only; _pivot_today's `agreement` field is
    # left intact for Summary section 4 + the pivot tests).
    pivoted["coint_window"] = pivoted["agreement"].map(_AGREEMENT_TO_WINDOW)

    # Friendly (performance) + persistence (durability) axes — see helper block.
    # Appended AFTER the existing columns so column D ("coint_window") and its
    # default regime filter are undisturbed.
    friendly = load_friendly()
    pivoted["tier"] = pivoted.apply(
        lambda r: friendly_tier(friendly, r["pair_a"], r["pair_b"]) or None, axis=1)
    _streaks = [coint_streak_days(conn, r["pair_a"], r["pair_b"], 252)
                for _, r in pivoted.iterrows()]
    pivoted["coint_streak_days"] = _streaks
    pivoted["persistence_tier"] = [persistence_tier(s) or None for s in _streaks]

    columns = [
        "pair_a", "pair_b",
        "pair_class",                   # NEW 2026-05-23 — FX / IX / CROSS
        "coint_window",                 # was "agreement" — relabeled col D selector
        "regime_252", "regime_504",
        "adf_pvalue_252", "adf_pvalue_504",
        "half_life_days_252", "half_life_days_504",
        "current_zscore_252", "current_zscore_504",
        "hedge_ratio_252", "hedge_ratio_504",
        "history_depth", "score",
        "methodology",  # NEW 2026-05-30 (C4) — cohort tag (v1_raw_adf / v2_log_eg)
        "tier",               # NEW — friendly performance tier (elite/friendly)
        "persistence_tier",   # NEW — cointegration durability band
        "coint_streak_days",  # NEW — current consecutive cointegrated days (1d×252)
    ]
    for c, h in enumerate(columns, start=1):
        _header_style(ws.cell(row=1, column=c, value=h))

    for i, r in pivoted.iterrows():
        excel_row = i + 2
        for c, col in enumerate(columns, start=1):
            v = r[col]
            cell = ws.cell(row=excel_row, column=c)
            if isinstance(v, float) and not pd.isna(v):
                cell.value = round(v, 4)
            elif pd.isna(v):
                cell.value = None
            else:
                cell.value = v
            cell.alignment = Alignment(horizontal="center")
            cell.border = BORDER
            # Conditional fills
            if col in ("regime_252", "regime_504") and v in _REGIME_FILL:
                cell.fill, cell.font = _REGIME_FILL[v]
            elif col == "coint_window" and v in _COINT_WINDOW_FILL:
                cell.fill, cell.font = _COINT_WINDOW_FILL[v]
            elif col in ("half_life_days_252", "half_life_days_504"):
                fc = _half_life_color(v)
                if fc:
                    cell.fill, cell.font = fc
            elif col in ("current_zscore_252", "current_zscore_504"):
                zc = _zscore_color(v)
                if zc:
                    cell.fill, cell.font = zc
            elif col == "history_depth":
                if isinstance(v, (int, float)) and v < HYSTERESIS_LOOKBACK:
                    cell.fill, cell.font = _fill(COLOR_BOOTSTRAP_BG, "000000")

    ws.freeze_panes = "D2"
    #         pair_a pair_b cls  cwin reg2 reg5 p25 p50 hl2 hl5 z25 z50 he2 he5 hd score meth tier ptier strk
    widths = [10,    10,    10,  12,  13,  13,  14, 14, 18, 18, 18, 18, 14, 14, 14, 10,   14,  10,  14,   12]
    for c, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(c)].width = w

    # All Pairs (Diagnostic) — column-D regime-selector default filter.
    # Column D ("coint_window") is the single default driver, pinned to
    # "Regime 252"; the operator switches via its dropdown (Both / Regime 504 /
    # Neither) or multi-checks to combine. See _apply_all_pairs_default_filter.
    _apply_all_pairs_default_filter(ws)


_ROUTE_FILL = {
    "direct":         _fill(COLOR_GREEN_BG,  COLOR_GREEN_FG),
    "synthesize":     _fill(COLOR_YELLOW_BG, COLOR_YELLOW_FG),
    "redundant_with": _fill(COLOR_GREY_BG,   "000000"),
}


def _write_asset_class_tab(
    wb: Workbook, conn: sqlite3.Connection, class_key: str, class_def: dict,
    df_pairs: pd.DataFrame, df_singles: pd.DataFrame, position: int,
) -> None:
    """Write one asset-class tab with curated candidates only."""
    label = class_def.get("label", class_key)
    description = class_def.get("description", "")
    deferred = bool(class_def.get("deferred", False))
    candidates = class_def.get("candidates") or []

    ws = wb.create_sheet(label, position)

    # Header row + description
    ws.cell(row=1, column=1, value=label).font = Font(bold=True, size=14)
    ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=19)
    ws.cell(row=2, column=1, value=description.replace("\n", " ").strip())
    ws.cell(row=2, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=2, end_row=2, start_column=1, end_column=19)
    ws.row_dimensions[2].height = 45

    if deferred:
        ws.cell(row=4, column=1, value="DEFERRED — no candidates active.")
        ws.cell(row=4, column=1).font = Font(italic=True, color="9C5700")
        ws.column_dimensions["A"].width = 80
        return

    if not candidates:
        ws.cell(row=4, column=1, value="No curated candidates configured.")
        ws.column_dimensions["A"].width = 80
        return

    # Layout (Tier B 2026-05-27): dual-TF columns (reg4h, dualTF, p4h)
    # interleaved with the existing 1d × {252, 504} cells.
    headers = [
        "key", "type", "subject", "route", "canonical",
        "reg252", "reg504", "reg4h", "dualTF",
        "p252", "p504", "p4h",
        "hl252", "hl504",
        "z252", "z504",
        "lot_a", "lot_b",
        "rationale",
    ]
    for c, h in enumerate(headers, start=1):
        _header_style(ws.cell(row=4, column=c, value=h))

    for i, candidate in enumerate(candidates):
        row = 5 + i
        cells = _query_candidate_rows(conn, candidate, df_pairs, df_singles)

        def put(col_idx, value):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal="center" if col_idx <= 18 else "left",
                                       vertical="top",
                                       wrap_text=col_idx == 19)
            cell.border = BORDER
            return cell

        put(1, cells["key"])
        put(2, cells["type"])
        put(3, cells["subject"])
        route_cell = put(4, cells["route"])
        if cells["route"] in _ROUTE_FILL:
            route_cell.fill, route_cell.font = _ROUTE_FILL[cells["route"]]
        put(5, cells.get("canonical") or "")

        # 1d regime cells (cols 6-7) — reg252 / reg504
        for col_offset, lb in enumerate((252, 504)):
            reg = cells.get(f"regime_{lb}")
            cell = put(6 + col_offset, reg)
            if reg in _REGIME_FILL:
                cell.fill, cell.font = _REGIME_FILL[reg]

        # 4h regime (col 8) + dual-TF agreement (col 9)
        reg_4h = cells.get("regime_4h")
        cell_4h = put(8, reg_4h)
        if reg_4h in _REGIME_FILL:
            cell_4h.fill, cell_4h.font = _REGIME_FILL[reg_4h]
        dual = cells.get("dual_tf_agreement")
        cell_dual = put(9, dual)
        if dual in _DUAL_TF_FILL:
            cell_dual.fill, cell_dual.font = _DUAL_TF_FILL[dual]

        # p-value cells (cols 10-12) — p252 / p504 / p4h
        for col_offset, lb in enumerate((252, 504)):
            p = cells.get(f"p_{lb}")
            put(10 + col_offset, round(p, 4) if p is not None else None)
        p_4h = cells.get("p_4h")
        put(12, round(p_4h, 4) if p_4h is not None else None)

        # Half-life cells (cols 13-14) — hl252 / hl504 (no 4h column;
        # dual-TF judgment uses p + regime, not 4h half-life directly)
        for col_offset, lb in enumerate((252, 504)):
            hl = cells.get(f"hl_{lb}")
            hl_cell = put(13 + col_offset, round(hl, 1) if hl is not None else None)
            fc = _half_life_color(hl)
            if fc:
                hl_cell.fill, hl_cell.font = fc

        # z-score cells (cols 15-16) — z252 / z504
        for col_offset, lb in enumerate((252, 504)):
            z = cells.get(f"z_{lb}")
            z_cell = put(15 + col_offset, round(z, 2) if z is not None else None)
            zc = _zscore_color(z)
            if zc:
                z_cell.fill, z_cell.font = zc

        # β-neutral lot basket (cols 17-18) — lot_a / lot_b
        put(17, cells.get("lot_a"))
        put(18, cells.get("lot_b"))

        # Rationale (col 19) — wrapped
        put(19, cells["rationale"])
        ws.row_dimensions[row].height = 50

    # Column widths
    #          key  type subj route can  r252 r504 r4h  dTF  p252 p504 p4h  hl252 hl504 z252 z504 ltA ltB ratio
    widths = [22,  16,  22,  14,  22,  12,  12,  12,  10,  9,   9,   9,   9,    9,    9,   9,   9,  9,  70]
    for c, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(c)].width = w

    # Auto-hide `canonical` column (col E) when no candidate uses route=
    # redundant_with on this tab. Cross-asset tabs without triangular
    # restatements (e.g. Indices & Stocks, where FX-equity pairs have no
    # shared token to cancel) get a cleaner view. Column un-hides
    # automatically the moment a redundant_with entry is added to this
    # class in the yaml.
    if not any(c.get("canonical") for c in candidates):
        ws.column_dimensions[get_column_letter(5)].hidden = True

    # Pre-applied cointegrated-only filter on reg252 (col F = index 6) for
    # crypto + indices_stocks tabs. Forex left unfiltered so all candidates
    # (incl. currently-broken ones like AUDNZD) stay visible by default.
    if class_key in ("crypto", "indices_stocks"):
        _apply_default_filter_cointegrated(ws, header_row=4, regime_col_idx=6)

    ws.freeze_panes = "F5"


def _write_dual_tf_shortlist(wb: Workbook, conn: sqlite3.Connection,
                              df_today: pd.DataFrame, position: int) -> None:
    """Pairs cointegrated on BOTH 1d × 252 AND 4h × 1500.

    Tier C of the 4h integration — the Thread B per-pair deep-dive surface
    from COINTEGRATION_FILTER_DESIGN_2026-05-26.md §5. Sorted by
    min(adf_pvalue_1d, adf_pvalue_4h) ascending so the strongest dual-TF
    candidates appear first. Distinct from the curated asset-class tabs
    (which include human-rationale candidates regardless of dual-TF status)
    and the All Pairs (Diagnostic) firehose (which carries every
    pair-pair × tf × lookback combination unfiltered).

    Sheet is empty when either surface has no cointegrated rows.
    """
    ws = wb.create_sheet("Dual-TF Shortlist", position)

    ws.cell(row=1, column=1, value=(
        "Dual-TF Shortlist — pairs cointegrated on BOTH 1d × 252 AND 4h × 1500"
    )).font = Font(bold=True, size=12)
    ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=12)
    ws.cell(row=2, column=1, value=(
        "Thread B per-pair deep-dive surface (design doc §5). Sorted by "
        "min(p_1d, p_4h) ascending — strongest dual-TF first. Treat raw "
        "p < 0.01 as the effective conventional threshold (≈ true p < 0.05)."
    ))
    ws.cell(row=2, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=2, end_row=2, start_column=1, end_column=12)
    ws.row_dimensions[2].height = 30

    df_1d = df_today[
        (df_today.tf == "1d") & (df_today.lookback_days == 252)
        & (df_today.regime == "cointegrated")
    ]
    df_4h = df_today[
        (df_today.tf == "4h") & (df_today.lookback_days == 1500)
        & (df_today.regime == "cointegrated")
    ]
    if df_1d.empty or df_4h.empty:
        ws.cell(row=4, column=1, value=(
            "(no dual-TF pairs — either 1d × 252 or 4h × 1500 surface has "
            "no cointegrated rows in the current data)"
        ))
        ws.cell(row=4, column=1).font = Font(italic=True, color="9C5700")
        ws.column_dimensions["A"].width = 80
        return

    merged = df_1d.merge(
        df_4h, on=["pair_a", "pair_b"], suffixes=("_1d", "_4h"),
    )
    if merged.empty:
        ws.cell(row=4, column=1, value="(no pairs cointegrated on both 1d × 252 and 4h × 1500)")
        ws.cell(row=4, column=1).font = Font(italic=True, color="9C5700")
        ws.column_dimensions["A"].width = 80
        return

    merged["min_p"] = merged[["adf_pvalue_1d", "adf_pvalue_4h"]].min(axis=1)
    merged = merged.sort_values("min_p").reset_index(drop=True)
    merged["pair_class"] = merged.apply(
        lambda r: _classify_pair(r["pair_a"], r["pair_b"]), axis=1,
    )

    headers = [
        "pair_a", "pair_b", "pair_class",
        "p_1d_252", "p_4h_1500",
        "hl_1d_252", "hl_4h_1500",
        "z_1d_252", "z_4h_1500",
        "hedge_ratio_1d", "hedge_ratio_4h",
        "history_depth",
        "tier", "persistence_tier", "coint_streak_days",
    ]
    for c, h in enumerate(headers, start=1):
        _header_style(ws.cell(row=4, column=c, value=h))

    friendly = load_friendly()
    for i, r in merged.iterrows():
        excel_row = 5 + i
        _streak = coint_streak_days(conn, r["pair_a"], r["pair_b"], 252)
        cells_data = [
            r["pair_a"], r["pair_b"], r["pair_class"],
            r["adf_pvalue_1d"], r["adf_pvalue_4h"],
            r["half_life_days_1d"], r["half_life_days_4h"],
            r["current_zscore_1d"], r["current_zscore_4h"],
            r["hedge_ratio_1d"], r["hedge_ratio_4h"],
            max(r["history_depth_1d"], r["history_depth_4h"]),
            friendly_tier(friendly, r["pair_a"], r["pair_b"]) or None,
            persistence_tier(_streak) or None,
            _streak,
        ]
        for c, v in enumerate(cells_data, start=1):
            cell = ws.cell(row=excel_row, column=c)
            if isinstance(v, float) and not pd.isna(v):
                if c in (4, 5):              # p-values: 4 decimals
                    cell.value = round(v, 4)
                elif c in (6, 7):            # half-life: 1 decimal
                    cell.value = round(v, 1)
                else:
                    cell.value = round(v, 2)
            elif pd.isna(v):
                cell.value = None
            else:
                cell.value = v
            cell.alignment = Alignment(horizontal="center")
            cell.border = BORDER
            if c in (6, 7):
                fc = _half_life_color(v)
                if fc:
                    cell.fill, cell.font = fc
            elif c in (8, 9):
                zc = _zscore_color(v)
                if zc:
                    cell.fill, cell.font = zc

    ws.freeze_panes = "D5"
    #          pa pb cls p1d p4h hl1d hl4h z1d z4h hr1d hr4h hd  tier ptier strk
    widths = [10, 10, 10, 10, 10, 10,  10,  10, 10, 14, 14,  12, 10,  14,   12]
    for c, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(c)].width = w


def _write_friendly_universe(wb: Workbook, conn: sqlite3.Connection,
                              df_today: pd.DataFrame, position: int) -> None:
    """Rule-derived FX-FX 'friendly' universe roster — ELITE / BROAD split.

    Reads governance/fx_fx_friendly.yaml (produced by tools/derive_friendly.py)
    and renders each pair with its backtest tier + Median Ret/DD AND its live
    1d × 252 state (regime, z, streak, persistence band), so the operator picks
    onboarding candidates from one view: filter to a cointegrated/persistent
    row, sort by z. `in_screen`=no flags a friendly pair absent from today's
    diagnostic universe. Empty (with a note) when the yaml is absent.
    """
    ws = wb.create_sheet("Friendly Universe", position)
    ws.cell(row=1, column=1, value=(
        "Friendly Universe — rule-derived FX-FX cointegration-friendly pairs"
    )).font = Font(bold=True, size=12)
    ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=7)
    ws.cell(row=2, column=1, value=(
        "Performance tier from MPS backtest (elite = Median Ret/DD >= 0.75, "
        "friendly = >= 0.50; FX-FX, Evaluable >= 5). The regime/z/streak columns "
        "are today's live 1d x 252 state. Rule-based & refreshed via "
        "tools/derive_friendly.py — see Notes."
    ))
    ws.cell(row=2, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=2, end_row=2, start_column=1, end_column=7)
    ws.row_dimensions[2].height = 30

    friendly = load_friendly()
    if not friendly:
        ws.cell(row=4, column=1, value=(
            "(governance/fx_fx_friendly.yaml not found — run "
            "python tools/derive_friendly.py)"
        ))
        ws.cell(row=4, column=1).font = Font(italic=True, color="9C5700")
        ws.column_dimensions["A"].width = 80
        return

    # Live 1d × 252 snapshot per canonical pair key.
    live: dict[str, object] = {}
    if not df_today.empty:
        d = df_today[(df_today.tf == "1d") & (df_today.lookback_days == 252)]
        for _, r in d.iterrows():
            live[_canonical_pair_key(r["pair_a"], r["pair_b"])] = r

    headers = ["pair", "median_ret_dd", "regime_252", "z_252",
               "coint_streak_days", "persistence_tier", "in_screen"]

    def _section(start_row: int, title: str, recs: list) -> int:
        ws.cell(row=start_row, column=1, value=title).font = Font(bold=True, size=11)
        hdr = start_row + 1
        for c, h in enumerate(headers, start=1):
            _header_style(ws.cell(row=hdr, column=c, value=h))
        for i, rec in enumerate(sorted(recs, key=lambda r: -r.get("median_ret_dd", 0))):
            row = hdr + 1 + i
            key = rec["pair"]
            a, b = key.split("/", 1)
            lr = live.get(key)
            streak = coint_streak_days(conn, a, b, 252)
            regime = lr["regime"] if lr is not None else None
            z = lr["current_zscore"] if lr is not None else None
            vals = [
                key, rec.get("median_ret_dd"), regime,
                round(float(z), 2) if z is not None and pd.notna(z) else None,
                streak, persistence_tier(streak) or None,
                "yes" if lr is not None else "no",
            ]
            for c, v in enumerate(vals, start=1):
                cell = ws.cell(row=row, column=c, value=v)
                cell.alignment = Alignment(horizontal="center")
                cell.border = BORDER
                if c == 3 and v in _REGIME_FILL:
                    cell.fill, cell.font = _REGIME_FILL[v]
                elif c == 4:
                    zc = _zscore_color(v)
                    if zc:
                        cell.fill, cell.font = zc
        return hdr + 1 + len(recs)

    elite = [r for r in friendly.values() if r.get("tier") == "elite"]
    broad = [r for r in friendly.values() if r.get("tier") == "friendly"]
    nxt = _section(4, f"ELITE  (Median Ret/DD >= 0.75)  —  {len(elite)} pairs", elite)
    _section(nxt + 2,
             f"BROAD / FRIENDLY  (0.50 <= Median Ret/DD < 0.75)  —  {len(broad)} pairs",
             broad)

    widths = [16, 14, 14, 10, 18, 16, 10]
    for c, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A4"


def _write_singles_diagnostic(wb: Workbook, conn: sqlite3.Connection,
                                df_singles: pd.DataFrame, position: int) -> None:
    """Full singles output (diagnostic / audit). Not operator-actionable
    unless an entry survives economic-rationale screening (then promote
    to candidates yaml).
    """
    ws = wb.create_sheet("Singles (Diagnostic)", position)
    ws.cell(row=1, column=1, value="Singles — full single-symbol ADF (diagnostic only)"
            ).font = Font(bold=True, size=12)
    ws.cell(row=2, column=1,
            value=("Each row: ADF on log-price of a single symbol (or synthetic "
                   "ratio). NOT operator-actionable without explicit economic "
                   "rationale — see Forex/Crypto tabs for curated subset."))
    ws.cell(row=2, column=1).alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=2, end_row=2, start_column=1, end_column=10)
    ws.row_dimensions[2].height = 30

    if df_singles.empty:
        ws.cell(row=4, column=1, value="(no singles data yet — daily runner pending)")
        return

    # Pivot per symbol: 252 + 504 side-by-side
    pivot = []
    for sym, sub in df_singles.groupby("symbol"):
        r252 = sub[sub.lookback_days == 252]
        r504 = sub[sub.lookback_days == 504]
        if r252.empty or r504.empty:
            continue
        pivot.append({
            "symbol": sym,
            "regime_252": r252.iloc[0]["regime"],
            "regime_504": r504.iloc[0]["regime"],
            "adf_pvalue_252": r252.iloc[0]["adf_pvalue"],
            "adf_pvalue_504": r504.iloc[0]["adf_pvalue"],
            "half_life_days_252": r252.iloc[0]["half_life_days"],
            "half_life_days_504": r504.iloc[0]["half_life_days"],
            "current_zscore_252": r252.iloc[0]["current_zscore"],
            "current_zscore_504": r504.iloc[0]["current_zscore"],
            "history_depth": max(r252.iloc[0]["history_depth"],
                                  r504.iloc[0]["history_depth"]),
        })
    df_pivot = pd.DataFrame(pivot).sort_values("adf_pvalue_252")

    headers = list(df_pivot.columns) if not df_pivot.empty else []
    for c, h in enumerate(headers, start=1):
        _header_style(ws.cell(row=4, column=c, value=h))

    for i, r in df_pivot.reset_index(drop=True).iterrows():
        excel_row = i + 5
        for c, col in enumerate(headers, start=1):
            v = r[col]
            cell = ws.cell(row=excel_row, column=c)
            cell.value = (round(float(v), 4) if isinstance(v, float) and pd.notna(v)
                          else (None if pd.isna(v) else v))
            cell.alignment = Alignment(horizontal="center")
            cell.border = BORDER
            if col in ("regime_252", "regime_504") and v in _REGIME_FILL:
                cell.fill, cell.font = _REGIME_FILL[v]
            elif col in ("half_life_days_252", "half_life_days_504"):
                fc = _half_life_color(v)
                if fc:
                    cell.fill, cell.font = fc
            elif col in ("current_zscore_252", "current_zscore_504"):
                zc = _zscore_color(v)
                if zc:
                    cell.fill, cell.font = zc

    ws.freeze_panes = "B5"
    widths = [22, 12, 12, 14, 14, 18, 18, 18, 18, 14]
    for c, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(c)].width = w

    # Singles (Diagnostic): pre-apply cointegrated filter on regime_252 (col B = 2)
    _apply_default_filter_cointegrated(ws, header_row=4, regime_col_idx=2)


def _apply_history_default_filter(ws, header_row: int = 1) -> None:
    """Multi-column synchronized default filter for the History sheet.

    Layout after `tf` was added: A=as_of | B=pair_a | C=pair_b |
    D=pair_class | E=tf | F=lookback_days | G=regime | H=adf_pvalue | ...

    Default visible set: `tf='1d' AND lookback_days=252 AND regime='cointegrated'`.
    Each filter renders as an independent dropdown the operator can adjust
    (e.g. swap `tf` to `4h` to inspect the 4h × 1500 surface, or expand
    `lookback_days` to also include 504).
    """
    last_row = ws.max_row
    if last_row <= header_row:
        return
    last_col = get_column_letter(ws.max_column)
    ws.auto_filter.ref = f"A{header_row}:{last_col}{last_row}"

    fc_tf = FilterColumn(colId=4)        # E = tf
    fc_tf.filters = Filters(filter=["1d"])
    ws.auto_filter.filterColumn.append(fc_tf)

    fc_lb = FilterColumn(colId=5)        # F = lookback_days
    fc_lb.filters = Filters(filter=["252"])
    ws.auto_filter.filterColumn.append(fc_lb)

    fc_reg = FilterColumn(colId=6)       # G = regime
    fc_reg.filters = Filters(filter=["cointegrated"])
    ws.auto_filter.filterColumn.append(fc_reg)

    for r in range(header_row + 1, last_row + 1):
        tf_val = ws.cell(row=r, column=5).value
        lb_val = ws.cell(row=r, column=6).value
        reg_val = ws.cell(row=r, column=7).value
        passes = (
            tf_val == "1d"
            and lb_val == 252
            and reg_val == "cointegrated"
        )
        if not passes:
            ws.row_dimensions[r].hidden = True


def _write_history(wb: Workbook, conn: sqlite3.Connection,
                    position: int | None = None) -> None:
    if position is None:
        position = len(wb.sheetnames)
    ws = wb.create_sheet("History", position)
    # Operator preference 2026-06-06: column A (as_of) defaults to Z→A so the
    # most-recent snapshot is on top when the sheet opens. Within each date
    # the rows stay grouped by (pair_a, pair_b, tf, lookback_days) so a single
    # pair-window's history is still contiguous when the operator narrows the
    # filter to one pair.
    df = pd.read_sql_query(
        f"""SELECT as_of, pair_a, pair_b, tf, lookback_days, regime,
                   adf_pvalue, pvalue_rolling_median_5d, half_life_days,
                   hedge_ratio, current_zscore, history_depth
            FROM {TABLE_NAME}
            WHERE as_of >= date('now', '-90 days')
            ORDER BY as_of DESC, pair_a, pair_b, tf, lookback_days""",
        conn,
    )
    # Insert pair_class as col D (between pair_b and tf) so the operator can
    # filter the 90-day-history firehose to a single asset class via the
    # dropdown, same UX as the All Pairs (Diagnostic) sheet.
    if not df.empty:
        df.insert(3, "pair_class",
                   df.apply(lambda r: _classify_pair(r["pair_a"], r["pair_b"]),
                            axis=1))
        # Friendly (performance) + persistence (durability) axes, appended at the
        # end so the tf/lookback/regime default filter (cols E/F/G) is undisturbed.
        # Streak is the pair-level 1d×252 durability — computed once per pair.
        friendly = load_friendly()
        _streak_cache: dict[tuple, int] = {}

        def _streak_for(a, b):
            if (a, b) not in _streak_cache:
                _streak_cache[(a, b)] = coint_streak_days(conn, a, b, 252)
            return _streak_cache[(a, b)]

        df["tier"] = df.apply(
            lambda r: friendly_tier(friendly, r["pair_a"], r["pair_b"]) or None, axis=1)
        df["coint_streak_days"] = df.apply(
            lambda r: _streak_for(r["pair_a"], r["pair_b"]), axis=1)
        df["persistence_tier"] = df["coint_streak_days"].apply(
            lambda s: persistence_tier(s) or None)
    columns = list(df.columns)
    for c, h in enumerate(columns, start=1):
        _header_style(ws.cell(row=1, column=c, value=h))
    for i, r in df.iterrows():
        excel_row = i + 2
        for c, col in enumerate(columns, start=1):
            v = r[col]
            cell = ws.cell(row=excel_row, column=c)
            if isinstance(v, float) and not pd.isna(v):
                cell.value = round(v, 4)
            elif pd.isna(v):
                cell.value = None
            else:
                cell.value = v
            cell.alignment = Alignment(horizontal="center")
            if col == "regime" and v in _REGIME_FILL:
                cell.fill, cell.font = _REGIME_FILL[v]
    ws.freeze_panes = "A2"
    #          as_of pa pb cls tf lb  reg p   pmed hl  hr  z   hd  tier strk ptier
    widths = [12,   10, 10, 10,  6, 10, 14, 12, 14,  14, 12, 12, 12, 10,  12,  14]
    for c, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(c)].width = w

    _apply_history_default_filter(ws, header_row=1)


def _write_notes(wb: Workbook) -> None:
    # Notes sheet appended last by export_excel orchestrator; position arg
    # is the new index after all earlier tabs have been written.
    position = len(wb.sheetnames)
    ws = wb.create_sheet("Notes", position)
    ws.cell(row=1, column=1, value="Cointegration Screener — Operator Notes").font = Font(bold=True, size=14)
    lines = [
        "",
        "METHODOLOGY",
        "  Pair-pair test:    Engle-Granger via statsmodels.tsa.stattools.coint(lb, la,",
        "                     trend='c', autolag='AIC') on log prices, MacKinnon (1996)",
        "                     critical values.",
        "  Single-series:     ADF on log price.",
        "  Methodology tag:   methodology_version column on every row = v2_log_eg (pair)",
        "                     / v2_log_adf (singles).",
        "  Field name note:   adf_pvalue and adf_statistic columns hold the EG / MacKinnon",
        "                     values, not plain-ADF. test_method='eg_mackinnon' disambiguates",
        "                     the semantics.",
        "",
        "HYPOTHESIS-LED CURATION",
        "  This workbook is organized by ASSET CLASS, with operator-actionable",
        "  candidates separated from the full diagnostic surface. Curated candidates",
        "  have explicit economic rationale (governance/cointegration_candidates.yaml).",
        "  Diagnostic tabs preserve the full universe output but are NOT operator-",
        "  actionable without economic-rationale screening.",
        "",
        "  Why: a blind statistical screen of ~465 pair-pairs × 2 windows is a",
        "  data-mining exercise. At α=0.05 you expect ~23 spurious 'cointegrated'",
        "  finds on pure random walks before bias. Hypothesis-led curation flips the",
        "  workflow: start from structural reasoning, use the screener to",
        "  confirm and monitor regime persistence — not to discover.",
        "",
        "TAB GUIDE",
        "  • Forex (incl. Metals)   : curated FX + XAU candidates only",
        "  • Crypto                 : curated BTC/ETH candidates only",
        "  • Indices & Stocks       : curated FX-equity cross-asset candidates",
        "  • Dual-TF Shortlist      : pairs cointegrated on BOTH 1d × 252 AND 4h × 1500",
        "                             (Thread B per-pair deep-dive surface)",
        "  • Friendly Universe      : rule-derived FX-FX roster, ELITE/BROAD split",
        "                             (backtest performance tier + live persistence)",
        "  • All Pairs (Diagnostic) : full pair-pair output, audit-only",
        "  • Singles (Diagnostic)   : full single-symbol output, audit-only",
        "  • History                : last 90 days per pair-window",
        "  • Notes                  : this sheet",
        "",
        "ROUTE COLUMN (curated tabs)",
        "  • direct           — trade the symbol/pair as listed (cheapest execution)",
        "  • synthesize       — trade as two-leg basket (no direct symbol available)",
        "  • redundant_with   — pointer to canonical candidate; this row is the",
        "                       triangular-synthesis alternative, generally not",
        "                       preferred over the canonical (see candidate row)",
        "",
        "SOURCE OF TRUTH",
        "  Parquet (data_root/SYSTEM_FACTORS/FX_COINTEGRATION/coint_1d_latest.parquet) is the deterministic",
        "  computation artifact. SQLite (DATA_ROOT/SYSTEM_FACTORS/FX_COINTEGRATION/cointegration.db) is the",
        "  longitudinal history sink. This Excel file is regenerated FROM SQLite on demand.",
        "",
        "REGIME CLASSIFIER (spec §7)",
        "  current_pvalue < 0.05  AND  ≥ 4 of last 5 priors also < 0.05   →  cointegrated",
        "  current_pvalue in [0.05, 0.10)  OR  insufficient persistence    →  breaking",
        "  current_pvalue ≥ 0.10                                            →  broken",
        f"  History depth required for hysteresis: {HYSTERESIS_LOOKBACK} prior daily snapshots.",
        "  Rows with history_depth < hysteresis threshold use the BOOTSTRAP classifier",
        "  (current p-value only) — flagged with orange highlight in the All Pairs (Diagnostic) sheet.",
        "",
        "RANKING SCORE (spec §8)",
        "  score = stability_persistence × half_life_quality × excursion_containment",
        "    stability_persistence  = fraction of last 90 days regime='cointegrated'",
        "    half_life_quality      = exp(−|log(half_life_days / 15)|)    (peaks at 15d, falls off below 3 or above 60)",
        "    excursion_containment  = fraction of last 252 days with |z-score| ≤ 3.0",
        "  Score helps SORTING. It does NOT replace diagnostic columns (p-value, half-life, z-score, regime, rolling-median).",
        "",
        "COINT WINDOW COLUMN (column D, All Pairs Diagnostic sheet)",
        "  Column D is the regime selector — its AutoFilter dropdown chooses which",
        "  cointegration window(s) to view. Default view = \"Regime 252\".",
        "  Both        — cointegrated on BOTH the 252d and 504d windows (strongest signal)",
        "  Regime 252  — cointegrated on the 252d window only (recent formation; treat with caution)",
        "  Regime 504  — cointegrated on the 504d window only (relationship may be degrading)",
        "  Neither     — cointegrated on neither window",
        "  Tip: multi-check \"Regime 252\" + \"Both\" in the dropdown to see every",
        "       252d-cointegrated pair. (The underlying field, also shown in the",
        "       Summary \"Window agreement\" section, keeps the BOTH / 252-only /",
        "       504-only / NEITHER labels.)",
        "",
        "FRIENDLY TIER + PERSISTENCE COLUMNS (Dual-TF, All Pairs Diagnostic, History)",
        "  Two ORTHOGONAL axes, each filterable via its own column dropdown:",
        "  tier (performance)        — rule-derived from the MPS backtest into",
        "                              governance/fx_fx_friendly.yaml by tools/derive_friendly.py:",
        "                              FX-FX AND Evaluable>=5 AND Median Ret/DD>=0.50 -> friendly;",
        "                              >=0.75 -> elite. Rule-based (no fixed N); RE-RUN MONTHLY or",
        "                              after new backtests. Joined by canonical sorted pair key.",
        "  persistence_tier          — current consecutive cointegrated-day streak on the 1d x 252",
        "  (durability)                window, computed LIVE over the COMPLETE universe (smoothed",
        "                              against 1-day flickers via the 5d rolling-median p-value).",
        "                              coint_streak_days carries the raw count for precise filtering.",
        "                              FIXED operational day-bands:",
        "                                emerging 10-14d | developing 15-29d | established 30-59d |",
        "                                mature 60-89d | entrenched 90d+",
        "                              Bands were originally INFORMED BY the ~7-day median half-life",
        "                              (2026-06) but are NOT mathematically tied to it — revisit them",
        "                              deliberately if the half-life regime shifts; they do not",
        "                              silently re-scale. Low bands double as a developing-",
        "                              relationship radar across the full universe.",
        "  The 'Friendly Universe' tab is the curated roster (ELITE/BROAD split) with both axes",
        "  plus live regime/z, for one-view onboarding selection (filter cointegrated, sort by z).",
        "",
        "CORRELATION ≠ COINTEGRATION (spec §1 doctrine)",
        "  Two high-correlation pairs can still have a non-stationary (non-mean-reverting) spread.",
        "  Trading mean-reversion on a non-cointegrated pair lets the spread drift indefinitely.",
        "  Empirical example from this corpus: EURUSD/NZDUSD at chart-TF correlation ≈ 0.79 but ADF",
        "  p-value 0.086 (252d) / 0.110 (504d) — regime='breaking'/'broken'. The screener catches this.",
        "",
        "STRATEGY-CONSTRUCTION CAVEAT (cleanup 2026-05-21)",
        "  Cointegration says: spread B − β·A is stationary for some β.",
        "  Trading that spread requires β-weighted lot sizing — NOT equal-lot.",
        "  An equal-lot trade on a cointegrated pair with |β| large is a",
        "  directional bet weighted by whichever leg has higher dollar-volatility,",
        "  not a mean-reversion trade. The retired COINTREV v1 strategy made",
        "  exactly this mistake and was deleted on 2026-05-21. Any future",
        "  strategy that consumes this screener must size legs by the screener's",
        "  hedge_ratio (β) column, not 1:1.",
        "",
        "PRODUCTION TRIGGER",
        "  Excel regenerated daily by DATA_INGRESS/engines/ops/invoke_daily_pipeline.ps1",
        "  via tools/cointegration_daily_runner.py, after the data update succeeds.",
        "  Smoke probe still available via tools/cointegration_screener_smoke.py.",
        "",
        f"  Generated: {datetime.now(timezone.utc).isoformat()}",
    ]
    for i, text in enumerate(lines, start=2):
        ws.cell(row=i, column=1, value=text)
    ws.column_dimensions["A"].width = 110


# ---------------------------------------------------------------------------
# Top-level
# ---------------------------------------------------------------------------


def _atomic_save_workbook(wb: Workbook, output_path: Path | str) -> None:
    """Resilient screener-xlsx save — delegates to the shared SSOT writer
    ``pipeline_utils.resilient_xlsx_write`` (temp-render → kill-Excel-if-locked →
    os.replace with backoff → loud-fail). Wired here after the 2026-06-05
    incident where a bare ``wb.save()`` silently deferred Phase 3 on a locked
    workbook (a transient post-boot Defender scan), leaving the History tab a
    full day stale. See COINTEGRATION_SCREENER_V1_SPEC.md §11.
    """
    from tools.pipeline_utils import resilient_xlsx_write
    resilient_xlsx_write(output_path, lambda p: wb.save(str(p)))


def export_excel(db_path: Path | str = SQLITE_DB,
                 output_path: Path | str = EXCEL_PATH) -> Path:
    """Read DB, build workbook, write to `output_path`.

    Sheet order:
      0  Summary
      1  Forex (incl. Metals)         [curated]
      2  Crypto                       [curated]
      3  Indices & Stocks             [curated FX-equity / deferred]
      4  Dual-TF Shortlist            [1d×252 ∩ 4h×1500]
      5  Friendly Universe            [rule-derived FX-FX roster; elite/broad split]
      6  All Pairs (Diagnostic)       [full pair-pair output]
      7  Singles (Diagnostic)         [full single-symbol output]
      8  History
      9  Notes
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    conn = connect(db_path)
    try:
        # PER-PAIR latest as_of (not global MAX). Different symbols have
        # different trading calendars: FX + most indices stop trading Friday,
        # but BTC/ETH trade weekends. A naive `MAX(as_of)` query on a Saturday
        # or Sunday returns only the crypto rows under today's date, leaving
        # FX/index pairs absent from `df_today` and blanking out their cells
        # in the curated asset-class tabs.
        df_today = pd.read_sql_query(
            f"""SELECT t.* FROM {TABLE_NAME} t
                WHERE t.as_of = (
                    SELECT MAX(as_of) FROM {TABLE_NAME}
                    WHERE pair_a = t.pair_a AND pair_b = t.pair_b
                      AND tf = t.tf AND lookback_days = t.lookback_days
                )
                ORDER BY t.pair_a, t.pair_b, t.lookback_days""",
            conn,
        )
        df_singles = pd.read_sql_query(
            f"""SELECT t.* FROM {SINGLES_TABLE_NAME} t
                WHERE t.as_of = (
                    SELECT MAX(as_of) FROM {SINGLES_TABLE_NAME}
                    WHERE symbol = t.symbol AND tf = t.tf
                      AND lookback_days = t.lookback_days
                )
                ORDER BY t.symbol, t.lookback_days""",
            conn,
        ) if _table_exists(conn, SINGLES_TABLE_NAME) else pd.DataFrame()
        candidates = load_candidates()

        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        # 0 — Summary (always first)
        _write_summary(wb, conn, df_today)

        # 1-3 — asset class tabs in canonical order
        class_order = ["forex", "crypto", "indices_stocks"]
        pos = 1
        for class_key in class_order:
            if class_key in candidates:
                _write_asset_class_tab(
                    wb, conn, class_key, candidates[class_key],
                    df_today, df_singles, position=pos,
                )
                pos += 1

        # 4 — Dual-TF Shortlist (Thread B per-pair deep-dive surface)
        _write_dual_tf_shortlist(wb, conn, df_today, position=pos)
        pos += 1

        # 5 — Friendly Universe (rule-derived FX-FX roster; elite/broad split)
        _write_friendly_universe(wb, conn, df_today, position=pos)
        pos += 1

        # 6 — All Pairs (Diagnostic) — appended after the operator tabs
        _write_today(wb, conn, df_today)
        wb["Today"].title = "All Pairs (Diagnostic)"

        # 7 — Singles (Diagnostic); 8 — History; 9 — Notes (all appended last)
        _write_singles_diagnostic(wb, conn, df_singles, position=len(wb.sheetnames))
        _write_history(wb, conn)
        _write_notes(wb)
        _atomic_save_workbook(wb, output_path)
    finally:
        conn.close()
    return output_path


def _table_exists(conn: sqlite3.Connection, name: str) -> bool:
    row = conn.execute(
        "SELECT name FROM sqlite_master WHERE type='table' AND name=?",
        (name,),
    ).fetchone()
    return row is not None


def _parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        description="Cointegration screener — Phase 3 SQLite → Excel."
    )
    p.add_argument("--export", action="store_true",
                   help="Regenerate Cointegration_Screener.xlsx from SQLite.")
    p.add_argument("--db", type=str, default=str(SQLITE_DB))
    p.add_argument("--output", type=str, default=str(EXCEL_PATH))
    return p


def main(argv: list[str] | None = None) -> int:
    args = _parser().parse_args(argv)
    if not args.export:
        print("Specify --export to regenerate the Excel file.")
        return 2
    out = export_excel(args.db, args.output)
    print(f"[cointegration_excel] wrote {out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
