"""
format_excel_artifact.py — Unified Excel Formatter & Rounding Governance

Applies strict styling and number formatting to Excel artifacts.
Logic is presentation-layer only. No data mutation.

Usage:
    python tools/format_excel_artifact.py --file <path> --profile <strategy|portfolio>
"""

import argparse
import sys
import openpyxl
from pathlib import Path
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ==========================================================
# CONSTANTS
# ==========================================================

# Styling
HEADER_FILL_COLOR = "4472C4"  # Lighter Blue (SOP Governance)
HEADER_FONT_COLOR = "FFFFFF"
ALT_ROW_FILL_COLOR = "DCE6F1"

# Number Formats
FMT_CURRENCY = "#,##0.00"   # USD (2 decimals, comma)
FMT_PERCENT = "0.00%"       # Percent (2 decimals: 0.1234 -> 12.34%)
FMT_FLOAT = "0.00"          # Ratio/Float (2 decimals: 1.23)
FMT_INT = "0"               # Integer

# Column Rules (Column Name -> Format)
# Keys should be lower-case for matching
FORMAT_MAP = {
    # Trades List Headers (Human Readable)
    "pnl (usd)": FMT_CURRENCY,
    "entry price": FMT_FLOAT,
    "exit price": FMT_FLOAT,
    "bars held": FMT_INT,
    
    # Summary Metrics (Human Readable - Lowercase)
    "net profit (usd)": FMT_CURRENCY,
    "gross profit (usd)": FMT_CURRENCY,
    "gross loss (usd)": FMT_CURRENCY,
    "starting capital": FMT_CURRENCY,
    "max drawdown (usd)": FMT_CURRENCY,
    "largest win (usd)": FMT_CURRENCY,
    "largest loss (usd)": FMT_CURRENCY,
    "expectancy (usd)": FMT_CURRENCY,
    "avg trade (usd)": FMT_CURRENCY,
    "avg win (usd)": FMT_CURRENCY,
    "avg loss (usd)": FMT_CURRENCY,
    
    "max drawdown (%)": FMT_PERCENT,
    "% profitable": FMT_PERCENT,
    "% of gross profit (top trades)": FMT_PERCENT,
    "worst 5 trades loss %": FMT_PERCENT,
    "return on capital": FMT_PERCENT,
    "% time in market": FMT_PERCENT,
    
    "profit factor": FMT_FLOAT,
    "sharpe ratio": FMT_FLOAT,
    "sortino ratio": FMT_FLOAT,
    "k-ratio": FMT_FLOAT,
    "sqn (system quality number)": FMT_FLOAT,
    "return / drawdown ratio": FMT_FLOAT,
    "return retracement ratio": FMT_FLOAT,
    "avg mfe (r)": FMT_FLOAT,
    "avg mae (r)": FMT_FLOAT,
    "edge ratio (mfe / mae)": FMT_FLOAT,
    "win/loss ratio": FMT_FLOAT,
    "avg bars in winning trades": FMT_FLOAT,
    "avg bars in losing trades": FMT_FLOAT,
    "avg bars per trade": FMT_FLOAT,
    
    # Strategy Master Filter Keys (Stage-3)
    "total_net_profit": FMT_CURRENCY,
    "expectancy": FMT_CURRENCY,
    "sharpe_ratio": FMT_FLOAT,
    "return_dd_ratio": FMT_FLOAT,
    "worst_5_loss_pct": FMT_PERCENT,
    "max_drawdown": FMT_CURRENCY, # "max_dd_usd" is legacy
    "longest_loss_streak": FMT_INT,
    "pct_time_in_market": FMT_PERCENT,
    "avg_bars_in_trade": FMT_FLOAT,
    "net_profit_high_vol": FMT_CURRENCY,
    "net_profit_normal_vol": FMT_CURRENCY,
    "net_profit_low_vol": FMT_CURRENCY,
    
    # Legacy / Internal Keys (Keep for compatibility)
    "net_pnl_usd": FMT_CURRENCY,
    "net_profit": FMT_CURRENCY, 
    "gross_profit": FMT_CURRENCY,
    "gross_loss": FMT_CURRENCY,
    "max_dd_usd": FMT_CURRENCY,
    "max_dd_pct": FMT_PERCENT,
    "win_rate": FMT_PERCENT,
    "sharpe": FMT_FLOAT,
    "profit_factor": FMT_FLOAT,
    "total_trades": FMT_INT,
    "winning_trades": FMT_INT,
    "losing_trades": FMT_INT,
    "max_consec_wins": FMT_INT,
    "max_consec_losses": FMT_INT,
    "consecutive_wins": FMT_INT,
    "consecutive_losses": FMT_INT
}

# Hidden Columns (SOP Auditing Fields)
HIDDEN_COLS = {
    "constituent_run_ids",
    "run_ids",
    "genome_id",
    "run_id",
    "creation_timestamp",
    "timestamp"
}

# ==========================================================
# LOGIC
# ==========================================================

def apply_formatting(file_path, profile):
    path = Path(file_path)
    if not path.exists():
        print(f"[ERROR] File not found: {file_path}")
        sys.exit(1)

    print(f"[INFO] Formatting {path.name} (Profile: {profile})...")
    
    try:
        wb = openpyxl.load_workbook(path)
        
        # Styles Objects
        header_fill = PatternFill(start_color=HEADER_FILL_COLOR, end_color=HEADER_FILL_COLOR, fill_type="solid")
        header_font = Font(bold=True, color=HEADER_FONT_COLOR)
        alt_row_fill = PatternFill(start_color=ALT_ROW_FILL_COLOR, end_color=ALT_ROW_FILL_COLOR, fill_type="solid")
        
        # Iterate over ALL sheets
        for ws in wb.worksheets:
            print(f"  [INFO] Processing sheet: {ws.title}")
            
            # 1. Header & Column ID Mapping
            col_map = {} # col_index -> col_name_lower
            max_col = ws.max_column
            max_row = ws.max_row
            
            for col_idx in range(1, max_col + 1):
                cell = ws.cell(row=1, column=col_idx)
                val = str(cell.value).lower().strip() if cell.value else ""
                col_map[col_idx] = val
                
                # Header Styling
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

            # 2. Row Iteration (Data)
            for row_idx in range(2, max_row + 1):
                is_alt = (row_idx % 2 == 0)
                
                # Special Logic for Transposed Summary
                # If Sheet is "Performance Summary", Column 1 is Metric Name.
                # Use that to determine format for Col 2, 3...
                row_metric_fmt = None
                if ws.title == "Performance Summary":
                     metric_name = str(ws.cell(row=row_idx, column=1).value).lower().strip()
                     row_metric_fmt = FORMAT_MAP.get(metric_name)
                    
                for col_idx in range(1, max_col + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    col_name = col_map.get(col_idx, "")
                    
                    # Alternating Row Shading
                    if is_alt:
                        cell.fill = alt_row_fill
                    
                    # Formatting & Alignment
                    fmt = None
                    
                    # Strategy A: Column Header Match (Normal Sheets)
                    if col_name in FORMAT_MAP:
                        fmt = FORMAT_MAP[col_name]
                        
                    # Strategy B: Row Metric Match (Transposed Summary) -- Value Columns Only (Col > 1)
                    elif row_metric_fmt and col_idx > 1:
                        fmt = row_metric_fmt
                        
                    if fmt:
                        cell.number_format = fmt
                        cell.alignment = Alignment(horizontal="right")
                    else:
                        # Default Left Align for text/other (unless it's a value col in summary, assume right? No, stick to default)
                        # Actually for Summary Column 1 (headers), left is good.
                        if row_metric_fmt and col_idx > 1:
                             cell.alignment = Alignment(horizontal="right") # Align unformatted numbers right too
                        else:
                             cell.alignment = Alignment(horizontal="left")

            # 3. Column Widths & Hiding
            for col_idx in range(1, max_col + 1):
                col_name = col_map.get(col_idx, "")
                col_letter = get_column_letter(col_idx)
                
                # Hide specific columns
                if col_name in HIDDEN_COLS:
                    ws.column_dimensions[col_letter].hidden = True
                    continue
                    
                # Auto-fit (Smarter Logic)
                max_len = 0
                sample_rows = list(range(2, min(max_row, 50) + 1))
                
                is_numeric_col = False
                
                for r in sample_rows:
                    cell = ws.cell(row=r, column=col_idx)
                    val = cell.value
                    if val is not None:
                        # Check if numeric
                        if isinstance(val, (int, float)):
                            is_numeric_col = True
                            # Estimate formatted length (approx 0.00 or 0.00%)
                            # Raw str(float) is too long (decimals). 
                            # Most metrics are < 10 chars formatted.
                            # Let's cap contribution of numbers to avoid 1.3333333 expanding width
                            s_val = f"{val:.2f}" 
                            max_len = max(max_len, len(s_val))
                        else:
                            max_len = max(max_len, len(str(val)))
                
                # Base padding
                width = max_len + 2
                
                # Constraints
                if is_numeric_col:
                    # Metrics usually don't need huge width, but header might wrap to 2 lines
                    # If max_len is small (e.g. "1.23"), width=6. header "Sharpe Ratio" wrapped might be ~6? No, "Sharpe" is 6.
                    # constant minimum for numeric to look good
                    if width < 10: width = 10
                    if width > 18: width = 18 # Cap numeric columns tighter
                else:
                    # Text columns
                    if width < 8: width = 8
                    if width > 40: width = 40
                    
                ws.column_dimensions[col_letter].width = width

            # 4. Freeze Header
            ws.freeze_panes = "A2"
        
        wb.save(path)
        print("[SUCCESS] Formatting complete.")
        
    except Exception as e:
        print(f"[FATAL] Formatting failed: {e}")
        sys.exit(1)

def main():
    parser = argparse.ArgumentParser(description="Unified Excel Formatter")
    parser.add_argument("--file", required=True, help="Path to Excel file")
    parser.add_argument("--profile", required=True, choices=["strategy", "portfolio"], help="Formatting profile")
    
    args = parser.parse_args()
    
    apply_formatting(args.file, args.profile)

if __name__ == "__main__":
    main()
