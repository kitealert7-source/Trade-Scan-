"""Notes-sheet generation for aggregate ledgers.

Public entry point: add_notes_sheet_to_ledger(file_path, sheet_type).
Private helpers (_build_notes_<sheet_type>, _notes_write_<section>) are dispatched
by add_notes_sheet_to_ledger.
"""

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font


def _build_notes_master_filter(ws, _w, r, fonts, thresholds):
    """Write ALL Notes sections for sheet_type='master_filter'. Returns the next row."""
    bold, header, normal = fonts["bold"], fonts["header"], fonts["normal"]
    CAND = thresholds["cand"]

    # Section 1 — Sheet Purpose
    _w(r, 1, "SECTION 1 — SHEET PURPOSE", header); r += 1
    _w(r, 1,
       "Strategy_Master_Filter.xlsx — Raw Stage-3 aggregation of all completed backtest runs. "
       "Every run that completes Stage 1→3 is written here. No promotion filter applied. "
       "Rows are keyed by run_id and are append-only (existing rows never modified).",
       normal); r += 2

    # Section 2 — Measurement Basis (shared with candidates)
    r = _notes_write_measurement_basis(ws, _w, r, fonts)

    # Section 3 — Inclusion Criteria
    _w(r, 1, "SECTION 3 — INCLUSION CRITERIA", header); r += 1
    _w(r, 1, "Condition", bold); _w(r, 2, "Threshold", bold); r += 1
    for cond, val in [
        ("All Stage-3 completed runs",          "No filter — everything included"),
        ("Deduplication key",                   "run_id (one row per backtest run)"),
        ("Row mutation policy",                 "Append-only — rows never modified after write"),
    ]:
        _w(r, 1, cond, normal); _w(r, 2, val, normal); r += 1
    r += 1

    # Section 4 — Classification Rules (shared with candidates) + master_filter footnote
    r = _notes_write_candidate_status_rules(ws, _w, r, fonts, CAND)
    r += 1
    _w(r, 1,
       "Note: candidate_status is computed in Filtered_Strategies_Passed.xlsx (filter_strategies.py). "
       "This sheet (Strategy_Master_Filter) holds raw aggregated metrics only.",
       Font(italic=True, size=9))
    r += 1
    r += 1

    # Section 5 — Key Column Glossary (shared filter glossary)
    r = _notes_write_filter_glossary(ws, _w, r, fonts)
    return r


def _build_notes_candidates(ws, _w, r, fonts, thresholds):
    """Write ALL Notes sections for sheet_type='candidates'. Returns the next row."""
    bold, header, normal = fonts["bold"], fonts["header"], fonts["normal"]
    PROMO = thresholds["promo"]
    CAND = thresholds["cand"]

    # Section 1 — Sheet Purpose
    _w(r, 1, "SECTION 1 — SHEET PURPOSE", header); r += 1
    _w(r, 1,
       "Filtered_Strategies_Passed.xlsx — Candidate ledger. Contains all strategies that passed "
       "the promotion criteria gate. Rows are append-only (strategies that once passed are "
       "never evicted). Each row carries a candidate_status classification.",
       normal); r += 2

    # Section 2 — Measurement Basis (shared with master_filter)
    r = _notes_write_measurement_basis(ws, _w, r, fonts)

    # Section 3 — Inclusion Criteria (promotion gate thresholds)
    _w(r, 1, "SECTION 3 — INCLUSION CRITERIA", header); r += 1
    _w(r, 1, "Condition", bold); _w(r, 2, "Threshold", bold); r += 1
    for cond, val in [
        ("Total Trades",         f">= {PROMO['MIN_TRADES']}"),
        ("Profit Factor",        f">= {PROMO['MIN_PF']}"),
        ("Return / DD Ratio",    f">= {PROMO['MIN_RET_DD']}"),
        ("Expectancy (USD)",     f">= {PROMO['MIN_EXP']}"),
        ("Sharpe Ratio",         f">= {PROMO['MIN_SHARPE']}"),
        ("Max Drawdown (%)",     f"<= {PROMO['MAX_DD_PCT']:.0f}%"),
    ]:
        _w(r, 1, cond, normal); _w(r, 2, val, normal); r += 1
    r += 1

    # Section 4 — Classification Rules (candidate_status)
    r = _notes_write_candidate_status_rules(ws, _w, r, fonts, CAND)
    r += 1

    # Section 5 — Key Column Glossary (shared filter glossary)
    r = _notes_write_filter_glossary(ws, _w, r, fonts)
    return r


def _build_notes_portfolio(ws, _w, r, fonts, thresholds):
    """Write ALL Notes sections for sheet_type='portfolio'. Returns the next row."""
    bold, header, normal = fonts["bold"], fonts["header"], fonts["normal"]
    PORT = thresholds["port"]

    # Section 1 — Sheet Purpose
    _w(r, 1, "SECTION 1 — SHEET PURPOSE", header); r += 1
    _w(r, 1,
       "Master_Portfolio_Sheet.xlsx — Portfolio evaluation ledger. Two data tabs: "
       "'Portfolios' (multi-asset, 2+ symbols) and 'Single-Asset Composites' "
       "(single symbol, multiple strategies). Each row is evaluated through the "
       "capital wrapper and deployed profile selector. Append-only; explicit human "
       "authorization required to modify an existing entry. "
       "Rankings are independent per tab — scores are not compared across types.",
       normal); r += 2

    # Section 2 — Inclusion Criteria
    _w(r, 1, "SECTION 2 — INCLUSION CRITERIA", header); r += 1
    _w(r, 1, "Condition", bold); _w(r, 2, "Threshold", bold); r += 1
    for cond, val in [
        ("All strategies evaluated through Stage 4+",  "No promotion filter"),
        ("Deduplication key",                          "portfolio_id (append-only)"),
        ("Overwrite policy",                           "BLOCKED — requires explicit human authorization"),
    ]:
        _w(r, 1, cond, normal); _w(r, 2, val, normal); r += 1
    r += 1

    # Section 3 — Classification Rules (portfolio_status)
    _w(r, 1, "SECTION 3 — CLASSIFICATION RULES", header); r += 1
    _w(r, 1, "portfolio_status", bold); r += 1
    _w(r, 1, "Class", bold); _w(r, 2, "Rule", bold); r += 1
    for cls, rule in [
        ("FAIL",
         f"Realized PnL <= ${PORT['FAIL_PNL']:.0f}  OR  Accepted Trades < {PORT['FAIL_TRADES']}  "
         f"OR  Trade Density < {PORT['FAIL_DENSITY']} (per-symbol)  "
         f"OR  expectancy below asset-class gate"),
        ("",
         f"  Expectancy FAIL gates (same as candidates, min-lot basis):  "
         f"FX < ${PORT['FAIL_EXP_FX']}  |  XAU < ${PORT['FAIL_EXP_XAU']}  |  BTC < ${PORT['FAIL_EXP_BTC']}  |  INDEX < ${PORT['FAIL_EXP_IDX']}  |  MIXED (PF_ composites): no gate"),
        ("",
         "  Note on PF_ composites: Asset class is detected from portfolio_id keywords. "
         "PF_ hash-based IDs (multi-strategy portfolios) cannot be classified from the name alone. "
         "Their expectancy is a trade-count-weighted average across mixed asset classes, so "
         "per-class gates do not apply. If needed, constituent asset classes can be inspected "
         "in portfolio_metadata.json under evaluated_assets."),
        ("CORE (Portfolios tab)",
         f"Realized PnL > ${PORT['CORE_PNL']:.0f}  AND  Accepted Trades >= {PORT['CORE_TRADES']}  AND  Rejection Rate <= {PORT['CORE_REJ_MAX']:.0f}%  AND  edge_quality >= 0.12"),
        ("CORE (Single-Asset tab)",
         f"Realized PnL > ${PORT['CORE_PNL']:.0f}  AND  Accepted Trades >= {PORT['CORE_TRADES']}  AND  Rejection Rate <= {PORT['CORE_REJ_MAX']:.0f}%  AND  SQN >= 2.5"),
        ("WATCH (Portfolios tab)",
         "Passes all FAIL gates  AND  edge_quality >= 0.08  (but not CORE)"),
        ("WATCH (Single-Asset tab)",
         "Passes all FAIL gates  AND  SQN >= 2.0  (but not CORE)"),
    ]:
        _w(r, 1, cls, bold); _w(r, 2, rule, normal); r += 1
    r += 1

    # Section 4 — Key Column Glossary (portfolio glossary)
    r = _notes_write_portfolio_glossary(ws, _w, r, fonts)

    # Section 5 — Deployed Profile Selection (portfolio-only)
    r = _notes_write_deployed_profile(ws, _w, r, fonts)

    # Section 6 — Cointegration tab (separate ontology; not deployment-classified)
    r = _notes_write_cointegration_section(ws, _w, r, fonts)

    # Section 7 — COINT TRADE CANDIDATES tab (pair-level shortlist from §6)
    r = _notes_write_trade_candidates_section(ws, _w, r, fonts)
    return r


def _notes_write_measurement_basis(ws, _w, r, fonts):
    """Section 2: Measurement Basis — shared by master_filter and candidates."""
    bold, header, normal = fonts["bold"], fonts["header"], fonts["normal"]
    _w(r, 1, "SECTION 2 — MEASUREMENT BASIS", header); r += 1
    _w(r, 1, "Aspect", bold); _w(r, 2, "Value", bold); r += 1
    for aspect, value in [
        ("Position size",
         "Fixed at min_lot = 0.01 for EVERY trade, EVERY asset class "
         "(FX, XAU, BTC, INDEX). Pulled from data_access/broker_specs/OctaFx/{SYMBOL}.yaml."),
        ("Starting capital",
         "None. No account-size concept exists at this stage. Every dollar figure "
         "is raw trade-level P&L summed across min-lot positions."),
        ("reference_capital_usd",
         "$1,000 constant (run_metadata.json). Used ONLY to convert max_drawdown to "
         "a percentage (max_dd_pct = max_drawdown / $1,000). NOT spent, NOT compounded, "
         "NOT sized against. Just a yardstick for reporting."),
        ("Leverage / heat / margin",
         "Not modelled. Every signal fires at 0.01 lot regardless of concurrent "
         "exposure. No heat cap, no leverage cap, no margin check."),
        ("Rejection",
         "None. Zero trades are filtered at this stage. All signals execute."),
        ("total_net_profit / gross_profit / gross_loss",
         "Sum of min-lot USD P&L across every trade."),
        ("expectancy",
         "total_net_profit / total_trades, in USD at 0.01 lot. Example: 'XAU "
         "expectancy = 0.80' means each XAU trade averaged $0.80 profit at min-lot."),
        ("max_drawdown",
         "Peak-to-trough decline of the cumulative min-lot equity curve, in USD."),
        ("max_dd_pct",
         "max_drawdown / $1,000. NOT percent-of-deployed-capital — it is "
         "percent-of-a-$1,000-yardstick."),
        ("return_dd_ratio",
         "total_net_profit / max_drawdown. Unitless. Scale-invariant — would be "
         "identical under any capital model."),
        ("sharpe_ratio / sqn",
         "Trade-level (NOT annualized), computed on per-trade min-lot P&L. "
         "Unitless, capital-independent."),
        ("Regime / session net_profit_* buckets",
         "Same USD, same min-lot basis as total_net_profit — just sliced by "
         "volatility regime, trend regime, or session."),
        ("Design intent",
         "FSP is an apples-to-apples ranking layer. Every strategy is measured "
         "under identical conditions (0.01 lot, no capital constraints) so XAU_MR "
         "and BTC_TREND can be compared without one winning by capital allocation. "
         "Capital sizing, heat caps, leverage, and rejection all live downstream "
         "in Master_Portfolio_Sheet.xlsx + capital_wrapper.py."),
    ]:
        _w(r, 1, aspect, bold); _w(r, 2, value, normal); r += 1
    r += 1
    return r


def _notes_write_candidate_status_rules(ws, _w, r, fonts, CAND):
    """Section 4 classification table: candidate_status rules — shared by
    master_filter and candidates. Caller writes the trailing blank row / footnote."""
    bold, header, normal = fonts["bold"], fonts["header"], fonts["normal"]
    _w(r, 1, "SECTION 4 — CLASSIFICATION RULES", header); r += 1
    _w(r, 1, "candidate_status", bold); r += 1
    _w(r, 1, "Class", bold); _w(r, 2, "Rule", bold); r += 1
    for cls, rule in [
        ("FAIL",
         f"Total Trades < {CAND['FAIL_TRADES']}  OR  Max Drawdown (%) > {CAND['FAIL_DD']:.0f}  "
         f"OR  expectancy below asset-class gate"),
        ("",
         f"  Expectancy FAIL gates (logic-driven from broker spread/slippage ratios):  "
         f"FX < ${CAND['FAIL_EXP_FX']}  |  XAU < ${CAND['FAIL_EXP_XAU']}  |  BTC < ${CAND['FAIL_EXP_BTC']}  |  INDEX < ${CAND['FAIL_EXP_IDX']}"),
        ("CORE",
         f"Total Trades >= {CAND['CORE_TRADES']}  AND  Return/DD >= {CAND['CORE_RET_DD']}  AND  Sharpe >= {CAND['CORE_SHARPE']}  AND  Max DD <= {CAND['CORE_MAX_DD']:.0f}%  AND  Trade Density >= {CAND['CORE_DENSITY']}  AND  PF >= {CAND['CORE_PF']}  (and not FAIL)"),
        ("WATCH",
         "Does not meet CORE criteria; not FAIL; not present in TS_Execution/portfolio.yaml"),
        ("LIVE",
         "Present in TS_Execution/portfolio.yaml with enabled=true — overrides computed status. "
         "LIVE promotion thresholds: FX >= $0.25  |  XAU >= $0.80  |  BTC >= $0.80  |  INDEX >= $0.80"),
        ("REMOVE",
         "Removed from LIVE deployment. Strategy entered production but failed the Edge Quality Gate "
         "(Section 18 robustness suite). Reverted to research pool for rework."),
    ]:
        _w(r, 1, cls, bold); _w(r, 2, rule, normal); r += 1
    return r


def _notes_write_filter_glossary(ws, _w, r, fonts):
    """Section 5: Key Column Glossary — filter/candidate sheets (master_filter, candidates)."""
    bold, header, normal = fonts["bold"], fonts["header"], fonts["normal"]
    _w(r, 1, "SECTION 5 — KEY COLUMN GLOSSARY", header); r += 1
    _w(r, 1, "Column", bold); _w(r, 2, "Definition", bold); r += 1
    glossary = [
            ("run_id",           "Unique identifier for the backtest run (immutable once written)"),
            ("strategy",         "Full strategy ID including symbol suffix"),
            ("profit_factor",    "Gross profit / gross loss across all trades"),
            ("sharpe_ratio",     "Risk-adjusted return (mean trade PnL / std dev), trade-level"),
            ("max_dd_pct",       "Maximum drawdown as positive percentage (e.g. 3.18 = 3.18%)"),
            ("return_dd_ratio",  "Net profit / max drawdown — primary risk-efficiency metric"),
            ("total_trades",     "Number of completed trades in the backtest period"),
            ("expectancy",       "Average expected PnL per trade (USD) at min-lot (0.01). "
                                  "Computed in Stage 1 as sum(pnl_usd)/total_trades where pnl_usd uses "
                                  "broker min_lot (0.01) x contract_size. This normalizes all strategies "
                                  "to the same position size for apples-to-apples comparison. "
                                  "Deployed lot sizes (risk-based) produce proportionally larger PnL."),
            ("sqn",              "System Quality Number: sqrt(N) × mean(R) / stdev(R, ddof=1). "
                                  "Van Tharp metric for individual trading systems. "
                                  "Thresholds: <1.5 FAIL, 1.5–2.5 WATCH, ≥2.5 required for CORE."),
            ("Analysis_selection",
                                 "0/1 — transient user-intent flag. Picks the rows "
                                 "that will form the next composite_portfolio_analysis "
                                 "run (>=2 required; correlation/concurrency need a pair). "
                                 "Source of truth is master_filter.Analysis_selection in "
                                 "ledger.db; the column here is a read-only projection "
                                 "refreshed on every filter_strategies.py run — hand-edits "
                                 "to the Excel are NOT round-tripped back to the DB. "
                                 "Write via `control_panel.py --select-analysis <run_id>` "
                                 "(or menu option 10). Auto-cleared after a successful "
                                 "analysis run so the next session starts fresh. "
                                 "NOT a membership flag — deployment membership lives in "
                                 "TS_Execution/portfolio.yaml (LIVE rows)."),
            ("candidate_status", "CORE / WATCH / FAIL / LIVE / REMOVE — see Section 4"),
            ("is_current",       "1 = this row represents the live result for its (strategy, run_id) "
                                  "combination; 0 = superseded by a later rerun. Written by "
                                  "tools/rerun_backtest.py finalize via ledger_db.mark_superseded(). "
                                  "Append-only invariant preserved — superseded rows are flagged, never "
                                  "deleted. Filter with is_current=1 for all downstream analytics; "
                                  "is_current=0 rows are retained only for provenance and audit. "
                                  "NULL on pre-2026-04-16 rows (treat as 1 for backward compatibility)."),
            ("superseded_by",    "run_id of the replacement run that retired this row. NULL for live "
                                  "rows (is_current=1). Traces the lineage chain — a strategy may be "
                                  "rerun multiple times, each old row pointing to the next one."),
            ("superseded_at",    "ISO-8601 UTC timestamp when this row was flagged superseded. NULL "
                                  "for live rows."),
            ("supersede_reason", "Category tag written at finalize time: DATA_FRESH | SIGNAL | "
                                  "PARAMETER | ENGINE | BUG_FIX. Mirrors the category passed to "
                                  "tools/rerun_backtest.py prepare. Free-form extension allowed — "
                                  "audit_log (outputs/logs/rerun_audit.jsonl) carries the human "
                                  "reason string. NULL for live rows."),
            ("quarantined",      "1 = this row's result was semantically wrong (BUG_FIX rerun) and "
                                  "must NEVER be resurrected for analytics. 0 = normal (default). "
                                  "Distinguishes 'this was an older view of a still-valid backtest' "
                                  "(is_current=0, quarantined=0) from 'this was a broken computation' "
                                  "(is_current=0, quarantined=1). Set by rerun_backtest.py finalize "
                                  "--quarantine. NULL on pre-2026-04-16 rows (treat as 0)."),
    ]
    for col_name, definition in glossary:
        _w(r, 1, col_name, normal); _w(r, 2, definition, normal); r += 1
    r += 1
    return r


def _notes_write_portfolio_glossary(ws, _w, r, fonts):
    """Section 4: Key Column Glossary — portfolio sheet."""
    bold, header, normal = fonts["bold"], fonts["header"], fonts["normal"]
    _w(r, 1, "SECTION 4 — KEY COLUMN GLOSSARY", header); r += 1
    _w(r, 1, "Column", bold); _w(r, 2, "Definition", bold); r += 1
    glossary = [
        ("portfolio_id",                "Unique portfolio identifier"),
        ("theoretical_pnl",             "Raw Stage-4 aggregated PnL before sizing/rejection"),
        ("realized_pnl",                "Actual PnL after deployed profile sizing rules"),
        ("deployed_profile",            "Capital sizing profile selected by profile_selector"),
        ("expectancy",                  "Trade-count-weighted average PnL per trade (USD) at min-lot (0.01). "
                                        "Equals sum(pnl_usd)/total_trades across ALL constituent symbols. "
                                        "NOT the simple average of per-symbol expectancies — symbols with "
                                        "more trades naturally weight more. Computed from Stage 1 results "
                                        "at broker min_lot; deployed profile lot sizes are much larger."),
        ("sharpe",                      "Portfolio-level Sharpe ratio"),
        ("max_dd_pct",                  "Maximum portfolio drawdown as positive percentage"),
        ("return_dd_ratio",             "Realized PnL / max drawdown"),
        ("edge_quality",                "mean(R) / stdev(R) — edge consistency metric for portfolios. "
                                        "Unlike SQN, omits sqrt(N) so portfolios with different trade "
                                        "counts are directly comparable. "
                                        "Thresholds: <0.05 weak (edge ≈ noise), 0.05–0.10 moderate, "
                                        "0.10–0.15 good, >0.15 strong. "
                                        "Values >0.25 with low trade counts may indicate overfitting."),
        ("sqn",                         "System Quality Number: sqrt(N) × mean(R) / stdev(R, ddof=1). "
                                        "Van Tharp metric for individual trading systems. "
                                        "Thresholds: <1.6 poor, 1.6–2.0 average, 2.0–3.0 good, >3.0 excellent. "
                                        "Single-Asset tab only — see edge_quality for Portfolios tab."),
        ("rejection_rate_pct",          "Percentage of trades rejected by leverage/heat cap"),
        ("symbol_count",                "Number of distinct symbols contributing to this portfolio row. "
                                        "Derived from master_filter via constituent_run_ids; for single-asset "
                                        "rows = 1. Used to expose unit mismatch in composite densities."),
        ("trade_density_total",         "Theoretical portfolio trade density (trades/year), SUMMED across "
                                        "symbols. Per-symbol density is taken as MAX across re-runs in "
                                        "master_filter (most favorable parameterization per symbol). "
                                        "Total = sum(per_symbol_max). Pre-capital, pre-rejection."),
        ("trade_density_min",           "Theoretical MINIMUM per-symbol trade density (trades/year) in the "
                                        "portfolio. Same per-symbol source as trade_density_total, then "
                                        "min() across symbols. This is the FAIL-gate quantity: a portfolio "
                                        "with total density 450 but min density 24 has one symbol that is "
                                        "too sparse to deploy. FAIL gate: trade_density_min < 50."),

        ("profile_trade_density_total", "POST-FILTER per-symbol density, summed. Two-stage derivation: "
                                        "(1) per-symbol raw density = count in portfolio_tradelevel.csv "
                                        "per symbol / deployed profile's simulation_years (captures the "
                                        "specific parameterization actually selected per symbol, not the "
                                        "master_filter MAX); (2) apply deployed profile's rejection_rate_pct "
                                        "uniformly (portfolio-wide — capital wrapper does not emit per-symbol "
                                        "rejection). total = sum(adjusted_per_symbol). Fallback when "
                                        "tradelevel.csv is unavailable: trade_density_total x (1 - rej/100)."),
        ("profile_trade_density_min",   "POST-FILTER MINIMUM per-symbol density. Same derivation as "
                                        "profile_trade_density_total but min() instead of sum(). Computed "
                                        "INDEPENDENTLY from profile_trade_density_total — never derived "
                                        "from it. This is the truthful deployment calendar viability metric: "
                                        "it reflects both the selected parameterization per symbol AND the "
                                        "deployed profile's rejection filter."),
        ("avg_concurrent",              "Average number of simultaneously open positions"),
        ("max_concurrent",              "Peak simultaneous open positions"),
        ("avg_pairwise_corr",           "Average correlation between constituent strategy equity curves"),
        ("portfolio_status",            "CORE / WATCH / FAIL — see Section 3"),
        ("parsed_fields",               "JSON decomposition of portfolio_id tokens "
                                        "(idea_id, family, symbol, timeframe, model, filter, sweep, "
                                        "variant, param_set, asset_class). Hidden column."),
        ("n_strategies",                "Number of constituent strategies (Single-Asset tab only)"),
    ]
    for col_name, definition in glossary:
        _w(r, 1, col_name, normal); _w(r, 2, definition, normal); r += 1
    r += 1
    return r


def _notes_write_deployed_profile(ws, _w, r, fonts):
    """Section 5: Deployed Profile Reference — portfolio sheet only."""
    bold, header, normal = fonts["bold"], fonts["header"], fonts["normal"]
    _w(r, 1, "SECTION 5 — DEPLOYED PROFILE SELECTION", header); r += 2

    # 5a. Available profiles
    _w(r, 1, "Available Capital Profiles", bold); r += 1
    _w(r, 1, "Profile", bold); _w(r, 2, "Lot Sizing Approach", bold); r += 1
    for pname, desc in [
        ("RAW_MIN_LOT_V1",
         "Baseline diagnostic. $1000 seed. Unconditional min-lot (0.01) on every "
         "signal — bypasses all risk/heat/leverage gates via raw_lot_mode. "
         "Shows the pure directional edge of the strategy independent of sizing."),
        ("FIXED_USD_V1",
         "Retail-amateur conservative. $1000 seed, risk = max(2% of current equity, "
         "$20 floor). Compounds as equity grows; floor preserves meaningful trade "
         "size if equity dips below start. No heat/leverage caps (real retail has "
         "no portfolio heat monitor). Trades below min_lot SKIP honestly."),
        ("REAL_MODEL_V1",
         "Retail-amateur aggressive. $1000 seed, tier-ramp risk: 2% base, +1% per "
         "2x equity doubling, capped at 5% (symmetric retrace). No heat/leverage "
         "caps. Trades below min_lot SKIP. retail_max_lot=10 enforces OctaFx-"
         "realistic ceiling — trades requiring more than 10 lots SKIP. "
         "Every capital_wrapper run also emits a linear-normalized "
         "overlay_comparison.png alongside per-profile equity_curve.png files."),
    ]:
        _w(r, 1, pname, normal); _w(r, 2, desc, normal); r += 1
    r += 1

    # 5b. Selection algorithm
    _w(r, 1, "Selection Algorithm", bold); r += 1
    _w(r, 1, "Step", bold); _w(r, 2, "Description", bold); r += 1
    for step, desc in [
        ("1. Hard Filter",
         "Remove profiles with realized_pnl <= 0, "
         "capital_validity_flag = False, or avg_risk_multiple > 1.5."),
        ("2. Reliability Pool",
         "Prefer profiles with >= 50 accepted trades AND >= 1.0 sim years (reliable pool). "
         "Fall back to hard-valid-only pool if none qualify."),
        ("3. Score",
         "base_score = realized_pnl / max(max_drawdown_usd, 1.0). "
         "Penalized by execution health: "
         "rejection <= 30% = 1.0x, 30-60% = 0.7x, > 60% = 0.4x."),
        ("4. Rank",
         "Sort by: (1) highest score, (2) lowest rejection rate, "
         "(3) most accepted trades, (4) alphabetical name."),
        ("5. Stabilize",
         "Profiles within 15% of best score are considered tied. "
         "Tied candidates re-ranked by rejection rate and trade count."),
        ("6. Persist",
         "If a previous profile exists and scores >= 85% of the new best, "
         "keep the previous profile to avoid flip-flopping."),
    ]:
        _w(r, 1, step, normal); _w(r, 2, desc, normal); r += 1
    r += 1
    return r


def _notes_write_trade_candidates_section(ws, _w, r, fonts):
    """Section 7: COINT TRADE CANDIDATES tab — pair-level decision-support
    shortlist projected from the Cointegration research rows (one row per pair
    vs one row per run). Reliability-under-exposure, not perfection."""
    bold, header, normal = fonts["bold"], fonts["header"], fonts["normal"]

    _w(r, 1, "SECTION 7 — COINT TRADE CANDIDATES TAB", header); r += 1
    _w(r, 1,
       "Decision-support shortlist: one row per PAIR (the Cointegration tab is "
       "one row per backtest run). Answers a single question — 'after thousands "
       "of backtests, which handful of pairs should I focus research or capital "
       "on next?' QUALIFICATION: a pair must have >= 5 current runs to appear "
       "here; thinner pairs are universe-explorer noise and remain on the "
       "Cointegration tab only. Criterion is reliability-under-exposure, NOT "
       "perfection: all_profitable is demoted to a badge so a heavily-tested "
       "near-perfect pair (e.g. 24 runs / 2 losses) stays VISIBLE instead of "
       "being banished for one loss. Because the ledger is append-only, a thin "
       "pair GRADUATES IN as it gathers runs, and a pair drifts DOWN the ranking "
       "as its robustness decays — it never vanishes on a single future loss.",
       normal); r += 2

    _w(r, 1, "Columns", bold); r += 1
    _w(r, 1, "Column", bold); _w(r, 2, "Definition", bold); r += 1
    for col, definition in [
        ("Pair",
         "pair_a / pair_b. Prefixed with the medal badge when the pair has zero "
         "losses so far (= the retired all_profitable 'never lost yet' flag, kept "
         "as an achievement, not a filter)."),
        ("Coint Status (252d)",
         "Current cointegration regime for the pair from the daily screener's "
         "252-day window (cointegration_daily — the source behind the screener's "
         "'All Pairs (Diagnostic)' sheet): cointegrated / breaking / broken. "
         "Refreshed on each MPS regeneration; blank if the pair is not in the "
         "current screen. NOTE: this is the standard 252d screen as of its latest "
         "run, NOT the per-pair continuous-cointegration window the candidate was "
         "actually backtested on."),
        ("Runs",
         "Total current (is_current=1) runs for the pair — every parameter "
         "variant and every test window."),
        ("Losses",
         "Count of runs with canonical_net_pct <= 0 (break-even counts as a loss). "
         "Shown as a RAW COUNT for instant legibility ('24 runs, 2 losses')."),
        ("Median Ret/DD",
         "Median canonical_ret_dd over ALL runs (including losers). Median, not "
         "mean, so one outlier run cannot dominate."),
    ]:
        _w(r, 1, col, bold); _w(r, 2, definition, normal); r += 1
    r += 1

    _w(r, 1, "Sort order", bold); r += 1
    _w(r, 1,
       "loss_rate ascending  ->  Median Ret/DD descending  ->  Runs descending. "
       "loss_rate (losses/runs) drives SORTING only; the displayed Losses column "
       "stays a raw count (24/2 and 12/1, both ~8%, rank together). The runs>=5 "
       "gate already certifies evidence, so once loss_rate ties, QUALITY (median "
       "Ret/DD) leads and runs is only the final tiebreak — a clean 6-run/3.9 "
       "pair outranks a clean 12-run/1.0 one.",
       normal); r += 2
    return r


def _notes_write_cointegration_section(ws, _w, r, fonts):
    """Section 6: Cointegration tab — separate ontology from Portfolios /
    Single-Asset Composites. Point-in-time research provenance, not a
    deployment ledger. Ranks by Ret/DD only; no verdict_status column."""
    bold, header, normal = fonts["bold"], fonts["header"], fonts["normal"]

    # 6a. Sheet Purpose
    _w(r, 1, "SECTION 6 — COINTEGRATION TAB", header); r += 1
    _w(r, 1,
       "Cointegration tab — research-stage record of regime-conditioned "
       "cointegration backtests. Distinct ontology from the deployment-classified "
       "Portfolios / Single-Asset Composites tabs: there is NO verdict_status "
       "(CORE / WATCH / FAIL) on this sheet. Rows are ranked by Ret/DD only; "
       "operator filtering carries the screening. Each row = one (pair, episode, "
       "variant) backtest stamped with full screener provenance at test_end. "
       "Two methodology cohorts coexist: v1_raw_adf (legacy, retired) and "
       "v2_log_eg (current, post-2026-05-30 EG/MacKinnon math correction). They "
       "are NOT comparable head-to-head — always filter by methodology before "
       "ranking.",
       normal); r += 2

    # 6b. Filter aids (the three columns operators use to narrow the universe)
    _w(r, 1, "Filter aids (columns 5-7 — added 2026-06-01)", bold); r += 1
    _w(r, 1, "Column", bold); _w(r, 2, "Definition", bold); r += 1
    filter_aids = [
        ("pair_class",
         "Structural taxonomy of the leg pair. Closed value set "
         "(no Unknown — silent expansion fails loudly). "
         "FX = both legs FX (majors or crosses). "
         "IDX = both legs equity indices. "
         "Cross = one FX leg + one IDX leg. "
         "Crypto = either leg in {BTCUSD, ETHUSD} — overrides FX/IDX. "
         "Metals = either leg in {XAUUSD, XAGUSD} — overrides FX/IDX. "
         "(BTCUSD/EURUSD lands in Crypto, NOT Cross.)"),
        ("coint_friendly",
         "Screener-side cointegration-strength band from continuous_span_obs "
         "at the run's test_end (point-in-time provenance, NOT lifetime-max). "
         "STRONG = >= 90 obs (top ~2% of corpus — matches 2026-05-28 cohort-shift "
         "survey). FRIENDLY = >= 30 obs (B-gate admissible today). WEAK = < 30 obs "
         "or NaN. SEMANTIC CAVEAT: episodes tested on short windows (e.g., "
         "post-onset 1-7 day windows from the v2 generation rule) carry low span "
         "at test_end and land as WEAK even when the underlying pair has had long "
         "stable cointegration arcs elsewhere. Do NOT dismiss WEAK wholesale — "
         "rank by return_dd_ratio to find robust short-window edges. "
         "Lifetime-peak alternative is a separate column candidate, not built."),
    ]
    for col_name, definition in filter_aids:
        _w(r, 1, col_name, normal); _w(r, 2, definition, normal); r += 1
    r += 1

    # 6c. Suggested filter sequence
    _w(r, 1, "Suggested filter sequence (Excel autofilter)", bold); r += 1
    _w(r, 1, "Step", bold); _w(r, 2, "Filter", bold); r += 1
    for step, desc in [
        ("1. Methodology",
         "methodology = v2_log_eg (drops legacy v1_raw_adf rows — different math)"),
        ("2. Robustness",
         "n_spans >= 2 (recurred at least twice — excludes one-off episodes like "
         "the BTCUSD/EUSTX50 single-span trap)"),
        ("3. Design target",
         "pair_class IN {FX, IDX} for the cointegration design target, "
         "OR Cross/Crypto/Metals for cross-class exploration"),
        ("4. Strength",
         "coint_friendly IN {STRONG, FRIENDLY} for B-gate admissible; "
         "OR include WEAK if testing short-window edges"),
        ("5. Rank",
         "Sort by return_dd_ratio descending (already the sheet's primary sort) "
         "— Ret/DD is the deployment-immune research evaluator; ignore CORE/WATCH/"
         "FAIL semantics on this sheet (they do not exist here by design)."),
    ]:
        _w(r, 1, step, normal); _w(r, 2, desc, normal); r += 1
    r += 1

    # 6d. Remaining column glossary
    _w(r, 1, "Cointegration column glossary", bold); r += 1
    _w(r, 1, "Column", bold); _w(r, 2, "Definition", bold); r += 1
    coint_glossary = [
        ("rank",
         "1-based row position after the deterministic sort "
         "(return_dd_ratio desc, completed_at_utc desc, run_id desc). "
         "Recomputed on every export — not stored."),
        ("pair",
         "Display string 'pair_a / pair_b'. Underlying legs are separate "
         "columns in the DB (pair_a, pair_b)."),
        ("timeframe / lookback",
         "Execution timeframe of the backtest (e.g., 15m); screener lookback "
         "in days (typically 252 = 1 trading year)."),
        ("run_date / test_start / test_end",
         "Date the run completed (UTC); episode start and end bounds. "
         "Episodes target the active cointegration arc — entry at onset+N+1, "
         "exit at last_coint_idx per the v2 generation rule. Test windows are "
         "structurally short (per [[feedback_test_window_must_match_signal_class]])."),
        ("return_dd_ratio",
         "PRIMARY METRIC. canonical_net_pct / canonical_max_dd_pct. "
         "Scale-invariant; capital-model-independent. The research evaluator "
         "per [[feedback_screening_rules_for_research]] — use this, NOT verdict "
         "thresholds, for short-window research tests."),
        ("net_pct / max drawdown %",
         "Canonical net % return and peak-relative drawdown % for the run "
         "(mark-to-market, includes floating). At fixed $1,000 stake, net_pct = "
         "5.0 means $50 USD net."),
        ("final_equity_usd",
         "stake_usd + realized PnL at test_end (USD)."),
        ("total_trades / cycles / win_rate",
         "Trade count, completed (open→close) cycles, and cycle-level win-rate %. "
         "Zcross variant typically produces MORE cycles than baseline (z-cross "
         "exit fires earlier and resets) — directly comparable cycle-mean PnL "
         "across variants understates the variant's trade pace."),
        ("regime",
         "Regime state as-of test_end (point-in-time provenance). 'cointegrated' "
         "is the expected/admissible state."),
        ("methodology",
         "Cohort tag: v1_raw_adf (legacy raw-spread + plain unit-root criticals; "
         "RETIRED 2026-05-30) | v2_log_eg (current, log prices + Engle-Granger "
         "MacKinnon criticals). Pair stats are NOT comparable across cohorts."),
        ("backtest",
         "Hyperlink-rendered run-folder identifier "
         "(e.g., 90_PORT_<pair>_<TF>_COINTREV_V3_L30__E<yyyymmdd>_<pair> for "
         "baseline; _ZCRS_ token marks zcross variant). Click in Excel to open."),
    ]
    for col_name, definition in coint_glossary:
        _w(r, 1, col_name, normal); _w(r, 2, definition, normal); r += 1
    r += 1
    return r


def add_notes_sheet_to_ledger(file_path: str, sheet_type: str) -> None:
    """
    Append a 'Notes' sheet to an aggregate ledger Excel file.

    sheet_type:
      "master_filter" — Strategy_Master_Filter.xlsx (Stage-3 raw aggregation)
      "candidates"    — Filtered_Strategies_Passed.xlsx (promotion-filtered + classified)
      "portfolio"     — Master_Portfolio_Sheet.xlsx (multi-symbol portfolio evaluations)

    Thresholds are sourced from:
      candidates  → filter_strategies._compute_candidate_status() and mask
      portfolio   → portfolio_evaluator._compute_portfolio_status()
    No logic is redefined; constants are reproduced here for display only.
    """
    # ── Classification thresholds per sheet type ──────────────────────────────
    # candidates: promotion mask (filter_strategies.py lines 165–172)
    thresholds = {
        "promo": {
            "MIN_TRADES":  40,
            "MIN_PF":      1.05,
            "MIN_RET_DD":  0.6,
            "MIN_EXP":     0.0,
            "MIN_SHARPE":  0.3,
            "MAX_DD_PCT":  80.0,
        },
        # candidates: candidate_status rules (filter_strategies._compute_candidate_status)
        "cand": {
            "FAIL_TRADES":  50,
            "FAIL_DD":      40.0,
            "FAIL_EXP_FX":  0.15,   # FX expectancy FAIL gate (USD/trade)
            "FAIL_EXP_XAU": 0.50,   # XAU expectancy FAIL gate (3.3x FX)
            "FAIL_EXP_BTC": 0.50,   # BTC/crypto expectancy FAIL gate (3.3x FX)
            "FAIL_EXP_IDX": 0.50,   # INDEX expectancy FAIL gate (3.0-3.8x FX)
            "CORE_TRADES":  200,
            "CORE_RET_DD":  2.0,
            "CORE_SHARPE":  1.5,
            "CORE_MAX_DD":  30.0,
            "CORE_DENSITY": 50,
            "CORE_PF":      1.25,
        },
        # portfolio: portfolio_status rules (portfolio_evaluator._compute_portfolio_status)
        "port": {
            "FAIL_PNL":     0.0,
            "FAIL_TRADES":  50,
            "FAIL_DENSITY": 50,
            "FAIL_EXP_FX":  0.15,   # Same gates as candidates (min-lot basis)
            "FAIL_EXP_XAU": 0.50,
            "FAIL_EXP_BTC": 0.50,
            "FAIL_EXP_IDX": 0.50,
            "CORE_PNL":     1000.0,
            "CORE_TRADES":  200,
            "CORE_REJ_MAX": 30.0,
        },
    }

    try:
        wb = load_workbook(file_path)
    except Exception as e:
        print(f"[WARN] Notes sheet skipped — cannot open {file_path}: {e}")
        return

    if "Notes" in wb.sheetnames:
        del wb["Notes"]
    ws = wb.create_sheet("Notes")

    fonts = {
        "bold":   Font(bold=True, size=10),
        "header": Font(bold=True, size=11),
        "normal": Font(size=10),
    }

    # Wrap long text in column B (definitions/rules); top-align so multi-line
    # cells read correctly next to short labels in column A.
    _wrap_align = Alignment(wrap_text=True, vertical="top")
    _top_align = Alignment(vertical="top")

    def _w(row, col, value, font=None):
        c = ws.cell(row=row, column=col, value=value)
        if font:
            c.font = font
        c.alignment = _wrap_align if col == 2 else _top_align

    r = 1
    if sheet_type == "master_filter":
        r = _build_notes_master_filter(ws, _w, r, fonts, thresholds)
    elif sheet_type == "candidates":
        r = _build_notes_candidates(ws, _w, r, fonts, thresholds)
    elif sheet_type == "portfolio":
        r = _build_notes_portfolio(ws, _w, r, fonts, thresholds)

    # ── Column widths ─────────────────────────────────────────────────────────
    # Column A must fit "SECTION 4 — KEY COLUMN GLOSSARY" (32 chars) and
    # the widest glossary labels ("profile_trade_density_total" = 27 chars).
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 144

    try:
        # Resilient SSOT save (kill-Excel-if-locked + atomic temp-swap + backoff).
        # The --notes-type step runs right after --profile on every refresh.
        from tools.pipeline_utils import resilient_xlsx_write
        resilient_xlsx_write(file_path, lambda p: wb.save(str(p)))
        print(f"[NOTES] Notes sheet added ({sheet_type}) -> {Path(file_path).name}")
    except Exception as e:
        print(f"[WARN] Notes sheet could not be saved: {e}")
