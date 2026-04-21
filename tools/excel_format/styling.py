"""Excel data-sheet styling — header fills, number formats, widths, dropdowns, per-profile actions.

Public entry point: apply_formatting(file_path, profile).
Private helpers (_preprocess_with_pandas … _restore_hyperlinks) are orchestrated by apply_formatting.
"""

import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from .rules import (
    ALT_ROW_FILL_COLOR,
    COLUMN_WIDTH_OVERRIDES,
    DROPDOWN_COLS,
    FORMAT_MAP,
    HEADER_FILL_COLOR,
    HEADER_FONT_COLOR,
    HEADER_ROW_HEIGHT,
    HIDDEN_COLS,
    NARROW_TEXT_COLS,
    PORTFOLIO_COLUMN_ORDER,
    PORTFOLIO_HEADER_DISPLAY,
    SHADOW_TRADES_COLUMN_ORDER,
    SINGLE_ASSET_COLUMN_ORDER,
    STRATEGY_COLUMN_ORDER,
)


def _preprocess_with_pandas(path, profile):
    """Pandas-based pre-step: sort by return_dd_ratio, inject missing columns, reorder.

    Only runs on single-sheet ledgers and recognised multi-sheet layouts (portfolio
    Portfolios/Single-Asset Composites, strategy Sheet1+Notes). Other multi-sheet
    workbooks are left untouched.
    """
    try:
        import pandas as pd

        xl = pd.ExcelFile(path)
        sheet_names = xl.sheet_names

        # Portfolio workbooks may have multiple data sheets (Portfolios,
        # Single-Asset Composites) — process each independently.
        _PORTFOLIO_DATA_SHEETS = {"Portfolios", "Single-Asset Composites"}
        _is_portfolio_multi = (profile == "portfolio"
                               and any(s in _PORTFOLIO_DATA_SHEETS for s in sheet_names))

        # Strategy ledgers (e.g. Filtered_Strategies_Passed) have Sheet1 as the
        # data sheet plus ancillary sheets (Notes, etc.) — process Sheet1 only.
        _is_strategy_with_notes = (profile == "strategy"
                                   and "Sheet1" in sheet_names
                                   and len(sheet_names) > 1)

        if _is_portfolio_multi or _is_strategy_with_notes or len(sheet_names) == 1:
            # Determine which sheets to process
            if _is_portfolio_multi:
                _sheets_to_process = [s for s in sheet_names
                                      if s in _PORTFOLIO_DATA_SHEETS]
            elif _is_strategy_with_notes:
                _sheets_to_process = ["Sheet1"]
            else:
                _sheets_to_process = [sheet_names[0]]

            _all_dfs = {}  # sheet_name -> processed df
            # Load non-data sheets (Notes, etc.) for preservation
            _preserve_sheets = [s for s in sheet_names
                                if s not in _sheets_to_process
                                and s not in _PORTFOLIO_DATA_SHEETS]

            for sheet_name in _sheets_to_process:
                df = pd.read_excel(path, sheet_name=sheet_name)

                # --- Sort by return_dd_ratio (descending) ---
                sort_col = None
                if "return_dd_ratio" in df.columns:
                    sort_col = "return_dd_ratio"

                if sort_col:
                    df[sort_col] = pd.to_numeric(df[sort_col], errors="coerce")
                    df = df.sort_values(by=sort_col, ascending=False, na_position="last")
                    df = df.reset_index(drop=True)

                    # Add/update rank column (1-based, independent per sheet)
                    if "rank" not in df.columns:
                        df.insert(0, "rank", range(1, len(df) + 1))
                    else:
                        df["rank"] = range(1, len(df) + 1)
                    print(f"  [RANKED] {sheet_name}: Sorted by {sort_col} descending ({len(df)} rows)")

                # --- Column Reorder ---
                col_order = None
                _fname_lower = Path(path).stem.lower()
                if "shadow_trade" in _fname_lower:
                    col_order = SHADOW_TRADES_COLUMN_ORDER
                elif profile == "strategy" and STRATEGY_COLUMN_ORDER:
                    col_order = ["rank"] + STRATEGY_COLUMN_ORDER
                elif profile == "portfolio":
                    if sheet_name == "Single-Asset Composites":
                        col_order = SINGLE_ASSET_COLUMN_ORDER
                    else:
                        col_order = PORTFOLIO_COLUMN_ORDER

                if col_order:
                    # Drop sqn from the MPS Portfolios tab only — it belongs
                    # exclusively to the Single-Asset Composites tab in MPS.
                    # Do NOT drop it for FSP (Sheet1) or any other strategy sheet.
                    if profile == "portfolio" and sheet_name == "Portfolios" and "sqn" in df.columns:
                        df = df.drop(columns=["sqn"])
                    # Inject missing ordered columns as 0.0 so they appear at the right position
                    for col in col_order:
                        if col not in df.columns:
                            df[col] = 0.0
                    existing = df.columns.tolist()
                    ordered = [c for c in col_order if c in existing]
                    remaining = [c for c in existing if c not in ordered]
                    new_order = ordered + remaining
                    if new_order != existing:
                        df = df[new_order]
                        print(f"  [REORDER] {sheet_name}: Columns reordered ({len(ordered)} matched, {len(remaining)} appended)")

                _all_dfs[sheet_name] = df

            # Save all processed sheets back.
            with pd.ExcelWriter(path, engine="openpyxl", mode="w") as writer:
                for sname, sdf in _all_dfs.items():
                    sdf.to_excel(writer, sheet_name=sname, index=False)
                # Preserve non-data sheets (will be re-created by notes pass, but
                # keep existing content if notes pass doesn't run)
                for ps in _preserve_sheets:
                    try:
                        pdf = pd.read_excel(xl, sheet_name=ps)
                        pdf.to_excel(writer, sheet_name=ps, index=False)
                    except Exception:
                        pass
        else:
            print(f"  [INFO] Pre-processing skipped (multi-sheet workbook: {len(sheet_names)} sheets)")
    except Exception as e:
        print(f"  [WARN] Pre-processing skipped: {e}")


def _apply_header_row(ws, col_map, max_col, profile, header_font, header_fill):
    """Style the header row and populate col_map with lower-cased column names."""
    for col_idx in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col_idx)
        val = str(cell.value).lower().strip() if cell.value else ""
        col_map[col_idx] = val

        # Portfolio profile: add clarification comments (NOT renames —
        # renaming headers causes column-name collisions when the evaluator
        # appends new rows with the canonical column names).
        if profile == "portfolio" and val in PORTFOLIO_HEADER_DISPLAY:
            from openpyxl.comments import Comment
            note = {"sharpe": "Annualized (x √252)", "equity_stability_k_ratio": "Computed on log(daily equity)"}
            cell.comment = Comment(note.get(val, PORTFOLIO_HEADER_DISPLAY[val]), "System")

        # Header Styling
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _format_data_rows(ws, col_map, max_col, max_row, alt_row_fill):
    """Apply alternating fill + number formatting + alignment to data rows."""
    for row_idx in range(2, max_row + 1):
        is_alt = (row_idx % 2 == 0)

        # Special Logic for Transposed Summary
        # If Sheet is "Performance Summary", Column 1 is Metric Name.
        # Use that to determine format for Col 2, 3...
        row_metric_fmt = None
        if ws.title == "Performance Summary":
             metric_name = str(ws.cell(row=row_idx, column=1).value).lower().strip()
             row_metric_fmt = FORMAT_MAP.get(metric_name)

        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            col_name = col_map.get(col_idx, "")

            # Alternating Row Shading
            if is_alt:
                cell.fill = alt_row_fill

            # Formatting & Alignment
            fmt = None

            # Strategy A: Column Header Match (Normal Sheets)
            if col_name in FORMAT_MAP:
                fmt = FORMAT_MAP[col_name]

            # Strategy B: Row Metric Match (Transposed Summary) -- Value Columns Only (Col > 1)
            elif row_metric_fmt and col_idx > 1:
                fmt = row_metric_fmt

            if fmt:
                cell.number_format = fmt
                cell.alignment = Alignment(horizontal="right")
            else:
                # Default Left Align for text/other (unless it's a value col in summary, assume right? No, stick to default)
                # Actually for Summary Column 1 (headers), left is good.
                if row_metric_fmt and col_idx > 1:
                     cell.alignment = Alignment(horizontal="right") # Align unformatted numbers right too
                else:
                     cell.alignment = Alignment(horizontal="left")


def _apply_column_widths(ws, col_map, max_col, max_row):
    """Set column widths (data-driven for numerics/narrow-text, header-floored for text). Hides HIDDEN_COLS."""
    for col_idx in range(1, max_col + 1):
        col_name = col_map.get(col_idx, "")
        col_letter = get_column_letter(col_idx)

        # Hide specific columns
        if col_name in HIDDEN_COLS:
            ws.column_dimensions[col_letter].hidden = True
            continue

        # Measure max data width from sample rows
        max_data_len = 0
        is_numeric_col = False
        sample_rows = list(range(2, min(max_row, 50) + 1))

        for r in sample_rows:
            cell = ws.cell(row=r, column=col_idx)
            val = cell.value
            if val is not None:
                if isinstance(val, (int, float)):
                    is_numeric_col = True
                    max_data_len = max(max_data_len, len(f"{val:.2f}"))
                else:
                    max_data_len = max(max_data_len, len(str(val)))

        # Header length — use longest word if wrapping
        header_text = str(ws.cell(row=1, column=col_idx).value or "")
        header_words = header_text.replace("_", " ").split()
        longest_word = max((len(w) for w in header_words), default=0)

        # Width: numeric columns are data-driven (headers wrap, so longest_word
        # is not a hard floor). Text columns retain header-word floor, except
        # known short-value text columns which are also data-driven.
        is_narrow_text = col_name in NARROW_TEXT_COLS
        if is_numeric_col or is_narrow_text:
            width = max(max_data_len + 2, 6)   # data-driven, small floor
            width = min(width, 16)              # cap tight for numbers
        else:
            width = max(max_data_len + 2, longest_word + 2)
            width = max(width, 8)    # minimum for text
            width = min(width, 28)   # cap for text

        # Apply explicit override if defined
        if col_name in COLUMN_WIDTH_OVERRIDES:
            width = COLUMN_WIDTH_OVERRIDES[col_name]

        ws.column_dimensions[col_letter].width = width


def _apply_dropdowns(ws, col_map, max_col, max_row):
    """Attach Excel data-validation dropdowns for columns listed in DROPDOWN_COLS."""
    for col_idx in range(1, max_col + 1):
        col_name = col_map.get(col_idx, "")
        if col_name in DROPDOWN_COLS:
            options = DROPDOWN_COLS[col_name]
            formula = '"' + ','.join(options) + '"'
            dv = DataValidation(
                type="list",
                formula1=formula,
                allow_blank=True,
                showDropDown=False,  # False = SHOW the dropdown arrow
            )
            dv.prompt = f"Select {col_name}"
            dv.promptTitle = col_name
            col_letter = get_column_letter(col_idx)
            dv.add(f"{col_letter}2:{col_letter}{max_row}")
            ws.add_data_validation(dv)
            print(f"    [DROPDOWN] {col_name} (col {col_letter}): {options}")


def _style_portfolio_sheet(ws, path, col_map, max_col, max_row):
    """Portfolio-profile per-sheet actions: column-A composition comments,
    freeze at B2, auto-filter with portfolio_status=CORE/WATCH pre-selected."""
    # Portfolio composition comments on column A (portfolio_id)
    if col_map.get(1, "") == "portfolio_id":
        import json as _json, re as _re, csv as _csv
        from openpyxl.comments import Comment as _Comment
        _strat_root = path.parent
        _commented = 0
        for row_idx in range(2, max_row + 1):
            pid = str(ws.cell(row=row_idx, column=1).value or "")
            if not pid:
                continue
            _eval_dir = _strat_root / pid / "portfolio_evaluation"
            _assets = None
            # Source 1: evaluated_assets from metadata/summary JSON
            for _fname in ("portfolio_metadata.json", "portfolio_summary.json"):
                _fp = _eval_dir / _fname
                if _fp.exists():
                    try:
                        with open(_fp, encoding="utf-8") as _f:
                            _d = _json.load(_f)
                        _assets = _d.get("evaluated_assets")
                        if _assets:
                            break
                    except Exception:
                        pass
            # Source 2: extract from trade-level CSV strategy_name (symbol = last segment)
            if not _assets:
                _tl = _eval_dir / "portfolio_tradelevel.csv"
                if _tl.exists():
                    try:
                        import pandas as _pd
                        _tdf = _pd.read_csv(_tl, usecols=["strategy_name"])
                        _syms = set()
                        for _sn in _tdf["strategy_name"].dropna().unique():
                            _last = str(_sn).split("_")[-1]
                            if _re.match(r"^[A-Za-z0-9]{4,}$", _last):
                                _syms.add(_last.upper())
                        if _syms:
                            _assets = sorted(_syms)
                    except Exception:
                        pass
            if _assets and isinstance(_assets, list):
                _text = " | ".join(str(a).upper() for a in _assets)
                ws.cell(row=row_idx, column=1).comment = _Comment(_text, "System")
                _commented += 1
        print(f"    [COMMENTS] Portfolio composition added to column A ({_commented}/{max_row - 1} rows)")

    # Freeze at B2 so portfolio_id stays visible; wire auto-filter + CORE/WATCH pre-selection.
    ws.freeze_panes = "B2"
    if max_col >= 1 and max_row >= 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"

        _headers_lower = [str(c.value).lower() if c.value else "" for c in ws[1]]
        if "portfolio_status" in _headers_lower:
            from openpyxl.worksheet.filters import FilterColumn, Filters
            _ps_idx = _headers_lower.index("portfolio_status")
            _fc = FilterColumn(colId=_ps_idx)
            _fc.filters = Filters(filter=["CORE", "WATCH"])
            ws.auto_filter.filterColumn.append(_fc)
            _ps_col = _ps_idx + 1  # 1-based
            _hidden_count = 0
            for _r in range(2, max_row + 1):
                _val = str(ws.cell(row=_r, column=_ps_col).value or "")
                if _val not in ("CORE", "WATCH"):
                    ws.row_dimensions[_r].hidden = True
                    _hidden_count += 1
            print(f"    [FILTER] Pre-selected portfolio_status=CORE/WATCH (hidden {_hidden_count} FAIL rows)")


def _style_strategy_sheet(ws, path, max_col, max_row):
    """Strategy-profile per-sheet actions: freeze at A2, auto-filter, and
    file-stem-conditional filters (filtered_strategies CORE/BURN_IN, shadow_trade EXIT)."""
    ws.freeze_panes = "A2"
    if max_col < 1 or max_row < 1:
        return
    ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"

    # Pre-select candidate_status=CORE/BURN_IN for Filtered_Strategies_Passed
    _fname_check = Path(path).stem.lower()
    if "filtered_strategies" in _fname_check:
        _headers_lower = [str(c.value).lower() if c.value else "" for c in ws[1]]
        if "candidate_status" in _headers_lower:
            from openpyxl.worksheet.filters import FilterColumn, Filters
            _cs_idx = _headers_lower.index("candidate_status")
            _fc = FilterColumn(colId=_cs_idx)
            _fc.filters = Filters(filter=["CORE", "BURN_IN"])
            ws.auto_filter.filterColumn.append(_fc)
            _cs_col = _cs_idx + 1  # 1-based
            _hidden_count = 0
            for _r in range(2, max_row + 1):
                _val = str(ws.cell(row=_r, column=_cs_col).value or "")
                if _val not in ("CORE", "BURN_IN"):
                    ws.row_dimensions[_r].hidden = True
                    _hidden_count += 1
            print(f"    [FILTER] Pre-selected candidate_status=CORE/BURN_IN (hidden {_hidden_count} rows)")

    # Pre-select event_type=EXIT filter for shadow_trades files
    _fname = Path(path).stem.lower()
    if "shadow_trade" in _fname:
        _headers_lower = [str(c.value).lower() if c.value else "" for c in ws[1]]
        if "event_type" in _headers_lower:
            from openpyxl.worksheet.filters import FilterColumn, Filters
            _et_idx = _headers_lower.index("event_type")
            _fc = FilterColumn(colId=_et_idx)
            _fc.filters = Filters(filter=["EXIT"])
            ws.auto_filter.filterColumn.append(_fc)
            # openpyxl sets filter metadata but does NOT hide rows —
            # Excel may or may not re-apply on open. Explicitly hide
            # non-EXIT rows so the filter is visually applied on save.
            _et_col = _et_idx + 1  # 1-based for cell access
            _hidden_count = 0
            for _r in range(2, max_row + 1):
                _val = str(ws.cell(row=_r, column=_et_col).value or "")
                if _val != "EXIT":
                    ws.row_dimensions[_r].hidden = True
                    _hidden_count += 1
            print(f"    [FILTER] Pre-selected event_type=EXIT on column {_et_idx} (hidden {_hidden_count} non-EXIT rows)")


def _ensure_legacy_portfolio_notes(wb):
    """Create a 4-line legacy Notes sheet for portfolio profile IFF one does not
    already exist (generated by --notes-type portfolio)."""
    if "Notes" in wb.sheetnames:
        return
    ns = wb.create_sheet("Notes")
    notes_lines = [
        "Metric Notes — Portfolio Sheet",
        "",
        "Sharpe: daily-return based, annualized (x sqrt(252)). Stage 2 Sharpe is trade-level (mean PnL / stdev), non-annualized. Values are not comparable.",
        "K-Ratio: computed on log(daily equity). Stage 2 K-Ratio uses cumulative trade PnL (linear). Values are not comparable.",
        "max_dd_pct: decimal fraction (0..1), negative sign. Filter sheets use positive 0..100 scale.",
    ]
    for row_idx, line in enumerate(notes_lines, start=1):
        ns.cell(row=row_idx, column=1, value=line)
    ns.cell(row=1, column=1).font = Font(bold=True)
    ns.column_dimensions["A"].width = 120


def _restore_hyperlinks(wb, path):
    """Apply hyperlinks (survives pandas rewrite). Config: column_name → link_prefix
    relative to file. Applied to ALL data sheets, skipping Notes."""
    _HYPERLINK_MAP = {
        "Filtered_Strategies_Passed": {"strategy": "../backtests/"},
        "Master_Portfolio_Sheet":     {"portfolio_id": ""},
    }
    _LINK_FONT = Font(color="0563C1", underline="single")
    _file_stem = Path(path).stem
    _hl_conf = _HYPERLINK_MAP.get(_file_stem, {})
    if not _hl_conf:
        return
    _data_sheets = [s for s in wb.sheetnames if s != "Notes"]
    for _sheet_name in _data_sheets:
        _ws = wb[_sheet_name]
        _max_row = _ws.max_row or 1
        _max_col = _ws.max_column or 1
        # Build header map for this sheet
        _hdr = {str(_ws.cell(row=1, column=c).value or "").strip(): c
                for c in range(1, _max_col + 1)}
        for _hl_col_name, _hl_prefix in _hl_conf.items():
            _hl_col_idx = _hdr.get(_hl_col_name)
            if _hl_col_idx is None:
                continue
            _hl_count = 0
            for _r in range(2, _max_row + 1):
                _cell = _ws.cell(row=_r, column=_hl_col_idx)
                _val = str(_cell.value or "").strip()
                if not _val:
                    continue
                _cell.hyperlink = f"{_hl_prefix}{_val}/"
                _cell.font = _LINK_FONT
                _hl_count += 1
            if _hl_count:
                print(f"    [HYPERLINKS] {_sheet_name} / {_hl_col_name}: {_hl_count} links applied")


def apply_formatting(file_path, profile):
    path = Path(file_path)
    if not path.exists():
        print(f"[ERROR] File not found: {file_path}")
        sys.exit(1)

    print(f"[INFO] Formatting {path.name} (Profile: {profile})...")

    # Pre-step: Column Reorder + Sort by Return/DD Ratio (pandas-based)
    # IMPORTANT: only run on single-sheet ledgers; multi-sheet AK reports must be preserved.
    _preprocess_with_pandas(path, profile)

    try:
        wb = openpyxl.load_workbook(path)

        # Styles Objects
        header_fill = PatternFill(start_color=HEADER_FILL_COLOR, end_color=HEADER_FILL_COLOR, fill_type="solid")
        header_font = Font(bold=True, color=HEADER_FONT_COLOR)
        alt_row_fill = PatternFill(start_color=ALT_ROW_FILL_COLOR, end_color=ALT_ROW_FILL_COLOR, fill_type="solid")

        # Iterate over ALL sheets (skip Notes — managed by --notes-type)
        for ws in wb.worksheets:
            if ws.title == "Notes":
                print(f"  [INFO] Skipping sheet: {ws.title} (managed by --notes-type)")
                continue
            print(f"  [INFO] Processing sheet: {ws.title}")

            # 1. Header & Column ID Mapping
            col_map = {}  # col_index -> col_name_lower
            max_col = ws.max_column
            max_row = ws.max_row

            _apply_header_row(ws, col_map, max_col, profile, header_font, header_fill)

            # 2. Row Iteration (Data)
            _format_data_rows(ws, col_map, max_col, max_row, alt_row_fill)

            # 3. Header Row Height (tall enough for 2-line wrapped headers)
            ws.row_dimensions[1].height = HEADER_ROW_HEIGHT

            # 4. Column Widths & Hiding
            _apply_column_widths(ws, col_map, max_col, max_row)

            # 5. Dropdown Data Validation
            _apply_dropdowns(ws, col_map, max_col, max_row)

            # 6. Profile-specific per-sheet operations (comments, freeze panes, filters)
            if profile == "portfolio":
                _style_portfolio_sheet(ws, path, col_map, max_col, max_row)
            else:
                _style_strategy_sheet(ws, path, max_col, max_row)

        # Portfolio profile: preserve existing Notes sheet if present (generated by
        # --notes-type portfolio). Only create the legacy 4-line Notes if none exists.
        if profile == "portfolio":
            _ensure_legacy_portfolio_notes(wb)

        # ── Hyperlink restoration ──────────────────────────────────────────
        # Applied here (last openpyxl step) so hyperlinks survive any upstream
        # pandas rewrite.
        _restore_hyperlinks(wb, path)

        wb.save(path)
        print("[SUCCESS] Formatting complete.")

    except Exception as e:
        print(f"[FATAL] Formatting failed: {e}")
        sys.exit(1)
