"""
Stage-2 Presentation Compiler — SOP-Complete with Full Metric Computation
Consumes Stage-1 authoritative artifacts and produces AK_Trade_Report Excel.
Implements SOP_OUTPUT §5.1 (Performance Summary) and §5.2 (Yearwise Performance).
All metrics computed from Stage-1 data. No placeholders.
Rewritten to use pandas and Unified Formatter (Zero OpenPyXL Styling / Imports).

Authority: SOP_OUTPUT, SOP_TESTING
State Gated: Yes (STAGE_2_START)
"""

from __future__ import annotations

import csv
import json
import sys
import subprocess
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any

import pandas as pd
import argparse

__all__ = [
    "load_stage1_artifacts",
    "compile_stage2",
    "generate_excel_report",
    "get_performance_summary_df",
    "get_trades_df",
]

# Config
PROJECT_ROOT = Path(__file__).parent.parent.parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

# Governance Imports
from tools.pipeline_utils import PipelineStateManager
from config.state_paths import BACKTESTS_DIR, RUNS_DIR

# ==================================================================
# Metrics Core — Single Source of Truth (extracted to tools/metrics_core.py)
# ==================================================================
from tools.metrics_core import (
    # Constants
    ASIA_START, ASIA_END, LONDON_START, LONDON_END, NY_START, NY_END,
    VOL_REGIME_BUCKETS as _VOL_REGIME_BUCKETS,
    TREND_LABEL_BUCKETS as _TREND_LABEL_BUCKETS,
    TF_BARS_PER_DAY as _TF_BARS_PER_DAY,
    # Utilities (internal — keep _ prefix)
    _safe_float, _safe_int, _parse_timestamp, _get_session,
    # Core metric functions (aliased for backward compat within this file)
    compute_pnl_basics as _compute_pnl_basics,
    compute_drawdown as _compute_drawdown,
    compute_streaks as _compute_streaks,
    compute_bars_stats as _compute_bars_stats,
    compute_trading_period as _compute_trading_period,
    compute_bars_per_day as _compute_bars_per_day,
    compute_risk_ratios as _compute_risk_ratios,
    compute_mfe_mae as _compute_mfe_mae,
    compute_k_ratio as _compute_k_ratio,
    compute_concentration as _compute_concentration,
    bucket_breakdown as _bucket_breakdown,
    summarize_buckets as _summarize_buckets,
    compute_session_breakdown as _compute_session_breakdown,
    compute_regime_age_breakdown as _compute_regime_age_breakdown,
    # Orchestrator
    compute_metrics_from_trades as _compute_metrics_from_trades,
    empty_metrics as _empty_metrics,
)


def load_stage1_artifacts(run_folder: Path) -> dict[str, Any]:
    execution_dir = run_folder / "raw"
    metadata_dir = run_folder / "metadata"
    
    required_files = [
        execution_dir / "results_tradelevel.csv",
        execution_dir / "results_standard.csv",
        execution_dir / "results_risk.csv",
        execution_dir / "results_yearwise.csv",
        execution_dir / "metrics_glossary.csv",
        metadata_dir / "run_metadata.json",
    ]
    
    for f in required_files:
        if not f.exists():
            raise FileNotFoundError(f"Required Stage-1 artifact missing: {f}")
    
    def read_csv(path):
        with open(path, "r", encoding="utf-8") as f:
            return list(csv.DictReader(f))
    
    tradelevel = read_csv(execution_dir / "results_tradelevel.csv")
    standard = read_csv(execution_dir / "results_standard.csv")
    risk = read_csv(execution_dir / "results_risk.csv")
    yearwise = read_csv(execution_dir / "results_yearwise.csv")
    glossary = read_csv(execution_dir / "metrics_glossary.csv")
    
    with open(metadata_dir / "run_metadata.json", "r", encoding="utf-8") as f:
        metadata = json.load(f)
        
    # v5 Metric Integrity: Load Bar Geometry if available
    geometry_path = execution_dir / "bar_geometry.json"
    if geometry_path.exists():
        try:
            with open(geometry_path, "r", encoding="utf-8") as f:
                geo = json.load(f)
                metadata["bar_geometry"] = geo
        except (OSError, json.JSONDecodeError) as e:
            print(f"  STAGE2_BAR_GEOMETRY_WARN  {type(e).__name__}: {e}  path={geometry_path}")
        
    starting_capital = metadata.get("reference_capital_usd")
    if not isinstance(starting_capital, (int, float)) or starting_capital <= 0:
        raise ValueError("Missing or invalid reference_capital_usd in metadata")

    return {
        "tradelevel": tradelevel,
        "standard": standard[0] if standard else {},
        "risk": risk[0] if risk else {},
        "yearwise": yearwise,
        "glossary": glossary,
        "metadata": metadata,
    }



def _round_val(val: Any, decimals: int = 2) -> Any:
    if isinstance(val, (int, float)):
        return round(val, decimals)
    return val



def _compute_yearwise_metrics(trades: list[dict[str, Any]], year: int, starting_capital: float, authoritative_data: dict[str, Any] | None = None) -> dict[str, Any] | None:
    if authoritative_data is None:
        authoritative_data = {}

    year_trades = []
    for t in trades:
        exit_dt = _parse_timestamp(t.get("exit_timestamp", ""))
        if exit_dt and exit_dt.year == year:
            year_trades.append(t)
    
    if not year_trades and not authoritative_data:
        return None
    
    pnls = [_safe_float(t.get("pnl_usd", 0)) for t in year_trades]
    bars_list = [_safe_int(t.get("bars_held", 0)) for t in year_trades if t.get("bars_held") not in (None, "", "None")]
    
    wins = [p for p in pnls if p > 0]
    losses = [p for p in pnls if p < 0]
    
    gross_profit = sum(wins)
    gross_loss = abs(sum(losses))
    
    net_profit = _safe_float(authoritative_data.get("net_pnl_usd")) if "net_pnl_usd" in authoritative_data else sum(pnls)
    trade_count = _safe_int(authoritative_data.get("trade_count")) if "trade_count" in authoritative_data else len(pnls)
    win_rate = _safe_float(authoritative_data.get("win_rate")) if "win_rate" in authoritative_data else ((len(wins) / len(pnls)) if pnls else 0.0)
    
    win_count = len(wins)
    loss_count = len(losses)
    
    profit_factor = (gross_profit / gross_loss) if gross_loss > 0 else gross_profit if gross_profit > 0 else 0.0
    avg_trade = (net_profit / trade_count) if trade_count > 0 else 0.0
    avg_bars = (sum(bars_list) / len(bars_list)) if bars_list else 0.0
    
    cumulative = 0.0
    peak = 0.0
    max_dd = 0.0
    for pnl in pnls:
        cumulative += pnl
        if cumulative > peak:
            peak = cumulative
        dd = peak - cumulative
        if dd > max_dd:
            max_dd = dd
    
    max_dd_pct = (max_dd / starting_capital) if starting_capital > 0 else 0.0
    return_dd_ratio = (net_profit / max_dd) if max_dd > 0 else net_profit if net_profit > 0 else 0.0
    
    return {
        "year": year,
        "net_profit": net_profit,
        "gross_profit": gross_profit,
        "gross_loss": gross_loss,
        "trade_count": trade_count,
        "win_rate": win_rate,
        "profit_factor": profit_factor,
        "avg_trade": avg_trade,
        "max_dd_usd": max_dd,
        "max_dd_pct": max_dd_pct,
        "return_dd_ratio": return_dd_ratio,
        "avg_bars": avg_bars,
        "win_count": win_count,
        "loss_count": loss_count,
    }


def _compute_buy_hold_benchmark(trades: list[dict[str, Any]]) -> dict[str, float] | None:
    if not trades or len(trades) < 2:
        return None
    try:
        sorted_trades = sorted(trades, key=lambda t: t.get("entry_timestamp", ""))
        first_price = _safe_float(sorted_trades[0].get("entry_price", 0))
        last_price = _safe_float(sorted_trades[-1].get("exit_price", 0))
        if first_price <= 0 or last_price <= 0:
            return None
        bh_return_pct = ((last_price - first_price) / first_price) * 100
        exit_prices = [_safe_float(t.get("exit_price", 0)) for t in sorted_trades]
        exit_prices = [p for p in exit_prices if p > 0]
        if not exit_prices:
            return None
        peak = exit_prices[0]
        max_dd_pct = 0.0
        for price in exit_prices:
            if price > peak:
                peak = price
            dd = ((peak - price) / peak) * 100 if peak > 0 else 0
            if dd > max_dd_pct:
                max_dd_pct = dd
        return {
            "return_pct": bh_return_pct,
            "max_drawdown_pct": max_dd_pct,
            "first_price": first_price,
            "last_price": last_price,
        }
    except (ValueError, TypeError, KeyError, ZeroDivisionError) as e:
        print(f"  STAGE2_BUY_HOLD_SKIP  {type(e).__name__}: {e}")
        return None


def get_runtime_engine_version() -> str:
    """Dynamically load version from validated engine manifest."""
    try:
        from tools.pipeline_utils import get_engine_version
        _ver = get_engine_version()
        manifest_path = PROJECT_ROOT / "engine_dev" / "universal_research_engine" / _ver / "VALIDATED_ENGINE.manifest.json"
        if manifest_path.exists():
            with open(manifest_path, "r", encoding="utf-8") as f:
                data = json.load(f)
                return data.get("engine_version", "UNKNOWN")
    except (ImportError, OSError, json.JSONDecodeError, KeyError) as e:
        print(f"  STAGE2_ENGINE_VERSION_WARN  {type(e).__name__}: {e}")
    return "UNKNOWN"

def get_settings_df(metadata: dict[str, Any]) -> pd.DataFrame:
    # Strict Version Validation (SOP v4.2 Remediation)
    meta_engine_ver = metadata.get("engine_version", "")
    runtime_ver = get_runtime_engine_version()
    
    # If runtime version is known, enforce strict match
    if runtime_ver != "UNKNOWN":
        if meta_engine_ver != runtime_ver:
             # Raise error to prevent silent corruption/mismatch
             raise ValueError(f"Stage-2 Metadata Mismatch: Metadata says {meta_engine_ver}, Runtime says {runtime_ver}. Strict enforcement enabled.")
    
    data = [
        {"Parameter": "Run ID", "Value": metadata.get("run_id", "")},
        {"Parameter": "Strategy Name", "Value": metadata.get("strategy_name", "")},
        {"Parameter": "Symbol", "Value": metadata.get("symbol", "")},
        {"Parameter": "Broker", "Value": metadata.get("broker", "")},
        {"Parameter": "Reference Capital (USD)", "Value": metadata.get("reference_capital_usd", "")},
        {"Parameter": "Position Sizing Basis", "Value": metadata.get("position_sizing_basis", "")},
        {"Parameter": "Timeframe", "Value": metadata.get("timeframe", "")},
        {"Parameter": "Date Range Start", "Value": metadata.get("date_range", {}).get("start", "")},
        {"Parameter": "Date Range End", "Value": metadata.get("date_range", {}).get("end", "")},
        {"Parameter": "Execution Timestamp", "Value": metadata.get("execution_timestamp_utc", "")},
        {"Parameter": "Engine Name", "Value": metadata.get("engine_name", "")},
        {"Parameter": "Engine Version", "Value": meta_engine_ver}, 
        {"Parameter": "Schema Version", "Value": metadata.get("schema_version", "")},
    ]
    return pd.DataFrame(data)

def get_performance_summary_df(trades: list[dict[str, Any]], starting_capital: float, standard_metrics: dict[str, Any], risk_metrics: dict[str, Any], metadata: dict[str, Any] | None = None) -> pd.DataFrame:
    all_metrics = _compute_metrics_from_trades(trades, starting_capital, None, metadata)
    long_metrics = _compute_metrics_from_trades(trades, starting_capital, 1, metadata)
    short_metrics = _compute_metrics_from_trades(trades, starting_capital, -1, metadata)
    
    # OVERRIDE All Trades
    all_metrics["net_profit"] = _safe_float(standard_metrics.get("net_pnl_usd", 0))
    all_metrics["gross_profit"] = _safe_float(standard_metrics.get("gross_profit", 0))
    all_metrics["gross_loss"] = _safe_float(standard_metrics.get("gross_loss", 0))
    # win_rate and max_drawdown_pct are stored as decimals (0..1) in Stage-1 CSVs.
    # AK Report convention is 0..100 (percentage). Multiply by 100 to match.
    _raw_win_rate = _safe_float(standard_metrics.get("win_rate", 0))
    assert _raw_win_rate <= 1.0 or _raw_win_rate == 0.0, (
        f"UNIT_GUARD: win_rate={_raw_win_rate} already in 0..100 scale — expected 0..1 from Stage-1"
    )
    all_metrics["pct_profitable"] = _raw_win_rate * 100
    all_metrics["profit_factor"] = _safe_float(standard_metrics.get("profit_factor", 0))
    all_metrics["max_dd_usd"] = _safe_float(risk_metrics.get("max_drawdown_usd", 0))
    _raw_dd_pct = _safe_float(risk_metrics.get("max_drawdown_pct", 0))
    if _raw_dd_pct > 1.0:
        print(f"  [UNIT_GUARD WARNING] max_drawdown_pct={_raw_dd_pct:.4f} exceeds 1.0 "
              f"(drawdown > reference capital) — converting as-is")
    all_metrics["max_dd_pct"] = _raw_dd_pct * 100
    all_metrics["return_dd_ratio"] = _safe_float(risk_metrics.get("return_dd_ratio", 0))
    all_metrics["sharpe_ratio"] = _safe_float(risk_metrics.get("sharpe_ratio", 0))
    all_metrics["sortino_ratio"] = _safe_float(risk_metrics.get("sortino_ratio", 0))
    # K-Ratio now computed by _compute_k_ratio() for all columns — no override needed
    
    rows = []
    def add_row(label, key):
        val_all = all_metrics.get(key)
        val_long = long_metrics.get(key)
        val_short = short_metrics.get(key)
        
        rows.append({
            "Metric": label,
            "All Trades": _round_val(val_all, 2),
            "Long Trades": _round_val(val_long, 2),
            "Short Trades": _round_val(val_short, 2)
        })

    add_row("Starting Capital", "starting_capital")
    add_row("Net Profit (USD)", "net_profit")
    add_row("Gross Profit (USD)", "gross_profit")
    add_row("Gross Loss (USD)", "gross_loss")
    add_row("Profit Factor", "profit_factor")
    add_row("Expectancy (USD)", "expectancy")
    add_row("Return / Drawdown Ratio", "return_dd_ratio")
    add_row("Total Trades", "total_trades")
    add_row("Winning Trades", "winning_trades")
    add_row("Losing Trades", "losing_trades")
    add_row("% Profitable", "pct_profitable")
    add_row("Trades per Month", "trades_per_month")
    add_row("Longest Flat Period (Days)", "longest_flat_days")
    add_row("Avg Trade (USD)", "avg_trade")
    add_row("Avg Win (USD)", "avg_win")
    add_row("Avg Loss (USD)", "avg_loss")
    add_row("Win/Loss Ratio", "win_loss_ratio")
    add_row("Avg MFE (R)", "avg_mfe_r")
    add_row("Avg MAE (R)", "avg_mae_r")
    add_row("Edge Ratio (MFE / MAE)", "edge_ratio")
    add_row("Largest Win (USD)", "largest_win")
    add_row("Largest Loss (USD)", "largest_loss")
    add_row("% of Gross Profit (Top Trades)", "top5_pct_gross_profit")
    add_row("Worst 5 Trades Loss %", "worst5_loss_pct")
    add_row("Max Consecutive Wins", "max_consec_wins")
    add_row("Max Consecutive Losses", "max_consec_losses")
    add_row("Max Drawdown (USD)", "max_dd_usd")
    add_row("Max Drawdown (%)", "max_dd_pct")
    add_row("Return on Capital", "return_on_capital")
    add_row("% Time in Market", "pct_time_in_market")
    add_row("Sharpe Ratio", "sharpe_ratio")
    add_row("Sortino Ratio", "sortino_ratio")
    add_row("K-Ratio", "k_ratio")
    add_row("SQN (System Quality Number)", "sqn")
    add_row("Return Retracement Ratio", "return_retracement_ratio")
    add_row("Avg Bars in Winning Trades", "avg_bars_win")
    add_row("Avg Bars in Losing Trades", "avg_bars_loss")
    add_row("Avg Bars per Trade", "avg_bars")
    add_row("Trading Period (Days)", "trading_period_days")
    add_row("Net Profit - Low Volatility", "net_profit_low_vol")
    add_row("Net Profit - Normal Volatility", "net_profit_normal_vol")
    add_row("Net Profit - High Volatility", "net_profit_high_vol")
    add_row("Trades - Low Volatility", "trades_low_vol")
    add_row("Trades - Normal Volatility", "trades_normal_vol")
    add_row("Trades - High Volatility", "trades_high_vol")
    add_row("Avg Trade - Low Volatility", "avg_trade_low_vol")
    add_row("Avg Trade - Normal Volatility", "avg_trade_normal_vol")
    add_row("Avg Trade - High Volatility", "avg_trade_high_vol")
    add_row("Net Profit - Asia Session", "net_profit_asia")
    add_row("Net Profit - London Session", "net_profit_london")
    add_row("Net Profit - New York Session", "net_profit_ny")
    add_row("Trades - Asia Session", "trades_asia")
    add_row("Trades - London Session", "trades_london")
    add_row("Trades - New York Session", "trades_ny")
    add_row("Avg Trade - Asia Session", "avg_trade_asia")
    add_row("Avg Trade - London Session", "avg_trade_london")
    add_row("Avg Trade - New York Session", "avg_trade_ny")
    add_row("Trade Density (Trades/Year)", "trade_density")
    add_row("Net Profit - Strong Up", "net_profit_strong_up")
    add_row("Net Profit - Weak Up", "net_profit_weak_up")
    add_row("Net Profit - Neutral", "net_profit_neutral")
    add_row("Net Profit - Weak Down", "net_profit_weak_down")
    add_row("Net Profit - Strong Down", "net_profit_strong_down")
    add_row("Trades - Strong Up", "trades_strong_up")
    add_row("Trades - Weak Up", "trades_weak_up")
    add_row("Trades - Neutral", "trades_neutral")
    add_row("Trades - Weak Down", "trades_weak_down")
    add_row("Trades - Strong Down", "trades_strong_down")

    return pd.DataFrame(rows)

def get_benchmark_df(trades: list[dict[str, Any]], starting_capital: float, all_net_profit: float) -> pd.DataFrame:
    bh = _compute_buy_hold_benchmark(trades)
    if not bh:
        return pd.DataFrame()
    
    strategy_return_pct = (all_net_profit / starting_capital) * 100 if starting_capital > 0 else 0.0
    relative_perf = strategy_return_pct - bh["return_pct"]
    
    data = [
        {"Metric": "First Price", "Value": bh["first_price"]},
        {"Metric": "Last Price", "Value": bh["last_price"]},
        {"Metric": "Buy & Hold Return (%)", "Value": bh["return_pct"]},
        {"Metric": "Buy & Hold Max Drawdown (%)", "Value": bh["max_drawdown_pct"]},
        {"Metric": "Strategy Net Return (%)", "Value": strategy_return_pct},
        {"Metric": "Relative Performance (%)", "Value": relative_perf},
    ]
    return pd.DataFrame(data)

def get_yearwise_df(trades: list[dict[str, Any]], starting_capital: float, yearwise_data: list[dict[str, Any]]) -> pd.DataFrame:
    year_lookup = {}
    if yearwise_data:
        for row in yearwise_data:
            try:
                y = int(row.get("year", 0))
                if y > 0: year_lookup[y] = row
            except (ValueError, TypeError):
                pass

    years = set()
    for t in trades:
        exit_dt = _parse_timestamp(t.get("exit_timestamp", ""))
        if exit_dt: years.add(exit_dt.year)
    years.update(year_lookup.keys())
    
    rows = []
    for year in sorted(years):
        auth_row = year_lookup.get(year, None)
        metrics = _compute_yearwise_metrics(trades, year, starting_capital, auth_row)
        if metrics:
            rows.append(metrics)
            
    if not rows:
        return pd.DataFrame()
        
    df = pd.DataFrame(rows)
    # Reorder columns matches headers
    cols = ["year", "net_profit", "gross_profit", "gross_loss", "trade_count", "win_rate", "profit_factor", "avg_trade", "max_dd_usd", "max_dd_pct", "return_dd_ratio", "avg_bars", "win_count", "loss_count"]
    # Ensure all cols exist
    for c in cols:
        if c not in df.columns: df[c] = 0
    return df[cols]

def get_trades_df(trades: list[dict[str, Any]]) -> pd.DataFrame:
    if not trades:
        return pd.DataFrame()
        
    # Get all unique keys from all trades to ensure we don't miss any columns
    all_keys = set()
    for t in trades:
        all_keys.update(t.keys())
    
    # Define a logical order for known columns, others appended at end
    priority_cols = [
        "strategy_name", "parent_trade_id", "sequence_index", "direction",
        "entry_timestamp", "exit_timestamp", "entry_price", "exit_price",
        "pnl_usd", "bars_held", "volatility_regime", "trend_score", "trend_regime", "trend_label"
    ]
    
    sorted_cols = [c for c in priority_cols if c in all_keys]
    remaining_cols = sorted([c for c in all_keys if c not in priority_cols])
    final_cols = sorted_cols + remaining_cols
    
    rows = []
    for t in trades:
        row = {}
        for k in final_cols:
            val = t.get(k, "")
            # Try to make numeric if possible for Excel
            try:
                if k in ["direction", "sequence_index", "bars_held"]:
                    row[k] = int(float(val))
                elif k in ["strategy_name", "parent_trade_id", "entry_timestamp", "exit_timestamp", "volatility_regime", "trend_label"]:
                    row[k] = str(val)
                else:
                    fval = float(val)
                    row[k] = _round_val(fval, 4) # 4 decimals for price/r-multiples
            except (ValueError, TypeError):
                row[k] = val
        rows.append(row)
        
    return pd.DataFrame(rows)


def get_regime_age_df(trades: list[dict[str, Any]]) -> pd.DataFrame:
    """Return a summary table of performance across regime_age buckets.

    Uses compute_regime_age_breakdown() from metrics_core — no custom calculations.
    Returns empty DataFrame if no trades carry regime_age (older strategies).
    """
    rows = _compute_regime_age_breakdown(trades)

    # If every bucket has 0 trades, regime_age is absent from this run — skip the sheet.
    if all(r["trades"] == 0 for r in rows):
        return pd.DataFrame()

    df = pd.DataFrame(rows)
    df = df.rename(columns={
        "label":         "Bucket",
        "trades":        "Trades",
        "net_pnl":       "Net PnL (USD)",
        "profit_factor": "Profit Factor",
        "win_rate":      "Win Rate (%)",
        "avg_trade":     "Avg Trade (USD)",
    })

    # Reconciliation footer: total / bucketed / missing
    total = len(trades)
    bucketed = int(df["Trades"].sum())
    missing = total - bucketed
    df = df.astype(object)
    footer = pd.DataFrame([
        {"Bucket": "",          "Trades": "",      "Net PnL (USD)": "", "Profit Factor": "", "Win Rate (%)": "", "Avg Trade (USD)": ""},
        {"Bucket": "Total",     "Trades": total,   "Net PnL (USD)": "", "Profit Factor": "", "Win Rate (%)": "", "Avg Trade (USD)": ""},
        {"Bucket": "Bucketed",  "Trades": bucketed,"Net PnL (USD)": "", "Profit Factor": "", "Win Rate (%)": "", "Avg Trade (USD)": ""},
        {"Bucket": "Missing",   "Trades": missing, "Net PnL (USD)": "", "Profit Factor": "", "Win Rate (%)": "", "Avg Trade (USD)": ""},
    ])
    df = pd.concat([df, footer], ignore_index=True)

    return df


def _add_notes_sheet(output_path: Path, df_summary: pd.DataFrame, metadata: dict[str, Any]) -> None:
    """
    Append a 'Notes' sheet to the AK_Trade_Report with classification transparency.

    Reads classification thresholds directly from the same logic as
    filter_strategies._compute_candidate_status() — no separate import needed.
    Called after the formatter subprocess so the sheet is not reformatted.
    """
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, Alignment
    except ImportError as e:
        print(f"[WARN] Notes sheet skipped — openpyxl not available: {e}")
        return

    # --- Classification thresholds (mirrors filter_strategies._compute_candidate_status) ---
    FAIL_MIN_TRADES = 50
    FAIL_MAX_DD_PCT = 40.0
    CORE_MIN_TRADES = 200
    CORE_MIN_RET_DD  = 2.0
    CORE_MIN_SHARPE  = 1.5

    # --- Extract metrics from Performance Summary ---
    def _get(metric_name: str) -> float | None:
        try:
            row = df_summary[df_summary["Metric"] == metric_name]
            if not row.empty:
                return float(row["All Trades"].values[0])
        except Exception:
            pass
        return None

    pf_val       = _get("Profit Factor")       or 0.0
    sharpe_val   = _get("Sharpe Ratio")        or 0.0
    dd_val       = _get("Max Drawdown (%)")    or 0.0
    ret_dd_val   = _get("Return / Drawdown Ratio") or 0.0
    trades_raw   = _get("Total Trades")
    trades_val   = int(trades_raw) if trades_raw is not None else 0
    net_pnl_val  = _get("Net Profit (USD)")    or 0.0

    strategy_name = metadata.get("strategy_name", "UNKNOWN")
    symbol        = metadata.get("symbol", "UNKNOWN")
    full_id       = f"{strategy_name}_{symbol}"

    # --- Compute classification (same precedence as filter_strategies) ---
    is_fail = (trades_val < FAIL_MIN_TRADES) or (dd_val > FAIL_MAX_DD_PCT)
    is_core = (not is_fail) and (
        trades_val >= CORE_MIN_TRADES
        and ret_dd_val >= CORE_MIN_RET_DD
        and sharpe_val >= CORE_MIN_SHARPE
    )

    if is_fail:
        classification = "FAIL"
    elif is_core:
        classification = "CORE"
    else:
        classification = "WATCH"

    # BURN_IN check: mirrors _load_burnin_ids() logic
    try:
        import yaml
        portfolio_path = PROJECT_ROOT.parent / "TS_Execution" / "portfolio.yaml"
        if portfolio_path.exists():
            with open(portfolio_path, encoding="utf-8") as _f:
                _port = yaml.safe_load(_f)
            _strategies = (_port.get("portfolio") or {}).get("strategies") or []
            _burnin_ids = {s["id"] for s in _strategies if s.get("enabled", True) and "id" in s}
            if any(full_id == bid or full_id.startswith(bid + "_") for bid in _burnin_ids):
                classification = "BURN_IN"
    except Exception:
        pass

    # --- Build Notes sheet ---
    try:
        wb = load_workbook(output_path)
    except Exception as e:
        print(f"[WARN] Notes sheet skipped — cannot open workbook: {e}")
        return

    if "Notes" in wb.sheetnames:
        del wb["Notes"]
    ws = wb.create_sheet("Notes")

    bold   = Font(bold=True, size=10)
    header = Font(bold=True, size=11)
    normal = Font(size=10)
    green  = Font(bold=True, size=10, color="1F6B1F")
    red    = Font(bold=True, size=10, color="A31515")

    r = 1

    def _write(row, col, value, font=None):
        c = ws.cell(row=row, column=col, value=value)
        if font:
            c.font = font
        return c

    # ── Section 1: Classification Result ─────────────────────────────────────
    _write(r, 1, "SECTION 1 — STRATEGY CLASSIFICATION RESULT", header)
    r += 1
    _write(r, 1, "Strategy Name",  bold); _write(r, 2, strategy_name, normal); r += 1
    _write(r, 1, "Symbol",         bold); _write(r, 2, symbol,         normal); r += 1
    _write(r, 1, "Classification", bold); _write(r, 2, classification,  bold);  r += 1
    r += 1

    _write(r, 1, "Metric",              bold); _write(r, 2, "Value", bold); r += 1
    for label, val in [
        ("Profit Factor",            f"{pf_val:.2f}"),
        ("Sharpe Ratio",             f"{sharpe_val:.2f}"),
        ("Max Drawdown (%)",         f"{dd_val:.2f}%"),
        ("Return / Drawdown Ratio",  f"{ret_dd_val:.2f}"),
        ("Total Trades",             str(trades_val)),
        ("Net Profit (USD)",         f"${net_pnl_val:.2f}"),
    ]:
        _write(r, 1, label, normal); _write(r, 2, val, normal); r += 1
    r += 1

    # ── Section 2: Classification Rules ──────────────────────────────────────
    _write(r, 1, "SECTION 2 — CLASSIFICATION RULES", header); r += 1
    _write(r, 1, "Class", bold); _write(r, 2, "Rule", bold); r += 1
    for cls, rule in [
        ("FAIL",
         f"Total Trades < {FAIL_MIN_TRADES}  OR  Max Drawdown (%) > {FAIL_MAX_DD_PCT:.0f}"),
        ("CORE",
         f"Total Trades >= {CORE_MIN_TRADES}  AND  Return/DD >= {CORE_MIN_RET_DD}  AND  Sharpe >= {CORE_MIN_SHARPE}  (and not FAIL)"),
        ("WATCH",
         "All other strategies (does not meet CORE; not FAIL; not in portfolio.yaml)"),
        ("BURN_IN",
         "Present in TS_Execution/portfolio.yaml with enabled=true — overrides computed status"),
    ]:
        _write(r, 1, cls, bold); _write(r, 2, rule, normal); r += 1
    r += 1

    # ── Section 3: Rule Evaluation ────────────────────────────────────────────
    _write(r, 1, "SECTION 3 — RULE EVALUATION (THIS STRATEGY)", header); r += 1
    for col, hdr in [(1, "Rule"), (2, "Condition"), (3, "Actual"), (4, "Result")]:
        _write(r, col, hdr, bold)
    r += 1

    eval_rows = [
        ("Min Trades — FAIL gate",
         f">= {FAIL_MIN_TRADES}",
         str(trades_val),
         trades_val >= FAIL_MIN_TRADES),
        ("Max DD % — FAIL gate",
         f"<= {FAIL_MAX_DD_PCT:.0f}%",
         f"{dd_val:.2f}%",
         dd_val <= FAIL_MAX_DD_PCT),
        ("Min Trades — CORE gate",
         f">= {CORE_MIN_TRADES}",
         str(trades_val),
         trades_val >= CORE_MIN_TRADES),
        ("Return/DD — CORE gate",
         f">= {CORE_MIN_RET_DD}",
         f"{ret_dd_val:.2f}",
         ret_dd_val >= CORE_MIN_RET_DD),
        ("Sharpe — CORE gate",
         f">= {CORE_MIN_SHARPE}",
         f"{sharpe_val:.2f}",
         sharpe_val >= CORE_MIN_SHARPE),
    ]
    for rule_name, condition, actual, passed in eval_rows:
        result_txt  = "PASS" if passed else "FAIL"
        result_font = green if passed else red
        _write(r, 1, rule_name,   normal)
        _write(r, 2, condition,   normal)
        _write(r, 3, actual,      normal)
        _write(r, 4, result_txt,  result_font)
        r += 1
    r += 1

    # ── Section 4: Remarks ───────────────────────────────────────────────────
    _write(r, 1, "SECTION 4 — NOTES / REMARKS", header); r += 1

    remarks = []
    if classification == "FAIL":
        if trades_val < FAIL_MIN_TRADES:
            remarks.append(
                f"Insufficient trades ({trades_val}); minimum {FAIL_MIN_TRADES} required to pass FAIL gate."
            )
        if dd_val > FAIL_MAX_DD_PCT:
            remarks.append(
                f"Excessive drawdown ({dd_val:.1f}%); maximum {FAIL_MAX_DD_PCT:.0f}% allowed."
            )
    elif classification == "BURN_IN":
        remarks.append("Strategy is currently active in the live portfolio (burn-in phase).")
    elif classification == "CORE":
        remarks.append("All CORE criteria satisfied. Eligible for promotion consideration.")
    else:  # WATCH
        gaps = []
        if trades_val < CORE_MIN_TRADES:
            gaps.append(f"trades ({trades_val} < {CORE_MIN_TRADES})")
        if ret_dd_val < CORE_MIN_RET_DD:
            gaps.append(f"Return/DD ({ret_dd_val:.2f} < {CORE_MIN_RET_DD})")
        if sharpe_val < CORE_MIN_SHARPE:
            gaps.append(f"Sharpe ({sharpe_val:.2f} < {CORE_MIN_SHARPE})")
        if gaps:
            remarks.append(f"CORE threshold not met on: {', '.join(gaps)}.")

    for remark in (remarks or ["No additional remarks."]):
        _write(r, 1, remark, normal); r += 1

    # ── Column widths ─────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 58
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 10

    try:
        wb.save(output_path)
        print(f"[NOTES] Notes sheet added to {output_path.name}")
    except Exception as e:
        print(f"[WARN] Notes sheet could not be saved: {e}")


def generate_excel_report(artifacts: dict[str, Any], output_path: Path) -> None:
    starting_capital = artifacts["metadata"]["reference_capital_usd"]
    
    df_settings = get_settings_df(artifacts["metadata"])
    df_summary = get_performance_summary_df(artifacts["tradelevel"], starting_capital, artifacts["standard"], artifacts["risk"], artifacts["metadata"])
    
    # Extract net profit from summary for benchmark
    try:
        net_profit_row = df_summary[df_summary["Metric"] == "Net Profit (USD)"]
        net_profit = float(net_profit_row["All Trades"].values[0])
    except (ValueError, TypeError, IndexError, KeyError) as e:
        print(f"  STAGE2_NET_PROFIT_EXTRACT_WARN  {type(e).__name__}: {e}  -> default=0.0")
        net_profit = 0.0
        
    df_benchmark = get_benchmark_df(artifacts["tradelevel"], starting_capital, net_profit)
    df_yearwise = get_yearwise_df(artifacts["tradelevel"], starting_capital, artifacts["yearwise"])
    df_trades = get_trades_df(artifacts["tradelevel"])
    df_age = get_regime_age_df(artifacts["tradelevel"])

    # Write to Excel (Raw Data)
    with pd.ExcelWriter(output_path, engine='xlsxwriter') as writer:
        df_settings.to_excel(writer, sheet_name="Settings", index=False)
        df_summary.to_excel(writer, sheet_name="Performance Summary", index=False)
        if not df_benchmark.empty:
            df_benchmark.to_excel(writer, sheet_name="Benchmark Analysis", index=False)
        df_yearwise.to_excel(writer, sheet_name="Yearwise Performance", index=False)
        if not df_age.empty:
            df_age.to_excel(writer, sheet_name="Regime Lifecycle (Age)", index=False)
        df_trades.to_excel(writer, sheet_name="Trades List", index=False)
        
    # Call Unified Formatter
    try:
        project_root = PROJECT_ROOT
        formatter = project_root / "tools" / "format_excel_artifact.py"
        cmd = [sys.executable, str(formatter), "--file", str(output_path), "--profile", "strategy"]
        subprocess.run(cmd, check=True)
        print(f"[SUCCESS] Formatted {output_path.name}")
    except (subprocess.CalledProcessError, OSError) as e:
        print(f"[WARN] Failed to format {output_path.name}: {type(e).__name__}: {e}")

    # Add Notes sheet after formatter (so it is not affected by the formatting pass)
    _add_notes_sheet(output_path, df_summary, artifacts["metadata"])


def compile_stage2(run_folder: Path) -> None:
    artifacts = load_stage1_artifacts(run_folder)
    strategy_name = artifacts["metadata"].get("strategy_name", "UNKNOWN")
    excel_filename = f"AK_Trade_Report_{strategy_name}.xlsx"
    excel_path = run_folder / excel_filename
    
    generate_excel_report(artifacts, excel_path)
    return run_folder, [excel_filename]

def main() -> None:
    parser = argparse.ArgumentParser(description="Stage-2 Presentation Compiler (v5 Multi-Asset) - Clean Engine")
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument("run_folder", nargs="?", help="Path to single Stage-1 run folder")
    group.add_argument("--scan", help="Scan backtests/ for folders matching DIRECTIVE_NAME_*")
    
    args = parser.parse_args()
    
    if args.scan:
        directive_name = args.scan
        backtests_root = BACKTESTS_DIR
        runs_root = RUNS_DIR # For governance
        
        if not backtests_root.exists():
            print(f"[FAIL] Backtests directory not found: {backtests_root}")
            sys.exit(1)
            
        candidates = sorted(list(backtests_root.glob(f"{directive_name}_*")))
        valid_runs = []
        for cand in candidates:
            if cand.is_dir() and (cand / "metadata" / "run_metadata.json").exists():
                valid_runs.append(cand)
        
        if not valid_runs:
            print(f"[SCAN] No valid run folders found for directive: {directive_name}")
            sys.exit(1)
            
        print(f"[SCAN] Found {len(valid_runs)} valid runs for '{directive_name}'")
        
        # BATCH EXECUTION — Per-asset error tracking
        batch_succeeded = []
        batch_failed = []
        for run_folder in valid_runs:
            try:
                # --- GOVERNANCE CHECK ---
                # We need Run ID to verify state.
                with open(run_folder / "metadata" / "run_metadata.json", "r", encoding="utf-8") as f:
                    meta = json.load(f)
                    run_id = meta.get("run_id")
                    
                if not run_id:
                     raise ValueError("Run ID missing in metadata.")
                     
                state_mgr = PipelineStateManager(run_id)
                # Verify we are cleared for Stage 2
                state_mgr.verify_state("STAGE_1_COMPLETE")
                print(f"[GOVERNANCE] Verified -> {run_folder.name}")
                # ------------------------
                
                compile_stage2(run_folder)
                
                # Verify critical artifact was actually written
                ak_reports = list(run_folder.glob("AK_Trade_Report_*.xlsx"))
                if not ak_reports:
                    raise RuntimeError(f"AK_Trade_Report not found after compile_stage2 for {run_folder.name}")
                
                batch_succeeded.append(run_folder.name)
            except Exception as e:
                print(f"[ERROR] Failed to compile {run_folder.name}: {e}")
                batch_failed.append({"folder": run_folder.name, "error": str(e)})
        
        # --- BATCH SUMMARY ---
        print(f"\n[SCAN SUMMARY] Succeeded: {len(batch_succeeded)} | Failed: {len(batch_failed)}")
        if batch_failed:
            for f_info in batch_failed:
                print(f"  FAILED: {f_info['folder']} — {f_info['error']}")
            sys.exit(1)
    else:
        # SINGLE EXECUTION
        run_folder = Path(args.run_folder).resolve()
        if not run_folder.exists():
            print(f"[FAIL] Run folder not found: {run_folder}")
            sys.exit(1)
            
        try:
             # --- GOVERNANCE CHECK ---
            with open(run_folder / "metadata" / "run_metadata.json", "r", encoding="utf-8") as f:
                meta = json.load(f)
                run_id = meta.get("run_id")
            
            if not run_id:
                 raise ValueError("Run ID missing in metadata.")
                 
            state_mgr = PipelineStateManager(run_id)
            state_mgr.verify_state("STAGE_1_COMPLETE")
            print(f"[GOVERNANCE] Verified -> {run_folder.name}")
            # ------------------------
            
            compile_stage2(run_folder)
        except Exception as e:
            print(f"[FATAL] {e}")
            sys.exit(1)

if __name__ == "__main__":
    main()
