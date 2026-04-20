"""
real_model_evaluator.py — REAL_MODEL_V1

Parallel always-on reference capital model for CORE-status portfolios.

Model (REAL_MODEL_V1):
  * Nominal per asset: FX / crypto = $1000, XAU / indices = $2000
  * Deployed capital = (max_concurrent / n_assets) * sum(per_asset_nominal)
  * Risk schedule: 3% of current pooled equity, stepping +1% each time equity
    crosses a 1.5x multiple of starting deployed capital. Capped at 10%.
    Symmetric — falls back as equity retraces.
  * Pooled equity across all constituent trades, merged chronologically.
  * Lot sizing: lots = floor_to_step(risk_capital / (risk_distance *
    usd_per_pu_per_lot)). Skip trade if lots < min_lot.
  * pnl_new = r_multiple * (lots * risk_distance * usd_per_pu_per_lot)
  * Unlimited leverage, no heat cap, no rejection-net.

Applied ONLY to rows where portfolio_status == 'CORE' in
Master_Portfolio_Sheet.xlsx. Writes TradeScan_State/strategies/
Real_Model_Evaluation.xlsx with two data tabs (Portfolios,
Single-Asset Composites) and a Notes tab.
"""

from __future__ import annotations

import math
import sys
from pathlib import Path

import pandas as pd
import yaml

PROJECT_ROOT = Path(__file__).resolve().parent.parent
STATE_ROOT = PROJECT_ROOT.parent / "TradeScan_State"
BROKER_SPECS_DIR = PROJECT_ROOT / "data_access" / "broker_specs" / "OctaFx"

MPS_PATH = STATE_ROOT / "strategies" / "Master_Portfolio_Sheet.xlsx"
OUT_PATH = STATE_ROOT / "strategies" / "Real_Model_Evaluation.xlsx"

sys.path.insert(0, str(PROJECT_ROOT))
from tools.ledger_db import _connect  # noqa: E402

# --------------------------------------------------------------------------
# Nominal capital classification
# --------------------------------------------------------------------------

INDEX_SYMBOLS = {
    "GER40", "US30", "US500", "NAS100", "UK100", "JP225", "FRA40", "AUS200",
    "HK50", "ESP35", "EU50", "SPX500", "NDX100", "DJI30",
}


def _nominal_usd_for_symbol(symbol: str) -> float:
    """All assets = $1000 per asset (flat nominal)."""
    return 1000.0


# --------------------------------------------------------------------------
# Broker spec loader
# --------------------------------------------------------------------------

_BROKER_CACHE: dict[str, dict] = {}


def _load_broker_spec(symbol: str) -> dict | None:
    s = (symbol or "").upper()
    if s in _BROKER_CACHE:
        return _BROKER_CACHE[s]
    p = BROKER_SPECS_DIR / f"{s}.yaml"
    if not p.exists():
        _BROKER_CACHE[s] = None
        return None
    d = yaml.safe_load(p.read_text(encoding="utf-8")) or {}
    _BROKER_CACHE[s] = d
    return d


def _lot_sizing(symbol: str, risk_capital_usd: float, risk_distance: float) -> tuple[float, float]:
    """Return (lots, actual_risk_usd). lots may be 0.0 meaning SKIP."""
    spec = _load_broker_spec(symbol)
    if not spec or risk_distance is None or risk_distance <= 0:
        return 0.0, 0.0
    calib = spec.get("calibration") or {}
    upp = float(calib.get("usd_per_pu_per_lot") or 0.0)
    min_lot = float(spec.get("min_lot") or 0.01)
    lot_step = float(spec.get("lot_step") or 0.01)
    if upp <= 0:
        return 0.0, 0.0
    raw = risk_capital_usd / (risk_distance * upp)
    # floor to lot_step
    if lot_step <= 0:
        lot_step = 0.01
    lots = math.floor(raw / lot_step) * lot_step
    # clean float noise
    lots = round(lots, 8)
    if lots < min_lot - 1e-9:
        return 0.0, 0.0
    actual_risk = lots * risk_distance * upp
    return lots, actual_risk


# --------------------------------------------------------------------------
# Risk tier schedule
# --------------------------------------------------------------------------

def _risk_pct_for_equity(equity: float, starting: float) -> float:
    """2% base; +1% for each 2x multiple of starting crossed; capped at 5%.

    Tier 0: ratio < 2x    -> 2%
    Tier 1: 2x  <= ratio  -> 3%
    Tier 2: 4x  <= ratio  -> 4%
    Tier 3: 8x  <= ratio  -> 5% (cap)
    Symmetric: retrace below threshold steps back down.
    """
    if starting <= 0 or equity <= 0:
        return 0.02
    ratio = equity / starting
    if ratio < 1.0:
        return 0.02
    tier = int(math.floor(math.log(ratio) / math.log(2.0)))
    if tier < 0:
        tier = 0
    risk = 0.02 + 0.01 * tier
    return min(risk, 0.05)


# --------------------------------------------------------------------------
# Run-id -> strategy_name resolver
# --------------------------------------------------------------------------

def _build_runid_to_strategy_map(run_ids: list[str]) -> dict[str, tuple[str, str]]:
    """Return {run_id: (strategy_name, symbol)}."""
    if not run_ids:
        return {}
    conn = _connect()
    try:
        placeholders = ",".join("?" * len(run_ids))
        rows = conn.execute(
            f"SELECT run_id, strategy, symbol FROM master_filter WHERE run_id IN ({placeholders})",
            run_ids,
        ).fetchall()
        return {r[0]: (r[1], r[2]) for r in rows}
    finally:
        conn.close()


def _load_trades(strategy_name: str) -> pd.DataFrame | None:
    p = STATE_ROOT / "backtests" / strategy_name / "raw" / "results_tradelevel.csv"
    if not p.exists():
        return None
    try:
        df = pd.read_csv(p)
    except Exception:
        return None
    return df


# --------------------------------------------------------------------------
# Core simulator
# --------------------------------------------------------------------------

def _simulate_portfolio(
    portfolio_id: str,
    constituent_run_ids: list[str],
    max_concurrent: int,
) -> dict:
    rid_map = _build_runid_to_strategy_map(constituent_run_ids)

    all_trades: list[dict] = []
    symbols_seen: set[str] = set()
    nominal_per_asset: dict[str, float] = {}

    for rid in constituent_run_ids:
        info = rid_map.get(rid)
        if not info:
            continue
        strat_name, sym = info
        if not sym:
            continue
        symbols_seen.add(sym)
        nominal_per_asset[sym] = _nominal_usd_for_symbol(sym)
        df = _load_trades(strat_name)
        if df is None or df.empty:
            continue
        need = {"entry_timestamp", "exit_timestamp", "r_multiple", "risk_distance", "symbol"}
        if not need.issubset(df.columns):
            continue
        for _, row in df.iterrows():
            try:
                et = pd.to_datetime(row["entry_timestamp"], utc=True, errors="coerce")
                xt = pd.to_datetime(row["exit_timestamp"], utc=True, errors="coerce")
            except Exception:
                continue
            if pd.isna(et) or pd.isna(xt):
                continue
            rd = row.get("risk_distance")
            rm = row.get("r_multiple")
            if pd.isna(rd) or pd.isna(rm):
                continue
            all_trades.append({
                "entry": et,
                "exit": xt,
                "symbol": str(row["symbol"]),
                "risk_distance": float(rd),
                "r_multiple": float(rm),
                "run_id": rid,
            })

    n_assets = max(len(symbols_seen), 1)
    total_nominal = sum(nominal_per_asset.values())
    mc = max(int(max_concurrent or 1), 1)
    deployed = (min(mc, n_assets) / n_assets) * total_nominal if total_nominal > 0 else 0.0
    starting = deployed

    # Sort entries chronologically; process exits via priority queue to update equity
    all_trades.sort(key=lambda t: t["entry"])

    equity = starting
    peak_equity = starting
    min_equity = starting
    max_dd_usd = 0.0
    trades_executed = 0
    trades_skipped = 0
    current_tier = 0
    max_tier_reached = 0
    tier_upgrades = 0
    tier_downgrades = 0

    # Event queue: pending exits (sorted by exit time)
    import heapq
    pending_exits: list[tuple[pd.Timestamp, float]] = []  # (exit_ts, pnl_usd)

    def _apply_exits_up_to(ts: pd.Timestamp) -> None:
        nonlocal equity, peak_equity, min_equity, max_dd_usd
        while pending_exits and pending_exits[0][0] <= ts:
            _, pnl = heapq.heappop(pending_exits)
            equity += pnl
            if equity > peak_equity:
                peak_equity = equity
            if equity < min_equity:
                min_equity = equity
            dd = peak_equity - equity
            if dd > max_dd_usd:
                max_dd_usd = dd

    def _update_tier() -> None:
        nonlocal current_tier, max_tier_reached, tier_upgrades, tier_downgrades
        new_risk = _risk_pct_for_equity(equity, starting)
        new_tier = int(round((new_risk - 0.03) * 100))
        if new_tier > current_tier:
            tier_upgrades += (new_tier - current_tier)
            current_tier = new_tier
            if new_tier > max_tier_reached:
                max_tier_reached = new_tier
        elif new_tier < current_tier:
            tier_downgrades += (current_tier - new_tier)
            current_tier = new_tier

    for t in all_trades:
        _apply_exits_up_to(t["entry"])
        _update_tier()
        if equity <= 0:
            trades_skipped += 1
            continue
        risk_pct = _risk_pct_for_equity(equity, starting)
        risk_capital = equity * risk_pct
        lots, actual_risk = _lot_sizing(t["symbol"], risk_capital, t["risk_distance"])
        if lots <= 0:
            trades_skipped += 1
            continue
        pnl = t["r_multiple"] * actual_risk
        heapq.heappush(pending_exits, (t["exit"], pnl))
        trades_executed += 1

    # Drain remaining exits
    while pending_exits:
        _, pnl = heapq.heappop(pending_exits)
        equity += pnl
        if equity > peak_equity:
            peak_equity = equity
        if equity < min_equity:
            min_equity = equity
        dd = peak_equity - equity
        if dd > max_dd_usd:
            max_dd_usd = dd
    _update_tier()

    total_trades = trades_executed + trades_skipped
    skip_rate = (trades_skipped / total_trades * 100.0) if total_trades else 0.0
    realized = equity - starting
    final_risk = _risk_pct_for_equity(equity, starting)
    max_dd_pct = (max_dd_usd / peak_equity * 100.0) if peak_equity > 0 else 0.0
    ret_dd = (realized / max_dd_usd) if max_dd_usd > 0 else (float("inf") if realized > 0 else 0.0)
    eq_mult = (equity / starting) if starting > 0 else 0.0

    return {
        "portfolio_id": portfolio_id,
        "real_starting_usd": round(starting, 2),
        "real_deployed_usd": round(deployed, 2),
        "real_realized_pnl": round(realized, 2),
        "real_ending_equity": round(equity, 2),
        "real_min_equity_usd": round(min_equity, 2),
        "real_equity_multiplier": round(eq_mult, 4),
        "real_max_dd_usd": round(max_dd_usd, 2),
        "real_max_dd_pct": round(max_dd_pct, 2),
        "real_return_dd_ratio": (round(ret_dd, 2) if ret_dd != float("inf") else None),
        "real_final_risk_pct": round(final_risk * 100.0, 2),
        "real_tier_upgrades": tier_upgrades,
        "real_tier_downgrades": tier_downgrades,
        "real_trades_executed": trades_executed,
        "real_trades_skipped": trades_skipped,
        "skip_rate_pct": round(skip_rate, 2),
    }


# --------------------------------------------------------------------------
# Notes tab content
# --------------------------------------------------------------------------

NOTES_ROWS = [
    ("SECTION 1 — MODEL OVERVIEW", ""),
    ("Model ID", "REAL_MODEL_V1"),
    ("Scope", "Applied only to portfolios with portfolio_status == 'CORE'. Non-CORE rows are not evaluated."),
    ("Purpose", "Always-on parallel reference model. Not a candidate — the truth column that re-scales existing backtest trades through a realistic capital + risk schedule."),
    ("", ""),
    ("SECTION 2 — CAPITAL ALLOCATION", ""),
    ("Nominal per asset", "USD 1000 per asset, flat across FX / crypto / metals / indices."),
    ("Deployed capital", "deployed = min(max_concurrent, n_assets) * 1000 USD."),
    ("Example — 3 assets, max_concurrent=2", "deployed = 2 * 1000 = 2000 USD."),
    ("Example — 9 assets, max_concurrent=7", "deployed = 7 * 1000 = 7000 USD."),
    ("Starting equity", "starting_usd = deployed_usd. All equity tracked in USD."),
    ("", ""),
    ("SECTION 3 — RISK SCHEDULE (STAIR-STEP)", ""),
    ("Base risk", "2% of current pooled equity per trade."),
    ("Upgrade trigger", "Each time equity crosses a 2x multiple of starting capital, risk increases by +1%."),
    ("Cap", "Risk is capped at 5% (tier 3, i.e. equity >= 8x starting)."),
    ("Schedule", "Tier 0 (<2x): 2%  |  Tier 1 (>=2x): 3%  |  Tier 2 (>=4x): 4%  |  Tier 3 (>=8x): 5% (cap)."),
    ("Symmetry", "If equity retraces below the last 2x band, risk steps back down. Never below 2%."),
    ("Tier formula", "tier = floor(log(equity/starting) / log(2)); risk_pct = min(2 + tier, 5) %."),
    ("", ""),
    ("SECTION 4 — LOT SIZING & EXECUTION", ""),
    ("Risk capital", "risk_capital = equity * risk_pct."),
    ("Raw lots", "raw_lots = risk_capital / (risk_distance * usd_per_pu_per_lot)  [broker_specs/OctaFx/<SYMBOL>.yaml]."),
    ("Rounding", "Floor to lot_step (0.01). No rounding up."),
    ("Skip rule", "If floored_lots < min_lot (0.01) the trade is SKIPPED and counted toward real_trades_skipped."),
    ("P&L re-scaling", "pnl_new = r_multiple * (lots * risk_distance * usd_per_pu_per_lot). Matches live broker math."),
    ("Leverage / heat", "Unlimited leverage. No per-trade heat cap. No rejection-net."),
    ("", ""),
    ("SECTION 5 — POOLING & CONCURRENCY", ""),
    ("Equity pool", "All constituent trades share one pooled equity account (single number, updated on each trade exit)."),
    ("Trade ordering", "Trades from all constituents merged and processed in chronological order of entry_timestamp."),
    ("Max concurrent", "Taken from MPS.max_concurrent (historical from backtest). Used for deployed-capital sizing; there is no further concurrent-trade cap at execution time."),
    ("", ""),
    ("SECTION 6 — COLUMN GLOSSARY", ""),
    ("real_starting_usd", "Starting pooled equity = deployed capital (USD)."),
    ("real_deployed_usd", "Nominal deployed capital after concurrency scaling (USD). Equals real_starting_usd."),
    ("real_realized_pnl", "Ending equity minus starting equity (USD)."),
    ("real_ending_equity", "Pooled equity after all trades (USD)."),
    ("real_equity_multiplier", "real_ending_equity / real_starting_usd."),
    ("real_max_dd_usd", "Largest peak-to-trough drawdown in pooled equity (USD)."),
    ("real_max_dd_pct", "real_max_dd_usd divided by the peak equity, as a percentage."),
    ("real_return_dd_ratio", "real_realized_pnl / real_max_dd_usd. Blank if DD = 0."),
    ("real_final_risk_pct", "Final risk tier at end of simulation, as percent (3–10)."),
    ("real_tier_upgrades", "Count of +1% tier upgrades experienced during the simulation."),
    ("real_tier_downgrades", "Count of −1% tier downgrades (retrace-driven)."),
    ("real_trades_executed", "Trades that passed the min-lot filter and were applied to equity."),
    ("real_trades_skipped", "Trades dropped because floored lots < min_lot (0.01)."),
    ("skip_rate_pct", "100 * real_trades_skipped / (executed + skipped)."),
    ("", ""),
    ("SECTION 7 — TAB LAYOUT", ""),
    ("Portfolios tab", "CORE rows whose portfolio_id starts with 'PF_' (multi-strategy composites)."),
    ("Single-Asset Composites tab", "CORE rows where portfolio_id == source_strategy (single strategy, multi-symbol)."),
    ("Refresh", "Regenerated by running tools/real_model_evaluator.py. Reads the live MPS; writes this file in place."),
]


# --------------------------------------------------------------------------
# Main
# --------------------------------------------------------------------------

COLUMNS = [
    "portfolio_id",
    "real_starting_usd",
    "real_deployed_usd",
    "real_realized_pnl",
    "real_ending_equity",
    "real_min_equity_usd",
    "real_equity_multiplier",
    "real_max_dd_usd",
    "real_max_dd_pct",
    "real_return_dd_ratio",
    "real_final_risk_pct",
    "real_tier_upgrades",
    "real_tier_downgrades",
    "real_trades_executed",
    "real_trades_skipped",
    "skip_rate_pct",
]


def main() -> int:
    if not MPS_PATH.exists():
        print(f"[FAIL] MPS not found: {MPS_PATH}")
        return 1
    mps = pd.read_excel(MPS_PATH)
    core = mps[mps["portfolio_status"].astype(str).str.upper() == "CORE"].copy()
    if core.empty:
        print("[INFO] No CORE rows in MPS.")
        # still write empty workbook so downstream never misses it
    print(f"[INFO] Evaluating {len(core)} CORE rows with REAL_MODEL_V1...")

    portfolio_rows = []
    single_rows = []

    for _, row in core.iterrows():
        pid = str(row.get("portfolio_id", "")).strip()
        if not pid:
            continue
        cids = [c.strip() for c in str(row.get("constituent_run_ids", "")).split(",") if c.strip()]
        mc = row.get("max_concurrent", 1)
        try:
            mc_int = int(mc) if not pd.isna(mc) else 1
        except Exception:
            mc_int = 1

        print(f"  - {pid}  constituents={len(cids)}  max_concurrent={mc_int}")
        res = _simulate_portfolio(pid, cids, mc_int)

        if pid.startswith("PF_"):
            portfolio_rows.append(res)
        else:
            single_rows.append(res)

    df_port = pd.DataFrame(portfolio_rows, columns=COLUMNS) if portfolio_rows else pd.DataFrame(columns=COLUMNS)
    df_single = pd.DataFrame(single_rows, columns=COLUMNS) if single_rows else pd.DataFrame(columns=COLUMNS)
    df_notes = pd.DataFrame(NOTES_ROWS, columns=["Field", "Description"])

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    try:
        with pd.ExcelWriter(OUT_PATH, engine="openpyxl") as w:
            df_port.to_excel(w, sheet_name="Portfolios", index=False)
            df_single.to_excel(w, sheet_name="Single-Asset Composites", index=False)
            df_notes.to_excel(w, sheet_name="Notes", index=False)
    except PermissionError:
        print(f"[BLOCK] Cannot write {OUT_PATH.name} — file is open in another application. Close Excel and retry.")
        return 1

    print(f"[OK] Wrote {OUT_PATH}")
    print(f"     Portfolios: {len(df_port)}   Single-Asset Composites: {len(df_single)}")

    _apply_inline_formatting(OUT_PATH)
    return 0


def _apply_inline_formatting(path: Path) -> None:
    """Light formatting: header style, column widths, number formats, freeze, Notes wrap."""
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    ALT_FILL = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")

    CURR_COLS = {"real_starting_usd", "real_deployed_usd", "real_realized_pnl",
                 "real_ending_equity", "real_min_equity_usd", "real_max_dd_usd"}
    FLOAT_COLS = {"real_equity_multiplier", "real_return_dd_ratio"}
    PCT_COLS = {"real_max_dd_pct", "real_final_risk_pct", "skip_rate_pct"}
    INT_COLS = {"real_tier_upgrades", "real_tier_downgrades",
                "real_trades_executed", "real_trades_skipped"}

    wb = openpyxl.load_workbook(path)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.max_row == 0:
            continue

        if sheet_name == "Notes":
            ws.column_dimensions["A"].width = 42
            ws.column_dimensions["B"].width = 144
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
            # Bold header
            for c in ws[1]:
                c.fill = HEADER_FILL
                c.font = HEADER_FONT
            ws.freeze_panes = "A2"
            continue

        # Data sheet
        headers = [c.value for c in ws[1]]
        for c in ws[1]:
            c.fill = HEADER_FILL
            c.font = HEADER_FONT
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[1].height = 32

        for idx, name in enumerate(headers, start=1):
            letter = openpyxl.utils.get_column_letter(idx)
            if name == "portfolio_id":
                ws.column_dimensions[letter].width = 44
            else:
                ws.column_dimensions[letter].width = 18

            fmt = None
            if name in CURR_COLS:
                fmt = "#,##0.00"
            elif name in FLOAT_COLS:
                fmt = "0.00"
            elif name in PCT_COLS:
                fmt = '0.00"%"'
            elif name in INT_COLS:
                fmt = "0"
            if fmt:
                for r in range(2, ws.max_row + 1):
                    ws.cell(row=r, column=idx).number_format = fmt

        # Alt row banding
        for r in range(2, ws.max_row + 1):
            if r % 2 == 0:
                for c in range(1, len(headers) + 1):
                    ws.cell(row=r, column=c).fill = ALT_FILL

        ws.freeze_panes = "B2"

    wb.save(path)
    print(f"[OK] Formatted {path.name}")


if __name__ == "__main__":
    raise SystemExit(main())
