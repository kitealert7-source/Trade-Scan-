"""
safe_append_excel.py — Non-Destructive Excel Appender

Appends a row of data to an existing Excel file using openpyxl, preserving
sheet structure and triggering the formatter to ensure strict styling.

Usage:
    python tools/safe_append_excel.py \
        --file strategies/Master_Portfolio_Sheet.xlsx \
        --data '{"portfolio_id": "IDX27", "net_pnl_usd": 100.0}' \
        --profile portfolio
"""

import argparse
import sys
import json
import subprocess
from pathlib import Path
import openpyxl

def safe_append(file_path, data_dict, profile):
    path = Path(file_path)
    if not path.exists():
        print(f"[ERROR] File not found: {file_path}")
        sys.exit(1)

    print(f"[INFO] Appending to {path.name}...")

    try:
        # 1. Load Workbook
        wb = openpyxl.load_workbook(path)
        ws = wb.active  # Assume first/active sheet is target

        # 2. Map Headers
        # Read header row (Row 1)
        headers = {}
        for cell in ws[1]:
            if cell.value:
                headers[str(cell.value).strip()] = cell.column

        if not headers:
            print("[ERROR] No headers found in row 1.")
            sys.exit(1)

        # 3. Prepare Row Data
        # Match data keys to headers
        # Use simple string matching (case-sensitive by default? converting inputs to match?)
        # Let's try to match intelligently (lower-case match)
        
        headers_lower = {k.lower(): v for k, v in headers.items()}
        
        # Determine target row (first empty)
        target_row = ws.max_row + 1
        
        # Write Data
        for key, value in data_dict.items():
            col_idx = headers_lower.get(str(key).lower())
            if col_idx:
                ws.cell(row=target_row, column=col_idx, value=value)
            else:
                print(f"[WARNING] Key '{key}' not found in headers. Skipping.")

        # 4. Save
        wb.save(path)
        print(f"[SUCCESS] Row appended to {path.name} (Row {target_row}).")

    except Exception as e:
        print(f"[FATAL] Append failed: {e}")
        sys.exit(1)

    # 5. Trigger Formatter
    print("[INFO] Triggering formatter...")
    tool_path = Path(__file__).parent / "format_excel_artifact.py"
    
    result = subprocess.run(
        [sys.executable, str(tool_path), "--file", str(path), "--profile", profile],
        capture_output=True,
        text=True
    )
    
    print(result.stdout)
    if result.returncode != 0:
        print(f"[ERROR] Formatter failed:\n{result.stderr}")
        sys.exit(result.returncode)

def main():
    parser = argparse.ArgumentParser(description="Safe Excel Appender")
    parser.add_argument("--file", required=True, help="Path to Excel file")
    parser.add_argument("--data", required=True, help="JSON string of row data")
    parser.add_argument("--profile", required=True, choices=["strategy", "portfolio"], help="Formatting profile")
    
    args = parser.parse_args()
    
    # Parse JSON data
    try:
        data = json.loads(args.data)
    except json.JSONDecodeError as e:
        print(f"[ERROR] Invalid JSON data: {e}")
        sys.exit(1)
        
    safe_append(args.file, data, args.profile)

if __name__ == "__main__":
    main()
