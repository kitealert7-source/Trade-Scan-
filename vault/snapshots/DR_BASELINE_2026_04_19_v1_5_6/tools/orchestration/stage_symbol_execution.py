"""Symbol execution stages (stage-0.9 through stage-3A).

Phase 6 Refactor: Split into 4 focused functions that map 1:1 to StageRegistry stages.
  run_stage1_execution()    — Strategy snapshots + backtest registry worker loop
  run_stage2_compilation()  — Engine resolution + Stage-2 --scan
  run_stage3_aggregation()  — Stage-3 aggregation compiler + cardinality gate
  run_manifest_binding()    — Per-run snapshot verify, artifact hash, manifest write, FSM close

run_symbol_execution_stages() preserved as a backward-compatible orchestrator that calls all 4.
"""

from __future__ import annotations

import hashlib
import json
import shutil
import importlib
import pandas as pd
from datetime import datetime, timezone
from pathlib import Path

from tools.orchestration.pipeline_errors import PipelineExecutionError
from tools.orchestration.run_registry import (
    claim_next_planned_run,
    ensure_registry,
    list_runs,
    requeue_running_runs,
    update_run_state,
)
from tools.orchestration.transition_service import (
    transition_directive_state,
    transition_run_state,
    transition_run_state_sequence,
)
from tools.pipeline_utils import PipelineContext, PipelineStateManager
from tools.system_registry import log_run_to_registry
from config.state_paths import RUNS_DIR, BACKTESTS_DIR, STRATEGIES_DIR, MASTER_FILTER_PATH
from config.engine_loader import get_active_engine


# ---------------------------------------------------------------------------
# Silent-zero guardrail helpers
# ---------------------------------------------------------------------------

_SILENT_ZERO_TOP_CAUSES = (
    "1) Default session_reset trap: 1D directives without "
    "`trade_management.session_reset: none` inherit `utc_day`, which clears "
    "pending entries each bar-close and zeroes trades.",
    "2) direction_restriction mismatch: strategy emits long-only (or short-only) "
    "signals but the directive's `state_machine.entry.direction` / "
    "`trade_management.direction_restriction` filters them out.",
    "3) volatility_regime encoding: strategy compares `volatility_regime` to "
    "string values ('low'/'normal'/'high') but the engine emits numeric -1/0/1 — "
    "the comparison is always False and every entry is rejected.",
)


def _load_dryrun_stats(directive_id: str) -> dict | None:
    """Read the dry-run stats sidecar written by strategy_dryrun_validator.

    Returns None if the file is absent or malformed — callers treat a missing
    sidecar as "no opinion" and fall back to the legacy soft NO_TRADES path.
    """
    path = RUNS_DIR / directive_id / "dryrun_stats.json"
    if not path.exists():
        return None
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception as _e:
        print(f"[SILENT-ZERO] WARNING: Failed to read {path}: {_e}")
        return None


def _silent_zero_message(directive_id: str, symbol: str, stats: dict) -> str:
    """Compose the SILENT_ZERO_TRADES failure message with top-3 causes."""
    sig = stats.get("signal_count", 0)
    bars = stats.get("sample_bars", 0)
    sample_sym = stats.get("sample_symbol", "?")
    causes = "\n    ".join(_SILENT_ZERO_TOP_CAUSES)
    return (
        f"SILENT_ZERO_TRADES: dry-run emitted {sig} signal(s) on {bars} sample "
        f"bars of {sample_sym}, but Stage-1 backtest for {symbol} produced 0 "
        f"trades. This is almost always a check_entry / engine-context mismatch. "
        f"Directive: {directive_id}.\n"
        f"  Top-3 likely causes:\n    {causes}\n"
        f"  Remediation: inspect the three items above against this directive + "
        f"strategy.py; if genuinely a no-edge outcome, confirm by widening the "
        f"dry-run sample (strategy_dryrun_validator sample_bars) and re-running."
    )


# ---------------------------------------------------------------------------
# Stage-1: Backtest Execution Loop
# ---------------------------------------------------------------------------

def run_stage1_execution(context: PipelineContext) -> None:
    """
    Stage-1: Strategy snapshot + registry worker loop.

    Iterates planned runs, executes backtests, transitions run FSM to STAGE_1_COMPLETE.
    Does NOT invoke Stage-2 or beyond.
    """
    clean_id = context.directive_id
    p_conf = context.directive_config
    run_ids = context.run_ids
    symbols = context.symbols
    project_root = context.project_root
    registry_path = context.registry_path

    strategy_id = p_conf.get("Strategy", p_conf.get("strategy")) or clean_id
    if registry_path is None:
        registry_path = RUNS_DIR / clean_id / "run_registry.json"

    if not registry_path.exists():
        ensure_registry(
            registry_path,
            clean_id,
            [
                {"run_id": rid, "strategy": strategy_id, "symbol": sym}
                for rid, sym in zip(run_ids, symbols)
            ],
        )

    reclaimed = requeue_running_runs(registry_path, clean_id)
    if reclaimed:
        print(f"[ORCHESTRATOR] Re-queued {reclaimed} interrupted RUNNING registry entries to PLANNED.")

    registry_runs = list_runs(registry_path, clean_id)
    if not registry_runs:
        raise PipelineExecutionError(
            f"Run registry has no planned runs: {registry_path}",
            directive_id=clean_id,
            run_ids=run_ids,
        )
    run_ids = [run["run_id"] for run in registry_runs]
    symbols = [run["symbol"] for run in registry_runs]

    # Store resolved run_ids/symbols back into context (registry may reorder)
    context.run_ids = run_ids
    context.symbols = symbols

    any_stage1_rerun = any(
        PipelineStateManager(rid).get_state_data()["current_state"]
        in ("IDLE", "PREFLIGHT_COMPLETE", "PREFLIGHT_COMPLETE_SEMANTICALLY_VALID")
        for rid in run_ids
    )
    summary_csv = BACKTESTS_DIR / f"batch_summary_{clean_id}.csv"
    if summary_csv.exists() and any_stage1_rerun:
        summary_csv.unlink()

    if strategy_id:
        source_strategy_path = project_root / "strategies" / strategy_id / "strategy.py"
        if source_strategy_path.exists():
            print("[ORCHESTRATOR] Performing atomic preemptive strategy snapshots...")
            for rid in run_ids:
                mgr = PipelineStateManager(rid)
                target_path = mgr.run_dir / "strategy.py"
                if not target_path.exists():
                    target_path.parent.mkdir(parents=True, exist_ok=True)
                    shutil.copy(str(source_strategy_path), str(target_path))

    print("[ORCHESTRATOR] Launching Stage-1 Generator (Registry Worker)...")
    skip_states = {
        "STAGE_1_COMPLETE",
        "STAGE_2_COMPLETE",
        "STAGE_3_COMPLETE",
        "STAGE_3A_COMPLETE",
        "COMPLETE",
    }
    while True:
        # Heartbeat all runs to prevent Watchdog timeouts during long sequential processing
        for r_id in run_ids:
            try:
                PipelineStateManager(r_id).record_heartbeat()
            except Exception:
                pass

        claim = claim_next_planned_run(registry_path, clean_id)
        if claim is None:
            break

        rid = claim["run_id"]
        symbol = claim["symbol"]
        mgr = PipelineStateManager(rid)
        current = mgr.get_state_data()["current_state"]

        if current in skip_states:
            update_run_state(registry_path, clean_id, rid, "COMPLETE")
            continue

        if current == "FAILED":
            update_run_state(
                registry_path,
                clean_id,
                rid,
                "FAILED",
                last_error="Run state already FAILED before Stage-1 execution.",
            )
            raise RuntimeError(f"Run {rid} is already FAILED before Stage-1.")

        try:
            from tools.skill_loader import run_skill

            run_skill("backtest_execution", strategy=clean_id, symbol=symbol, run_id=rid)

            out_folder = BACKTESTS_DIR / f"{clean_id}_{symbol}"
            if not (out_folder / "raw" / "results_tradelevel.csv").exists():
                # Engine exited cleanly (no exception) but produced no trade data.
                # Silent-zero guardrail: if Stage-0.75 dry-run proved check_entry()
                # can emit signals on a real data sample, 0 trades at Stage-1 is a
                # regression (engine-context mismatch, session_reset trap, etc.) —
                # hard-fail instead of soft-skipping.
                _dry_stats = _load_dryrun_stats(clean_id)
                if _dry_stats and _dry_stats.get("executed") and _dry_stats.get("signal_count", 0) > 0:
                    _msg = _silent_zero_message(clean_id, symbol, _dry_stats)
                    print(f"[STAGE-1] {_msg}")
                    try:
                        transition_run_state(rid, "FAILED")
                        update_run_state(registry_path, clean_id, rid, "FAILED",
                                         last_error="SILENT_ZERO_TRADES")
                    except Exception as _cleanup_err:
                        print(f"[WARN] Failed to mark {symbol} FAILED: {_cleanup_err}")
                    log_run_to_registry(rid, "silent_zero", clean_id)
                    raise PipelineExecutionError(
                        _msg, directive_id=clean_id, run_ids=[rid],
                    )

                # Write a persistent marker so the cardinality gate and cleanup tools
                # can identify and correctly handle this run (not a crash).
                marker = {
                    "run_id": rid,
                    "symbol": symbol,
                    "status": "NO_TRADES",
                    "valid": False,
                    "timestamp": datetime.now(timezone.utc).isoformat(),
                }
                raw_dir = out_folder / "raw"
                raw_dir.mkdir(parents=True, exist_ok=True)
                with open(raw_dir / "status_no_trades.json", "w", encoding="utf-8") as mf:
                    json.dump(marker, mf, indent=2)
                with open(RUNS_DIR / rid / "status_no_trades.json", "w", encoding="utf-8") as mf:
                    json.dump(marker, mf, indent=2)
                print(f"[STAGE-1] NO_TRADES: {symbol} ({rid[:8]}) produced 0 trades. Run skipped (not a crash).")
                transition_run_state(rid, "FAILED")
                update_run_state(registry_path, clean_id, rid, "FAILED", last_error="NO_TRADES")
                log_run_to_registry(rid, "no_trades", clean_id)
                continue  # Do NOT raise — proceed to next symbol

            transition_run_state(rid, "STAGE_1_COMPLETE")
            update_run_state(registry_path, clean_id, rid, "COMPLETE")
            try:
                from tools.run_index import append_run_to_index
                append_run_to_index(clean_id, symbol)
            except Exception as idx_err:
                print(f"[INDEX] append failed (non-blocking): {idx_err}")
        except Exception as err:
            print(f"[ERROR] Stage-1 Failed for {symbol}: {err}")
            try:
                transition_run_state(rid, "FAILED")
            except Exception as cleanup_err:
                print(f"[WARN] Failed to mark {symbol} run as FAILED: {cleanup_err}")
            try:
                update_run_state(registry_path, clean_id, rid, "FAILED", last_error=str(err))
            except Exception as reg_err:
                print(f"[WARN] Failed to update registry state for {rid}: {reg_err}")

            # Master Ledger Update
            log_run_to_registry(rid, "failed", clean_id)
            raise err


# ---------------------------------------------------------------------------
# Stage-2: Compilation
# ---------------------------------------------------------------------------

def run_stage2_compilation(context: PipelineContext) -> None:
    """
    Stage-2: Engine resolution + compilation scan.

    Resolves the active engine module, validates its existence,
    then invokes the stage-2 compiler across all symbols.
    Transitions per-run FSM to STAGE_2_COMPLETE.
    """
    clean_id = context.directive_id
    run_ids = context.run_ids
    symbols = context.symbols
    python_exe = context.python_exe
    registry_path = context.registry_path
    if registry_path is None:
        registry_path = RUNS_DIR / clean_id / "run_registry.json"

    from tools.orchestration.execution_adapter import run_command

    # Engine Version Registry Resolution
    try:
        active_engine = get_active_engine()
    except Exception as e:
        raise PipelineExecutionError(f"Engine Resolution Failed: {e}", directive_id=clean_id)

    engine_module = f"engine_dev.universal_research_engine.{active_engine}.stage2_compiler"

    # Safety Check: Verify module existence via importlib
    try:
        importlib.import_module(engine_module)
    except ImportError:
        raise PipelineExecutionError(
            f"Configured engine version '{active_engine}' does not exist. (Module {engine_module} not found).",
            directive_id=clean_id
        )

    # --- REC-04: Pre-entry artifact check ---
    # Verify results_tradelevel.csv exists for every run that should have it before
    # invoking stage2_compiler. Catches filesystem inconsistency or state drift where
    # a run's FSM was advanced past FAILED without Stage-1 actually completing.
    # Skips runs already FAILED (no-trades or prior crash) and STAGE_2_COMPLETE (resume).
    _artifact_failures: list[tuple[str, str]] = []
    for rid, symbol in zip(run_ids, symbols):
        _run_state = PipelineStateManager(rid).get_state_data().get("current_state", "IDLE")
        if _run_state in ("FAILED", "STAGE_2_COMPLETE"):
            continue
        _csv = BACKTESTS_DIR / f"{clean_id}_{symbol}" / "raw" / "results_tradelevel.csv"
        if not _csv.exists():
            _artifact_failures.append((rid, symbol))
            print(
                f"[STAGE-2] Stage-1 artifact missing for {symbol} ({rid[:8]}): "
                f"results_tradelevel.csv not found — marking FAILED."
            )
            try:
                transition_run_state(rid, "FAILED")
            except Exception:
                pass
            update_run_state(
                registry_path,
                clean_id,
                rid,
                "FAILED",
                last_error="Stage-1 artifact missing before Stage-2 invocation (results_tradelevel.csv not found).",
            )
            log_run_to_registry(rid, "failed", clean_id)
    if _artifact_failures:
        raise PipelineExecutionError(
            f"[STAGE-2] Stage-1 artifacts missing for {len(_artifact_failures)} run(s) — "
            "cannot invoke stage2_compiler:\n"
            + "\n".join(
                f"  {sym} ({rid[:8]}): results_tradelevel.csv not found"
                for rid, sym in _artifact_failures
            ),
            directive_id=clean_id,
            run_ids=[rid for rid, _ in _artifact_failures],
        )

    run_command(
        [python_exe, "-m", engine_module, "--scan", clean_id],
        "Stage-2 Compilation",
    )

    for rid, symbol in zip(run_ids, symbols):
        mgr = PipelineStateManager(rid)
        current = mgr.get_state_data()["current_state"]
        if current in ("STAGE_1_COMPLETE", "STAGE_2_COMPLETE"):
            run_folder = BACKTESTS_DIR / f"{clean_id}_{symbol}"
            ak_reports = list(run_folder.glob("AK_Trade_Report_*.xlsx"))
            if ak_reports:
                if current != "STAGE_2_COMPLETE":
                    transition_run_state(rid, "STAGE_2_COMPLETE")
            else:
                print(f"[WARN] Stage-2 artifact missing for {symbol} ({rid[:8]}). Marking FAILED.")
                transition_run_state(rid, "FAILED")
                update_run_state(
                    registry_path,
                    clean_id,
                    rid,
                    "FAILED",
                    last_error="Stage-2 artifact missing (AK_Trade_Report_*.xlsx).",
                )
                log_run_to_registry(rid, "failed", clean_id)


# ---------------------------------------------------------------------------
# Stage-3: Aggregation + Cardinality Gate
# ---------------------------------------------------------------------------

def run_stage3_aggregation(context: PipelineContext) -> None:
    """
    Stage-3: Portfolio aggregation compiler + cardinality enforcement gate.

    Guardrail: cardinality gate is INSIDE this function.
    If cardinality fails, a PipelineExecutionError is raised → ManifestBindingStage is never reached.
    This preserves portfolio integrity ordering: aggregate → validate → bind.
    """
    clean_id = context.directive_id
    run_ids = context.run_ids
    symbols = context.symbols
    python_exe = context.python_exe

    from tools.orchestration.execution_adapter import run_command

    # --- IDEMPOTENCY PRE-CHECK ---
    # If Master Filter already contains the expected number of rows for this
    # directive, skip the stage3_compiler write entirely. Prevents duplicate row
    # appends when resuming after a downstream failure (e.g. Portfolio failed after
    # Aggregation already succeeded). The cardinality gate below still verifies.
    master_filter_path = MASTER_FILTER_PATH
    _skip_write = False
    if master_filter_path.exists():
        # Expected symbol set: all symbols except those with a NO_TRADES marker.
        _expected_symbols = {
            sym for rid, sym in zip(run_ids, symbols)
            if not (RUNS_DIR / rid / "status_no_trades.json").exists()
        }
        if _expected_symbols:
            import openpyxl as _oxl
            _wb = None
            try:
                _wb = _oxl.load_workbook(master_filter_path, read_only=True)
            except Exception as _load_err:
                print(f"[AGGREGATION] Master Filter unreadable ({_load_err}) — falling through to full write.")
            if _wb is not None:
                _ws = _wb.active
                try:
                    _headers = list(next(_ws.iter_rows(min_row=1, max_row=1, values_only=True)))
                    _strat_idx = _headers.index("strategy")
                    _sym_idx = _headers.index("symbol")
                    # Extract symbol strings only — list preserves duplicates for count check.
                    # _present_syms: List[str], one entry per matching row, NOT full row objects.
                    _min_col = max(_strat_idx, _sym_idx)
                    # Also collect run_ids for present rows so reruns (new run_ids,
                    # same symbols) are not falsely treated as already-written.
                    _run_id_idx = _headers.index("run_id") if "run_id" in _headers else None
                    _min_col2 = max(_strat_idx, _sym_idx, _run_id_idx if _run_id_idx is not None else 0)
                    _matching_rows = [
                        row
                        for row in _ws.iter_rows(min_row=2, values_only=True)
                        if row
                        and len(row) > _min_col2
                        and row[_strat_idx]
                        and str(row[_strat_idx]).startswith(clean_id)
                        and row[_sym_idx]
                    ]
                    _present_syms = [str(row[_sym_idx]) for row in _matching_rows]
                    _present_run_ids = (
                        {str(row[_run_id_idx]) for row in _matching_rows if row[_run_id_idx]}
                        if _run_id_idx is not None else set()
                    )
                    _current_run_ids = {
                        rid for rid, sym in zip(run_ids, symbols)
                        if not (RUNS_DIR / rid / "status_no_trades.json").exists()
                    }
                    _present_set = set(_present_syms)
                    _dup_count = len(_present_syms) - len(_present_set)

                    # Guard: row count must equal expected symbol count.
                    # Checked BEFORE set comparison — a set collapses duplicates and
                    # would incorrectly approve a skip when duplicates are present.
                    if len(_present_syms) != len(_expected_symbols):
                        _missing = _expected_symbols - _present_set
                        _extra = _present_set - _expected_symbols
                        if _missing:
                            print(f"[AGGREGATION] Missing symbols: {sorted(_missing)} — will write.")
                        if _extra:
                            print(f"[AGGREGATION] Unexpected symbols: {sorted(_extra)} — will write.")
                        if _dup_count > 0:
                            from collections import Counter as _Counter
                            _dupes = {s: c for s, c in _Counter(_present_syms).items() if c > 1}
                            print(f"[AGGREGATION] Duplicate rows ({_dup_count} extra): {_dupes} — will write.")
                    elif _present_set == _expected_symbols:
                        # Skip only when the current run_ids are already in master_filter.
                        # If run_ids differ (e.g. ENGINE rerun), the symbols match but the
                        # rows belong to the old run — we must write the new ones.
                        _run_ids_present = bool(_current_run_ids) and _current_run_ids.issubset(_present_run_ids)
                        if _run_ids_present:
                            _skip_write = True
                            print(
                                f"[AGGREGATION] Symbol set already complete for {clean_id} "
                                f"({sorted(_present_set)}) — skipping stage3_compiler write."
                            )
                        else:
                            _new_ids = _current_run_ids - _present_run_ids
                            print(
                                f"[AGGREGATION] Symbol set complete for {clean_id} but "
                                f"{len(_new_ids)} new run_id(s) not yet in Master Filter "
                                f"(e.g. ENGINE rerun) — removing {len(_present_syms)} stale row(s) before write."
                            )
                            # ENGINE rerun: stale rows exist for the same symbols but belong to
                            # an old run_id. Cardinality gate counts ALL rows for clean_id, so
                            # old rows must be cleared BEFORE stage3_compiler runs — from BOTH
                            # SQLite (source of truth) and Excel (derived view).
                            # stage3_compiler reads df_master from SQLite via read_master_filter();
                            # if old rows remain there, they are concat'd with new rows → 2x rows
                            # → cardinality gate fires.
                            try:
                                # 1. SQLite purge — must come first (stage3_compiler reads from here)
                                from tools.ledger_db import _connect as _ldb_connect
                                _ldb = _ldb_connect()
                                _del_result = _ldb.execute(
                                    'DELETE FROM master_filter WHERE "strategy" LIKE ?',
                                    (f"{clean_id}%",),
                                )
                                _deleted_db = _del_result.rowcount
                                _ldb.commit()
                                _ldb.close()
                                print(
                                    f"[AGGREGATION] Purged {_deleted_db} stale row(s) from SQLite for "
                                    f"{clean_id}."
                                )
                            except Exception as _db_cleanup_err:
                                print(
                                    f"[AGGREGATION][WARN] SQLite purge failed for "
                                    f"{clean_id}: {_db_cleanup_err} — proceeding; cardinality gate "
                                    f"may fail if duplicates remain."
                                )
                            try:
                                # 2. Excel purge — derived view; best-effort after SQLite
                                import openpyxl as _oxl2
                                _wb2 = _oxl2.load_workbook(master_filter_path)
                                _ws2 = _wb2.active
                                _hdr2 = list(next(_ws2.iter_rows(min_row=1, max_row=1, values_only=True)))
                                _si2 = _hdr2.index("strategy")
                                _stale_rows = [
                                    row[0].row
                                    for row in _ws2.iter_rows(min_row=2)
                                    if row[_si2].value and str(row[_si2].value).startswith(clean_id)
                                ]
                                for _rn in sorted(_stale_rows, reverse=True):
                                    _ws2.delete_rows(_rn)
                                _wb2.save(master_filter_path)
                                _wb2.close()
                                print(
                                    f"[AGGREGATION] Cleared {len(_stale_rows)} stale Excel row(s) for "
                                    f"{clean_id} — stage3_compiler will write fresh entries."
                                )
                            except Exception as _cleanup_err:
                                print(
                                    f"[AGGREGATION][WARN] Excel purge failed for "
                                    f"{clean_id}: {_cleanup_err} — proceeding; SQLite was already cleaned."
                                )
                    else:
                        # Count matches but symbols differ (wrong symbols in filter).
                        _missing = _expected_symbols - _present_set
                        _extra = _present_set - _expected_symbols
                        if _missing:
                            print(f"[AGGREGATION] Missing symbols: {sorted(_missing)} — will write.")
                        if _extra:
                            print(f"[AGGREGATION] Unexpected symbols: {sorted(_extra)} — will write.")
                except (StopIteration, ValueError):
                    pass  # Malformed or empty file — fall through to full write
                finally:
                    _wb.close()

    if not _skip_write:
        run_command([python_exe, "tools/stage3_compiler.py", clean_id], "Stage-3 Aggregation")
    if not master_filter_path.exists():
        raise PipelineExecutionError(
            f"Stage-3 artifact missing: {master_filter_path}",
            directive_id=clean_id,
            run_ids=run_ids,
        )

    import openpyxl

    wb = openpyxl.load_workbook(master_filter_path, read_only=True)
    ws = wb.active
    try:
        headers = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        strategy_idx = headers.index("strategy")
    except Exception as err:
        wb.close()
        raise PipelineExecutionError(
            f"Failed to resolve 'strategy' column in Master Filter: {err}",
            directive_id=clean_id,
            run_ids=run_ids,
        ) from err

    actual_count = sum(
        1
        for row in ws.iter_rows(min_row=2, values_only=True)
        if row and len(row) > strategy_idx and row[strategy_idx] and str(row[strategy_idx]).startswith(clean_id)
    )
    wb.close()
    no_trades_count = sum(
        1 for rid in run_ids
        if (RUNS_DIR / rid / "status_no_trades.json").exists()
    )
    expected_count = len(symbols) - no_trades_count
    if no_trades_count:
        print(f"[GATE] {no_trades_count} run(s) had NO_TRADES — excluded from cardinality check.")
    if actual_count != expected_count:
        raise PipelineExecutionError(
            f"Stage-3 cardinality mismatch: expected {expected_count}, found {actual_count} for {clean_id}",
            directive_id=clean_id,
            run_ids=run_ids,
        )

    print(f"[GATE] Stage-3 artifact verified: {actual_count}/{expected_count} rows for {clean_id}")


# ---------------------------------------------------------------------------
# Stage-3a: Manifest Binding + Run Close
# ---------------------------------------------------------------------------

def run_manifest_binding(context: PipelineContext) -> None:
    """
    Stage-3a: Per-run snapshot verification, artifact hashing, manifest write, and FSM close.

    Only runs after Stage-3 aggregation + cardinality gate succeed (StageRunner fail-fast guarantees this).
    Transitions each run: STAGE_2_COMPLETE -> STAGE_3_COMPLETE -> STAGE_3A_COMPLETE -> COMPLETE.
    Emits directive FSM transition: SYMBOL_RUNS_COMPLETE.
    """
    clean_id = context.directive_id
    p_conf = context.directive_config
    run_ids = context.run_ids
    symbols = context.symbols
    project_root = context.project_root
    registry_path = context.registry_path
    if registry_path is None:
        registry_path = RUNS_DIR / clean_id / "run_registry.json"

    strategy_id = p_conf.get("Strategy", p_conf.get("strategy")) or clean_id

    def get_file_hash(path: Path) -> str:
        return hashlib.sha256(path.read_bytes()).hexdigest()

    for rid, symbol in zip(run_ids, symbols):
        mgr = PipelineStateManager(rid)
        current = mgr.get_state_data()["current_state"]
        if current == "STAGE_2_COMPLETE":
            transition_run_state(rid, "STAGE_3_COMPLETE")

            snapshot_path = mgr.run_dir / "strategy.py"
            if not snapshot_path.exists():
                transition_run_state(rid, "FAILED")
                update_run_state(registry_path, clean_id, rid, "FAILED", last_error="Snapshot file missing.")
                log_run_to_registry(rid, "failed", clean_id)
                continue

            source_path = project_root / "strategies" / strategy_id / "strategy.py"
            if not source_path.exists():
                transition_run_state(rid, "FAILED")
                update_run_state(registry_path, clean_id, rid, "FAILED", last_error="Source strategy missing.")
                log_run_to_registry(rid, "failed", clean_id)
                raise RuntimeError(f"Source strategy missing: {source_path}")

            if get_file_hash(snapshot_path) != get_file_hash(source_path):
                transition_run_state(rid, "FAILED")
                update_run_state(
                    registry_path,
                    clean_id,
                    rid,
                    "FAILED",
                    last_error="Snapshot integrity mismatch against source strategy.",
                )
                log_run_to_registry(rid, "failed", clean_id)
                raise RuntimeError(
                    f"Snapshot Integrity Mismatch! {rid}/strategy.py != strategies/{strategy_id}/strategy.py"
                )

            print(f"[ORCHESTRATOR] Snapshot Verified: {rid} matches source.")
            mgr._append_audit_log(
                "SNAPSHOT_VERIFIED",
                {"strategy_hash": get_file_hash(snapshot_path), "source_hash": get_file_hash(source_path)},
            )

            bt_dir = RUNS_DIR / rid / "data"

            # --- MANDATORY ARTIFACT: EQUITY CURVE GENERATION ---
            trade_file = bt_dir / "results_tradelevel.csv"
            equity_file = bt_dir / "equity_curve.csv"
            if trade_file.exists() and not equity_file.exists():
                try:
                    # Guard: truncated or zero-byte files crash pd.read_csv with
                    # 'Truncated file header' — catch and skip gracefully.
                    if trade_file.stat().st_size == 0:
                        raise pd.errors.EmptyDataError("File is empty (0 bytes)")
                    df_t = pd.read_csv(trade_file)
                    if "pnl_usd" in df_t.columns:
                        pnl_series = df_t["pnl_usd"].fillna(0)
                        equity_series = 10000.0 + pnl_series.cumsum()
                        df_eq = pd.DataFrame({
                            "exit_timestamp": df_t.get("exit_timestamp", []),
                            "equity": equity_series
                        })
                        df_eq.to_csv(equity_file, index=False)
                        print(f"[ORCHESTRATOR] Generated local equity curve for {rid}")
                        
                        bt_dest_raw = BACKTESTS_DIR / f"{clean_id}_{symbol}" / "raw"
                        if bt_dest_raw.exists():
                            import shutil
                            shutil.copy2(equity_file, bt_dest_raw / "equity_curve.csv")
                except pd.errors.EmptyDataError as e:
                    print(f"[WARN] Skipping equity curve generation for {rid}: file is empty or truncated — {e}")
                except Exception as e:
                    print(f"[WARN] Failed to auto-generate equity curve for {rid}: {e}")

            required_artifacts = {
                "results_tradelevel.csv": bt_dir / "results_tradelevel.csv",
                "results_standard.csv": bt_dir / "results_standard.csv",
                "equity_curve.csv": bt_dir / "equity_curve.csv",
            }
            artifacts_manifest: dict[str, str] = {}
            for name, path in required_artifacts.items():
                if not path.exists():
                    transition_run_state(rid, "FAILED")
                    update_run_state(
                        registry_path,
                        clean_id,
                        rid,
                        "FAILED",
                        last_error=f"Missing required artifact for binding: {path}",
                    )
                    log_run_to_registry(rid, "failed", clean_id)
                    raise RuntimeError(f"Missing required artifact for binding: {path}")
                # Guard: zero-byte or truncated files would pass existence check
                # but crash the hasher / downstream CSV readers with ParserError.
                if path.stat().st_size == 0:
                    _err = f"Artifact is empty (0 bytes) — likely from a mid-run crash: {path}"
                    print(f"[WARN] {_err}")
                    transition_run_state(rid, "FAILED")
                    update_run_state(
                        registry_path, clean_id, rid, "FAILED",
                        last_error=_err,
                    )
                    log_run_to_registry(rid, "failed", clean_id)
                    raise RuntimeError(_err)
                artifacts_manifest[name] = get_file_hash(path)

            manifest = {
                "run_id": rid,
                "strategy_hash": get_file_hash(snapshot_path),
                "artifacts": artifacts_manifest,
                "timestamp": datetime.now(timezone.utc).isoformat(),
            }
            manifest_path = mgr.run_dir / "manifest.json"

            # Manifest Freeze Guard
            if manifest_path.exists() and current == "COMPLETE":
                with open(manifest_path, "r", encoding="utf-8") as f:
                    existing_manifest = json.load(f)
                if existing_manifest.get("run_id") == rid and existing_manifest.get("artifacts") == artifacts_manifest:
                    print(f"[ORCHESTRATOR] Manifest verified (frozen): {manifest_path}")
                else:
                    raise RuntimeError(f"[FATAL] Manifest Immutability Violation: Attempted to modify manifest of COMPLETED run {rid}.")
            else:
                with open(manifest_path, "w", encoding="utf-8") as f:
                    json.dump(manifest, f, indent=4)
                print(f"[ORCHESTRATOR] Manifest Bound: {manifest_path}")
            mgr._append_audit_log(
                "ARTIFACT_BOUND",
                {"manifest_path": str(manifest_path), "artifact_hashes": artifacts_manifest},
            )

            transition_run_state_sequence(rid, ["STAGE_3A_COMPLETE", "COMPLETE"])
            mgr._append_audit_log("RUN_COMPLETE", {"status": "SUCCESS"})

            # SUCCESS Master Ledger update
            log_run_to_registry(rid, "complete", clean_id)

    transition_directive_state(clean_id, "SYMBOL_RUNS_COMPLETE")


# ---------------------------------------------------------------------------
# Backward-Compatible Orchestrator Wrapper
# ---------------------------------------------------------------------------

def run_symbol_execution_stages(
    *,
    clean_id: str,
    p_conf: dict,
    run_ids: list[str],
    symbols: list[str],
    project_root: Path,
    python_exe: str,
    run_command,
    registry_path: Path | None = None,
) -> None:
    """
    Backward-compatible wrapper that calls all 4 stage functions in sequence.

    Used by any callers that have not yet migrated to PipelineContext.
    Internally constructs a minimal PipelineContext to satisfy the stage function signatures.
    """
    # Build a minimal context for the stage functions
    ctx = PipelineContext(
        directive_id=clean_id,
        directive_path=project_root / "backtest_directives" / "active" / f"{clean_id}.txt",
        project_root=project_root,
        python_exe=python_exe,
        provision_only=False,
    )
    ctx.directive_config = p_conf
    ctx.run_ids = run_ids
    ctx.symbols = symbols
    ctx.registry_path = registry_path

    run_stage1_execution(ctx)
    run_stage2_compilation(ctx)
    run_stage3_aggregation(ctx)
    run_manifest_binding(ctx)
