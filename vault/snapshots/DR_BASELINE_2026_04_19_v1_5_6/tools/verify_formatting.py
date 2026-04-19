import openpyxl
from pathlib import Path

# Path to check
fpath = Path("backtests/Range_Breakout01_USDJPY/AK_Trade_Report_Range_Breakout01_USDJPY.xlsx")

if not fpath.exists():
    print(f"File not found: {fpath}")
    exit(1)

wb = openpyxl.load_workbook(fpath)

# check "Trades List" formatting
ws = wb["Trades List"]
# PnL (USD) is typically column I (9th col)? Let's find header
pnl_col = None
for cell in ws[1]:
    if "pnl (usd)" in str(cell.value).lower():
        pnl_col = cell.column
        break

if pnl_col:
    # Check row 2 format
    cell = ws.cell(row=2, column=pnl_col)
    fmt = cell.number_format
    print(f"Trades List 'PnL (USD)' Format: {fmt}")
else:
    print("Trades List 'PnL (USD)' Column Not Found")

# check "Performance Summary" formatting (Transposed)
ws_sum = wb["Performance Summary"]
# find "Net Profit (USD)" in Column A
net_profit_row = None
for row in ws_sum.iter_rows(min_col=1, max_col=1):
    val = str(row[0].value).lower().strip()
    if "net profit (usd)" in val:
        net_profit_row = row[0].row
        break

if net_profit_row:
    # Check Col B (Value) format
    cell = ws_sum.cell(row=net_profit_row, column=2)
    fmt = cell.number_format
    print(f"Performance Summary 'Net Profit (USD)' Format: {fmt}")
else:
    print("Performance Summary 'Net Profit (USD)' Not Found")

print("-" * 20)

# check "Strategy_Master_Filter.xlsx" formatting
master_path = Path("backtests/Strategy_Master_Filter.xlsx")
if not master_path.exists():
    print(f"File not found: {master_path}")
    exit(1)

wb_master = openpyxl.load_workbook(master_path)
ws_master = wb_master.active

# Find "max_dd_pct" (or "Max Drawdown (%)"?)
# Stage-3 writes "max_dd_pct" as header? No, Stage-3 writes MASTER_FILTER_COLUMNS which are snake_case.
# Let's check headers in Row 1.
dd_pct_col = None
profit_col = None

for cell in ws_master[1]:
    val = str(cell.value).lower().strip()
    if val == "max_dd_pct":
        dd_pct_col = cell.column
    if val == "total_net_profit": 
        profit_col = cell.column

if dd_pct_col:
    # Check row 2
    cell = ws_master.cell(row=2, column=dd_pct_col)
    fmt = cell.number_format
    val = cell.value
    print(f"Master Filter 'max_dd_pct' Format: {fmt}, Value: {val}")
else:
    print("Master Filter 'max_dd_pct' Column Not Found")

if profit_col:
    cell = ws_master.cell(row=2, column=profit_col)
    fmt = cell.number_format
    print(f"Master Filter 'total_net_profit' Format: {fmt}")
else:
    print("Master Filter 'total_net_profit' Column Not Found")
