"""
promote_readiness.py -- Promotion readiness dashboard.

Scans all CORE + WATCH strategies from Filtered_Strategies_Passed.xlsx (FSP)
and PF_* composites from Master_Portfolio_Sheet.xlsx (MPS), then reports
promotion readiness for each.

Usage:
    python tools/promote_readiness.py              # full dashboard
    python tools/promote_readiness.py --core-only  # only CORE strategies
    python tools/promote_readiness.py --json       # machine-readable output
"""

import argparse
import json
import sys
from collections import OrderedDict
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PROJECT_ROOT))

from config.state_paths import (
    STATE_ROOT, BACKTESTS_DIR, STRATEGIES_DIR,
    CANDIDATE_FILTER_PATH,
)
from tools.promote_to_burnin import (
    _compute_quality_gate, _load_portfolio_yaml, _get_existing_ids,
    decompose_portfolio,
)
from tools.pipeline_utils import find_run_id_for_directive
from tools.baseline_freshness_gate import compute_baseline_age

import openpyxl

MPS_PATH = STRATEGIES_DIR / "Master_Portfolio_Sheet.xlsx"


def _check_freshness(strategy_id: str) -> tuple[str, int | None, bool]:
    """Return (display, numeric_age, hard_fail).
    Never aborts the dashboard; FAIL path returns marker string + hard_fail=True."""
    try:
        r = compute_baseline_age(strategy_id)
    except Exception:
        return "ERR", None, True
    if r.status == "FAIL":
        return "FAIL", None, True
    if r.worst_age_days is None:
        return "N/A", None, False
    age = r.worst_age_days
    if age > 14:
        marker = f"{age}d!!"   # promote-to-burnin threshold breached
    elif age > 7:
        marker = f"{age}d!"    # tighter watermark (not enforced)
    else:
        marker = f"{age}d"
    return marker, age, False


def _check_strategy_py(strategy_id: str) -> tuple[str, str]:
    """Check if strategy.py exists or is recoverable."""
    base_spy = PROJECT_ROOT / "strategies" / strategy_id / "strategy.py"
    if base_spy.exists():
        return "OK", ""
    # Check run snapshot
    run_id = find_run_id_for_directive(strategy_id)
    if run_id:
        snapshot = STATE_ROOT / "runs" / run_id / "strategy.py"
        if snapshot.exists():
            return "OK(run)", f"recoverable from {run_id[:12]}"
    return "MISSING", ""


def _check_run_id(strategy_id: str) -> tuple[str, str]:
    """Check if run_id is resolvable."""
    run_id = find_run_id_for_directive(strategy_id)
    if run_id:
        return "OK", run_id[:12]
    return "MISSING", ""


def _check_deployable(strategy_id: str) -> tuple[str, str]:
    """Check if deployable/ artifacts exist."""
    deploy = STRATEGIES_DIR / strategy_id / "deployable"
    if deploy.exists() and any(deploy.iterdir()):
        return "OK", ""
    return "MISSING", ""


def _check_portfolio_yaml(strategy_id: str, existing_ids: set) -> tuple[str, str]:
    """Check portfolio.yaml presence and lifecycle."""
    # Check exact ID or per-symbol variants
    if strategy_id in existing_ids:
        return "IN_YAML", ""
    for eid in existing_ids:
        if eid.startswith(strategy_id + "_"):
            return "IN_YAML", ""
    return "not present", ""


def _check_quality_gate(strategy_id: str) -> tuple[str, str]:
    """Run quality gate and return status."""
    qg = _compute_quality_gate(strategy_id)
    if not qg["metrics"]:
        return "N/A", "no trade data"
    if qg["passed"] and not qg["warns"]:
        return "PASS", ""
    if qg["passed"]:
        return "WARN", f"{len(qg['warns'])} warn"
    return "FAIL", f"{len(qg['hard_fails'])} hard"


def scan_fsp_strategies(core_only: bool = False) -> list[dict]:
    """Scan Filtered_Strategies_Passed.xlsx for CORE + WATCH strategies."""
    if not CANDIDATE_FILTER_PATH.exists():
        print(f"[WARN] FSP not found: {CANDIDATE_FILTER_PATH}")
        return []

    wb = openpyxl.load_workbook(str(CANDIDATE_FILTER_PATH), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=False))
    headers = [c.value for c in rows[0]]

    stat_idx = headers.index("candidate_status")
    base_idx = headers.index("base_strategy_id")
    sym_idx = headers.index("symbol")

    # Group by base_strategy_id
    bases = OrderedDict()
    target_statuses = {"CORE"} if core_only else {"CORE", "WATCH"}
    for row in rows[1:]:
        status = row[stat_idx].value
        if status not in target_statuses:
            continue
        base = row[base_idx].value
        sym = row[sym_idx].value
        if base not in bases:
            bases[base] = {"status": status, "symbols": []}
        bases[base]["symbols"].append(sym)

    wb.close()
    return [{"strategy_id": k, **v, "source": "FSP"} for k, v in bases.items()]


def scan_mps_composites(core_only: bool = False) -> list[dict]:
    """Scan Master_Portfolio_Sheet.xlsx for CORE + WATCH PF_* composites."""
    if not MPS_PATH.exists():
        print(f"[WARN] MPS not found: {MPS_PATH}")
        return []

    target_statuses = {"CORE"} if core_only else {"CORE", "WATCH"}
    composites = []

    wb = openpyxl.load_workbook(str(MPS_PATH), read_only=True, data_only=True)
    for sheet_name in ["Portfolios", "Single-Asset Composites"]:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=False))
        if not rows:
            continue
        headers = [c.value for c in rows[0]]
        pid_idx = headers.index("portfolio_id")
        stat_idx = headers.index("portfolio_status")
        for row in rows[1:]:
            pid = row[pid_idx].value
            stat = row[stat_idx].value
            if pid and str(pid).startswith("PF_") and stat in target_statuses:
                composites.append({
                    "strategy_id": pid,
                    "status": stat,
                    "symbols": [],
                    "source": f"MPS:{sheet_name}",
                })
    wb.close()
    return composites


def build_readiness_report(core_only: bool = False) -> list[dict]:
    """Build the full readiness report."""
    # Load portfolio.yaml once
    data = _load_portfolio_yaml()
    existing_ids = _get_existing_ids(data)

    # Collect all candidates
    strategies = scan_fsp_strategies(core_only)
    composites = scan_mps_composites(core_only)

    report = []

    # 1. Individual strategies from FSP
    for entry in strategies:
        sid = entry["strategy_id"]
        spy_status, spy_detail = _check_strategy_py(sid)
        rid_status, rid_detail = _check_run_id(sid)
        dep_status, dep_detail = _check_deployable(sid)
        port_status, port_detail = _check_portfolio_yaml(sid, existing_ids)
        qg_status, qg_detail = _check_quality_gate(sid)
        age_display, age_numeric, age_fail = _check_freshness(sid)

        ready = all([
            spy_status in ("OK", "OK(run)"),
            rid_status == "OK",
            qg_status in ("PASS", "WARN"),
            port_status == "not present",
            not age_fail,
        ])

        report.append({
            "strategy_id": sid,
            "classification": entry["status"],
            "source": entry["source"],
            "symbols": entry["symbols"],
            "checks": {
                "strategy_py": spy_status,
                "run_id": rid_status,
                "deployable": dep_status,
                "quality_gate": qg_status,
                "baseline_age": age_display,
                "baseline_age_numeric": age_numeric,
                "portfolio_yaml": port_status,
            },
            "ready": ready,
            "is_composite": False,
        })

    # 2. PF_* composites from MPS
    for entry in composites:
        pid = entry["strategy_id"]
        port_status, _ = _check_portfolio_yaml(pid, existing_ids)

        # Decompose
        constituents = []
        decompose_error = None
        try:
            raw = decompose_portfolio(pid)
            # Group by base strategy
            seen = {}
            for c in raw:
                csid = c["strategy_id"]
                if csid not in seen:
                    spy_s, _ = _check_strategy_py(csid)
                    rid_s, _ = _check_run_id(csid)
                    dep_s, _ = _check_deployable(csid)
                    qg_s, _ = _check_quality_gate(csid)
                    cp_status, _ = _check_portfolio_yaml(csid, existing_ids)
                    age_s, _, _ = _check_freshness(csid)
                    seen[csid] = {
                        "strategy_id": csid,
                        "symbols": [],
                        "checks": {
                            "strategy_py": spy_s,
                            "run_id": rid_s,
                            "deployable": dep_s,
                            "quality_gate": qg_s,
                            "baseline_age": age_s,
                            "portfolio_yaml": cp_status,
                        },
                    }
                seen[csid]["symbols"].append(c["symbol"])
            constituents = list(seen.values())
        except RuntimeError as e:
            decompose_error = str(e)

        report.append({
            "strategy_id": pid,
            "classification": entry["status"],
            "source": entry["source"],
            "symbols": [],
            "checks": {
                "strategy_py": "COMPOSITE",
                "run_id": "N/A",
                "deployable": "N/A",
                "quality_gate": "N/A",
                "baseline_age": "COMPOSITE",
                "portfolio_yaml": port_status,
            },
            "ready": False,  # composites promoted via --composite
            "is_composite": True,
            "constituents": constituents,
            "decompose_error": decompose_error,
        })

    return report


def print_report(report: list[dict]) -> None:
    """Print the readiness dashboard."""
    # Sort: CORE first, then by readiness, then alphabetical
    report.sort(key=lambda r: (
        0 if r["classification"] == "CORE" else 1,
        0 if r["ready"] else 1,
        -(r["checks"].get("baseline_age_numeric") or 0),
        r["strategy_id"],
    ))

    print()
    print(f"{'='*100}")
    print(f"  PROMOTION READINESS DASHBOARD")
    print(f"{'='*100}")
    print()

    # Header
    hdr = (
        f"  {'Strategy':<52s}  {'Class':5s}  {'strategy.py':11s}  "
        f"{'run_id':8s}  {'deploy':8s}  {'QG':6s}  {'age':7s}  {'portfolio':12s}"
    )
    print(hdr)
    print(f"  {'-'*98}")

    ready_count = 0
    in_portfolio_count = 0
    blocked_count = 0

    for entry in report:
        sid = entry["strategy_id"]
        cls = entry["classification"]
        c = entry["checks"]

        # Truncate long IDs
        display_id = sid if len(sid) <= 52 else sid[:49] + "..."

        # Color hints via markers
        port_display = c["portfolio_yaml"]
        if port_display == "IN_YAML":
            in_portfolio_count += 1
        elif entry["ready"]:
            ready_count += 1
        else:
            blocked_count += 1

        ready_marker = ">>>" if entry["ready"] else "   "

        age_display = c.get("baseline_age", "")
        print(
            f"{ready_marker}{display_id:<52s}  {cls:5s}  {c['strategy_py']:11s}  "
            f"{c['run_id']:8s}  {c['deployable']:8s}  {c['quality_gate']:6s}  "
            f"{age_display:7s}  {port_display:12s}"
        )

        # Composite constituents
        if entry.get("is_composite"):
            if entry.get("decompose_error"):
                print(f"      [DECOMPOSE ERROR] {entry['decompose_error']}")
            for con in entry.get("constituents", []):
                cc = con["checks"]
                csid = con["strategy_id"]
                syms = ",".join(con["symbols"])
                print(
                    f"      -> {csid:<48s}  {syms:10s}  {cc['strategy_py']:11s}  "
                    f"{cc['run_id']:8s}  {cc['deployable']:8s}  {cc['quality_gate']:6s}  "
                    f"{cc['portfolio_yaml']:12s}"
                )

    print(f"  {'-'*98}")
    total = len([r for r in report if not r.get("is_composite")])
    composites = len([r for r in report if r.get("is_composite")])
    print(f"\n  Summary: {total} strategies + {composites} composites")
    print(f"    READY to promote:    {ready_count}")
    print(f"    Already in portfolio: {in_portfolio_count}")
    print(f"    Blocked:             {blocked_count}")
    print(f"\n  >>> = ready for promotion")
    print(f"  Use: python tools/promote_to_burnin.py <ID> --profile <PROFILE> --dry-run")
    print(f"  Use: python tools/promote_to_burnin.py <PF_ID> --composite --profile <PROFILE> --dry-run")
    print()


def main():
    parser = argparse.ArgumentParser(
        description="Promotion readiness dashboard for CORE + WATCH strategies.",
    )
    parser.add_argument("--core-only", action="store_true",
                        help="Show only CORE-classified strategies")
    parser.add_argument("--json", action="store_true",
                        help="Output machine-readable JSON instead of table")
    args = parser.parse_args()

    report = build_readiness_report(core_only=args.core_only)

    if args.json:
        print(json.dumps(report, indent=2, default=str))
    else:
        print_report(report)


if __name__ == "__main__":
    main()
