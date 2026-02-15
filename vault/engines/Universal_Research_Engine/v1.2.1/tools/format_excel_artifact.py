"""
format_excel_artifact.py — Unified Excel Formatter & Rounding Governance

Applies strict styling and number formatting to Excel artifacts.
Logic is presentation-layer only. No data mutation.

Usage:
    python tools/format_excel_artifact.py --file <path> --profile <strategy|portfolio>
"""

import argparse
import sys
import openpyxl
from pathlib import Path
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ==========================================================
# CONSTANTS
# ==========================================================

# Styling
HEADER_FILL_COLOR = "4472C4"  # Lighter Blue (SOP Governance)
HEADER_FONT_COLOR = "FFFFFF"
ALT_ROW_FILL_COLOR = "DCE6F1"

# Number Formats
FMT_CURRENCY = "#,##0.00"   # USD (2 decimals, comma)
FMT_PERCENT = "0.00%"       # Percent (2 decimals: 0.1234 -> 12.34%)
FMT_FLOAT = "0.00"          # Ratio/Float (2 decimals: 1.23)
FMT_INT = "0"               # Integer

# Column Rules (Column Name -> Format)
# Keys should be lower-case for matching
FORMAT_MAP = {
    # Currency
    "net_pnl_usd": FMT_CURRENCY,
    "peak_capital_deployed": FMT_CURRENCY,
    "reference_capital_usd": FMT_CURRENCY,
    "max_dd_usd": FMT_CURRENCY,
    "gross_profit_usd": FMT_CURRENCY,
    "gross_loss_usd": FMT_CURRENCY,
    
    # Percent
    "max_dd_pct": FMT_PERCENT,
    "capital_overextension_ratio": FMT_PERCENT, # Ratio treated as % ? User said "Percent Fields... applies to capital_overextension_ratio". 
    # Usually ratio > 1 is bad. If it's 1.5, format 150.00%? Or 1.50? 
    # User listed it under "Percent Fields". I will trust instructions.
    "avg_pairwise_corr": FMT_PERCENT, 
    "max_pairwise_corr_stress": FMT_PERCENT,
    "win_rate": FMT_PERCENT,
    "cagr": FMT_PERCENT,
    
    # Float / Ratio
    "sharpe": FMT_FLOAT,
    "return_dd_ratio": FMT_FLOAT,
    "avg_concurrent": FMT_FLOAT,
    "p95_concurrent": FMT_FLOAT,
    "profit_factor": FMT_FLOAT,
    "avg_r_multiple": FMT_FLOAT,
    "sqn": FMT_FLOAT,
    
    # Integer
    "max_concurrent": FMT_INT,
    "dd_max_concurrent": FMT_INT,
    "total_trades": FMT_INT,
    "total_wins": FMT_INT,
    "total_losses": FMT_INT,
    "consecutive_wins": FMT_INT,
    "consecutive_losses": FMT_INT
}

# Hidden Columns (SOP Auditing Fields)
HIDDEN_COLS = {
    "constituent_run_ids",
    "run_ids",
    "genome_id" 
}

# ==========================================================
# LOGIC
# ==========================================================

def apply_formatting(file_path, profile):
    path = Path(file_path)
    if not path.exists():
        print(f"[ERROR] File not found: {file_path}")
        sys.exit(1)

    print(f"[INFO] Formatting {path.name} (Profile: {profile})...")
    
    try:
        wb = openpyxl.load_workbook(path)
        
        # Styles Objects
        header_fill = PatternFill(start_color=HEADER_FILL_COLOR, end_color=HEADER_FILL_COLOR, fill_type="solid")
        header_font = Font(bold=True, color=HEADER_FONT_COLOR)
        alt_row_fill = PatternFill(start_color=ALT_ROW_FILL_COLOR, end_color=ALT_ROW_FILL_COLOR, fill_type="solid")
        
        # Iterate over ALL sheets
        for ws in wb.worksheets:
            print(f"  [INFO] Processing sheet: {ws.title}")
            
            # 1. Header & Column ID Mapping
            col_map = {} # col_index -> col_name_lower
            max_col = ws.max_column
            max_row = ws.max_row
            
            for col_idx in range(1, max_col + 1):
                cell = ws.cell(row=1, column=col_idx)
                val = str(cell.value).lower().strip() if cell.value else ""
                col_map[col_idx] = val
                
                # Header Styling
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="left", vertical="center")

            # 2. Row Iteration (Data)
            for row_idx in range(2, max_row + 1):
                is_alt = (row_idx % 2 == 0)
                
                for col_idx in range(1, max_col + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    col_name = col_map.get(col_idx, "")
                    
                    # Alternating Row Shading
                    if is_alt:
                        cell.fill = alt_row_fill
                    
                    # Formatting & Alignment
                    if col_name in FORMAT_MAP:
                        fmt = FORMAT_MAP[col_name]
                        cell.number_format = fmt
                        
                        # Right align numerics
                        cell.alignment = Alignment(horizontal="right")
                    else:
                        # Default Left Align for text/other
                        cell.alignment = Alignment(horizontal="left")

            # 3. Column Widths & Hiding
            for col_idx in range(1, max_col + 1):
                col_name = col_map.get(col_idx, "")
                col_letter = get_column_letter(col_idx)
                
                # Hide specific columns
                if col_name in HIDDEN_COLS:
                    ws.column_dimensions[col_letter].hidden = True
                    continue
                    
                # Auto-fit (Simple approximation)
                max_len = 0
                # Sample first 20 rows + header for speed
                sample_rows = list(range(1, min(max_row, 50) + 1))
                for r in sample_rows:
                    val = ws.cell(row=r, column=col_idx).value
                    if val:
                        max_len = max(max_len, len(str(val)))
                
                width = max_len + 2
                if width < 10: width = 10
                if width > 50: width = 50
                ws.column_dimensions[col_letter].width = width

            # 4. Freeze Header
            ws.freeze_panes = "A2"
        
        wb.save(path)
        print("[SUCCESS] Formatting complete.")
        
    except Exception as e:
        print(f"[FATAL] Formatting failed: {e}")
        sys.exit(1)

def main():
    parser = argparse.ArgumentParser(description="Unified Excel Formatter")
    parser.add_argument("--file", required=True, help="Path to Excel file")
    parser.add_argument("--profile", required=True, choices=["strategy", "portfolio"], help="Formatting profile")
    
    args = parser.parse_args()
    
    apply_formatting(args.file, args.profile)

if __name__ == "__main__":
    main()
