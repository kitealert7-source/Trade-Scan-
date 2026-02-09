"""
Stage-3 Aggregation Engine — Strategy Master Filter Population
Governed by: SOP_OUTPUT §6

This engine is READ-ONLY, COPY-ONLY, and APPEND-ONLY.
No computation, derivation, inference, or ranking is performed.
Values are treated as opaque scalars — no conversion or cleaning.
"""

import json
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# Constants
PROJECT_ROOT = Path(__file__).parent.parent
BACKTESTS_ROOT = PROJECT_ROOT / "backtests"

# Column schema per SOP_OUTPUT §6 (exact order)
MASTER_FILTER_COLUMNS = [
    "run_id",
    "strategy",
    "symbol",
    "timeframe",
    "test_start",
    "test_end",
    "trading_period",
    "total_trades",
    "total_net_profit",
    "gross_profit",
    "gross_loss",
    "profit_factor",
    "expectancy",
    "sharpe_ratio",
    "max_drawdown",
    "max_dd_pct",
    "return_dd_ratio",
    "worst_5_loss_pct",
    "longest_loss_streak",
    "pct_time_in_market",
    "avg_bars_in_trade",
    "net_profit_high_vol",
    "net_profit_low_vol",
]

# Required Performance Summary metric labels (EXACT, no fuzzy matching)
# Maps: Master Filter column -> Exact label in AK Trade Report Performance Summary
REQUIRED_METRICS = {
    "trading_period": "Trading Period (Days)",
    "total_trades": "Total Trades",
    "total_net_profit": "Net Profit (USD)",
    "gross_profit": "Gross Profit (USD)",
    "gross_loss": "Gross Loss (USD)",
    "profit_factor": "Profit Factor",
    "expectancy": "Expectancy (USD)",
    "sharpe_ratio": "Sharpe Ratio",
    "max_drawdown": "Max Drawdown (USD)",
    "max_dd_pct": "Max Drawdown (%)",
    "return_dd_ratio": "Return / Drawdown Ratio",
    "worst_5_loss_pct": "Worst 5 Trades Loss %",
    "longest_loss_streak": "Max Consecutive Losses",
    "pct_time_in_market": "% Time in Market",
    "avg_bars_in_trade": "Avg Bars per Trade",
}

# Volatility metric labels (EXACT, hardcoded per Stage-2 output)
VOLATILITY_METRICS = {
    "net_profit_high_vol": "Net Profit - High Volatility",
    "net_profit_low_vol": "Net Profit - Low Volatility",
}

# Required metadata fields
REQUIRED_METADATA_FIELDS = [
    "run_id",
    "strategy_name",
    "symbol",
    "timeframe",
]

# Styles per SOP_OUTPUT §6
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
ALT_ROW_FILL = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")


def load_run_metadata(run_folder):
    """Load run_metadata.json from a run folder."""
    metadata_path = run_folder / "metadata" / "run_metadata.json"
    if not metadata_path.exists():
        return None
    
    with open(metadata_path, "r", encoding="utf-8") as f:
        return json.load(f)


def validate_metadata(metadata, run_folder):
    """
    Validate that all required metadata fields are present.
    Returns (is_valid, error_message).
    """
    if not metadata:
        return False, "run_metadata.json not found or empty"
    
    # Check run_id is mandatory and non-empty
    run_id = metadata.get("run_id")
    if not run_id or run_id == "":
        return False, "run_id is missing or empty"
    
    # Check required fields
    for field in REQUIRED_METADATA_FIELDS:
        if field not in metadata or metadata[field] is None or metadata[field] == "":
            return False, f"Missing required metadata field: {field}"
    
    # Check date_range
    date_range = metadata.get("date_range", {})
    if not date_range.get("start"):
        return False, "Missing date_range.start"
    if not date_range.get("end"):
        return False, "Missing date_range.end"
    
    return True, None


def find_ak_trade_report(run_folder):
    """Find the AK_Trade_Report Excel file in a run folder."""
    pattern = "AK_Trade_Report_*.xlsx"
    matches = list(run_folder.glob(pattern))
    
    # Filter out temp files
    matches = [m for m in matches if not m.name.startswith("~$")]
    
    if not matches:
        return None
    return matches[0]


def extract_performance_metrics(report_path):
    """
    Extract metrics from AK_Trade_Report Performance Summary sheet.
    Returns dict of metric_name -> value (opaque scalar, no conversion).
    """
    wb = openpyxl.load_workbook(report_path, data_only=True)
    
    if "Performance Summary" not in wb.sheetnames:
        wb.close()
        return {}
    
    ws = wb["Performance Summary"]
    
    # Build a lookup of metric name -> value (All column)
    metrics = {}
    
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        if len(row) >= 2:
            metric_name = row[0].value
            all_value = row[1].value if len(row) > 1 else None
            
            if metric_name and isinstance(metric_name, str):
                # Store exactly as-is (opaque scalar)
                metrics[metric_name.strip()] = all_value
    
    wb.close()
    return metrics


def validate_required_metrics(metrics, run_id):
    """
    Validate that all required metrics are present.
    Returns (is_valid, missing_metrics_list).
    """
    missing = []
    
    # Check standard required metrics
    for col_name, label in REQUIRED_METRICS.items():
        if label not in metrics or metrics[label] is None:
            missing.append(label)
    
    # Check volatility metrics (exact labels)
    for col_name, label in VOLATILITY_METRICS.items():
        if label not in metrics or metrics[label] is None:
            missing.append(label)
    
    if missing:
        return False, missing
    return True, []


def extract_from_report(report_path, metadata):
    """
    Extract all required fields from AK_Trade_Report.
    Maps to Master Filter columns per SOP_OUTPUT §6.
    Returns (row_data, error_message).
    """
    metrics = extract_performance_metrics(report_path)
    run_id = metadata.get("run_id", "UNKNOWN")
    
    # Validate all required metrics are present
    is_valid, missing = validate_required_metrics(metrics, run_id)
    if not is_valid:
        return None, f"Missing required metrics: {', '.join(missing)}"
    
    # Build row data — values are opaque scalars, no conversion
    row_data = {
        "run_id": metadata.get("run_id"),
        "strategy": metadata.get("strategy_name"),
        "symbol": metadata.get("symbol"),
        "timeframe": metadata.get("timeframe"),
        "test_start": metadata.get("date_range", {}).get("start"),
        "test_end": metadata.get("date_range", {}).get("end"),
    }
    
    # Map Performance Summary metrics (exact labels)
    for col_name, label in REQUIRED_METRICS.items():
        row_data[col_name] = metrics.get(label)
    
    # Map volatility metrics (exact labels)
    for col_name, label in VOLATILITY_METRICS.items():
        row_data[col_name] = metrics.get(label)
    
    return row_data, None


def get_existing_run_ids(master_filter_path):
    """Get set of run_ids already in the Master Filter."""
    if not master_filter_path.exists():
        return set()
    
    wb = openpyxl.load_workbook(master_filter_path)
    ws = wb.active
    
    run_ids = set()
    for row in ws.iter_rows(min_row=2, max_col=1):  # Skip header, column A = run_id
        if row[0].value:
            run_ids.add(str(row[0].value))
    
    wb.close()
    return run_ids


def create_master_filter(master_filter_path):
    """Create new Master Filter with headers."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Master Filter"
    
    # Write headers
    for col_idx, col_name in enumerate(MASTER_FILTER_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="left")
    
    # Freeze header row
    ws.freeze_panes = "A2"
    
    wb.save(master_filter_path)
    wb.close()


def append_row_to_master_filter(master_filter_path, row_data):
    """Append a single row to Master Filter."""
    wb = openpyxl.load_workbook(master_filter_path)
    ws = wb.active
    
    # Find next empty row
    next_row = ws.max_row + 1
    
    # Write data (opaque scalars, no conversion)
    for col_idx, col_name in enumerate(MASTER_FILTER_COLUMNS, 1):
        value = row_data.get(col_name)
        cell = ws.cell(row=next_row, column=col_idx, value=value)
        cell.alignment = Alignment(horizontal="left")
        
        # Alternate row shading
        if next_row % 2 == 0:
            cell.fill = ALT_ROW_FILL
    
    # Auto-fit columns
    for col_idx in range(1, len(MASTER_FILTER_COLUMNS) + 1):
        col_letter = get_column_letter(col_idx)
        max_width = 0
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value is not None:
                    max_width = max(max_width, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_width + 2, 12)
    
    while True:
        try:
            wb.save(master_filter_path)
            break
        except PermissionError:
            print(f"[WARN] Output file is open: {master_filter_path}")
            input("Output file is open. Close it and press Enter to retry...")
    wb.close()


def enforce_strategy_persistence(run_id: str):
    """
    Enforce Stage-5 Strategy Persistence invariant.
    """
    strategies_root = PROJECT_ROOT / "strategies"
    strategy_dir = strategies_root / run_id

    if not strategy_dir.exists():
        raise RuntimeError(f"Strategy folder missing: strategies/{run_id}")

    allowed = {"strategy.py", "__pycache__"}
    found = {p.name for p in strategy_dir.iterdir()}

    if "strategy.py" not in found:
        raise RuntimeError(f"strategy.py missing in strategies/{run_id}")

    extra = found - allowed
    if extra:
        raise RuntimeError(
            f"Non-compliant files in strategies/{run_id}: {sorted(extra)}"
        )


def discover_completed_runs():
    """Discover all completed Stage-2 runs in backtests/."""
    runs = []
    rejected = []
    
    for run_folder in BACKTESTS_ROOT.iterdir():
        if not run_folder.is_dir():
            continue
        
        # Skip staging folders and master filter
        if run_folder.name.startswith("."):
            continue
        
        # Load and validate metadata
        metadata = load_run_metadata(run_folder)
        is_valid, error = validate_metadata(metadata, run_folder)
        
        if not is_valid:
            rejected.append({"folder": run_folder.name, "reason": error})
            continue
        
        # Check for AK Trade Report
        report_path = find_ak_trade_report(run_folder)
        if not report_path:
            rejected.append({"folder": run_folder.name, "reason": "AK_Trade_Report not found"})
            continue
        
        runs.append({
            "folder": run_folder,
            "metadata": metadata,
            "report_path": report_path,
        })
    
    return runs, rejected


def compile_stage3(strategy_filter=None):
    """
    Main Stage-3 compilation entry point.
    
    Args:
        strategy_filter: Optional strategy name to filter runs
    
    Returns:
        dict with results
    """
    print("Stage-3 Aggregation Engine")
    print("=" * 60)
    
    # Discover completed runs
    runs, discovery_rejected = discover_completed_runs()
    
    if strategy_filter:
        runs = [r for r in runs if r["metadata"].get("strategy_name", "").startswith(strategy_filter)]
    
    print(f"Discovered {len(runs)} valid runs")
    if discovery_rejected:
        print(f"Rejected at discovery: {len(discovery_rejected)}")
        for r in discovery_rejected:
            print(f"  - {r['folder']}: {r['reason']}")
    
    master_filter_path = BACKTESTS_ROOT / "Strategy_Master_Filter.xlsx"
    
    # Create if doesn't exist
    if not master_filter_path.exists():
        print(f"Creating new Master Filter: {master_filter_path}")
        create_master_filter(master_filter_path)
    
    # Get existing run IDs for idempotency
    existing_ids = get_existing_run_ids(master_filter_path)
    print(f"Existing runs in Master Filter: {len(existing_ids)}")
    
    # Process each run
    added = []
    skipped = []
    
    for run in runs:
        run_id = run["metadata"].get("run_id")
        strategy = run["metadata"].get("strategy_name")
        
        # Idempotency check (run_id based only)
        if run_id in existing_ids:
            skipped.append({"strategy": strategy, "run_id": run_id, "reason": "Already exists"})
            continue
        
        # Extract data from report with validation
        row_data, error = extract_from_report(run["report_path"], run["metadata"])
        
        if error:
            skipped.append({"strategy": strategy, "run_id": run_id, "reason": error})
            print(f"  REJECTED: {strategy} [{run_id[:8]}] - {error}")
            continue
        
        # Append to Master Filter
        append_row_to_master_filter(master_filter_path, row_data)
        
        # Enforce Strategy Persistence (Stage-4A Invariant)
        enforce_strategy_persistence(run_id)
        
        added.append({"strategy": strategy, "run_id": run_id})
        print(f"  Added: {strategy} [{run_id[:8]}]")
    
    print("\n" + "=" * 60)
    print("STAGE-3 SUMMARY")
    print("=" * 60)
    print(f"Rows added: {len(added)}")
    print(f"Rows skipped: {len(skipped)}")
    print(f"Output: {master_filter_path}")
    
    return {
        "added": added,
        "skipped": skipped,
        "discovery_rejected": discovery_rejected,
        "output_path": str(master_filter_path),
    }


def main():
    """CLI entry point."""
    import sys
    
    strategy_filter = None
    if len(sys.argv) > 1:
        strategy_filter = sys.argv[1]
    
    result = compile_stage3(strategy_filter)
    
    print("\nAdded runs:")
    for r in result["added"]:
        print(f"  - {r['strategy']} ({r['run_id'][:8]}...)")
    
    if result["skipped"]:
        print("\nSkipped runs:")
        for r in result["skipped"]:
            print(f"  - {r['strategy']}: {r['reason']}")


if __name__ == "__main__":
    main()
