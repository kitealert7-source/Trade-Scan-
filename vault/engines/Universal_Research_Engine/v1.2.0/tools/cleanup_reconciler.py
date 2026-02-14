"""
cleanup_reconciler.py — SOP_CLEANUP Enforcement (Dry-Run)

Authority: SOP_CLEANUP (FINAL)
Mode: DRY-RUN ONLY
"""

import sys
import shutil
import argparse
from pathlib import Path
import openpyxl

# Constants
PROJECT_ROOT = Path(__file__).parent.parent
BACKTESTS_ROOT = PROJECT_ROOT / "backtests"
RUNS_ROOT = PROJECT_ROOT / "runs"
MASTER_SHEET_PATH = BACKTESTS_ROOT / "Strategy_Master_Filter.xlsx"

def load_master_index():
    """
    Read Strategy Master Sheet once.
    Returns:
        rows (list): list of (run_id, strategy_name)
        valid_run_ids (set)
        valid_strategies (set)
    """
    if not MASTER_SHEET_PATH.exists():
        print(f"[ERROR] Master Sheet not found: {MASTER_SHEET_PATH}")
        sys.exit(1)

    try:
        wb = openpyxl.load_workbook(MASTER_SHEET_PATH, data_only=True)
        ws = wb.active
        
        rows = []
        valid_run_ids = set()
        valid_strategies = set()
        
        # Iterating from row 2 to skip header
        # Column A = run_id, Column B = strategy_name
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
                
            r_id = row[0]
            s_name = row[1]
            
            if r_id and s_name:
                r_id = str(r_id).strip()
                s_name = str(s_name).strip()
                rows.append((r_id, s_name))
                # Canonicalize set contents for case-insensitive lookup
                valid_run_ids.add(r_id.casefold())
                valid_strategies.add(s_name.casefold())
                 
        wb.close()
        return rows, valid_run_ids, valid_strategies
        
    except Exception as e:
        print(f"[ERROR] Failed to read Master Sheet: {e}")
        sys.exit(1)

def scan_runs(valid_run_ids):
    """
    Scan runs/ directory.
    Each runs/<run_id>/ folder may contain authoritative snapshots.
    Entire folder MUST be deleted if run_id is invalid.
    Match ONLY against valid_run_ids.
    Returns list of action strings.
    """
    actions = []
    if not RUNS_ROOT.exists():
        return actions

    for item in RUNS_ROOT.iterdir():
        if item.is_dir():
            run_id = item.name
            # Rule: run_id is matched only against runs/<run_id>/
            # Case-insensitive Match
            if run_id.casefold() not in valid_run_ids:
                actions.append(f"DELETE_RUN_SNAPSHOT runs/{run_id}/")
    return actions

def scan_backtests(valid_strategies):
    """
    Scan backtests/ directory.
    Match ONLY against valid_strategies.
    Returns list of action strings.
    """
    actions = []
    if not BACKTESTS_ROOT.exists():
        return actions

    # Whitelist of preserved files/folders per SOP_CLEANUP §5
    # - Strategy_Master_Filter.xlsx
    # - batch_summary_*.csv
    # - (implied) .gitignore or similar dotfiles if any (SOP says 'backtest output' is under <strategy>)
    
    for item in BACKTESTS_ROOT.iterdir():
        name = item.name
        
        # Skipping files explicitly preserved
        if name == "Strategy_Master_Filter.xlsx":
            continue
        if name.startswith("batch_summary_") and name.endswith(".csv"):
            continue
        if name.startswith("."): # skip hidden
            continue
            
        # Strategy folders
        if item.is_dir():
            # Rule: strategy_name is matched only against backtests/<strategy_name>/
            # Case-insensitive Match
            if name.casefold() not in valid_strategies:
                actions.append(f"DELETE_BACKTEST_STATE backtests/{name}/")
        
        # Any other files in backtests/ root are technically not governed/allowed by strict interpretation,
        # but SOP strictly targets "backtests/<strategy>/".
        # We will focus on folders as per "Derive valid {strategy} ... Scan filesystem ... backtests/*"
        # The prompt implies cleaning up unindexed strategies.
    return actions
        
def reconcile_orphaned_rows(rows):
    """
    Identify rows in master sheet that point to non-existent artifacts.
    Also identify any lingering artifacts associated with those rows.
    Returns:
        actions_strategies (list): extra strategy deletes
        actions_rows (list): row removals
    """
    actions_strategies = []
    actions_rows = []
    
    for r_id, s_name in rows:
        run_path = RUNS_ROOT / r_id
        backtest_path = BACKTESTS_ROOT / s_name
        
        exists_run = run_path.exists()
        exists_bt = backtest_path.exists()
        
        # Rule: if ANY authoritative artifact is missing, remove row
        # AND cleanup any partial leftovers
        if not exists_run or not exists_bt:
            # If run exists, queue for deletion (Run-scoped)
            if exists_run:
                actions_strategies.append(f"DELETE_RUN_SNAPSHOT runs/{r_id}/")
            
            # DO NOT DELETE BACKTEST FOLDERS HERE
            # Backtest folders are strategy-scoped and may be shared by other valid rows.
            # Only scan_backtests() has the authority to delete unreferenced backtests.
                
            # Log removal
            # Format: REMOVE_MASTER_ROW run_id=<run_id> strategy=<strategy>
            actions_rows.append(f"REMOVE_MASTER_ROW run_id={r_id} strategy={s_name}")
            
    return actions_strategies, actions_rows

def execute_removals(actions_rows):
    """
    Remove rows from Master Sheet for REMOVE_MASTER_ROW actions.
    Must re-open workbook in read/write mode.
    """
    if not actions_rows:
        return
        
    try:
        wb = openpyxl.load_workbook(MASTER_SHEET_PATH)
        ws = wb.active
        
        # Parse targets from actions
        # Action format: REMOVE_MASTER_ROW run_id=<run_id> strategy=<strategy>
        targets = []
        for action in actions_rows:
            parts = action.split()
            r_id = parts[1].split('=')[1]
            s_name = parts[2].split('=')[1]
            targets.append((r_id, s_name))
            
        # Iterate backwards to safely delete
        # We need to find rows matching targets
        rows_to_delete = []
        
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row:
                continue
            r_id = str(row[0]).strip() if row[0] else ""
            s_name = str(row[1]).strip() if row[1] else ""
            
            if (r_id, s_name) in targets:
                rows_to_delete.append(i)
                
        # Delete from bottom up
        for idx in sorted(rows_to_delete, reverse=True):
            ws.delete_rows(idx)
            
        wb.save(MASTER_SHEET_PATH)
        wb.close()
        
    except Exception as e:
        print(f"[ERROR] Failed to execute row removals: {e}")

def main():
    parser = argparse.ArgumentParser(description="Cleanup Reconciler")
    parser.add_argument("--execute", action="store_true", help="Execute planned deletions")
    args = parser.parse_args()

    # 1. Read Master Sheet
    rows, valid_run_ids, valid_strategies = load_master_index()
    
    # Map for easy lookup of strategy -> run_id
    strategy_to_run_id = {s: r for r, s in rows}
    
    # 2. Key Actions Lists
    all_strategies = []
    all_backtests = []
    all_rows = []
    
    # Scan Filesystem (Zombies)
    zombie_runs = scan_runs(valid_run_ids)
    zombie_bt = scan_backtests(valid_strategies)
    
    all_strategies.extend(zombie_runs)
    all_backtests.extend(zombie_bt)
    
    # Check Rows (Orphans + Leftovers)
    leftover_runs, orphan_rows = reconcile_orphaned_rows(rows)
    
    all_strategies.extend(leftover_runs)
    # No backtest leftovers from reconcile_orphaned_rows
    all_rows.extend(orphan_rows)
    
    # Track failures to prevent row removal
    failed_run_ids = set()
    
    # 3. Process Actions
    
    # DELETE_RUN_SNAPSHOT
    # Use simple list dedup in case of overlap (though uncommon)
    for action in sorted(list(set(all_strategies))):
        print(action)
        if args.execute:
            parts = action.split() # DELETE_RUN_SNAPSHOT runs/<id>/
            path_str = parts[1]
            try:
                # Extract run_id for tracking failure
                # runs/<id>/ -> id
                r_id = path_str.split('/')[1]
                target_path = PROJECT_ROOT / path_str
                
                if target_path.exists():
                    shutil.rmtree(target_path)
            except Exception as e:
                print(f"[ERROR] Failed to delete {path_str}: {e}")
                failed_run_ids.add(r_id)

    # DELETE_BACKTEST_STATE
    for action in sorted(list(set(all_backtests))):
        print(action)
        if args.execute:
            parts = action.split() # DELETE_BACKTEST_STATE backtests/<name>/
            path_str = parts[1]
            try:
                # Extract strategy name
                # backtests/<name>/ -> name
                s_name = path_str.split('/')[1]
                target_path = PROJECT_ROOT / path_str
                
                if target_path.exists():
                    shutil.rmtree(target_path)
            except Exception as e:
                print(f"[ERROR] Failed to delete {path_str}: {e}")
                # Mark associated run_id as failed
                if s_name in strategy_to_run_id:
                    failed_run_ids.add(strategy_to_run_id[s_name])

    # REMOVE_MASTER_ROW
    # Condition: Only remove if run_id NOT in failed_run_ids
    valid_row_actions = []
    for action in all_rows:
        parts = action.split()
        r_id = parts[1].split('=')[1]
        
        if r_id in failed_run_ids:
            print(f"[SKIP] Keeping row for failed cleanup: {r_id}")
            continue
            
        print(action)
        valid_row_actions.append(action)
        
    if args.execute and valid_row_actions:
        execute_removals(valid_row_actions)

    sys.exit(0)

if __name__ == "__main__":
    main()
