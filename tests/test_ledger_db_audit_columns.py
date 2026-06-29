"""Audit-column preservation across export_mps regeneration.

Operator-added columns on the MPS Baskets sheet (e.g. quarantine_status
from the 2026-05-24 leg_direction_flip_bug cleanup) used to be wiped on
the next basket pipeline run: export_mps rewrites the workbook from DB
columns only, and the DB doesn't know these annotations.

The fix in `tools/ledger_db.py::_merge_audit_columns` reads the existing
xlsx, joins whitelisted audit columns back onto the DB-sourced rows via
the sheet's natural key, then writes. These tests pin that behaviour.
"""
from __future__ import annotations

import sys
from pathlib import Path

import pandas as pd
import pytest


REPO_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO_ROOT))


@pytest.fixture
def fake_state(tmp_path, monkeypatch):
    """Redirect TRADE_SCAN_STATE to a tmp dir; create the strategies/ subdir
    where Master_Portfolio_Sheet.xlsx will land. ledger_db's reader guards
    resolve the DB path dynamically, so a single monkeypatch on
    config.path_authority.TRADE_SCAN_STATE is sufficient — no need to
    chase the (legacy, still-exposed) module-level LEDGER_DB_PATH alias.
    """
    import config.path_authority as pa
    monkeypatch.setattr(pa, "TRADE_SCAN_STATE", tmp_path)
    (tmp_path / "strategies").mkdir()
    return tmp_path


def _seed_basket_row(run_id: str, directive_id: str = "dir_test",
                     basket_id: str = "H2") -> None:
    """Insert a minimal basket_sheet row via the DB API. The writer's full
    contract is not exercised here — we just need a row with a stable
    run_id so export_mps has something to emit."""
    from tools.ledger_db import _connect, create_tables, upsert_basket_row
    conn = _connect()
    try:
        create_tables(conn)
        upsert_basket_row(conn, {
            "run_id": run_id,
            "directive_id": directive_id,
            "basket_id": basket_id,
            "execution_mode": "basket",
            "rule_name": "H2_recycle",
            "rule_version": 1,
            "leg_count": 2,
            "leg_specs": "EURUSD:0.02:long;USDJPY:0.01:short",
            "trades_total": 0,
            "recycle_event_count": 0,
            "harvested_total_usd": 0.0,
            "final_realized_usd": 0.0,
            "schema_version": "1.3.0-basket",
            "is_current": 1,
        })
    finally:
        conn.close()


def _seed_mps_xlsx_with_audit_cols(mps_path: Path, rows: list[dict]) -> None:
    """Write an MPS xlsx with a Baskets sheet whose rows carry the
    operator audit columns. Used to simulate the state after a human
    has annotated the file but before the next pipeline run."""
    df = pd.DataFrame(rows)
    with pd.ExcelWriter(str(mps_path), engine="openpyxl") as w:
        # Two empty data sheets — export_mps drops empty Portfolios/
        # Single-Asset rows, but the file must look like a valid MPS.
        pd.DataFrame(columns=["portfolio_id"]).to_excel(
            w, sheet_name="Portfolios", index=False)
        pd.DataFrame(columns=["portfolio_id"]).to_excel(
            w, sheet_name="Single-Asset Composites", index=False)
        df.to_excel(w, sheet_name="Baskets", index=False)


# ---- _merge_audit_columns unit tests ----------------------------------------


def test_merge_audit_columns_carries_over_whitelisted_columns(fake_state):
    """Audit columns present in the xlsx land on the matching df row."""
    from tools.ledger_db import _merge_audit_columns
    mps = fake_state / "strategies" / "Master_Portfolio_Sheet.xlsx"
    _seed_mps_xlsx_with_audit_cols(mps, [
        {"run_id": "RID01",
         "quarantine_status": "SUPERSEDED",
         "superseded_by_run_id": "RID02",
         "quarantine_reason": "leg_direction_flip_bug"},
    ])

    df = pd.DataFrame({"run_id": ["RID01"], "trades_total": [42]})
    merged = _merge_audit_columns(df, "Baskets", mps)

    assert "quarantine_status" in merged.columns
    assert "superseded_by_run_id" in merged.columns
    assert "quarantine_reason" in merged.columns
    row = merged[merged["run_id"] == "RID01"].iloc[0]
    assert row["quarantine_status"] == "SUPERSEDED"
    assert row["superseded_by_run_id"] == "RID02"
    assert row["quarantine_reason"] == "leg_direction_flip_bug"
    assert int(row["trades_total"]) == 42


def test_merge_audit_columns_missing_xlsx_is_noop(fake_state):
    """First run, no xlsx yet — helper returns df unchanged."""
    from tools.ledger_db import _merge_audit_columns
    mps = fake_state / "strategies" / "Master_Portfolio_Sheet.xlsx"
    assert not mps.exists()
    df = pd.DataFrame({"run_id": ["RID01"], "trades_total": [42]})
    merged = _merge_audit_columns(df, "Baskets", mps)
    pd.testing.assert_frame_equal(merged, df)


def test_merge_audit_columns_no_audit_cols_in_xlsx_is_noop(fake_state):
    """Existing xlsx without audit columns — helper returns df unchanged."""
    from tools.ledger_db import _merge_audit_columns
    mps = fake_state / "strategies" / "Master_Portfolio_Sheet.xlsx"
    _seed_mps_xlsx_with_audit_cols(mps, [
        {"run_id": "RID01", "trades_total": 7},
    ])
    df = pd.DataFrame({"run_id": ["RID01"], "trades_total": [42]})
    merged = _merge_audit_columns(df, "Baskets", mps)
    assert "quarantine_status" not in merged.columns
    assert list(merged.columns) == ["run_id", "trades_total"]


def test_merge_audit_columns_skips_already_in_df(fake_state):
    """If an audit column ever migrates into the DB schema, the DB value
    wins and the merge must not create _x/_y suffixed duplicates."""
    from tools.ledger_db import _merge_audit_columns
    mps = fake_state / "strategies" / "Master_Portfolio_Sheet.xlsx"
    _seed_mps_xlsx_with_audit_cols(mps, [
        {"run_id": "RID01", "quarantine_status": "OLD_VALUE"},
    ])
    df = pd.DataFrame({"run_id": ["RID01"], "quarantine_status": ["DB_VALUE"]})
    merged = _merge_audit_columns(df, "Baskets", mps)
    assert "quarantine_status_x" not in merged.columns
    assert "quarantine_status_y" not in merged.columns
    assert merged["quarantine_status"].iloc[0] == "DB_VALUE"


def test_merge_audit_columns_unmatched_xlsx_rows_dropped(fake_state):
    """xlsx may carry audit cols for run_ids the DB doesn't know about
    (e.g. archived rows); left-join must not resurrect them."""
    from tools.ledger_db import _merge_audit_columns
    mps = fake_state / "strategies" / "Master_Portfolio_Sheet.xlsx"
    _seed_mps_xlsx_with_audit_cols(mps, [
        {"run_id": "RID01", "quarantine_status": "SUPERSEDED"},
        {"run_id": "RID_GHOST", "quarantine_status": "SUPERSEDED"},
    ])
    df = pd.DataFrame({"run_id": ["RID01"]})
    merged = _merge_audit_columns(df, "Baskets", mps)
    assert set(merged["run_id"]) == {"RID01"}
    assert merged["quarantine_status"].iloc[0] == "SUPERSEDED"


# ---- export_mps end-to-end --------------------------------------------------


def test_export_mps_preserves_audit_columns_round_trip(fake_state):
    """The full Baskets bug repro: pre-populate xlsx with audit columns,
    seed the DB with a row sharing that run_id, call export_mps, and
    verify the audit columns survived the rewrite."""
    from tools.ledger_db import export_mps

    mps = fake_state / "strategies" / "Master_Portfolio_Sheet.xlsx"

    # 1) Seed DB with one basket row.
    _seed_basket_row(run_id="RID_AUDIT")

    # 2) Pre-emit MPS xlsx so it exists; export_mps will overwrite it.
    export_mps()
    assert mps.is_file()

    # 3) Operator adds audit columns to the Baskets sheet.
    with pd.ExcelFile(mps) as xls:
        df_b = pd.read_excel(xls, sheet_name="Baskets")
        df_p = pd.read_excel(xls, sheet_name="Portfolios")
        df_s = pd.read_excel(xls, sheet_name="Single-Asset Composites")
    df_b["quarantine_status"] = "SUPERSEDED"
    df_b["superseded_by_run_id"] = "RID_NEW"
    df_b["quarantine_reason"] = "leg_direction_flip_bug"
    with pd.ExcelWriter(str(mps), engine="openpyxl") as w:
        df_p.to_excel(w, sheet_name="Portfolios", index=False)
        df_s.to_excel(w, sheet_name="Single-Asset Composites", index=False)
        df_b.to_excel(w, sheet_name="Baskets", index=False)

    # 4) Simulate the next basket pipeline run — re-export from DB.
    export_mps()

    # 5) Audit columns survived, with correct values joined by run_id.
    with pd.ExcelFile(mps) as xls:
        out = pd.read_excel(xls, sheet_name="Baskets")
    assert "quarantine_status" in out.columns, (
        "quarantine_status was stripped by export_mps — preservation regressed")
    assert "superseded_by_run_id" in out.columns
    assert "quarantine_reason" in out.columns
    row = out[out["run_id"] == "RID_AUDIT"].iloc[0]
    assert row["quarantine_status"] == "SUPERSEDED"
    assert row["superseded_by_run_id"] == "RID_NEW"
    assert row["quarantine_reason"] == "leg_direction_flip_bug"


def test_export_mps_no_existing_xlsx_writes_clean(fake_state):
    """Fresh install: no MPS xlsx, no audit columns to merge.
    export_mps must not raise and must emit the DB schema cleanly."""
    from tools.ledger_db import export_mps
    mps = fake_state / "strategies" / "Master_Portfolio_Sheet.xlsx"
    assert not mps.exists()
    _seed_basket_row(run_id="RID_FRESH")
    export_mps()
    assert mps.is_file()
    with pd.ExcelFile(mps) as xls:
        out = pd.read_excel(xls, sheet_name="Baskets")
    assert "quarantine_status" not in out.columns
    assert out["run_id"].iloc[0] == "RID_FRESH"


def test_export_mps_atomic_no_leftover_temp(fake_state):
    """Atomic write: export renders to a per-process temp then os.replace()s it
    into place, so concurrent exports can't leave a half-written workbook. After
    a normal export there must be NO .tmp.* artifact and the file must be valid."""
    from openpyxl import load_workbook
    from tools.ledger_db import export_mps
    strat = fake_state / "strategies"
    _seed_basket_row(run_id="RID_ATOMIC")
    out = export_mps()
    leftovers = list(strat.glob("Master_Portfolio_Sheet.tmp.*"))
    assert not leftovers, f"atomic export left temp artifacts: {leftovers}"
    assert out.is_file()
    wb = load_workbook(out)
    try:
        assert "Baskets" in wb.sheetnames
    finally:
        wb.close()


def test_export_mps_audit_col_only_on_some_rows(fake_state):
    """Two DB rows; xlsx carries audit metadata for only one of them.
    Annotated row keeps its values; unannotated row gets NaN, not dropped."""
    from tools.ledger_db import export_mps
    mps = fake_state / "strategies" / "Master_Portfolio_Sheet.xlsx"

    _seed_basket_row(run_id="RID_A")
    _seed_basket_row(run_id="RID_B")
    export_mps()

    # Annotate RID_A only.
    with pd.ExcelFile(mps) as xls:
        df_b = pd.read_excel(xls, sheet_name="Baskets")
        df_p = pd.read_excel(xls, sheet_name="Portfolios")
        df_s = pd.read_excel(xls, sheet_name="Single-Asset Composites")
    df_b["quarantine_status"] = [
        "SUPERSEDED" if rid == "RID_A" else None for rid in df_b["run_id"]
    ]
    with pd.ExcelWriter(str(mps), engine="openpyxl") as w:
        df_p.to_excel(w, sheet_name="Portfolios", index=False)
        df_s.to_excel(w, sheet_name="Single-Asset Composites", index=False)
        df_b.to_excel(w, sheet_name="Baskets", index=False)

    export_mps()
    with pd.ExcelFile(mps) as xls:
        out = pd.read_excel(xls, sheet_name="Baskets")
    a = out[out["run_id"] == "RID_A"].iloc[0]
    b = out[out["run_id"] == "RID_B"].iloc[0]
    assert a["quarantine_status"] == "SUPERSEDED"
    assert pd.isna(b["quarantine_status"])


def test_export_mps_regenerates_notes_sheet(fake_state):
    """export_mps regenerates the portfolio Notes glossary intrinsically.

    A bare export (no separate format_excel_artifact.py --profile portfolio
    pass) must never leave the operator MPS without its Notes sheet. Regression
    guard for the recurring sheet-coverage break (RESOLVED 2026-06-25, resurfaced
    2026-06-29) where a pipeline MPS export stripped Notes until the next manual
    format step. Pins the fix in ledger_db.export_mps that calls
    add_notes_sheet_to_ledger after the data-sheet render."""
    from openpyxl import load_workbook
    from tools.ledger_db import export_mps

    _seed_basket_row(run_id="RID_NOTES")
    out = export_mps()

    wb = load_workbook(out)
    try:
        assert "Notes" in wb.sheetnames, (
            "export_mps did not regenerate the Notes sheet — the operator MPS "
            "would be missing its glossary until the next manual format step")
    finally:
        wb.close()


# ---- format_excel_artifact passes audit columns through ---------------------


def test_format_excel_preserves_audit_columns_in_baskets_sheet(fake_state):
    """The formatter's column-reorder logic must keep unknown columns
    (audit columns are not in BASKETS_COLUMN_ORDER). They land in the
    `remaining` tail at the right edge and survive formatting."""
    from tools.excel_format.styling import apply_formatting

    mps = fake_state / "strategies" / "Master_Portfolio_Sheet.xlsx"
    _seed_mps_xlsx_with_audit_cols(mps, [
        {"run_id": "RID01",
         "directive_id": "dir_test",
         "basket_id": "H2",
         "canonical_ret_dd": 2.5,  # needed for sort/rank by formatter
         "quarantine_status": "SUPERSEDED",
         "superseded_by_run_id": "RID02",
         "quarantine_reason": "leg_direction_flip_bug"},
    ])

    apply_formatting(str(mps), "portfolio")

    with pd.ExcelFile(mps) as xls:
        out = pd.read_excel(xls, sheet_name="Baskets")
    assert "quarantine_status" in out.columns
    assert "superseded_by_run_id" in out.columns
    assert "quarantine_reason" in out.columns
    row = out[out["run_id"] == "RID01"].iloc[0]
    assert row["quarantine_status"] == "SUPERSEDED"
