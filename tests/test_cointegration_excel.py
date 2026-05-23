"""test_cointegration_excel.py — Phase 3 unit tests.

Validates the Excel renderer:
  * pivot logic (windows-as-columns with agreement flag)
  * half-life quality scoring formula (spec §8)
  * composite score is a float in [0, 1]
  * export_excel produces a 4-sheet valid .xlsx

No MASTER_DATA dependency.
"""
from __future__ import annotations

import math
import sys
from datetime import datetime, timezone
from pathlib import Path

import openpyxl
import pandas as pd
import pytest

PROJECT_ROOT = Path(__file__).resolve().parent.parent
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from tools.cointegration_db import (
    DB_COLUMNS,
    TABLE_NAME,
    connect,
    create_tables,
    upsert_from_parquet,
)
from tools.cointegration_excel import (
    _half_life_quality,
    _pivot_today,
    composite_score,
    export_excel,
)
from tools.cointegration_screen import PARQUET_COLUMNS


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------


def _make_row(pair_a, pair_b, lookback, pvalue, regime,
              half_life=10.0, zscore=0.5, history_depth=0):
    return {
        "as_of": "2026-05-15",
        "pair_a": pair_a, "pair_b": pair_b,
        "tf": "1d", "lookback_days": lookback,
        "window_start": "2025-08-25", "window_end": "2026-05-15",
        "sample_size": lookback,
        "adf_pvalue": pvalue,
        "pvalue_rolling_median_5d": None,
        "history_depth": history_depth,
        "adf_statistic": -3.0,
        "half_life_days": half_life,
        "hedge_ratio": 1.5,
        "beta_method": "ols_static", "test_method": "adf",
        "current_zscore": zscore,
        "regime": regime,
        "data_version": "v0",
        "inserted_at": datetime.now(timezone.utc).isoformat(),
    }


@pytest.fixture
def conn_with_4_pairs(tmp_path):
    """In-DB seeded with 4 (pair_a, pair_b) examples covering all 4 agreement cases."""
    db = tmp_path / "test_excel.db"
    c = connect(db)
    create_tables(c)
    # GBPUSD/USDCHF: BOTH cointegrated
    # EURAUD/USDCHF: 252-only
    # CADJPY/GBPJPY: 504-only
    # EURUSD/USDJPY: NEITHER
    rows = [
        _make_row("GBPUSD", "USDCHF", 252, 0.001, "cointegrated"),
        _make_row("GBPUSD", "USDCHF", 504, 0.002, "cointegrated"),
        _make_row("EURAUD", "USDCHF", 252, 0.01,  "cointegrated"),
        _make_row("EURAUD", "USDCHF", 504, 0.60,  "broken"),
        _make_row("CADJPY", "GBPJPY", 252, 0.15,  "broken"),
        _make_row("CADJPY", "GBPJPY", 504, 0.001, "cointegrated"),
        _make_row("EURUSD", "USDJPY", 252, 0.50,  "broken"),
        _make_row("EURUSD", "USDJPY", 504, 0.50,  "broken"),
    ]
    for r in rows:
        c.execute(
            f"INSERT INTO {TABLE_NAME} ({', '.join(DB_COLUMNS)}) "
            f"VALUES ({', '.join(['?']*len(DB_COLUMNS))})",
            tuple(r[col] for col in DB_COLUMNS),
        )
    c.commit()
    yield c
    c.close()


# ---------------------------------------------------------------------------
# Pivot
# ---------------------------------------------------------------------------


class TestPivot:

    def test_all_four_agreement_cases(self, conn_with_4_pairs):
        df = pd.read_sql_query(
            f"SELECT * FROM {TABLE_NAME}", conn_with_4_pairs)
        pivoted = _pivot_today(df)
        # 4 unique (pair_a, pair_b)
        assert len(pivoted) == 4
        agreements = dict(zip(
            pivoted["pair_a"] + "/" + pivoted["pair_b"],
            pivoted["agreement"],
        ))
        assert agreements["GBPUSD/USDCHF"] == "BOTH"
        assert agreements["EURAUD/USDCHF"] == "252-only"
        assert agreements["CADJPY/GBPJPY"] == "504-only"
        assert agreements["EURUSD/USDJPY"] == "NEITHER"

    def test_window_columns_adjacent(self, conn_with_4_pairs):
        df = pd.read_sql_query(
            f"SELECT * FROM {TABLE_NAME}", conn_with_4_pairs)
        pivoted = _pivot_today(df)
        # Every pivoted column either has _252 / _504 suffix or is a key/agreement column
        for col in pivoted.columns:
            assert (col.endswith("_252") or col.endswith("_504")
                    or col in ("pair_a", "pair_b", "pair_class", "agreement",
                                "tradability", "history_depth", "corr_504d"))


# ---------------------------------------------------------------------------
# Score components (spec §8)
# ---------------------------------------------------------------------------


class TestHalfLifeQuality:

    def test_peaks_at_15_days(self):
        # Maximum should be at hl=15 (the spec's target).
        peak = _half_life_quality(15.0)
        for hl in (3, 5, 10, 20, 30, 60, 100):
            assert _half_life_quality(hl) <= peak

    def test_symmetric_in_log_space(self):
        # exp(-|log(hl/15)|) — log-symmetric around 15.
        # So hl=7.5 (half of 15) and hl=30 (double of 15) should be roughly equal.
        assert abs(_half_life_quality(7.5) - _half_life_quality(30.0)) < 1e-9

    def test_nan_returns_zero(self):
        assert _half_life_quality(None) == 0.0
        assert _half_life_quality(float("nan")) == 0.0
        assert _half_life_quality(0) == 0.0
        assert _half_life_quality(-5) == 0.0


class TestCompositeScore:

    def test_returns_float_in_unit_interval(self, conn_with_4_pairs):
        row = {"pair_a": "GBPUSD", "pair_b": "USDCHF", "lookback_days": 252,
               "half_life_days": 10.0}
        score = composite_score(conn_with_4_pairs, row)
        assert isinstance(score, float)
        assert 0.0 <= score <= 1.0

    def test_score_nonzero_when_cointegrated(self, conn_with_4_pairs):
        # GBPUSD/USDCHF is cointegrated + has half-life => score > 0
        row = {"pair_a": "GBPUSD", "pair_b": "USDCHF", "lookback_days": 252,
               "half_life_days": 10.0}
        assert composite_score(conn_with_4_pairs, row) > 0


# ---------------------------------------------------------------------------
# End-to-end Excel output
# ---------------------------------------------------------------------------


class TestExportExcel:

    def test_produces_expected_sheet_layout(self, conn_with_4_pairs, tmp_path):
        """2026-05-21: layout extended from 4 tabs to 8 (Summary, asset-
        class tabs, diagnostics, History, Notes). Singles tab may be
        empty if the test fixture's DB has no singles_daily table."""
        out = tmp_path / "out.xlsx"
        export_excel(db_path=conn_with_4_pairs.execute("PRAGMA database_list").fetchone()[2],
                     output_path=out)
        assert out.exists()
        wb = openpyxl.load_workbook(out, read_only=True)
        try:
            names = wb.sheetnames
            # Summary first, Notes last, History before Notes.
            assert names[0] == "Summary"
            assert names[-1] == "Notes"
            assert "History" in names
            # Diagnostic tab renamed to "All Pairs (Diagnostic)".
            assert "All Pairs (Diagnostic)" in names
            # Asset-class tabs present (from candidates yaml).
            assert "Forex (incl. Metals)" in names
            assert "Crypto" in names
            assert "Indices & Stocks" in names
            # All Pairs has the existing 4-pair test fixture data
            ws = wb["All Pairs (Diagnostic)"]
            assert ws.max_row >= 5
            # Summary still has multiple sections
            ws = wb["Summary"]
            assert ws.max_row >= 25
        finally:
            wb.close()
