"""test_quarantine_integrity.py — CI guard for quarantine governance.

Two checks:

  Check A (rule-side): every rule in governance/recycle_rules/registry.yaml
    flagged with deprecated=true (or future quarantine_required=true) must
    have ZERO untagged rows in MPS Baskets. Catches retirements that skip
    MPS row tagging.

  Check B (consumer-side): each known MPS Baskets default-projection
    consumer must produce a quarantine-free output by default. Catches
    regressions where a consumer accidentally drops its filter — even
    when the rows themselves are tagged correctly.

If MPS doesn't exist (e.g., CI ephemeral env without the state repo),
both checks skip rather than fail. Real-data checks need real data.
"""
from __future__ import annotations

import sys
from pathlib import Path

import pandas as pd
import pytest
import yaml

_REPO_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(_REPO_ROOT))

from config.path_authority import TRADE_SCAN_STATE  # noqa: E402

MPS_PATH = TRADE_SCAN_STATE / "strategies" / "Master_Portfolio_Sheet.xlsx"
REGISTRY_PATH = _REPO_ROOT / "governance" / "recycle_rules" / "registry.yaml"


def _untagged_mask(df: pd.DataFrame) -> pd.Series:
    """True where quarantine_status is null/empty."""
    if "quarantine_status" not in df.columns:
        return pd.Series([True] * len(df), index=df.index)
    qs = df["quarantine_status"]
    return qs.isna() | (qs.astype(str).str.strip() == "")


@pytest.fixture(scope="module")
def baskets_df() -> pd.DataFrame:
    if not MPS_PATH.is_file():
        pytest.skip(f"MPS not found at {MPS_PATH}; quarantine integrity check requires state repo")
    return pd.read_excel(MPS_PATH, sheet_name="Baskets")


@pytest.fixture(scope="module")
def deprecated_rules() -> list[dict]:
    """Returns the registry entries flagged deprecated=true (or quarantine_required=true)."""
    if not REGISTRY_PATH.is_file():
        pytest.skip(f"Registry not found at {REGISTRY_PATH}")
    with open(REGISTRY_PATH, "r", encoding="utf-8") as f:
        data = yaml.safe_load(f)
    deprecated = [
        r for r in data.get("rules", [])
        if r.get("deprecated", False) or r.get("quarantine_required", False)
    ]
    return deprecated


# ════════════════════════════════════════════════════════════════════
# CHECK A — Rule-side integrity
# ════════════════════════════════════════════════════════════════════

def test_deprecated_rules_have_all_mps_rows_tagged(baskets_df, deprecated_rules):
    """Every rule flagged deprecated=true in the registry must have all its
    MPS Baskets rows tagged with quarantine_status. Plus a hardcoded list
    of retired rule_name strings whose code has been deleted (so they
    won't appear in the registry but still have historical MPS rows)."""

    # Hardcoded historical retirements (rule code deleted, not in registry):
    HISTORICAL_RETIREMENTS = {
        # 2026-05-21: COINTREV v1 equal-lot conflation retirement (commit 605317c).
        # Rule file (cointegration_meanrev_v1.py) deleted; directives quarantined.
        # MPS rule_name stays as the historical string.
        "COINTREV_meanrev",
    }

    retired_names: set[str] = set(HISTORICAL_RETIREMENTS)
    for r in deprecated_rules:
        retired_names.add(r["name"])

    failures: list[str] = []
    for name in sorted(retired_names):
        sub = baskets_df[baskets_df["rule_name"] == name]
        if sub.empty:
            # Rule never produced any MPS rows — fine
            continue
        untagged = sub[_untagged_mask(sub)]
        if not untagged.empty:
            samples = untagged["directive_id"].head(3).tolist()
            failures.append(
                f"  rule_name={name!r}: {len(untagged)}/{len(sub)} rows missing quarantine_status. "
                f"Examples: {samples}"
            )

    assert not failures, (
        "Retired rules have untagged MPS Baskets rows:\n"
        + "\n".join(failures)
        + "\n\nFix: tag the rows with quarantine_status, superseded_by_run_id (if applicable), "
        + "and quarantine_reason. See tmp/tag_retired_cointrev_v1.py for the pattern."
    )


# ════════════════════════════════════════════════════════════════════
# CHECK B — Consumer-side integrity (default projections)
# ════════════════════════════════════════════════════════════════════

def test_aggregator_default_excludes_quarantined(baskets_df):
    """cointrev_v1_2_aggregator._filter_cointrev_real must drop quarantined
    rows by default. Regression guard for the patch landed 2026-05-25."""
    from tools.cointrev_v1_2_aggregator import _filter_cointrev_real

    out = _filter_cointrev_real(baskets_df, lookback=252)  # default include_superseded=False
    leaked = out[~_untagged_mask(out)]
    assert leaked.empty, (
        f"cointrev_v1_2_aggregator leaked {len(leaked)} quarantined row(s) in default output:\n"
        f"{leaked[['directive_id', 'quarantine_status']].to_string(index=False)}"
    )


def test_h2_parity_default_excludes_quarantined(baskets_df):
    """h2_parity_run.aggregate_baskets_sheet must drop quarantined rows by
    default. Regression guard for the patch landed 2026-05-25."""
    from tools.h2_parity_run import aggregate_baskets_sheet
    if (baskets_df["basket_id"] == "H2").sum() == 0:
        pytest.skip("No H2 basket_id rows in MPS to exercise the path")

    out = aggregate_baskets_sheet()  # default include_quarantined=False
    leaked = out[~_untagged_mask(out)]
    assert leaked.empty, (
        f"h2_parity_run leaked {len(leaked)} quarantined row(s) in default output:\n"
        f"{leaked[['directive_id', 'quarantine_status']].to_string(index=False)}"
    )


def test_formatter_baskets_default_filter_logic(tmp_path):
    """Black-box: create a tiny xlsx with Baskets sheet containing a mix of
    visible / quarantined / FAIL rows, run the format_excel_artifact.py CLI
    end-to-end, then re-open the saved file and assert quarantined rows have
    row_dimensions.hidden=True. Catches regressions in the entire format
    pipeline, not just one helper."""
    import openpyxl
    import subprocess

    fixture = tmp_path / "Master_Portfolio_Sheet.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Baskets"
    # portfolio_id is needed at column 1 for the portfolio-profile path
    headers = ["portfolio_id", "directive_id", "verdict_status", "quarantine_status",
               "canonical_ret_dd"]
    ws.append(headers)
    rows = [
        ("p_alive_1",       "alive_1",       "CORE",  "",           1.5),
        ("p_quarantined_1", "quarantined_1", "CORE",  "SUPERSEDED", 2.0),
        ("p_quarantined_2", "quarantined_2", "WATCH", "RETIRED",    1.0),
        ("p_fail_1",        "fail_1",        "FAIL",  "",           0.5),
        ("p_alive_2",       "alive_2",       "WATCH", "",           1.2),
    ]
    for r in rows:
        ws.append(r)
    wb.save(fixture)

    cli = _REPO_ROOT / "tools" / "format_excel_artifact.py"
    result = subprocess.run(
        [sys.executable, str(cli), "--file", str(fixture), "--profile", "portfolio"],
        capture_output=True, text=True, cwd=str(_REPO_ROOT),
    )
    assert result.returncode == 0, (
        f"Formatter exited {result.returncode}\nSTDOUT:\n{result.stdout}\nSTDERR:\n{result.stderr}"
    )

    # Re-open the saved file and locate the Baskets sheet
    wb2 = openpyxl.load_workbook(fixture)
    assert "Baskets" in wb2.sheetnames
    ws2 = wb2["Baskets"]

    # Build map from directive_id (which we can find by header lookup) to row index
    hdr_row = [c.value for c in ws2[1]]
    did_col = hdr_row.index("directive_id") + 1
    did_to_row = {}
    for r in range(2, ws2.max_row + 1):
        did = ws2.cell(row=r, column=did_col).value
        did_to_row[did] = r

    def is_hidden(did: str) -> bool:
        return bool(ws2.row_dimensions[did_to_row[did]].hidden)

    # Quarantined rows MUST be hidden regardless of verdict
    assert is_hidden("quarantined_1"), "SUPERSEDED row was NOT hidden"
    assert is_hidden("quarantined_2"), "RETIRED row was NOT hidden"
    # FAIL row hidden via verdict (existing CORE/WATCH filter)
    assert is_hidden("fail_1"), "FAIL row was NOT hidden by verdict filter"
    # Clean CORE/WATCH rows must remain visible
    assert not is_hidden("alive_1"), "CORE non-quarantined row was hidden — over-broad filter"
    assert not is_hidden("alive_2"), "WATCH non-quarantined row was hidden — over-broad filter"
