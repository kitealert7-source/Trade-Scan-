"""Regression — Stage-3 aggregation idempotency must key on run_id, not strategy prefix.

Background
----------
`run_stage3_aggregation` had two checks that compared Master Filter rows by
`strategy.startswith(clean_id)`:

  1. **Idempotency pre-check** (decides whether to invoke stage3_compiler).
  2. **Cardinality check** (verifies the expected row count after write).

A re-run of any directive produces a fresh `run_id` but the same `clean_id`.
An older row from a prior run trivially matched the prefix check, so the
idempotency pre-check declared "already present", skipped the write, and
Master Filter stayed pinned to the first run's metrics. All 6 PSBRK
finalists re-run on 2026-05-11 hit this; their new run_ids never landed in
MF until the fix.

This regression pins the run_id-keyed semantics so the bug cannot return.
"""

from __future__ import annotations

import sys
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import patch

import openpyxl
import pytest


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

DIRECTIVE_ID = "65_BRK_XAUUSD_5M_PSBRK_S01_V4_TEST"
OLD_RUN_ID = "aaaaaaaaaaaaaaaaaaaaaaaa"
NEW_RUN_ID = "bbbbbbbbbbbbbbbbbbbbbbbb"
SYMBOL = "XAUUSD"


def _seed_master_filter(path: Path, rows: list[tuple[str, str, str]]) -> None:
    """Write a tiny xlsx with columns (run_id, strategy, symbol)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["run_id", "strategy", "symbol"])
    for rid, strat, sym in rows:
        ws.append([rid, strat, sym])
    wb.save(path)


@pytest.fixture()
def isolated_pipeline(tmp_path, monkeypatch):
    """Redirect MASTER_FILTER_PATH and RUNS_DIR to tmp tree; stub state checks."""
    mf_path = tmp_path / "Strategy_Master_Filter.xlsx"
    runs_dir = tmp_path / "runs"
    runs_dir.mkdir()
    (runs_dir / NEW_RUN_ID).mkdir()
    (runs_dir / OLD_RUN_ID).mkdir()

    import tools.orchestration.stage_symbol_execution as sse
    monkeypatch.setattr(sse, "MASTER_FILTER_PATH", mf_path, raising=True)
    monkeypatch.setattr(sse, "RUNS_DIR", runs_dir, raising=True)

    # Stub PipelineStateManager so the FAILED-run guard at the top of
    # run_stage3_aggregation does not block our synthetic state.
    class _StubStateMgr:
        def __init__(self, rid):
            self.rid = rid
        def get_state_data(self):
            return {"current_state": "STAGE_2_COMPLETE"}
    monkeypatch.setattr(sse, "PipelineStateManager", _StubStateMgr, raising=True)

    # Build a minimal PipelineContext with just the fields the function reads
    ctx = SimpleNamespace(
        directive_id=DIRECTIVE_ID,
        run_ids=[NEW_RUN_ID],
        symbols=[SYMBOL],
        python_exe="python",
    )
    return {"ctx": ctx, "mf_path": mf_path, "runs_dir": runs_dir}


# ---------------------------------------------------------------------------
# Regression tests
# ---------------------------------------------------------------------------

def test_re_run_with_fresh_run_id_invokes_stage3_compiler(isolated_pipeline):
    """A re-run of a directive that already has a stale MF row must still
    invoke stage3_compiler. This is the bug the fix addresses.
    """
    mf_path = isolated_pipeline["mf_path"]
    # Seed MF with ONLY the OLD run's row — directive matches, run_id does not.
    _seed_master_filter(mf_path, [
        (OLD_RUN_ID, f"{DIRECTIVE_ID}_{SYMBOL}", SYMBOL),
    ])

    invocations: list[list[str]] = []

    def _fake_run_command(argv, *_a, **_kw):
        """When stage3_compiler is invoked, simulate it by appending the new
        run_id row to MF so the subsequent cardinality check passes."""
        invocations.append(argv)
        if argv[1] == "tools/stage3_compiler.py":
            wb = openpyxl.load_workbook(mf_path)
            ws = wb.active
            ws.append([NEW_RUN_ID, f"{DIRECTIVE_ID}_{SYMBOL}", SYMBOL])
            wb.save(mf_path)

    with patch("tools.orchestration.execution_adapter.run_command",
               side_effect=_fake_run_command):
        from tools.orchestration.stage_symbol_execution import run_stage3_aggregation
        # Should not raise — cardinality should pass after our fake write
        run_stage3_aggregation(isolated_pipeline["ctx"])

    # Stage 3 compiler WAS invoked despite the stale MF row sitting there.
    invoked_stage3 = any(
        len(a) >= 2 and a[1] == "tools/stage3_compiler.py" for a in invocations
    )
    assert invoked_stage3, "stage3_compiler must run when the new run_id is missing from MF"


def test_resume_with_current_run_id_present_skips_stage3_compiler(isolated_pipeline):
    """Idempotency invariant — when MF already contains the CURRENT run_id
    (e.g. resume-after-Stage-4-failure), stage3_compiler must NOT re-run.
    """
    mf_path = isolated_pipeline["mf_path"]
    # Seed MF with both old AND the current new run_id — simulates the
    # resume-after-stage-4 scenario the idempotency check is meant to handle.
    _seed_master_filter(mf_path, [
        (OLD_RUN_ID, f"{DIRECTIVE_ID}_{SYMBOL}", SYMBOL),
        (NEW_RUN_ID, f"{DIRECTIVE_ID}_{SYMBOL}", SYMBOL),
    ])

    invocations: list[list[str]] = []

    def _fake_run_command(argv, *_a, **_kw):
        invocations.append(argv)

    with patch("tools.orchestration.execution_adapter.run_command",
               side_effect=_fake_run_command):
        from tools.orchestration.stage_symbol_execution import run_stage3_aggregation
        run_stage3_aggregation(isolated_pipeline["ctx"])

    invoked_stage3 = any(
        len(a) >= 2 and a[1] == "tools/stage3_compiler.py" for a in invocations
    )
    assert not invoked_stage3, (
        "stage3_compiler must NOT be invoked when current run_id is already in MF "
        "(idempotency invariant preserved)"
    )


def test_no_trades_run_passes_cardinality(isolated_pipeline):
    """NO_TRADES runs are excluded from expected_count, so cardinality must
    pass with zero current-run_id rows in MF. (Preserves original behavior:
    when all runs are NO_TRADES, expected_count = 0 and actual_count = 0.)
    """
    runs_dir = isolated_pipeline["runs_dir"]
    mf_path = isolated_pipeline["mf_path"]

    # Mark NEW_RUN_ID as NO_TRADES — pipeline expects no MF row for it.
    (runs_dir / NEW_RUN_ID / "status_no_trades.json").write_text("{}", encoding="utf-8")

    # MF carries only an old row; the new run is NO_TRADES so it does not
    # need a row of its own.
    _seed_master_filter(mf_path, [
        (OLD_RUN_ID, f"{DIRECTIVE_ID}_{SYMBOL}", SYMBOL),
    ])

    def _fake_run_command(argv, *_a, **_kw):
        # No-op — stage3_compiler call is harmless here; we only care that
        # the cardinality check downstream passes.
        return None

    with patch("tools.orchestration.execution_adapter.run_command",
               side_effect=_fake_run_command):
        from tools.orchestration.stage_symbol_execution import run_stage3_aggregation
        # Must not raise — expected_count=0 (NO_TRADES excluded),
        # actual_count=0 (no NEW_RUN_ID row), so cardinality passes.
        run_stage3_aggregation(isolated_pipeline["ctx"])


def test_cardinality_check_uses_run_id_not_strategy_prefix(isolated_pipeline):
    """Even after the write, the cardinality check must count rows by
    run_id membership, not by `strategy.startswith(clean_id)`. Otherwise
    older rows from prior runs of the same directive inflate the count.
    """
    mf_path = isolated_pipeline["mf_path"]
    # Pre-seed MF with 2 OLD rows for the same directive (both stale).
    # Then have stage3_compiler add the NEW row. Cardinality should count
    # only the NEW row (== expected_count=1) — not all 3 prefix-matching rows.
    _seed_master_filter(mf_path, [
        ("old_rid_001",         f"{DIRECTIVE_ID}_{SYMBOL}", SYMBOL),
        ("old_rid_002",         f"{DIRECTIVE_ID}_{SYMBOL}", SYMBOL),
    ])

    def _fake_run_command(argv, *_a, **_kw):
        if argv[1] == "tools/stage3_compiler.py":
            wb = openpyxl.load_workbook(mf_path)
            ws = wb.active
            ws.append([NEW_RUN_ID, f"{DIRECTIVE_ID}_{SYMBOL}", SYMBOL])
            wb.save(mf_path)

    with patch("tools.orchestration.execution_adapter.run_command",
               side_effect=_fake_run_command):
        from tools.orchestration.stage_symbol_execution import run_stage3_aggregation
        # Must not raise — actual_count == 1 (NEW_RUN_ID only), expected_count == 1
        run_stage3_aggregation(isolated_pipeline["ctx"])
